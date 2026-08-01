import sys
from pathlib import Path

import openpyxl


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import update_cf_statements as cf


def workbook_with_mortgage_rows():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "2025"
    sheet.cell(1, 2, "Jan")
    sheet.cell(2, 1, "Mortgage Principal Payments")
    sheet.cell(2, 2, -350)
    sheet.cell(3, 1, "Mortgage Interest Payments")
    sheet.cell(3, 2, -100)
    sheet.cell(4, 1, "Mortgage Principal Balance")
    sheet.cell(4, 2, 100000)
    return workbook


def test_90_madison_preserves_only_approved_principal_row():
    path = Path(
        "/mnt/c/Users/digit/Dropbox/Real Estate/NY/90 Madison Ave Public/"
        "07 - P&L & Owner Statements/Cash Flow Statement - 90 Madison Ave.xlsx"
    )
    workbook = workbook_with_mortgage_rows()

    result = cf.clear_no_mortgage_debt_rows(workbook, path)

    sheet = workbook["2025"]
    assert result["cleared_cell_count"] == 2
    assert sheet.cell(2, 2).value == -350
    assert sheet.cell(3, 2).value == 0
    assert sheet.cell(4, 2).value == 0


def test_other_no_dao_property_still_clears_principal_row():
    path = Path(
        "/mnt/c/Users/digit/Dropbox/Real Estate/NY/86 Madison Ave Public/"
        "07 - P&L & Owner Statements/Cash Flow Statement - 86 Madison Ave.xlsx"
    )
    workbook = workbook_with_mortgage_rows()

    cf.clear_no_mortgage_debt_rows(workbook, path)

    assert workbook["2025"].cell(2, 2).value == 0


def test_90_madison_legacy_interest_only_row_can_be_reused():
    path = Path(
        "/mnt/c/Users/digit/Dropbox/Real Estate/NY/90 Madison Ave Public/"
        "07 - P&L & Owner Statements/Cash Flow Statement - 90 Madison Ave.xlsx"
    )
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "2024"
    sheet.cell(2, 1, "Mortgage Interest-Only Payments")
    sheet.cell(2, 2, 0)

    change = cf.ensure_madison_90_principal_payments_row(
        sheet, path, dry_run=False
    )

    assert change["action"] == "rename_legacy_90_madison_principal_row"
    assert sheet.cell(2, 1).value == "Mortgage Principal Payments"
