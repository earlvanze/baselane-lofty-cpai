import csv
import copy
import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import baselane_monthly_accruals_idempotent as accruals


class BaselaneMonthlyAccrualsIdempotentTests(unittest.TestCase):
    def test_dao_llc_admin_fee_policy_keeps_eco_costs_separate(self):
        self.assertEqual(accruals.DAO_LLC_ADMIN_ANNUAL_CHARGE, accruals.Decimal("750.00"))
        self.assertEqual(accruals.DAO_LLC_ADMIN_MONTHLY_ACCRUAL * 12, accruals.Decimal("750.00"))
        self.assertEqual(accruals.DAO_LLC_ADMIN_ECO_LOFTY_PAYABLE, accruals.Decimal("200.00"))
        self.assertEqual(accruals.DAO_LLC_ADMIN_ECO_FILING_COST_REFERENCE, accruals.Decimal("125.00"))
        self.assertIn("never reduce the DAO's fee or payable", accruals.DAO_LLC_ADMIN_ECO_REVENUE_NOTE)

    def test_categoryless_booking_split_child_is_rent_but_refund_is_not(self):
        row = {
            "Merchant": "Booking.com",
            "Description": "Booking payout split child",
            "Type": "",
            "Category": "",
            "Sub-category": "",
            "Notes": "",
        }

        self.assertTrue(accruals.is_rent_revenue(row, 2478.72))
        self.assertFalse(accruals.is_rent_revenue(row, -2478.72))

    def test_aligned_import_rent_note_survives_categoryless_live_export(self):
        rent_row = {
            "Date": "July 07, 2026",
            "Property": "1456 W 85th St.",
            "Merchant": "Angela Heath",
            "Amount": "500.00",
            "Type": "",
            "Category": "",
            "Sub-category": "",
            "Notes": "Aligned clearing detail import | Rent or tenant receipt | post_yhome_transition | accounting/manual detail only",
        }
        management_row = {
            **rent_row,
            "Amount": "124.13",
            "Notes": "Aligned clearing detail import | Management fee | post_yhome_transition | accounting/manual detail only",
        }

        self.assertTrue(accruals.is_rent_revenue(rent_row, 500.0))
        self.assertFalse(accruals.is_rent_revenue(management_row, 124.13))

    def test_1456_aligned_import_rows_compute_correct_july_pm_basis(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        try:
            accruals.PM_FEE_PROPERTIES = [
                ("1456 W 85th St, Cleveland, OH 44102", 0.10, "AOPS-OHIL-ACCRUAL"),
            ]
            rows = [
                {
                    "Date": date,
                    "Property": "1456 W 85th St.",
                    "Merchant": merchant,
                    "Amount": amount,
                    "Type": "",
                    "Category": "",
                    "Sub-category": "",
                    "Notes": "Aligned clearing detail import | Rent or tenant receipt | post_yhome_transition | accounting/manual detail only",
                }
                for date, merchant, amount in (
                    ("July 03, 2026", "Ajanae N. Barrett", "299.00"),
                    ("July 04, 2026", "Ajanae N. Barrett", "289.39"),
                    ("July 07, 2026", "Angela Heath", "500.00"),
                    ("July 07, 2026", "Angela Heath", "435.18"),
                )
            ]

            fees = accruals.compute_pm_fees(rows, "2026-07")

            self.assertEqual(fees["1456 W 85th St, Cleveland, OH 44102"], 152.36)
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm

    def test_reporting_cutoff_keeps_target_month_aops_and_excludes_later_transactions(self):
        rows = [
            {"Date": "2026-07-29", "Notes": "", "id": "through-cutoff"},
            {"Date": "2026-07-30", "Notes": "", "id": "later-ordinary"},
            {
                "Date": "2026-07-31",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|85-104 Alawa Pl|2026-07|1597.86",
                "id": "july-aops",
            },
            {
                "Date": "2026-07-31",
                "Notes": "AOPS-84-TAXES-INSURANCE|84 Madison Ave|2026-07",
                "id": "july-custom-aops",
            },
            {
                "Date": "2026-08-31",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|85-104 Alawa Pl|2026-08|100.00",
                "id": "august-aops",
            },
        ]

        included, excluded = accruals.rows_through_reporting_cutoff(
            rows,
            accruals.parse_reporting_cutoff("2026-07-29"),
            ["2026-07"],
        )

        self.assertEqual(
            [row["id"] for row in included],
            ["through-cutoff", "july-aops", "july-custom-aops"],
        )
        self.assertEqual(
            [row["id"] for row in excluded],
            ["later-ordinary", "august-aops"],
        )

    def test_reporting_cutoff_rejects_invalid_date(self):
        with self.assertRaisesRegex(ValueError, "must be YYYY-MM-DD"):
            accruals.parse_reporting_cutoff("07/29/2026")

    def test_default_lofty_reserve_snapshot_uses_invoked_workspace_path(self):
        snapshot = accruals.default_lofty_reserve_snapshot()

        self.assertEqual(snapshot, Path(accruals.__file__).absolute().parents[1] / "reports" / "lofty-pm-current" / "get-manager-properties.full-response.json")
        self.assertNotIn("cyber-gateway", str(snapshot))

    def test_zero_dollar_coverage_marker_is_non_blocking(self):
        self.assertTrue(accruals.is_zero_amount_text("0.00"))
        self.assertTrue(accruals.is_zero_amount_text("-0.00"))
        self.assertFalse(accruals.is_zero_amount_text("-0.01"))

    def test_nonzero_missing_accrual_rows_require_review(self):
        self.assertTrue(accruals.has_nonzero_accrual_rows([{"Amount": "62.50"}]))
        self.assertFalse(accruals.has_nonzero_accrual_rows([{"Amount": "0.00"}]))

    def test_preflight_failure_overwrites_stale_success_report(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            report_path = Path(tmpdir) / "monthly.json"
            review_path = Path(tmpdir) / "monthly.md"
            report_path.write_text('{"status":"ok"}\n', encoding="utf-8")

            report = accruals.write_preflight_failure_report(
                report_path,
                review_path,
                target_month="2026-06",
                code="lofty_reserve_properties_missing",
                message="missing current live reserve rows",
                missing_lofty_reserves=["724 3rd Ave"],
            )

            written = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(report["status"], "blocked")
            self.assertEqual(written["status"], "blocked")
            self.assertEqual(
                written["preflight_failure"]["missing_lofty_reserves"],
                ["724 3rd Ave"],
            )
            self.assertIn("Status: `blocked`", review_path.read_text(encoding="utf-8"))

    def test_preflight_failure_review_includes_live_roster_remediation(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            review_path = Path(tmpdir) / "monthly.md"

            accruals.write_preflight_failure_report(
                None,
                review_path,
                target_month="2026-06",
                code="lofty_reserve_properties_missing",
                message="missing current live reserve rows",
                missing_lofty_reserves=["724 3rd Ave"],
                missing_lofty_reserve_diagnostics=[
                    {
                        "property": "724 3rd Ave",
                        "status": "absent_from_live_manager_roster",
                        "required_action": "Restore the property to Lofty manager access.",
                    }
                ],
            )

            review = review_path.read_text(encoding="utf-8")
            self.assertIn("Missing live reserve properties: `724 3rd Ave`", review)
            self.assertIn("absent_from_live_manager_roster", review)
            self.assertIn("Restore the property to Lofty manager access.", review)

    def test_missing_lofty_reserve_does_not_reprice_historical_retained_capital_row(self):
        rows = [
            {
                "Date": "2026-06-28",
                "Property": "724 3rd Ave",
                "Amount": "-100.00",
                "Merchant": "OR Replenishment | 724 3rd Ave | June 2026",
                "Description": "OR Replenishment | 724 3rd Ave | June 2026",
                "Type": "Transfers & Other",
                "Category": "Owner Contributions/Distributions",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|724 3rd Ave|2026-06|100.00",
            }
        ]

        self.assertEqual(accruals.find_amount_mismatches(rows, "2026-06", lofty_reserves={}), [])

    def test_read_and_write_gl_ignores_extra_unnamed_csv_cells(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            gl_path = Path(tmpdir) / "gl.csv"
            gl_path.write_text(
                "Date,Property,Amount,Notes\n"
                "\"June 28, 2026\",86 Madison Ave,-62.50,AOPS-PNL-ACCRUAL|dao|86 Madison Ave|2026-06|62.50,extra-cell\n",
                encoding="utf-8",
            )

            fieldnames, rows = accruals.read_gl(gl_path)

            self.assertEqual(fieldnames, ["Date", "Property", "Amount", "Notes"])
            self.assertEqual(len(rows), 1)
            self.assertNotIn(None, rows[0])
            self.assertEqual(rows[0]["Notes"], "AOPS-PNL-ACCRUAL|dao|86 Madison Ave|2026-06|62.50")

            accruals.write_gl(gl_path, rows, fieldnames)
            rewritten = list(csv.DictReader(gl_path.open(newline="", encoding="utf-8")))

            self.assertEqual(len(rewritten), 1)
            self.assertNotIn(None, rewritten[0])
            self.assertEqual(rewritten[0]["Notes"], "AOPS-PNL-ACCRUAL|dao|86 Madison Ave|2026-06|62.50")

    def test_90_madison_or_replenishment_uses_same_month_noi_and_middle_tier(self):
        rows = [
            {"Date": "2026-05-10", "Property": "90 Madison Ave", "Amount": "7856.27", "Category": "Revenue", "Notes": ""},
            {"Date": "2026-05-15", "Property": "90 Madison Ave", "Amount": "-6132.39", "Category": "Operating Expenses", "Notes": ""},
            {"Date": "2026-06-10", "Property": "90 Madison Ave", "Amount": "8703.82", "Category": "Revenue", "Notes": ""},
            {"Date": "2026-06-15", "Property": "90 Madison Ave", "Amount": "-7832.16", "Category": "Operating Expenses", "Notes": ""},
            {
                "Date": "2026-06-28",
                "Property": "90 Madison Ave",
                "Amount": "-850.00",
                "Category": "Operating Expenses",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|90 Madison Ave|2026-06|850.00 | stale reserve",
            },
        ]

        summary = accruals.retained_capital_amount(
            rows,
            "2026-06",
            "90 Madison Ave",
            {"90 Madison Ave": 0.0},
        )

        self.assertEqual(summary["basis_month"], "2026-06")
        self.assertEqual(summary["noi"], 871.66)
        self.assertEqual(summary["combined_reserve_basis"], 2595.54)
        self.assertEqual(summary["replenishment_rate"], 0.5)
        self.assertEqual(summary["amount"], 435.83)

    def test_90_madison_retained_capital_uses_type_when_prior_month_categories_are_detailed(self):
        rows = [
            {"Date": "May 29, 2026", "Property": "90 Madison Ave", "Amount": "7856.27", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
            {"Date": "May 28, 2026", "Property": "90 Madison Ave", "Amount": "-6132.39", "Type": "Operating Expenses", "Category": "Property Management", "Notes": ""},
            {"Date": "June 29, 2026", "Property": "90 Madison Ave", "Amount": "8703.82", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
            {"Date": "June 28, 2026", "Property": "90 Madison Ave", "Amount": "-7832.16", "Type": "Operating Expenses", "Category": "Property Management", "Notes": ""},
            {
                "Date": "June 28, 2026",
                "Property": "90 Madison Ave",
                "Amount": "-850.00",
                "Type": "Transfers & Other",
                "Category": "Owner Contributions/Distributions",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|90 Madison Ave|2026-06|850.00 | legacy retained-capital pseudo-entry",
            },
        ]

        summary = accruals.retained_capital_amount(
            rows,
            "2026-06",
            "90 Madison Ave",
            {"90 Madison Ave": 0.0},
        )

        self.assertEqual(summary["basis_month"], "2026-06")
        self.assertEqual(summary["noi"], 871.66)
        self.assertEqual(summary["amount"], 435.83)

    def test_90_madison_retained_capital_amount_update_rebuilds_note_basis(self):
        rows = [
            {"Date": "May 29, 2026", "Property": "90 Madison Ave", "Amount": "7856.27", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
            {"Date": "May 28, 2026", "Property": "90 Madison Ave", "Amount": "-6132.39", "Type": "Operating Expenses", "Category": "Property Management", "Notes": ""},
            {
                "Date": "June 28, 2026",
                "Property": "90 Madison Ave",
                "Merchant": "Retained Capital | 90 Madison Ave | June 2026",
                "Description": "Retained Capital | 90 Madison Ave | June 2026",
                "Amount": "-1000.00",
                "Type": "Transfers & Other",
                "Category": "Owner Contributions/Distributions",
                "Sub-category": "",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|90 Madison Ave|2026-06|1000.00 | stale governance reserve.",
            },
            {"Date": "June 29, 2026", "Property": "90 Madison Ave", "Amount": "8703.82", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
            {"Date": "June 28, 2026", "Property": "90 Madison Ave", "Amount": "-7832.16", "Type": "Operating Expenses", "Category": "Property Management", "Notes": ""},
        ]

        reserves = {"90 Madison Ave": 0.0}
        mismatches = accruals.find_amount_mismatches(rows, "2026-06", lofty_reserves=reserves)
        applied = accruals.apply_amount_mismatch_updates(rows, mismatches, reserves)

        self.assertEqual(applied[0]["new_amount"], 435.83)
        self.assertEqual(rows[2]["Amount"], "-435.83")
        self.assertIn("|retained_capital|90 Madison Ave|2026-06|435.83", rows[2]["Notes"])
        self.assertIn("NOI $871.66", rows[2]["Notes"])
        self.assertEqual(rows[2]["Type"], "Transfers & Other")
        self.assertEqual(rows[2]["Category"], "Owner Contributions/Distributions")
        self.assertEqual(rows[2]["Sub-category"], "")
        self.assertEqual(rows[2]["Merchant"], "OR Replenishment | 90 Madison Ave | June 2026")
        self.assertEqual(rows[2]["Description"], "OR Replenishment | 90 Madison Ave | June 2026")
        rows[2]["Description"] = ""
        rows[2]["Type"] = "Manual"
        rows[2]["Category"] = "Transfers & Other"
        rows[2]["Sub-category"] = "Owner Contributions/Distributions"
        self.assertEqual(
            accruals.find_amount_mismatches(rows, "2026-06", lofty_reserves=reserves),
            [],
        )

    def test_retained_capital_policy_snapshot_survives_explanatory_basis_drift(self):
        rows = [
            {
                "Date": "June 29, 2026",
                "Property": "90 Madison Ave",
                "Amount": "-5000.00",
                "Type": "Operating Expenses",
                "Category": "Repairs & Maintenance",
                "Notes": "",
            },
            {
                "Date": "July 29, 2026",
                "Property": "90 Madison Ave",
                "Amount": "536.28",
                "Type": "Revenue",
                "Category": "Short Term Rents",
                "Notes": "",
            },
            {
                "Date": "July 28, 2026",
                "Property": "90 Madison Ave",
                "Merchant": "OR Replenishment | 90 Madison Ave | July 2026",
                "Description": "OR Replenishment | 90 Madison Ave | July 2026",
                "Amount": "-536.28",
                "Type": "Manual",
                "Category": "Transfers & Other",
                "Sub-category": "Owner Contributions/Distributions",
                "Notes": (
                    "AOPS-PNL-ACCRUAL|retained_capital|90 Madison Ave|2026-07|536.28 | "
                    "OR Replenishment per March 2026 governance vote: 100% of 2026-07 NOI "
                    "$536.28. Reserve basis $100.00 = ECO GL net of accruals $50.00 + Lofty "
                    "OR $50.00. Outstanding cash reserve settlement requirement after $0.00 "
                    "of verified same-month cash settlement."
                ),
            },
        ]

        self.assertEqual(
            accruals.find_amount_mismatches(
                rows,
                "2026-07",
                lofty_reserves={"90 Madison Ave": 1946.77},
                property_filters=["90 Madison Ave"],
                kind_filters={"retained_capital"},
            ),
            [],
        )

    def test_90_madison_retained_capital_recomputes_after_generated_pm_fee(self):
        rows = [
            {"Date": "June 29, 2026", "Property": "90 Madison Ave", "Amount": "8703.82", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
            {"Date": "June 28, 2026", "Property": "90 Madison Ave", "Amount": "-7832.16", "Type": "Operating Expenses", "Category": "Repairs & Maintenance", "Notes": ""},
            {
                "Date": "June 28, 2026",
                "Property": "90 Madison Ave",
                "Merchant": "Retained Capital | 90 Madison Ave | June 2026",
                "Description": "Retained Capital | 90 Madison Ave | June 2026",
                "Amount": "-435.83",
                "Type": "Transfers & Other",
                "Category": "Owner Contributions/Distributions",
                "Sub-category": "",
                "Notes": "AOPS-PNL-ACCRUAL|retained_capital|90 Madison Ave|2026-06|435.83 | pre-PM reserve amount.",
            },
        ]
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]
        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            lofty_reserves={"90 Madison Ave": 0.0},
            property_filters=["90 Madison Ave"],
            kind_filters={"pm"},
        )
        rows.extend(generated)
        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            lofty_reserves={"90 Madison Ave": 0.0},
            property_filters=["90 Madison Ave"],
            kind_filters={"retained_capital"},
        )
        applied = accruals.apply_amount_mismatch_updates(rows, mismatches, {"90 Madison Ave": 0.0})
        remaining = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            lofty_reserves={"90 Madison Ave": 0.0},
            property_filters=["90 Madison Ave"],
            kind_filters={"retained_capital"},
        )

        self.assertEqual(len(generated), 2)
        self.assertEqual(applied[0]["new_amount"], 0.0)
        self.assertEqual(remaining, [])

    def test_90_madison_retained_capital_generates_utilities_tagged_idempotent_row(self):
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]
        rows = [
            {"Date": "2026-05-10", "Property": "90 Madison Ave", "Amount": "7856.27", "Category": "Revenue", "Notes": ""},
            {"Date": "2026-05-15", "Property": "90 Madison Ave", "Amount": "-6132.39", "Category": "Operating Expenses", "Notes": ""},
            {"Date": "2026-06-10", "Property": "90 Madison Ave", "Amount": "8703.82", "Category": "Revenue", "Notes": ""},
            {"Date": "2026-06-15", "Property": "90 Madison Ave", "Amount": "-7832.16", "Category": "Operating Expenses", "Notes": ""},
        ]

        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            lofty_reserves={"90 Madison Ave": 0.0},
            property_filters=["90 Madison Ave"],
            kind_filters={"retained_capital"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "-435.83")
        self.assertEqual(generated[0]["Category"], "Owner Contributions/Distributions")
        self.assertEqual(generated[0]["Sub-category"], "")
        self.assertIn("|retained_capital|90 Madison Ave|2026-06|435.83", generated[0]["Notes"])
        self.assertEqual(
            accruals.generate_missing_accruals(
                rows + generated,
                "2026-06",
                fieldnames,
                lofty_reserves={"90 Madison Ave": 0.0},
                property_filters=["90 Madison Ave"],
                kind_filters={"retained_capital"},
            ),
            [],
        )

    def test_direct_pm_marker_on_28th_counts_as_existing_accrual(self):
        rows = [
            {
                "Date": "2026-06-28",
                "Category": "Operating Expenses",
                "Sub-category": "Property Management",
                "Property": "85-104 Alawa Pl",
                "Notes": "AOPS-PM-FEE|85-104 Alawa Pl|2026-06|500.18 | monthly PM fee",
            }
        ]

        self.assertIn("85-104 Alawa Pl|pm", accruals.find_existing_accruals(rows, "2026-06"))

    def test_madison_historical_pm_rates_are_effective_dated(self):
        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave"):
            for year in (2025, 2026):
                for month in range(1, 13):
                    target_month = f"{year:04d}-{month:02d}"
                    self.assertEqual(accruals.effective_pm_rate(property_name, target_month, 0.22), 0.25)
        self.assertEqual(accruals.effective_pm_rate("724 3rd Ave", "2026-04", 0.06), 0.05)
        self.assertEqual(accruals.effective_pm_rate("724 3rd Ave", "2026-05", 0.05), 0.06)
        self.assertEqual(accruals.effective_pm_rate("724 3rd Ave", "2026-06", 0.05), 0.06)

    def test_governed_madison_and_724_rates_override_loaded_schedule_components(self):
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        try:
            accruals.PM_FEE_COMPONENTS = {
                "86 Madison Ave": [{"rate": 0.22, "deduction": 0.0, "basis_token": "86"}],
                "724 3rd Ave": [{"rate": 0.06, "deduction": 0.0, "basis_token": "724"}],
            }
            rows = [
                {"Date": "August 15, 2025", "Property": "86 Madison Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Revenue", "Category": "Short Term Rents"},
                {"Date": "April 15, 2026", "Property": "724 3rd Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Revenue", "Category": "Short Term Rents"},
                {"Date": "May 15, 2026", "Property": "724 3rd Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Revenue", "Category": "Short Term Rents"},
            ]

            self.assertEqual(accruals.compute_pm_fees(rows, "2025-08")["86 Madison Ave"], 250.00)
            self.assertEqual(accruals.compute_pm_fees(rows, "2026-04")["724 3rd Ave"], 50.00)
            self.assertEqual(accruals.compute_pm_fees(rows, "2026-05")["724 3rd Ave"], 60.00)
        finally:
            accruals.PM_FEE_COMPONENTS = original_components

    def test_madison_jazmine_three_percent_reduces_eco_component_only(self):
        rows = []
        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave"):
            rows.extend([
                {"Date": "2025-08-15", "Property": property_name, "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": ""},
                {"Date": "2025-09-10", "Property": property_name, "Merchant": "CASH APP*JAZMINE MONEY", "Amount": "-30.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": "August 2025 Guest Specialist (Co-Host) PM Fee 3 percent"},
            ])

        fees = accruals.compute_pm_fees(rows, "2025-08")

        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave"):
            self.assertEqual(fees[property_name], 220.00)

    def test_madison_pm_fee_uses_financial_half_up_rounding(self):
        rows = [
            {"Date": "2025-09-15", "Property": "90 Madison Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "5732.98", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": ""},
        ]

        fees = accruals.compute_pm_fees(rows, "2025-09")

        self.assertEqual(fees["90 Madison Ave"], 1433.25)

    def test_madison_accrual_note_does_not_count_as_jazmine_charge(self):
        rows = [
            {"Date": "2026-06-15", "Property": "90 Madison Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": ""},
            {"Date": "2026-06-28", "Property": "90 Madison Ave", "Merchant": "PM Fee Accrual | 90 Madison Ave | June 2026", "Amount": "-250.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": "AOPS-PNL-ACCRUAL|pm|90 Madison Ave|2026-06|250.00 | 25% less any separately booked Jazmine 3% co-host PM fee"},
        ]

        fees = accruals.compute_pm_fees(rows, "2026-06")

        self.assertEqual(fees["90 Madison Ave"], 250.00)

    def test_madison_generic_jazmine_expenses_do_not_reduce_twenty_five_percent_fee(self):
        rows = []
        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave"):
            rows.extend([
                {"Date": "2026-06-15", "Property": property_name, "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": ""},
                {"Date": "2026-06-19", "Property": property_name, "Merchant": "CASH APP*JAZMINE MONEY", "Amount": "-40.00", "Type": "Transaction", "Category": "Cleaning & Janitorial", "Sub-category": "", "Notes": "turnover cleaning"},
            ])

        fees = accruals.compute_pm_fees(rows, "2026-06")

        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave"):
            self.assertEqual(fees[property_name], 250.00)

    def test_madison_jazmine_deduction_audit_only_reports_explicit_three_percent_fee(self):
        rows = [
            {"Date": "2025-09-10", "Property": "90 Madison Ave", "Merchant": "CASH APP*JAZMINE MONEY", "Amount": "-108.10", "Type": "Transaction", "Category": "Service Calls", "Sub-category": "", "Notes": "August 2025 Guest Specialist (Co-Host) PM Fee 3 percent"},
            {"Date": "2025-08-15", "Property": "88 Madison Ave", "Merchant": "CASH APP*JAZMINE MONEY", "Amount": "-231.42", "Type": "Transaction", "Category": "C & M Labor", "Sub-category": "", "Notes": ""},
        ]

        deductions = accruals.madison_jazmine_pm_fee_deductions(rows, "2025-08")

        self.assertEqual(deductions, {"90 Madison Ave": 108.10})

    def test_9ccln_evolve_net_and_non_evolve_gross_pm_math(self):
        rows = [
            {"Date": "2026-03-15", "Property": "9 Country Club Ln N", "Merchant": "EVOLVE VACATION", "Amount": "900.00", "Type": "Transaction", "Category": "", "Sub-category": ""},
            {"Date": "March 20, 2026", "Property": "9 Country Club Ln N", "Merchant": "Hospitable, Inc", "Amount": "1000.00", "Type": "Revenue", "Category": "Short Term Rents", "Sub-category": ""},
        ]

        fees = accruals.compute_pm_fees(rows, "2026-03")

        self.assertEqual(fees["9 Country Club Ln N"], 300.00)

    def test_9ccln_stale_rule_text_is_refreshed_when_amount_matches(self):
        rows = [
            {"Date": "2026-06-24", "Property": "9 Country Club Ln N", "Merchant": "EVOLVE VACATION", "Amount": "900.00", "Type": "Transaction", "Category": "", "Sub-category": "", "Notes": ""},
            {"Date": "2026-06-28", "Property": "9 Country Club Ln N", "Merchant": "PM Fee Accrual", "Amount": "-100.00", "Type": "Manual", "Category": "Property Management", "Sub-category": "", "Notes": "AOPS-PNL-ACCRUAL|pm|9 Country Club Ln N|2026-06|100.00 | stale 10% rule"},
        ]

        mismatches = accruals.find_amount_mismatches(rows, "2026-06", ["9 Country Club"], {"pm"})

        self.assertEqual(len(mismatches), 1)
        self.assertTrue(mismatches[0]["stale_pm_rule"])

    def test_iso_export_schema_detects_unclassified_known_rent_merchants(self):
        rows = [
            {"Date": "2026-06-22", "Property": "86 Madison Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "1000.00", "Type": "Transaction", "Category": "", "Sub-category": ""},
            {"Date": "2026-06-23", "Property": "86 Madison Ave", "Merchant": "AIRBNB PAYMENTS", "Amount": "250.00", "Type": "Transaction", "Category": "Revenue", "Sub-category": "Fees & Other Revenue"},
        ]

        fees = accruals.compute_pm_fees(rows, "2026-06")

        self.assertEqual(fees["86 Madison Ave"], 250.00)

    def test_full_schedule_adds_active_pm_and_dao_and_skips_sold(self):
        from openpyxl import Workbook

        original_templates = copy.deepcopy(accruals.ACCRUAL_TEMPLATES)
        original_pm = list(accruals.PM_FEE_PROPERTIES)
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        original_aliases = copy.deepcopy(accruals.PROPERTY_ALIASES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "schedule.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Address", "PM / Sub-PM", "On Lofty?", "DAO", "Current Status (Occupied Units)", "Total Units", "PM Fee (% of Gross Rents)"])
                sheet.append(["1432 Sara Avenue, Akron, OH 44305", "Hemlane (sub)", "Yes", "Example DAO LLC", 1, 1, "10% (minus Hemlane fee - $57/unit/month)"])
                sheet.append(["9919 South Oglesby Avenue, Chicago, IL 60617", "Sold", "Yes", "Sold DAO LLC", "Sold", 1, 0.10])
                workbook.save(path)
                workbook.close()
                rows = [{"Date": "June 04, 2026", "Property": "1432 Sara Ave.", "Amount": "1035", "Type": "Revenue", "Category": "Rents", "Notes": ""}]

                resolved, evidence = accruals.pm_fee_properties_from_schedule(path, gl_rows=rows)
                accruals.PM_FEE_PROPERTIES = resolved

            self.assertIn(("1432 Sara Ave.", 0.10, accruals.SCHEDULE_SOURCE_PREFIX), resolved)
            self.assertNotIn("9919 S Oglesby Ave, Chicago, IL 60617", {item[0] for item in resolved})
            self.assertTrue(any(item["property"] == "1432 Sara Ave." and item["kind"] == "dao" for item in accruals.ACCRUAL_TEMPLATES))
            self.assertEqual(accruals.compute_pm_fees(rows, "2026-06")["1432 Sara Ave."], 103.50)
            self.assertEqual(evidence["excluded_property_count"], 1)
        finally:
            accruals.ACCRUAL_TEMPLATES = original_templates
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components
            accruals.PROPERTY_ALIASES = original_aliases

    def test_sold_listing_update_policy_excludes_schedule_pm_and_dao(self):
        from openpyxl import Workbook

        original_templates = copy.deepcopy(accruals.ACCRUAL_TEMPLATES)
        original_pm = list(accruals.PM_FEE_PROPERTIES)
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        original_aliases = copy.deepcopy(accruals.PROPERTY_ALIASES)
        original_policy = os.environ.get("LOFTY_LISTING_UPDATE_POLICY")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                tmp_path = Path(tmp)
                policy = tmp_path / "policy.json"
                policy.write_text(
                    json.dumps({"sold_ignore_listing_updates": ["402 N Wild Olive Ave, Daytona Beach, FL 32118"]}),
                    encoding="utf-8",
                )
                os.environ["LOFTY_LISTING_UPDATE_POLICY"] = str(policy)
                accruals.sold_listing_update_policy_excluded_schedule_names.cache_clear()

                schedule = tmp_path / "schedule.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Address", "PM / Sub-PM", "On Lofty?", "DAO", "Current Status (Occupied Units)", "Total Units", "PM Fee (% of Gross Rents)"])
                sheet.append(["402 N Wild Olive Ave, Daytona Beach, FL 32118", "Self", "Yes", "402 DAO LLC", 1, 1, "15%"])
                workbook.save(schedule)
                workbook.close()
                rows = [{"Date": "June 04, 2026", "Property": "402 N Wild Olive Ave", "Amount": "1000", "Type": "Revenue", "Category": "Rents", "Notes": ""}]

                resolved, evidence = accruals.pm_fee_properties_from_schedule(schedule, gl_rows=rows)

            self.assertNotIn("402 N Wild Olive Ave", {item[0] for item in resolved})
            self.assertFalse(any(item["property"] == "402 N Wild Olive Ave" and item["kind"] == "dao" for item in accruals.ACCRUAL_TEMPLATES))
            self.assertEqual(evidence["excluded_property_count"], 1)
        finally:
            if original_policy is None:
                os.environ.pop("LOFTY_LISTING_UPDATE_POLICY", None)
            else:
                os.environ["LOFTY_LISTING_UPDATE_POLICY"] = original_policy
            accruals.sold_listing_update_policy_excluded_schedule_names.cache_clear()
            accruals.ACCRUAL_TEMPLATES = original_templates
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components
            accruals.PROPERTY_ALIASES = original_aliases

    def test_manual_workspace_exclusion_skips_coolwood_schedule_pm_and_dao(self):
        from openpyxl import Workbook

        original_templates = copy.deepcopy(accruals.ACCRUAL_TEMPLATES)
        original_pm = list(accruals.PM_FEE_PROPERTIES)
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        original_aliases = copy.deepcopy(accruals.PROPERTY_ALIASES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                schedule = Path(tmp) / "schedule.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Address", "PM / Sub-PM", "On Lofty?", "DAO", "Current Status (Occupied Units)", "Total Units", "PM Fee (% of Gross Rents)"])
                sheet.append(["1 Coolwood Dr., Little Rock, AR 72202", "ECO Systems LLC", "Yes", "Coolwood DAO LLC", 1, 1, "10%"])
                workbook.save(schedule)
                workbook.close()
                rows = [{"Date": "July 04, 2026", "Property": "1 Coolwood Dr.", "Amount": "1000", "Type": "Revenue", "Category": "Rents", "Notes": ""}]

                resolved, evidence = accruals.pm_fee_properties_from_schedule(schedule, gl_rows=rows)

            self.assertNotIn("1 Coolwood Dr.", {item[0] for item in resolved})
            self.assertFalse(any(item["property"] == "1 Coolwood Dr." and item["kind"] == "dao" for item in accruals.ACCRUAL_TEMPLATES))
            self.assertEqual(evidence["excluded_property_count"], 1)
        finally:
            accruals.ACCRUAL_TEMPLATES = original_templates
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components
            accruals.PROPERTY_ALIASES = original_aliases

    def test_schedule_without_pm_fee_column_preserves_existing_pm_rates(self):
        from openpyxl import Workbook

        original_templates = copy.deepcopy(accruals.ACCRUAL_TEMPLATES)
        original_pm = list(accruals.PM_FEE_PROPERTIES)
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        original_aliases = copy.deepcopy(accruals.PROPERTY_ALIASES)
        try:
            with tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "schedule.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Address", "PM / Sub-PM", "On Lofty?", "DAO", "Current Status (Occupied Units)", "Total Units"])
                sheet.append(["86 Madison Avenue, Albany, NY 12202", "ECO Systems LLC", "Yes", "Snow Leopard DAO LLC", "STR", "2"])
                workbook.save(path)
                workbook.close()

                resolved, evidence = accruals.pm_fee_properties_from_schedule(
                    path,
                    base=[("86 Madison Ave", 0.25, "AOPS-PNL-ACCRUAL")],
                    gl_rows=[],
                )

            self.assertEqual(resolved, [("86 Madison Ave", 0.25, "AOPS-PNL-ACCRUAL")])
            self.assertEqual(evidence["fallback_rate_property_count"], 1)
            self.assertEqual(evidence["matched_property_count"], 0)
        finally:
            accruals.ACCRUAL_TEMPLATES = original_templates
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components
            accruals.PROPERTY_ALIASES = original_aliases

    def test_stale_86_pm_amount_is_detected_from_schedule_rate(self):
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        try:
            accruals.PM_FEE_COMPONENTS = {"86 Madison Ave": [{"rate": 0.22, "deduction": 0.0, "basis_token": "86"}]}
            rows = [
                {"Date": "June 15, 2026", "Property": "86 Madison Ave", "Amount": "4969.72", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
                {"Date": "June 28, 2026", "Property": "86 Madison Ave", "Amount": "-496.97", "Type": "Operating Expenses", "Category": "Property Management", "Notes": "AOPS-PNL-ACCRUAL|pm|86 Madison Ave|2026-06|496.97 | stale 10%"},
            ]

            mismatches = accruals.find_amount_mismatches(rows, "2026-06", property_filters=["86 Madison Ave"], kind_filters={"pm"})

            self.assertEqual(len(mismatches), 1)
            self.assertEqual(mismatches[0]["expected_amount"], 1242.43)
        finally:
            accruals.PM_FEE_COMPONENTS = original_components

    def test_applied_pm_amount_update_clears_remaining_mismatch(self):
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        try:
            accruals.PM_FEE_COMPONENTS = {"86 Madison Ave": [{"rate": 0.22, "deduction": 0.0, "basis_token": "86"}]}
            rows = [
                {"Date": "June 15, 2026", "Property": "86 Madison Ave", "Amount": "4969.72", "Type": "Revenue", "Category": "Short Term Rents", "Notes": ""},
                {"Date": "June 28, 2026", "Property": "86 Madison Ave", "Amount": "-496.97", "Type": "Operating Expenses", "Category": "Property Management", "Notes": "AOPS-PNL-ACCRUAL|pm|86 Madison Ave|2026-06|496.97 | stale 10%"},
            ]

            mismatches = accruals.find_amount_mismatches(rows, "2026-06", property_filters=["86 Madison Ave"], kind_filters={"pm"})
            applied = accruals.apply_amount_mismatch_updates(rows, mismatches)
            remaining = accruals.find_amount_mismatches(rows, "2026-06", property_filters=["86 Madison Ave"], kind_filters={"pm"})

            self.assertEqual(applied[0]["new_amount"], 1242.43)
            self.assertEqual(remaining, [])
        finally:
            accruals.PM_FEE_COMPONENTS = original_components

    def test_hemlane_direct_split_pm_accrual_is_voided_from_transaction_source(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Property": "566 Nash Rd, Northfield, OH 44067",
                "Merchant": "Hemlane",
                "Description": "Hemlane | rent remittance",
                "Amount": "1200.00",
                "Type": "Revenue",
                "Category": "Long Term Rents",
                "Notes": "",
            },
            {
                "Date": "June 28, 2026",
                "Property": "566 Nash Rd, Northfield, OH 44067",
                "Merchant": "PM Fee Accrual",
                "Description": "PM Fee Accrual",
                "Amount": "-82.71",
                "Type": "Operating Expenses",
                "Category": "Property Management",
                "Notes": "AOPS-OHIL-ACCRUAL|pm|566 Nash Rd, Northfield, OH 44067|2026-06|82.71 | legacy PM accrual",
            },
        ]

        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            property_filters=["566 Nash Rd"],
            kind_filters={"pm"},
        )
        applied = accruals.apply_amount_mismatch_updates(rows, mismatches)

        self.assertEqual(len(mismatches), 1)
        self.assertTrue(mismatches[0]["void_hemlane_direct_split"])
        self.assertEqual(applied[0]["new_amount"], 0.0)
        self.assertEqual(rows[1]["Amount"], "0.00")
        self.assertIn("|pm|566 Nash Rd, Northfield, OH 44067|2026-06|0.00", rows[1]["Notes"])
        self.assertIn("transaction-level Hemlane rent evidence", rows[1]["Notes"])
        self.assertIn("no cash movement", rows[1]["Notes"])
        self.assertEqual(
            accruals.find_amount_mismatches(
                rows,
                "2026-06",
                property_filters=["566 Nash Rd"],
                kind_filters={"pm"},
            ),
            [],
        )

    def test_hemlane_exclusion_is_transaction_based_across_states(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        properties = [
            "10 New York Test Rd, Albany, NY 12207",
            "20 Arkansas Test Rd, Little Rock, AR 72201",
            "30 Colorado Test Rd, Denver, CO 80202",
        ]
        try:
            accruals.PM_FEE_PROPERTIES = [
                (property_name, 0.10, "AOPS-TEST-ACCRUAL")
                for property_name in properties
            ]
            accruals.PM_FEE_COMPONENTS = {}
            rows = []
            for property_name in properties:
                rows.extend(
                    [
                        {
                            "Date": "June 10, 2026",
                            "Property": property_name,
                            "Merchant": "Hemlane",
                            "Description": "Hemlane net rent remittance",
                            "Amount": "900.00",
                            "Type": "Revenue",
                            "Category": "Long Term Rents",
                            "Notes": "",
                        },
                        {
                            "Date": "June 12, 2026",
                            "Property": property_name,
                            "Merchant": "Tenant direct ACH",
                            "Description": "June gross rent",
                            "Amount": "1000.00",
                            "Type": "Revenue",
                            "Category": "Long Term Rents",
                            "Notes": "",
                        },
                    ]
                )

            fees = accruals.compute_pm_fees(rows, "2026-06")

            self.assertEqual(fees, {property_name: 100.0 for property_name in properties})
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components

    def test_zero_hemlane_void_marker_legacy_wording_is_nonblocking(self):
        rows = [
            {
                "Date": "July 15, 2026",
                "Property": "566 Nash Rd, Northfield, OH 44067",
                "Merchant": "Hemlane",
                "Description": "Hemlane | rent remittance",
                "Amount": "1200.00",
                "Type": "Revenue",
                "Category": "Long Term Rents",
                "Notes": "",
            },
            {
                "Date": "July 31, 2026",
                "Property": "566 Nash Rd, Northfield, OH 44067",
                "Amount": "0.00",
                "Type": "Operating Expenses",
                "Category": "Property Management",
                "Notes": "AOPS-OHIL-ACCRUAL|pm_dao|566 Nash Rd, Northfield, OH 44067|2026-07|0.00 | older void explanation",
            },
        ]

        self.assertEqual(accruals.find_amount_mismatches(rows, "2026-07"), [])

    def test_hemlane_direct_split_pm_settlement_is_voided_too(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Property": "1456 W 85th St, Cleveland, OH 44102",
                "Merchant": "Hemlane",
                "Description": "Hemlane | rent remittance",
                "Amount": "1500.00",
                "Type": "Revenue",
                "Category": "Long Term Rents",
                "Notes": "",
            },
            {
                "Date": "June 28, 2026",
                "Property": "1456 W 85th St, Cleveland, OH 44102",
                "Merchant": "PM Fee Settlement",
                "Description": "PM Fee Settlement",
                "Amount": "161.67",
                "Type": "Operating Expenses",
                "Category": "Property Management",
                "Notes": "AOPS-OHIL-ACCRUAL|pm_settlement|1456 W 85th St, Cleveland, OH 44102|2026-06|161.67 | legacy settlement",
            }
        ]

        mismatches = accruals.find_amount_mismatches(rows, "2026-06")
        accruals.apply_amount_mismatch_updates(rows, mismatches)

        self.assertEqual(len(mismatches), 1)
        self.assertTrue(mismatches[0]["void_hemlane_direct_split"])
        self.assertEqual(rows[1]["Amount"], "0.00")
        self.assertIn("|pm_settlement|1456 W 85th St, Cleveland, OH 44102|2026-06|0.00", rows[1]["Notes"])

    def test_direct_gross_rent_accrues_regardless_of_property_state(self):
        original_pm_properties = list(accruals.PM_FEE_PROPERTIES)
        accruals.PM_FEE_PROPERTIES.append(
            ("566 Nash Rd, Northfield, OH 44067", 0.10, "AOPS-OHIL-ACCRUAL")
        )
        try:
            rows = [
                {
                    "Date": "June 15, 2026",
                    "Property": "566 Nash Rd, Northfield, OH 44067",
                    "Amount": "1200.00",
                    "Type": "Revenue",
                    "Category": "Long Term Rents",
                    "Notes": "",
                },
                {
                    "Date": "June 28, 2026",
                    "Property": "566 Nash Rd, Northfield, OH 44067",
                    "Amount": "-82.71",
                    "Type": "Operating Expenses",
                    "Category": "Property Management",
                    "Notes": "AOPS-OHIL-ACCRUAL|pm|566 Nash Rd, Northfield, OH 44067|2026-06|82.71 | reviewed unpaid exception",
                },
            ]

            mismatches = accruals.find_amount_mismatches(rows, "2026-06")

            self.assertEqual(len(mismatches), 1)
            self.assertFalse(mismatches[0]["void_hemlane_direct_split"])
            self.assertEqual(mismatches[0]["expected_amount"], 120.0)
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm_properties

    def test_bookkeeping_schedule_overrides_pm_rates(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "schedule.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Address", "PM Fee (% of Gross Rents)"])
            sheet.append(["86 Madison Avenue, Albany, NY 12202", 0.22])
            sheet.append(["724 3rd Avenue, Watervliet, NY 12189", 0.05])
            workbook.save(path)
            workbook.close()

            resolved, evidence = accruals.pm_fee_properties_from_schedule(
                path,
                [
                    ("86 Madison Ave", 0.10, "AOPS-PNL-ACCRUAL"),
                    ("724 3rd Ave", 0.10, "AOPS-PNL-ACCRUAL"),
                ],
            )

        self.assertEqual(resolved[0][1], 0.22)
        self.assertEqual(resolved[1][1], 0.05)
        self.assertEqual(evidence["matched_property_count"], 2)

    def test_legacy_pnl_legal_marker_counts_as_dao_llc_coverage(self):
        rows = [{"Notes": "AOPS-PNL-ACCRUAL|legal|86 Madison Ave|2026-06|62.50 | legacy label"}]

        coverage = accruals.find_existing_coverage(rows, "2026-06")

        self.assertIn("86 Madison Ave|dao", coverage)

    def test_legacy_pnl_legal_marker_with_correct_amount_is_not_a_mismatch(self):
        rows = [
            {
                "Amount": "-62.50",
                "Notes": "AOPS-PNL-ACCRUAL|legal|86 Madison Ave|2026-06|62.50 | legacy label",
            }
        ]

        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            property_filters=["86 Madison Ave"],
            kind_filters={"dao"},
        )

        self.assertEqual(mismatches, [])

    def test_legacy_pnl_legal_marker_with_wrong_amount_remains_a_mismatch(self):
        rows = [
            {
                "Amount": "-50.00",
                "Notes": "AOPS-PNL-ACCRUAL|legal|86 Madison Ave|2026-06|50.00 | legacy label",
            }
        ]

        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            property_filters=["86 Madison Ave"],
            kind_filters={"dao"},
        )

        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0]["expected_amount"], 62.50)

    def test_annual_reference_tax_template_generates_monthly_posting_amount(self):
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            [],
            "2026-06",
            fieldnames,
            property_filters=["5541 S Peoria St"],
            kind_filters={"taxes"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "-219.00")
        self.assertIn("|taxes|5541 S Peoria St, Chicago, IL 60621|2026-06|219.00", generated[0]["Notes"])
        self.assertIn("annual reference $2628.00", generated[0]["Notes"])

    def test_sold_property_manual_tax_accruals_are_not_regenerated(self):
        sold_properties = {
            "1315 E 114th St, Cleveland, OH 44106",
            "13806 Coit Rd, Cleveland, OH 44110",
            "3024 W 103rd St, Cleveland, OH 44111",
            "9919 S Oglesby Ave, Chicago, IL 60617",
        }
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            [],
            "2026-06",
            fieldnames,
            property_filters=list(sold_properties),
            kind_filters={"taxes"},
        )

        self.assertEqual(generated, [])
        tax_templates = {
            template["property"]
            for template in accruals.ACCRUAL_TEMPLATES
            if template["kind"] == "taxes"
        }
        self.assertFalse(sold_properties & tax_templates)
        self.assertFalse(
            sold_properties
            & {property_name for property_name, kind in accruals.ANNUAL_REFERENCE_FIXED_ACCRUALS if kind == "taxes"}
        )

    def test_annual_reference_tax_template_repairs_stale_full_year_posting(self):
        rows = [
            {
                "Amount": "-2628.00",
                "Notes": "AOPS-OHIL-ACCRUAL|taxes|5541 S Peoria St, Chicago, IL 60621|2026-06|2628.00 | stale annual amount",
            }
        ]

        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            property_filters=["5541 S Peoria St"],
            kind_filters={"taxes"},
        )
        applied = accruals.apply_amount_mismatch_updates(rows, mismatches)

        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0]["expected_amount"], 219.00)
        self.assertEqual(applied[0]["new_amount"], 219.00)
        self.assertEqual(rows[0]["Amount"], "-219.00")
        self.assertIn("|taxes|5541 S Peoria St, Chicago, IL 60621|2026-06|219.00", rows[0]["Notes"])
        self.assertIn("annual reference $2628.00", rows[0]["Notes"])

    def test_monthly_tax_template_stays_monthly_for_9_country_club(self):
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            [],
            "2026-06",
            fieldnames,
            property_filters=["9 Country Club"],
            kind_filters={"taxes"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "-3225.01")
        self.assertIn("|taxes|9 Country Club Ln N|2026-06|3225.01", generated[0]["Notes"])

    def test_9_country_club_mortgage_interest_starts_september_2025(self):
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        august = accruals.generate_missing_accruals(
            [],
            "2025-08",
            fieldnames,
            property_filters=["9 Country Club"],
            kind_filters={"mortgage_interest"},
        )
        september = accruals.generate_missing_accruals(
            [],
            "2025-09",
            fieldnames,
            property_filters=["9 Country Club"],
            kind_filters={"mortgage_interest"},
        )

        self.assertEqual(august, [])
        self.assertEqual(len(september), 1)
        self.assertEqual(september[0]["Amount"], "-2875.00")

    def test_address_abbreviation_marker_counts_as_canonical_coverage(self):
        rows = [{"Notes": "AOPS-MONTHLY-ACCRUAL|dao|86 Madison Avenue|2026-06|62.50 | canonical alias"}]

        coverage = accruals.find_existing_coverage(rows, "2026-06")

        self.assertIn("86 Madison Ave|dao", coverage)

    def test_86_madison_generates_june_28_pm_and_dao_llc_accruals(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Category": "Short Term Rents",
                "Property": "86 Madison Ave",
                "Amount": "4969.72",
                "Type": "Revenue",
                "Notes": "",
            }
        ]
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            property_filters=["86 Madison Ave"],
            kind_filters={"pm", "dao"},
        )

        self.assertEqual(len(generated), 4)
        self.assertEqual({row["Date"] for row in generated}, {"June 28, 2026"})
        self.assertEqual({row["Amount"] for row in generated}, {"-62.50", "62.50", "-1242.43", "1242.43"})
        self.assertTrue(any("DAO LLC Fee Accrual" in row["Description"] for row in generated))
        eco_row = next(row for row in generated if row["Notes"].split("|")[1] == "pm_eco")
        self.assertEqual(eco_row["Type"], "Revenue")
        self.assertEqual(eco_row["Category"], "Fees & Other Revenue")
        self.assertEqual(eco_row["Sub-category"], "Property Management")
        dao_eco_row = next(row for row in generated if row["Notes"].split("|")[1] == "dao_eco")
        self.assertEqual(dao_eco_row["Amount"], "62.50")
        self.assertEqual(dao_eco_row["Type"], "Revenue")
        self.assertEqual(dao_eco_row["Category"], "Fees & Other Revenue")
        self.assertEqual(dao_eco_row["Property"], "Mining, Sales, Consulting, and PM")
        self.assertIn("full $750.00 annual charge", dao_eco_row["Notes"])
        self.assertIn("separate $200.00 annual payable to Lofty", dao_eco_row["Notes"])
        self.assertIn("never reduce the DAO's fee or payable", dao_eco_row["Notes"])

    def test_existing_legacy_dao_fee_expense_only_backfills_eco_revenue_side(self):
        rows = [{
            "Date": "June 28, 2026",
            "Property": "724 3rd Ave",
            "Amount": "-62.50",
            "Notes": "AOPS-PNL-ACCRUAL|dao|724 3rd Ave|2026-06|62.50 | legacy one-sided fee",
        }]
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            property_filters=["724 3rd Ave"],
            kind_filters={"dao"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "62.50")
        self.assertIn("|dao_eco|724 3rd Ave|2026-06|62.50", generated[0]["Notes"])

    def test_dao_eco_only_backfills_existing_non_template_expense(self):
        rows = [{
            "Date": "June 28, 2026",
            "Property": "1 Coolwood Dr.",
            "Amount": "-62.50",
            "Notes": "AOPS-MONTHLY-ACCRUAL|dao|1 Coolwood Dr.|2026-06|62.50 | existing fee",
        }]
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]

        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            kind_filters={"dao_eco"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "62.50")
        self.assertEqual(generated[0]["Category"], "Fees & Other Revenue")
        self.assertEqual(generated[0]["Property"], "Mining, Sales, Consulting, and PM")
        self.assertIn(
            "AOPS-MONTHLY-ACCRUAL|dao_eco|1 Coolwood Dr.|2026-06|62.50",
            generated[0]["Notes"],
        )

    def test_pm_eco_metadata_normalization_converges_when_amount_already_matches(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Property": "86 Madison Ave",
                "Amount": "4969.72",
                "Type": "Revenue",
                "Category": "Short Term Rents",
                "Notes": "",
            },
            {
                "Date": "June 28, 2026",
                "Property": "86 Madison Ave",
                "Amount": "1242.43",
                "Type": "Revenue",
                "Category": "Fees & Other Revenue",
                "Sub-category": "Property Management",
                "Notes": "AOPS-PNL-ACCRUAL|pm_eco|86 Madison Ave|2026-06|1242.43 | legacy PM ECO note",
            },
        ]

        mismatches = accruals.find_amount_mismatches(rows, "2026-06")
        self.assertEqual(len(mismatches), 1)
        self.assertTrue(mismatches[0]["stale_pm_rule"])

        accruals.apply_amount_mismatch_updates(rows, mismatches)

        self.assertEqual(accruals.find_amount_mismatches(rows, "2026-06"), [])
        self.assertIn("25% of direct gross rent", rows[1]["Notes"])

    def test_month_end_pm_fee_marker_counts_as_coverage(self):
        rows = [
            {
                "Date": "June 28, 2026",
                "Category": "Property Management",
                "Notes": "AOPS-PM-FEE|85-104 Alawa Pl|2026-06|345.23 | current-month PM fee accrual",
            }
        ]

        coverage = accruals.find_existing_coverage(rows, "2026-06")

        self.assertIn("85-104 Alawa Pl|pm", coverage)

    def test_explicit_partial_paired_pm_run_completes_only_missing_eco_side(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Property": "90 Madison Ave",
                "Amount": "10000.00",
                "Type": "Revenue",
                "Category": "Short Term Rents",
                "Notes": "",
            },
            {
                "Date": "June 28, 2026",
                "Property": "90 Madison Ave",
                "Amount": "-2500.00",
                "Type": "Operating Expenses",
                "Category": "Property Management",
                "Notes": "AOPS-PNL-ACCRUAL|pm_dao|90 Madison Ave|2026-06|2500.00 | Accrual ID AOPS-PNL-ACCRUAL|pm|90 Madison Ave|2026-06|2500.00",
            },
        ]

        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
            property_filters=["90 Madison Ave"],
            kind_filters={"pm"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "2500.00")
        self.assertEqual(generated[0]["Type"], "Revenue")
        self.assertIn("|pm_eco|90 Madison Ave|2026-06|2500.00", generated[0]["Notes"])

    def test_first_day_pm_fee_marker_does_not_count_as_coverage(self):
        rows = [
            {
                "Date": "June 01, 2026",
                "Category": "Property Management",
                "Notes": "AOPS-PM-FEE|90 Madison Ave|2026-06|1053.05 | stale first-day PM fee",
            }
        ]

        coverage = accruals.find_existing_coverage(rows, "2026-06")

        self.assertNotIn("90 Madison Ave|pm", coverage)

    def test_first_day_pm_fee_marker_blocks_generated_pm_duplicate(self):
        rows = [
            {
                "Date": "June 01, 2026",
                "Category": "Property Management",
                "Property": "90 Madison Ave",
                "Notes": "AOPS-PM-FEE|90 Madison Ave|2026-06|1053.05 | stale first-day PM fee",
            },
            {
                "Date": "June 15, 2026",
                "Category": "Short Term Rents",
                "Property": "90 Madison Ave",
                "Amount": "10000.00",
                "Type": "Revenue",
                "Notes": "",
            },
        ]

        blocked = accruals.find_blocked_first_day_pm_fees(rows, "2026-06")
        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
            property_filters=["90 Madison Ave"],
            kind_filters={"pm"},
        )

        self.assertEqual(1, len(blocked))
        self.assertEqual([], generated)

    def test_duplicate_month_end_pm_markers_are_reported(self):
        rows = [
            {
                "Date": "August 28, 2025",
                "Property": "90 Madison Ave",
                "Amount": "-1526.08",
                "Category": "Property Management",
                "Notes": "AOPS-PNL-ACCRUAL|pm|90 Madison Ave|2025-08|1526.08",
            },
            {
                "Date": "August 28, 2025",
                "Property": "90 Madison Ave",
                "Amount": "-1526.08",
                "Category": "Property Management",
                "Notes": "AOPS-PNL-ACCRUAL|pm|90 Madison Ave|2025-08|1526.08",
            },
        ]

        duplicates = accruals.find_duplicate_accrual_markers(
            rows,
            "2025-08",
            property_filters=["90 Madison Ave"],
            kind_filters={"pm"},
        )

        self.assertEqual(len(duplicates), 1)
        self.assertEqual(duplicates[0]["key"], "90 Madison Ave|pm|2025-08")
        self.assertEqual(duplicates[0]["row_count"], 2)

    def test_legacy_pm_accrual_is_not_silently_backfilled_or_settled(self):
        rows = [
            {
                "Date": "August 15, 2025",
                "Property": "90 Madison Ave",
                "Amount": "8000.00",
                "Type": "Revenue",
                "Category": "Short Term Rents",
            },
            {
                "Date": "September 10, 2025",
                "Property": "90 Madison Ave",
                "Merchant": "ECO Systems, LLC | INTERNAL_TRANSFER",
                "Amount": "-1760.00",
                "Category": "Property Management",
                "Notes": "August 2025 PM Fee 22 percent",
            },
            {
                "Date": "September 10, 2025",
                "Property": "90 Madison Ave",
                "Merchant": "CASH APP*JAZMINE MONEY",
                "Amount": "-240.00",
                "Category": "Service Calls",
                "Notes": "August 2025 Guest Specialist (Co-Host) PM Fee 3 percent",
            },
        ]

        fees = accruals.compute_pm_fees(rows, "2025-08")
        settlements = accruals.compute_pm_settlements(rows, "2025-08")

        self.assertEqual(fees["90 Madison Ave"], 1760.0)
        self.assertEqual(settlements["90 Madison Ave"], 1760.0)

        rows.append(
            {
                "Date": "August 28, 2025",
                "Property": "90 Madison Ave",
                "Amount": "-1760.00",
                "Category": "Property Management",
                "Notes": "AOPS-PNL-ACCRUAL|pm|90 Madison Ave|2025-08|1760.00",
            }
        )
        generated = accruals.generate_missing_accruals(
            rows,
            "2025-08",
            ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
            property_filters=["90 Madison Ave"],
            kind_filters={"pm"},
        )

        self.assertEqual(generated, [])

    def test_madison_pm_accrual_keeps_unpaid_three_percent(self):
        rows = [
            {
                "Date": "August 15, 2025",
                "Property": "86 Madison Ave",
                "Amount": "8000.00",
                "Type": "Revenue",
                "Category": "Short Term Rents",
            },
            {
                "Date": "September 10, 2025",
                "Property": "86 Madison Ave",
                "Merchant": "ECO Systems, LLC | INTERNAL_TRANSFER",
                "Amount": "-1760.00",
                "Category": "Property Management",
                "Notes": "August 2025 PM Fees 22 percent",
            },
        ]

        fees = accruals.compute_pm_fees(rows, "2025-08")
        settlements = accruals.compute_pm_settlements(rows, "2025-08")

        self.assertEqual(fees["86 Madison Ave"], 2000.0)
        self.assertEqual(settlements["86 Madison Ave"], 1760.0)

    def test_third_party_management_charge_does_not_settle_eco_pm_accrual(self):
        rows = [
            {
                "Date": "June 15, 2026",
                "Property": "8708 Willard Ave",
                "Amount": "1595.00",
                "Type": "Revenue",
                "Category": "Long Term Rents",
            },
            {
                "Date": "June 28, 2026",
                "Property": "8708 Willard Ave",
                "Merchant": "Hemlane",
                "Amount": "-159.50",
                "Category": "Property Management",
                "Notes": "June 2026 management fee",
            },
        ]

        fees = accruals.compute_pm_fees(rows, "2026-06")
        settlements = accruals.compute_pm_settlements(rows, "2026-06")

        self.assertEqual(fees["8708 Willard Ave, Cleveland, OH 44102"], 159.5)
        self.assertNotIn("8708 Willard Ave, Cleveland, OH 44102", settlements)

    def test_explicit_property_prevents_account_name_cross_match(self):
        row = {
            "Account": "Heron LFTY0314 DAO LLC-88 Madison Ave Operations-0181",
            "Property": "86 Madison Ave",
            "Merchant": "Morgan Linen Service - 86 Madison Ave",
        }

        self.assertTrue(accruals.row_matches_property(row, "86 Madison Ave"))
        self.assertFalse(accruals.row_matches_property(row, "88 Madison Ave"))

    def test_cron_hook_applies_by_default_with_dry_run_override(self):
        hook = Path("/home/digit/.openclaw/workspace/scripts/baselane_monthly_accruals_cron.sh").read_text(
            encoding="utf-8"
        )

        self.assertIn('APPLY="${BASELANE_MONTHLY_ACCRUALS_APPLY:-0}"', hook)
        self.assertIn('DRY_RUN="${BASELANE_MONTHLY_ACCRUALS_DRY_RUN:-0}"', hook)
        self.assertIn("--gap-approval-review-csv", hook)
        self.assertIn("--gap-approval-import-commands", hook)
        self.assertIn("--import-gap-approval-csv", hook)
        self.assertIn("--update-amount-mismatches", hook)
        self.assertIn("obie_cash_basis_insurance_cleanup.py", hook)
        self.assertIn("OH-IL-TN Cash-Basis Insurance Duplicate Audit.md", hook)
        self.assertIn("BASELANE_OBIE_INSURANCE_CLEANUP_APPLY_LIVE:-0", hook)
        self.assertIn('echo "[baselane-monthly-accruals] Verifying post-apply accrual state"', hook)
        self.assertIn('if [ "$arg" != "--apply" ]; then', hook)
        self.assertIn('BASELANE_MONTHLY_ACCRUALS_IMPORT_GAP_APPROVAL_CSV:-0', hook)
        self.assertIn(
            'if [ "$DRY_RUN" = "1" ]; then\n'
            '  :\n'
            'elif [ "$APPLY" = "1" ]; then',
            hook,
        )

    def test_append_rows_to_gl_creates_recoverable_backup(self):
        with tempfile.TemporaryDirectory() as tmp:
            gl_path = Path(tmp) / "ECO Systems General Ledger.csv"
            fieldnames = ["Date", "Amount", "Notes"]
            gl_path.write_text("Date,Amount,Notes\nJune 1 2026,-1.00,existing\n", encoding="utf-8")

            accruals.append_rows_to_gl(
                gl_path,
                [{"Date": "June 28 2026", "Amount": "-62.50", "Notes": "AOPS-test"}],
                fieldnames,
            )

            backups = list(gl_path.parent.glob("ECO Systems General Ledger.csv.bak-*"))
            self.assertEqual(1, len(backups))
            self.assertIn("existing", backups[0].read_text(encoding="utf-8"))
            self.assertIn("AOPS-test", gl_path.read_text(encoding="utf-8"))

    def test_changed_ledger_is_rejected_before_accrual_write(self):
        with tempfile.TemporaryDirectory() as tmp:
            gl_path = Path(tmp) / "ECO Systems General Ledger.csv"
            gl_path.write_text("Date,Amount,Notes\nJune 1 2026,-1.00,existing\n", encoding="utf-8")
            expected_sha256 = accruals.ledger_sha256(gl_path)
            gl_path.write_text("Date,Amount,Notes\nJune 2 2026,-2.00,concurrent change\n", encoding="utf-8")

            with self.assertRaisesRegex(RuntimeError, "GL changed during accrual run"):
                accruals.append_rows_to_gl(
                    gl_path,
                    [{"Date": "June 28 2026", "Amount": "-62.50", "Notes": "AOPS-test"}],
                    ["Date", "Amount", "Notes"],
                    expected_sha256=expected_sha256,
                )

            self.assertNotIn("AOPS-test", gl_path.read_text(encoding="utf-8"))

    def test_full_address_templates_report_to_existing_property_names(self):
        self.assertEqual(
            "10724 Gooding Ave",
            accruals.reporting_property_name("10724 Gooding Ave, Cleveland, OH 44108"),
        )
        self.assertEqual(
            "1456 W 85th St.",
            accruals.reporting_property_name("1456 W 85th St, Cleveland, OH 44102"),
        )
        self.assertEqual(
            "9634 S Green St",
            accruals.reporting_property_name("9634 S Green St, Chicago, IL 60643"),
        )
        self.assertEqual(
            "85-104 Alawa Pl",
            accruals.reporting_property_name("85-104 Alawa Pl"),
        )


    def test_expected_fixed_accrual_coverage_keys_include_templates_and_pm(self):
        keys = accruals.expected_fixed_accrual_keys(property_filters=["1456 W 85th St"])

        self.assertIn("1456 W 85th St, Cleveland, OH 44102|dao", keys)
        self.assertNotIn("1456 W 85th St, Cleveland, OH 44102|insurance", keys)
        self.assertIn("1456 W 85th St, Cleveland, OH 44102|taxes", keys)
        self.assertIn("1456 W 85th St, Cleveland, OH 44102|pm", keys)
        self.assertEqual(accruals.coverage_by_kind(keys), {"dao": 1, "pm": 1, "taxes": 1})
        self.assertEqual(accruals.coverage_by_kind({"1456 W 85th St, Cleveland, OH 44102|pm"}), {"pm": 1})

    def test_expected_fixed_accrual_coverage_excludes_sold_policy_properties(self):
        keys = accruals.expected_fixed_accrual_keys(property_filters=["9919 S Oglesby"])

        self.assertEqual(keys, set())

    def test_cash_basis_insurance_states_are_not_generated_as_accruals(self):
        rows = []
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]
        generated = accruals.generate_missing_accruals(
            rows,
            "2026-06",
            fieldnames,
            property_filters=["917 Pawnee Ave"],
            kind_filters={"insurance"},
        )

        self.assertEqual([], generated)
        self.assertTrue(accruals.is_cash_basis_insurance_accrual("917 Pawnee Ave, Memphis, TN 38109", "insurance"))
        self.assertFalse(accruals.is_cash_basis_insurance_accrual("326-332 S Alcott St, Denver, CO 80219", "insurance"))

    def test_cash_basis_insurance_states_have_no_fixed_insurance_templates(self):
        offenders = [
            template["property"]
            for template in accruals.ACCRUAL_TEMPLATES
            if template["kind"] == "insurance"
            and accruals.property_state(str(template["property"])) in accruals.CASH_BASIS_INSURANCE_STATES
        ]

        self.assertEqual([], offenders)

    def test_named_no_dao_mortgage_properties_skip_tax_and_insurance_accruals(self):
        fieldnames = ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"]
        original_templates = accruals.ACCRUAL_TEMPLATES
        try:
            accruals.ACCRUAL_TEMPLATES = [
                {
                    "property": "724 3rd Ave",
                    "kind": "taxes",
                    "amount": 572.08,
                    "category": "City, State, & Local Taxes",
                    "sub_category": "",
                    "source_prefix": "AOPS-PNL-ACCRUAL",
                    "description": "Tax Accrual | {property} | {month_label}",
                    "merchant": "Tax Accrual | {property} | {month_label}",
                },
                {
                    "property": "724 3rd Ave",
                    "kind": "insurance",
                    "amount": 68.13,
                    "category": "Rental Dwelling",
                    "sub_category": "",
                    "source_prefix": "AOPS-PNL-ACCRUAL",
                    "description": "Insurance Accrual | {property} | {month_label}",
                    "merchant": "Insurance Accrual | {property} | {month_label}",
                },
            ]
            self.assertEqual(
                accruals.generate_missing_accruals(
                    [],
                    "2026-06",
                    fieldnames,
                    property_filters=["724 3rd Ave"],
                    kind_filters={"taxes", "insurance"},
                ),
                [],
            )
            self.assertEqual(
                accruals.expected_fixed_accrual_keys(
                    kind_filters={"taxes", "insurance"},
                    property_filters=["724 3rd Ave"],
                ),
                set(),
            )
        finally:
            accruals.ACCRUAL_TEMPLATES = original_templates

    def test_existing_724_manual_tax_accrual_is_zeroed(self):
        rows = [{
            "Amount": "-572.08",
            "Notes": "AOPS-PNL-ACCRUAL|taxes|724 3rd Ave|2026-06|572.08 | Monthly accrual entry.",
        }]

        mismatches = accruals.find_amount_mismatches(
            rows,
            "2026-06",
            property_filters=["724 3rd Ave"],
            kind_filters={"taxes"},
        )

        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0]["expected_amount"], 0.0)

    def test_coverage_key_details_names_missing_property_kind_month(self):
        details = accruals.coverage_key_details(
            {"1456 W 85th St, Cleveland, OH 44102|pm"},
            "2026-06",
        )

        self.assertEqual(
            details,
            [
                {
                    "property": "1456 W 85th St, Cleveland, OH 44102",
                    "kind": "pm",
                    "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                    "month": "2026-06",
                }
            ],
        )

    def test_pm_fee_basis_gap_accepts_zero_when_no_prior_rent_exists(self):
        gaps = accruals.pm_fee_basis_gaps(
            [],
            "2026-06",
            set(),
            property_filters=["1456 W 85th St"],
        )

        self.assertEqual(gaps, [])

    def test_pm_fee_basis_gap_accepts_hemlane_net_rent_without_second_pm_accrual(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        try:
            accruals.PM_FEE_PROPERTIES = [
                ("724 3rd Ave", 0.06, "AOPS-PNL-ACCRUAL"),
            ]
            rows = [
                {
                    "Date": "June 06, 2026",
                    "Merchant": "Hemlane",
                    "Description": "Hemlane rent remittance",
                    "Amount": "997.50",
                    "Type": "Revenue",
                    "Category": "Long Term Rents",
                    "Property": "724 3rd Ave",
                    "Notes": "",
                },
                {
                    "Date": "July 06, 2026",
                    "Merchant": "Hemlane",
                    "Description": "Hemlane rent remittance",
                    "Amount": "997.50",
                    "Type": "Revenue",
                    "Category": "Long Term Rents",
                    "Property": "724 3rd Ave",
                    "Notes": "",
                },
            ]

            gaps = accruals.pm_fee_basis_gaps(rows, "2026-07", set())

            self.assertEqual(gaps, [])
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm

    def test_pm_fee_basis_gap_accepts_explicit_hemlane_zero_collection(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        original_basis = accruals.HEMLANE_PM_FEE_BASIS
        try:
            property_name = "917 Pawnee Ave, Memphis, TN 38109"
            accruals.PM_FEE_PROPERTIES = [
                (property_name, 0.10, "AOPS-OHIL-ACCRUAL"),
            ]
            accruals.HEMLANE_PM_FEE_BASIS = {
                property_name: {
                    "month": "2026-07",
                    "rent_request_amount": 1145.0,
                    "rent_success_amount": 0.0,
                    "rent_success_evidence_count": 1,
                    "rent_transaction_ids": ["rent-pawnee-july"],
                }
            }
            rows = [
                {
                    "Date": "June 08, 2026",
                    "Merchant": "REMOTE ONLINE DEPOSIT # 1",
                    "Amount": "2335.00",
                    "Type": "Revenue",
                    "Category": "Long Term Rents",
                    "Property": "917 Pawnee Ave",
                    "Notes": "",
                }
            ]

            gaps = accruals.pm_fee_basis_gaps(rows, "2026-07", set())

            self.assertEqual(gaps, [])
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.HEMLANE_PM_FEE_BASIS = original_basis

    def test_zero_pm_fee_generates_local_coverage_marker_without_cash_effect(self):
        generated = accruals.generate_missing_accruals(
            [],
            "2026-06",
            ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
            property_filters=["1456 W 85th St"],
            kind_filters={"pm"},
        )

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0]["Amount"], "0.00")
        self.assertIn("Zero-dollar coverage marker only", generated[0]["Notes"])

    def test_zero_pm_fee_marker_never_uses_negative_zero_amount(self):
        original_pm = list(accruals.PM_FEE_PROPERTIES)
        original_components = copy.deepcopy(accruals.PM_FEE_COMPONENTS)
        try:
            accruals.PM_FEE_PROPERTIES = [("Deducted PM Property", 0.10, accruals.SCHEDULE_SOURCE_PREFIX)]
            accruals.PM_FEE_COMPONENTS = {"Deducted PM Property": [{"rate": 0.10, "deduction": 57.0, "basis_token": None}]}

            generated = accruals.generate_missing_accruals(
                [],
                "2026-06",
                ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
                property_filters=["Deducted PM Property"],
                kind_filters={"pm"},
            )

            self.assertEqual(1, len(generated))
            self.assertEqual(generated[0]["Amount"], "0.00")
            self.assertNotEqual(generated[0]["Amount"], "-0.00")
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.PM_FEE_COMPONENTS = original_components

    def test_accrual_gap_action_queue_blocks_pm_basis_and_missing_fixed(self):
        queue = accruals.accrual_gap_action_queue(
            [
                {
                    "property": "1456 W 85th St, Cleveland, OH 44102",
                    "kind": "pm",
                    "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                    "month": "2026-06",
                },
                {
                    "property": "1456 W 85th St, Cleveland, OH 44102",
                    "kind": "taxes",
                    "key": "1456 W 85th St, Cleveland, OH 44102|taxes",
                    "month": "2026-06",
                },
            ],
            [
                {
                    "property": "1456 W 85th St, Cleveland, OH 44102",
                    "kind": "pm",
                    "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                    "month": "2026-06",
                    "reason": "no_current_month_revenue_basis_for_pm_fee",
                    "review_action": "Verify rent basis",
                    "current_month_gross_rent": 0.0,
                    "previous_month": "2026-05",
                    "previous_month_gross_rent": 1794.93,
                }
            ],
            [{"property_name": "804 S Quitman St, Denver", "reason": "development property"}],
        )

        self.assertEqual(
            [item["action"] for item in queue],
            [
                "verify_missing_rent_or_approve_zero_pm",
                "generate_or_verify_missing_fixed_accrual",
                "document_or_add_fixed_accrual_requirement",
            ],
        )
        self.assertEqual(queue[0]["severity"], "blocks_monthly_close")
        self.assertEqual(queue[1]["kind"], "taxes")
        self.assertEqual(queue[2]["severity"], "audit_only")
        self.assertEqual(accruals.coverage_by_kind({item["key"] for item in queue[:2]}), {"pm": 1, "taxes": 1})

    def test_pm_fee_basis_gap_marks_missing_aligned_detail_when_net_transfer_exists(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        try:
            accruals.PM_FEE_PROPERTIES = [("1456 W 85th St, Cleveland, OH 44102", 0.10, "AOPS-OHIL-ACCRUAL")]
            rows = [
                {
                    "Date": "May 05, 2026",
                    "Merchant": "Angela Heath",
                    "Description": "Rent Income - May 2026",
                    "Amount": "1794.93",
                    "Type": "Revenue",
                    "Category": "Long Term Rents",
                    "Property": "1456 W 85th St.",
                    "Notes": "",
                },
                {
                    "Date": "June 15, 2026",
                    "Merchant": "Aligned Properti",
                    "Description": "Aligned Properti | SIGONFILE",
                    "Amount": "1543.77",
                    "Type": "Transfers & Other",
                    "Category": "Transfers Between Accounts",
                    "Property": "1456 W 85th St.",
                    "Notes": "",
                },
            ]

            gaps = accruals.pm_fee_basis_gaps(rows, "2026-06", set())

            self.assertEqual(len(gaps), 1)
            self.assertEqual(gaps[0]["reason"], "aligned_owner_statement_detail_missing_for_pm_fee")
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm

    def test_aligned_management_fee_rows_offset_pm_accrual_by_service_month(self):
        rows = [
            {
                "Date": "June 04, 2026",
                "Merchant": "Tenant",
                "Description": "Rent Income - June 2026",
                "Amount": "1000.00",
                "Type": "Revenue",
                "Category": "Long Term Rents",
                "Property": "1456 W 85th St.",
                "Notes": "",
            },
            {
                "Date": "July 01, 2026",
                "Merchant": "Aligned Properties",
                "Description": "Management Fees for 06/2026",
                "Amount": "-100.00",
                "Type": "Operating Expenses",
                "Category": "Management Fees",
                "Property": "1456 W 85th St.",
                "Notes": "Aligned clearing detail import | key=aligned-pm-fee",
            },
        ]

        self.assertEqual(accruals.infer_service_month(rows[1]), "2026-06")
        self.assertEqual(
            accruals.separately_booked_eco_pm_fee_payment(
                rows,
                "1456 W 85th St, Cleveland, OH 44102",
                "2026-06",
            ),
            100.0,
        )

    def test_explicit_pm_transfer_offsets_accrual_without_changing_cashflow_category(self):
        rows = [
            {
                "Date": "March 16, 2026",
                "Merchant": "Transfer to ECO Systems LLC",
                "Description": "Internal account transfer",
                "Amount": "-1338.63",
                "Type": "Transfers & Other",
                "Category": "Transfers Between Accounts",
                "Property": "88 Madison Ave",
                "Notes": "September 2025 PM fee",
            },
            {
                "Date": "March 16, 2026",
                "Merchant": "Transfer to ECO Systems LLC",
                "Description": "Internal account transfer",
                "Amount": "-500.00",
                "Type": "Transfers & Other",
                "Category": "Transfers Between Accounts",
                "Property": "88 Madison Ave",
                "Notes": "September 2025 owner distribution",
            },
        ]

        self.assertEqual(accruals.infer_service_month(rows[0]), "2025-09")
        self.assertEqual(
            accruals.separately_booked_eco_pm_fee_payment(
                rows,
                "88 Madison Ave",
                "2025-09",
            ),
            1338.63,
        )
        self.assertEqual(
            accruals.separately_booked_eco_pm_fee_payment_date(
                rows,
                "88 Madison Ave",
                "2025-09",
            ),
            "March 16, 2026",
        )

    def test_pending_hemlane_request_can_supply_pm_fee_accrual_basis(self):
        original_pm = accruals.PM_FEE_PROPERTIES
        original_basis = accruals.HEMLANE_PM_FEE_BASIS
        original_exceptions = set(accruals.PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS)
        try:
            accruals.PM_FEE_PROPERTIES = [("9634 S Green St, Chicago, IL 60643", 0.10, "AOPS-OHIL-ACCRUAL")]
            accruals.PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS.add(
                ("9634 S Green St, Chicago, IL 60643", "2026-06")
            )
            with tempfile.TemporaryDirectory() as tmpdir:
                report_path = Path(tmpdir) / "hemlane_live_transactions.json"
                report_path.write_text(
                    json.dumps(
                        {
                            "status": "ok",
                            "transactions": [
                                {
                                    "id": "rent-9634",
                                    "request_amount": 2000,
                                    "due_date": "2026-06-01",
                                    "property": "9634 South Green Street",
                                    "property_address": "9634 South Green Street, Chicago, IL 60643",
                                    "payment_category": "Rental Income",
                                    "payment_subcategory": "Rent",
                                },
                                {
                                    "id": "pm-9634",
                                    "request_amount": 200,
                                    "due_date": "2026-06-01",
                                    "property": "9634 South Green Street",
                                    "property_address": "9634 South Green Street, Chicago, IL 60643",
                                    "payment_category": "Leasing & Management Fees",
                                    "payment_subcategory": "Property Management Fees",
                                },
                                {
                                    "id": "late-fee-9634",
                                    "request_amount": 85,
                                    "due_date": "2026-06-05",
                                    "property": "9634 South Green Street",
                                    "property_address": "9634 South Green Street, Chicago, IL 60643",
                                    "payment_category": "Non-Rental Income",
                                    "payment_subcategory": "Late Fees",
                                },
                            ],
                        }
                    ),
                    encoding="utf-8",
                )
                basis, report = accruals.load_hemlane_pm_fee_basis(report_path, "2026-06")
            self.assertEqual(report["status"], "ok")
            self.assertEqual(basis["9634 S Green St, Chicago, IL 60643"]["rent_request_amount"], 2000.0)
            self.assertEqual(basis["9634 S Green St, Chicago, IL 60643"]["rent_success_evidence_count"], 0)
            self.assertEqual(basis["9634 S Green St, Chicago, IL 60643"]["pm_fee_request_amount"], 200.0)
            accruals.HEMLANE_PM_FEE_BASIS = basis

            pm_fees = accruals.compute_pm_fees([], "2026-06")

            self.assertEqual(pm_fees["9634 S Green St, Chicago, IL 60643"], 200.0)
            self.assertIn(
                "Hemlane live request basis",
                accruals.pm_fee_rule_summary("9634 S Green St, Chicago, IL 60643", "2026-06", 0.10, []),
            )
        finally:
            accruals.PM_FEE_PROPERTIES = original_pm
            accruals.HEMLANE_PM_FEE_BASIS = original_basis
            accruals.PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS.clear()
            accruals.PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS.update(original_exceptions)

    def test_build_review_markdown_lists_blocking_accrual_actions(self):
        markdown = accruals.build_review_markdown(
            {
                "generated_at": "2026-07-13T00:00:00Z",
                "target_month": "2026-06",
                "status": "review",
                "missing_fixed_accrual_coverage_count": 5,
                "pm_fee_basis_gap_count": 5,
                "blocking_gap_action_count": 1,
                "gap_approval_scaffold": {"path": "config/baselane_monthly_accrual_gap_approvals.json"},
                "gap_approval_review_csv": {"path": "reports/baselane_monthly_accrual_gap_approvals_review.csv"},
                "gap_approval_import_commands": {
                    "path": "reports/baselane_monthly_accrual_gap_approvals_import.requires-explicit-approval.sh"
                },
                "gap_action_queue": [
                    {
                        "property": "1456 W 85th St, Cleveland, OH 44102",
                        "kind": "pm",
                        "severity": "blocks_monthly_close",
                        "action": "verify_missing_rent_or_approve_zero_pm",
                        "current_month_gross_rent": 0,
                        "previous_month_gross_rent": 1794.93,
                        "review_action": "Verify rent basis",
                    }
                ],
            }
        )

        self.assertIn("# Monthly Accrual Completeness Review", markdown)
        self.assertIn("verify_missing_rent_or_approve_zero_pm", markdown)
        self.assertIn("approve_zero_pm_fee", markdown)
        self.assertIn("$1,794.93", markdown)
        self.assertIn("Evidence Digest", markdown)
        self.assertIn("Current PM", markdown)
        self.assertIn("evidence_digest", markdown)
        self.assertIn("substantive `note`", markdown)
        self.assertIn("## Approval Artifacts", markdown)
        self.assertIn("config/baselane_monthly_accrual_gap_approvals.json", markdown)
        self.assertIn("reports/baselane_monthly_accrual_gap_approvals_review.csv", markdown)
        self.assertIn("bash reports/baselane_monthly_accrual_gap_approvals_import.requires-explicit-approval.sh", markdown)

    def test_gap_approvals_make_reviewed_zero_pm_nonblocking(self):
        queue = [
            {
                "property": "1456 W 85th St, Cleveland, OH 44102",
                "kind": "pm",
                "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
            }
        ]

        approval = {
            "key": "1456 W 85th St, Cleveland, OH 44102|pm",
            "decision": "approve_zero_pm_fee",
            "reviewed": True,
            "reviewed_at": "2026-07-13T00:00:00Z",
            "note": "No rent collected in June.",
            "evidence_digest": accruals.gap_approval_evidence_digest(queue[0]),
        }

        validation = accruals.apply_gap_approvals(
            queue,
            {
                "status": "loaded",
                "path": "/tmp/approvals.json",
                "approvals": [approval],
            },
        )

        self.assertEqual(validation["status"], "ok")
        self.assertEqual(validation["approved_count"], 1)
        self.assertEqual(queue[0]["severity"], "reviewed_nonblocking")
        self.assertEqual(queue[0]["approval_decision"], "approve_zero_pm_fee")
        self.assertEqual(accruals.blocking_gap_actions(queue), [])

    def test_gap_approvals_reject_stale_or_unreviewed_decisions(self):
        queue = [
            {
                "key": "current|pm",
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
            }
        ]

        validation = accruals.apply_gap_approvals(
            queue,
            {
                "status": "loaded",
                "approvals": [
                    {"key": "current|pm", "decision": "approve_zero_pm_fee", "reviewed": False},
                    {
                        "key": "stale|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": True,
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "evidence_digest": "stale",
                        "note": "Checked all rent systems and confirmed no current-month rent.",
                    },
                ],
            },
        )

        self.assertEqual(validation["status"], "review")
        self.assertEqual(
            {issue["code"] for issue in validation["issues"]},
            {"approval_not_reviewed"},
        )
        self.assertEqual(validation["ignored_stale_approval_count"], 1)
        self.assertEqual(validation["ignored_stale_approval_keys"], ["stale|pm"])

    def test_gap_approvals_ignore_stale_keys_when_no_current_gap_remains(self):
        validation = accruals.apply_gap_approvals(
            [],
            {
                "status": "loaded",
                "approvals": [
                    {
                        "key": "old|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": True,
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "evidence_digest": "old",
                        "note": "Previously reviewed zero-rent PM gap that is no longer present.",
                    }
                ],
            },
        )

        self.assertEqual(validation["status"], "ok")
        self.assertEqual(validation["issue_count"], 0)
        self.assertEqual(validation["ignored_stale_approval_count"], 1)
        self.assertEqual(validation["ignored_stale_approval_keys"], ["old|pm"])


    def test_gap_approvals_reject_stale_evidence_digest(self):
        queue = [
            {
                "key": "current|pm",
                "property": "Current",
                "kind": "pm",
                "month": "2026-06",
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
                "current_month_gross_rent": 0,
                "previous_month": "2026-05",
                "previous_month_gross_rent": 1000,
            }
        ]

        validation = accruals.apply_gap_approvals(
            queue,
            {
                "status": "loaded",
                "approvals": [
                    {
                        "key": "current|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": True,
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "evidence_digest": "not-current",
                        "note": "Checked all rent systems and confirmed no current-month rent.",
                    }
                ],
            },
        )

        self.assertEqual(validation["status"], "review")
        self.assertEqual(validation["issues"][0]["code"], "approval_evidence_digest_mismatch")
        self.assertEqual(queue[0]["severity"], "blocks_monthly_close")

    def test_gap_approvals_reject_reviewed_without_substantive_note(self):
        queue = [
            {
                "key": "current|pm",
                "property": "Current",
                "kind": "pm",
                "month": "2026-06",
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
                "current_month_gross_rent": 0,
                "previous_month": "2026-05",
                "previous_month_gross_rent": 1000,
            }
        ]

        validation = accruals.apply_gap_approvals(
            queue,
            {
                "status": "loaded",
                "approvals": [
                    {
                        "key": "current|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": True,
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "evidence_digest": accruals.gap_approval_evidence_digest(queue[0]),
                        "note": "ok",
                    }
                ],
            },
        )

        self.assertEqual(validation["status"], "review")
        self.assertEqual(validation["issues"][0]["code"], "approval_missing_substantive_note")
        self.assertEqual(queue[0]["severity"], "blocks_monthly_close")

    def test_gap_approval_scaffold_preserves_reviewed_approvals(self):
        queue = [
            {
                "property": "1456 W 85th St, Cleveland, OH 44102",
                "kind": "pm",
                "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                "month": "2026-06",
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
                "current_month_gross_rent": 0,
                "previous_month": "2026-05",
                "previous_month_gross_rent": 1794.93,
            }
        ]

        scaffold = accruals.build_gap_approval_scaffold(
            queue,
            {
                "approvals": [
                    {
                        "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": True,
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "note": "confirmed zero rent",
                    }
                ]
            },
        )

        self.assertEqual(scaffold["approval_count"], 1)
        self.assertFalse(scaffold["approvals"][0]["reviewed"])
        self.assertIn("evidence_digest", scaffold["approvals"][0])
        self.assertIn("required_review_evidence", scaffold["approvals"][0])
        self.assertIn("review_note_template", scaffold["approvals"][0])
        self.assertEqual(scaffold["approvals"][0]["current_implied_pm_fee_amount"], 0.0)
        self.assertEqual(scaffold["approvals"][0]["previous_month_implied_pm_fee_amount"], 179.49)

    def test_write_gap_approval_scaffold_creates_review_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "approvals.json"
            result = accruals.write_gap_approval_scaffold(
                path,
                [
                    {
                        "property": "1456 W 85th St, Cleveland, OH 44102",
                        "kind": "pm",
                        "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                        "month": "2026-06",
                        "action": "verify_missing_rent_or_approve_zero_pm",
                    }
                ],
            )

            payload = json.loads(path.read_text(encoding="utf-8"))
            self.assertEqual(result["status"], "written")
            self.assertEqual(payload["approval_count"], 1)
            self.assertFalse(payload["approvals"][0]["reviewed"])
            self.assertIn("evidence_digest", payload["approvals"][0])
            self.assertIn("required_review_evidence", payload["approvals"][0])
            self.assertEqual(payload["approvals"][0]["approval_effect"], "waive_pm_accrual_for_month_only_no_cash_transfer")

    def test_gap_approval_review_csv_and_import_command_are_written(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            scaffold = accruals.build_gap_approval_scaffold(
                [
                    {
                        "property": "1456 W 85th St, Cleveland, OH 44102",
                        "kind": "pm",
                        "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                        "month": "2026-06",
                        "action": "verify_missing_rent_or_approve_zero_pm",
                        "current_month_gross_rent": 0,
                        "previous_month": "2026-05",
                        "previous_month_gross_rent": 1794.93,
                    }
                ]
            )
            csv_path = root / "review.csv"
            command_path = root / "import.sh"

            csv_result = accruals.write_gap_approval_review_csv(csv_path, scaffold)
            command_result = accruals.write_gap_approval_import_commands(command_path)

            self.assertEqual(csv_result["approval_count"], 1)
            self.assertIn("key,property,kind,month,decision,reviewed,reviewed_at,note", csv_path.read_text(encoding="utf-8"))
            command_text = command_path.read_text(encoding="utf-8")
            self.assertIn("BASELANE_MONTHLY_ACCRUALS_IMPORT_GAP_APPROVAL_CSV=1", command_text)
            self.assertIn("missing_fixed_accrual_coverage_count=", command_text)
            self.assertIn("gap_approval_import_issue=", command_text)
            self.assertIn("gap_action=", command_text)
            self.assertIn("${STATUS%%$'\\n'*}", command_text)
            self.assertEqual(command_result["status"], "written")

    def test_import_gap_approval_review_csv_updates_reviewed_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            approvals_path = root / "approvals.json"
            csv_path = root / "review.csv"
            queue = [
                {
                    "property": "1456 W 85th St, Cleveland, OH 44102",
                    "kind": "pm",
                    "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                    "month": "2026-06",
                    "action": "verify_missing_rent_or_approve_zero_pm",
                    "current_month_gross_rent": 0,
                    "previous_month": "2026-05",
                    "previous_month_gross_rent": 1794.93,
                }
            ]
            approvals_path.write_text(json.dumps(accruals.build_gap_approval_scaffold(queue)) + "\n", encoding="utf-8")
            digest = accruals.gap_approval_evidence_digest(queue[0])
            with csv_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=["key", "decision", "reviewed", "reviewed_at", "note", "evidence_digest"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                        "decision": "approve_zero_pm_fee",
                        "reviewed": "true",
                        "reviewed_at": "2026-07-13T00:00:00Z",
                        "note": "Checked Baselane and PM portal; no rent rows were missing or mistagged",
                        "evidence_digest": digest,
                    }
                )

            result = accruals.import_gap_approval_review_csv(csv_path, approvals_path)
            payload = json.loads(approvals_path.read_text(encoding="utf-8"))

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["imported_count"], 1)
            self.assertTrue(payload["approvals"][0]["reviewed"])

    def test_write_gap_approval_scaffold_recovers_placeholder_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "approvals.json"
            path.write_text(json.dumps({"status": "review_required"}) + "\n", encoding="utf-8")

            result = accruals.write_gap_approval_scaffold(
                path,
                [
                    {
                        "property": "1456 W 85th St, Cleveland, OH 44102",
                        "kind": "pm",
                        "key": "1456 W 85th St, Cleveland, OH 44102|pm",
                        "month": "2026-06",
                        "action": "verify_missing_rent_or_approve_zero_pm",
                    }
                ],
            )

            payload = json.loads(path.read_text(encoding="utf-8"))
            self.assertTrue(result["existing_recovered"])
            self.assertEqual(result["existing_loaded_status"], "invalid")
            self.assertEqual(payload["approval_count"], 1)
            self.assertEqual(payload["approvals"][0]["key"], "1456 W 85th St, Cleveland, OH 44102|pm")

    def test_active_property_without_accrual_template_is_reported(self):
        active = [
            {"property_name": "326-332 S Alcott", "full_address": "326-332 S Alcott St, Denver, CO 80219"},
            {"property_name": "49 Bannbury Ln, Palm Coast", "full_address": "49 Bannbury Ln, Palm Coast, FL 32137"},
            {"property_name": "Unmapped DAO", "full_address": "1 Missing Template Rd, Test, TX 75001"},
        ]

        missing = accruals.find_active_properties_without_accrual_templates(active)
        non_managed = accruals.active_properties_without_fixed_accrual_requirement(active)

        self.assertEqual([active[2]], missing)
        self.assertEqual("49 Bannbury Ln, Palm Coast", non_managed[0]["property_name"])


if __name__ == "__main__":
    unittest.main()
