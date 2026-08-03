#!/usr/bin/env python3
"""Guarded updater for CF balance-sheet cash rows.

This updates literal current-month balance-sheet values from the monthly
financial summary packet:
- Lofty Operating Cash from Lofty ``curr_maintenance_reserve``
- ECO General Ledger from the complete per-property ECO GL Column E sum
- ECO Net DAO Funds from verified nonnegative spendable cash in ECO custody

ECO Net DAO Funds is never sourced from the general ledger. It is updated only
when the explicit ``--apply-eco-net-dao-funds`` gate is enabled. Cash-settlement
basis is also kept as a separate transfer-review measure.

The script is dry-run by default and refuses workbook writes while known source
cleanup blockers are present unless explicitly overridden.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).absolute().parents[1]
AUDIT_SCRIPT = ROOT / "scripts/baselane_cf_balance_sheet_consistency_audit.py"
DEFAULT_CANDIDATE_PACKET = ROOT / "reports/baselane_financials_monthly_review_candidate_packet.json"
DEFAULT_SOURCE_CLEANUP_QUEUE = ROOT / "reports/baselane_source_cleanup_queue.json"
DEFAULT_SOURCE_CASH_REPORT = ROOT / "reports/baselane_daily_source_cash_balance_report.json"
DEFAULT_DATA_QUALITY_REPORT = ROOT / "reports/baselane_ecogl_data_quality_autonomy.json"
DEFAULT_YHOME_CSV = ROOT / "reports/yhome_transition_reconciliation.csv"
DEFAULT_REPORT = ROOT / "reports/baselane_cf_balance_sheet_cash_apply_report.json"
LOFTY_LABEL = "Lofty Operating Reserve (OR) Balance"
LEGACY_ECO_LABEL = "ECO GL Net Cash Balance (excl. EARLDAO Interest)"
CONFLICT_THRESHOLD = 0.01


try:
    from openpyxl import load_workbook
except Exception as exc:  # noqa: BLE001
    load_workbook = None
    OPENPYXL_IMPORT_ERROR = f"{type(exc).__name__}: {exc}"
else:
    OPENPYXL_IMPORT_ERROR = None


def load_audit_module():
    spec = importlib.util.spec_from_file_location("baselane_cf_balance_sheet_consistency_audit", AUDIT_SCRIPT)
    if not spec or not spec.loader:
        raise RuntimeError(f"Cannot load audit helpers: {AUDIT_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def iso_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {"status": "missing", "path": str(path)}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return {"status": "unreadable", "path": str(path), "error": f"{type(exc).__name__}: {exc}"}
    return payload if isinstance(payload, dict) else {"status": "unreadable", "path": str(path), "error": "root is not object"}


def parse_money(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return round(-number if negative else number, 2)


def parse_month(value: str) -> tuple[int, int]:
    year, month = value.split("-", 1)
    return int(year), int(month)


def candidate_records(path: Path) -> list[dict[str, Any]]:
    payload = read_json(path)
    records = payload.get("records")
    return [record for record in records if isinstance(record, dict)] if isinstance(records, list) else []


def source_cash_correction_records(
    path: Path,
    existing_records: list[dict[str, Any]],
    audit: Any,
) -> list[dict[str, Any]]:
    payload = read_json(path)
    normalize = getattr(audit, "normalize_property_name", None)
    existing_keys = {
        normalize(str(record.get("property_name") or record.get("input_property_name") or "").strip())
        if callable(normalize)
        else str(record.get("property_name") or record.get("input_property_name") or "").strip().casefold()
        for record in existing_records
    }
    corrections: list[dict[str, Any]] = []
    for violation in payload.get("violations_bounded") or []:
        if not isinstance(violation, dict) or violation.get("action") != "replace_with_source_gl_cumulative_balance":
            continue
        property_name = str(violation.get("property") or "").strip()
        expected = parse_money(violation.get("expected"))
        workbook = str(violation.get("file") or "").strip()
        if not property_name or expected is None or not workbook:
            continue
        key = normalize(property_name) if callable(normalize) else property_name.casefold()
        if key in existing_keys:
            continue
        corrections.append(
            {
                "property_name": property_name,
                "property_path": str(Path(workbook).parent),
                "monthly_financial_summary": {
                    "eco_operating_cash": expected,
                    "eco_general_ledger_sum": expected,
                    "eco_gl_column_e_sum": expected,
                },
                "source_cash_correction": True,
            }
        )
        existing_keys.add(key)
    return corrections


def filter_excluded_candidate_records(
    audit: Any,
    records: list[dict[str, Any]],
    yhome_csv: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    guards, _yhome_guard, _manual_exclusions = (
        audit.monthly_exclusion_guards(yhome_csv)
        if callable(getattr(audit, "monthly_exclusion_guards", None))
        else ([], {}, [])
    )
    included: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for record in records:
        property_name = str(record.get("property_name") or record.get("input_property_name") or "").strip()
        property_path = str(record.get("property_path") or record.get("input_property_path") or "").strip()
        manual_match = bool(
            callable(getattr(audit, "is_manually_excluded_property", None))
            and audit.is_manually_excluded_property(property_name)
        )
        guard = (
            audit.exclusion_guard_for_record(record, guards)
            if callable(getattr(audit, "exclusion_guard_for_record", None))
            else None
        )
        if not manual_match and not guard:
            included.append(record)
            continue
        excluded.append(
            {
                "property": property_name,
                "property_path": property_path,
                "source": "manual_exclusion" if manual_match else guard.get("source"),
                "reason": "manual do-not-update exclusion" if manual_match else guard.get("exclude_reason"),
            }
        )
    return included, excluded


def source_blockers(
    cleanup_path: Path,
    source_cash_path: Path,
    data_quality_path: Path | None = None,
) -> list[str]:
    cleanup = read_json(cleanup_path)
    source_cash = read_json(source_cash_path)
    blockers = []
    if int(cleanup.get("action_count") or 0) > 0:
        blockers.append(f"source_cleanup_queue_actions={int(cleanup.get('action_count') or 0)}")
    if int(cleanup.get("missing_id_count") or 0) > 0:
        blockers.append(f"source_cleanup_queue_missing_ids={int(cleanup.get('missing_id_count') or 0)}")
    if bool(source_cash.get("apply_blocked_by_raw_no_dao_mortgage_guard")):
        blockers.append("source_cash_apply_blocked_by_raw_no_dao_mortgage_guard")
    if bool(source_cash.get("apply_blocked_by_source_ledger_quality_guard")):
        blockers.append("source_cash_apply_blocked_by_source_ledger_quality_guard")
    if data_quality_path is not None:
        data_quality = read_json(data_quality_path)
        count = int(data_quality.get("exception_count") or 0)
        if data_quality.get("status") == "unreadable":
            blockers.append("source_ledger_quality_report_unreadable")
        elif count > 0:
            blockers.append(f"source_ledger_quality_exceptions={count}")
    if int(source_cash.get("violation_count") or 0) > 0:
        blockers.append(f"source_cash_balance_violations={int(source_cash.get('violation_count') or 0)}")
    return blockers


def downstream_balance_correction_gate(
    cleanup_path: Path,
    source_cash_path: Path,
    data_quality_path: Path | None = None,
) -> dict[str, Any]:
    source_cash = read_json(source_cash_path)
    blockers = source_blockers(cleanup_path, source_cash_path, data_quality_path)
    violation_count = int(source_cash.get("violation_count") or 0)
    structural_fields = (
        "missing_row_count",
        "missing_month_column_count",
        "unreadable_count",
        "noncanonical_source_count",
        "blocking_no_match_count",
        "split_scope_missing_property_count",
    )
    structural_blockers = {
        field: int(source_cash.get(field) or 0)
        for field in structural_fields
        if int(source_cash.get(field) or 0) > 0
    }
    allowed = bool(
        str(source_cash.get("status") or "") in {"ok", "review"}
        and violation_count > 0
        and not structural_blockers
        and not source_cash.get("apply_blocked_by_raw_no_dao_mortgage_guard")
        and not source_cash.get("apply_blocked_by_source_ledger_quality_guard")
        and blockers == [f"source_cash_balance_violations={violation_count}"]
    )
    return {
        "allowed": allowed,
        "reason": "canonical_source_only_downstream_balance_correction"
        if allowed
        else "source_cash_correction_gate_not_clean",
        "status": source_cash.get("status"),
        "violation_count": violation_count,
        "structural_blockers": structural_blockers,
        "source_blockers": blockers,
    }


def canonical_workbook_map(audit: Any, source_cash_path: Path) -> dict[str, Path]:
    source_cash = read_json(source_cash_path)
    normalize = getattr(audit, "normalize_property_name", None)
    mapping: dict[str, Path] = {}
    for item in source_cash.get("checked_workbooks_bounded") or []:
        if not isinstance(item, dict):
            continue
        property_name = str(item.get("property") or "").strip()
        workbook = Path(str(item.get("file") or ""))
        if not property_name or not workbook.is_file():
            continue
        key = normalize(property_name) if callable(normalize) else property_name.casefold()
        mapping.setdefault(key, workbook)
    return mapping


def sheet_for_year(workbook: Any, year: int):
    if str(year) in workbook.sheetnames:
        return workbook[str(year)]
    for sheet in workbook.worksheets:
        if str(year) in str(sheet.title):
            return sheet
    return None


def month_column(audit: Any, sheet: Any, year: int, month: int) -> int | None:
    for row in (1, 4):
        for column in range(2, 41):
            if audit.parse_month_header(sheet.cell(row=row, column=column).value) == (year, month):
                return column
    return None


def find_or_create_row(sheet: Any, labels: tuple[str, ...], create_missing: bool) -> tuple[int | None, bool]:
    max_row = sheet.max_row or 0
    for wanted in (label.strip() for label in labels):
        for row in range(1, max_row + 1):
            if str(sheet.cell(row=row, column=1).value or "").strip() == wanted:
                return row, False
    if not create_missing:
        return None, False
    row = max_row + 1
    sheet.cell(row=row, column=1).value = labels[0]
    return row, True


def clear_legacy_eco_cash_rows(
    sheet: Any,
    *,
    canonical_row: int | None,
    property_name: str,
    apply: bool,
    legacy_labels: tuple[str, ...] = (LEGACY_ECO_LABEL,),
) -> list[dict[str, Any]]:
    if canonical_row is None:
        return []
    changes = []
    for row in range(1, (sheet.max_row or 0) + 1):
        if row == canonical_row:
            continue
        if str(sheet.cell(row=row, column=1).value or "").strip() not in set(legacy_labels):
            continue
        populated_cells = [
            sheet.cell(row=row, column=column).coordinate
            for column in range(1, (sheet.max_column or 1) + 1)
            if sheet.cell(row=row, column=column).value is not None
        ]
        changes.append(
            {
                "property": property_name,
                "source": "ECO General Ledger",
                "action": "clear_legacy_eco_cash_row",
                "label": LEGACY_ECO_LABEL,
                "row": row,
                "populated_cells": populated_cells,
            }
        )
        if apply:
            for column in range(1, (sheet.max_column or 1) + 1):
                sheet.cell(row=row, column=column).value = None
    return changes


def update_workbook(
    audit: Any,
    record: dict[str, Any],
    *,
    year: int,
    month: int,
    apply: bool,
    create_missing_rows: bool,
    apply_eco_net_dao_funds: bool,
    canonical_workbooks: dict[str, Path] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    property_name = str(record.get("property_name") or record.get("input_property_name") or "").strip()
    property_path = Path(str(record.get("property_path") or record.get("input_property_path") or ""))
    summary = (
        audit.candidate_source(record)
        if hasattr(audit, "candidate_source")
        else record.get("monthly_financial_summary")
        if isinstance(record.get("monthly_financial_summary"), dict)
        else {}
    )
    normalize = getattr(audit, "normalize_property_name", None)
    property_key = normalize(property_name) if callable(normalize) else property_name.casefold()
    workbook_path = (canonical_workbooks or {}).get(property_key) or audit.find_cf_workbook(property_path)
    changes: list[dict[str, Any]] = []
    result = {"property": property_name, "property_path": str(property_path), "workbook": str(workbook_path) if workbook_path else None}
    if workbook_path is None:
        result["status"] = "cf_workbook_missing"
        return changes, result
    if load_workbook is None:
        result["status"] = "openpyxl_missing"
        result["error"] = OPENPYXL_IMPORT_ERROR
        return changes, result
    workbook = load_workbook(workbook_path)
    try:
        sheet = sheet_for_year(workbook, year)
        if sheet is None:
            result["status"] = "year_sheet_missing"
            return changes, result
        column = month_column(audit, sheet, year, month)
        if column is None:
            result["status"] = "month_column_missing"
            return changes, result
        retained_earnings_exemption = (
            property_key,
            f"{year}-{month:02d}",
        ) in getattr(audit, "RETAINED_EARNINGS_LOFTY_EXEMPTIONS", set())
        lofty_reserve = parse_money(summary.get("lofty_curr_maintenance_reserve"))
        specs = [
            ("ECO General Ledger", getattr(audit, "ECO_GL_LABELS", (audit.ECO_GL_LABEL,)), parse_money(summary.get("eco_general_ledger_sum", summary.get("eco_gl_column_e_sum")))),
        ]
        if apply_eco_net_dao_funds:
            specs.insert(
                0,
                (
                    "ECO Net DAO Funds",
                    getattr(audit, "ECO_CASH_LABELS", (audit.ECO_CASH_LABEL,)),
                    parse_money(summary.get("eco_operating_cash")),
                ),
            )
        if retained_earnings_exemption:
            specs.insert(
                0,
                (
                    "Undistributed Cash Flow Retained Earnings",
                    (audit.RETAINED_EARNINGS_LABEL,),
                    -lofty_reserve if lofty_reserve is not None else None,
                ),
            )
        else:
            specs.insert(0, ("Lofty Operating Cash", audit.LOFTY_OR_LABELS, lofty_reserve))
        canonical_gl_row = None
        for source_name, labels, expected in specs:
            row, created = find_or_create_row(sheet, labels, create_missing_rows)
            if source_name == "ECO General Ledger":
                canonical_gl_row = row
            if row is None:
                changes.append(
                    {
                        "property": property_name,
                        "source": source_name,
                        "action": "row_missing",
                        "label": labels[0],
                        "expected": expected,
                    }
                )
                continue
            if source_name == "ECO General Ledger" and sheet.cell(row=row, column=1).value != audit.ECO_GL_LABEL:
                changes.append(
                    {
                        "property": property_name,
                        "source": source_name,
                        "action": "relabel_general_ledger_row",
                        "row": row,
                        "old_label": sheet.cell(row=row, column=1).value,
                        "label": audit.ECO_GL_LABEL,
                    }
                )
                if apply:
                    sheet.cell(row=row, column=1).value = audit.ECO_GL_LABEL
            if source_name == "ECO Net DAO Funds" and sheet.cell(row=row, column=1).value != audit.ECO_CASH_LABEL:
                changes.append(
                    {
                        "property": property_name,
                        "source": source_name,
                        "action": "relabel_eco_net_dao_funds_row",
                        "row": row,
                        "old_label": sheet.cell(row=row, column=1).value,
                        "label": audit.ECO_CASH_LABEL,
                    }
                )
                if apply:
                    sheet.cell(row=row, column=1).value = audit.ECO_CASH_LABEL
            cell = sheet.cell(row=row, column=column)
            old_value = cell.value
            if expected is None:
                changes.append(
                    {
                        "property": property_name,
                        "source": source_name,
                        "action": "source_missing",
                        "label": labels[0],
                        "cell": cell.coordinate,
                        "old_value": old_value,
                        "expected": None,
                    }
                )
                continue
            old_numeric = parse_money(old_value)
            if old_numeric is not None and abs(round(old_numeric - expected, 2)) <= CONFLICT_THRESHOLD and not str(old_value).strip().startswith("="):
                action = "no_change"
            elif str(old_value or "").strip().startswith("="):
                action = "overwrite_formula"
            elif old_numeric is None:
                action = "set_non_numeric_cell"
            else:
                action = "set_mismatched_cell"
            if action != "no_change" and apply:
                cell.value = expected
            changes.append(
                {
                    "property": property_name,
                    "source": source_name,
                    "action": action,
                    "label": labels[0],
                    "cell": cell.coordinate,
                    "row_created": created,
                    "old_value": old_value,
                    "new_value": expected,
                }
            )
        changes.extend(
            clear_legacy_eco_cash_rows(
                sheet,
                canonical_row=canonical_gl_row,
                property_name=property_name,
                legacy_labels=tuple(
                    {
                        LEGACY_ECO_LABEL,
                        *getattr(audit, "ECO_GL_LABELS", ()),
                    }
                    - {
                        getattr(audit, "ECO_GL_LABEL", "ECO General Ledger (ECO GL Column E Total)"),
                        "ECO Operating Cash",
                    }
                ),
                apply=apply,
            )
        )
        if apply and any(change.get("action") not in {"no_change", "source_missing", "row_missing"} for change in changes):
            workbook.save(workbook_path)
        result["status"] = "ok"
        return changes, result
    finally:
        workbook.close()


def build_report(args: argparse.Namespace) -> dict[str, Any]:
    audit = load_audit_module()
    year, month = parse_month(args.month)
    blockers = source_blockers(args.source_cleanup_queue, args.source_cash_report, args.data_quality_report)
    apply_requested = bool(args.apply)
    correction_gate = downstream_balance_correction_gate(
        args.source_cleanup_queue,
        args.source_cash_report,
        args.data_quality_report,
    )
    hard_source_quality_blocker = any(
        blocker.startswith("source_cash_apply_blocked_by_source_ledger_quality_guard")
        or blocker.startswith("source_ledger_quality_")
        for blocker in blockers
    )
    apply_blocked = bool(
        hard_source_quality_blocker
        or (
            blockers
            and not args.ignore_source_blockers
            and not (args.allow_downstream_balance_correction and correction_gate["allowed"])
        )
    )
    effective_apply = apply_requested and not apply_blocked
    all_records = candidate_records(args.candidate_packet)
    source_correction_records = source_cash_correction_records(args.source_cash_report, all_records, audit)
    all_apply_records = all_records + source_correction_records
    records, excluded_records = filter_excluded_candidate_records(audit, all_apply_records, args.yhome_csv)
    canonical_workbooks = canonical_workbook_map(audit, args.source_cash_report)
    all_changes: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    if load_workbook is None:
        return {
            "job": "baselane-cf-balance-sheet-cash-apply",
            "generated_at": iso_z(),
            "status": "openpyxl_missing",
            "openpyxl_import_error": OPENPYXL_IMPORT_ERROR,
            "mode": "apply" if apply_requested else "dry_run",
            "effective_mode": "blocked" if apply_blocked else "dry_run",
            "source_blockers": blockers,
            "candidate_packet_property_count": len(all_records),
            "source_cash_correction_property_count": len(source_correction_records),
            "property_count": len(records),
            "policy_excluded_candidate_count": len(excluded_records),
            "excluded_candidate_properties": excluded_records,
            "changes": [],
            "summaries": [],
        }
    for record in records:
        changes, summary = update_workbook(
            audit,
            record,
            year=year,
            month=month,
            apply=effective_apply,
            create_missing_rows=bool(args.create_missing_rows),
            apply_eco_net_dao_funds=bool(args.apply_eco_net_dao_funds),
            canonical_workbooks=canonical_workbooks,
        )
        all_changes.extend(changes)
        summary["change_count"] = len([change for change in changes if change.get("action") != "no_change"])
        summaries.append(summary)
    action_counts = Counter(change.get("action") or "unknown" for change in all_changes)
    blocking_change_count = sum(
        count for action, count in action_counts.items() if action not in {"no_change"}
    )
    status = "ok"
    if apply_blocked:
        status = "blocked_source_not_clean"
    elif blocking_change_count and not effective_apply:
        status = "review"
    elif any(summary.get("status") != "ok" for summary in summaries):
        status = "review"
    return {
        "job": "baselane-cf-balance-sheet-cash-apply",
        "generated_at": iso_z(),
        "status": status,
        "month": args.month,
        "mode": "apply" if apply_requested else "dry_run",
        "effective_mode": "apply" if effective_apply else "blocked" if apply_blocked else "dry_run",
        "apply_requested": apply_requested,
        "apply_blocked_by_source_guard": apply_blocked,
        "apply_eco_net_dao_funds": bool(args.apply_eco_net_dao_funds),
        "allow_downstream_balance_correction": bool(args.allow_downstream_balance_correction),
        "downstream_balance_correction_gate": correction_gate,
        "source_blockers": blockers,
        "candidate_packet": str(args.candidate_packet),
        "yhome_csv": str(args.yhome_csv),
        "candidate_packet_property_count": len(all_records),
        "source_cash_correction_property_count": len(source_correction_records),
        "property_count": len(records),
        "policy_excluded_candidate_count": len(excluded_records),
        "excluded_candidate_properties": excluded_records,
        "change_count": blocking_change_count,
        "action_counts": dict(sorted(action_counts.items())),
        "summaries": summaries,
        "changes": all_changes[:500],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply current cash-position values to CF balance-sheet rows.")
    parser.add_argument("--month", default=datetime.now().strftime("%Y-%m"))
    parser.add_argument("--candidate-packet", type=Path, default=DEFAULT_CANDIDATE_PACKET)
    parser.add_argument("--source-cleanup-queue", type=Path, default=DEFAULT_SOURCE_CLEANUP_QUEUE)
    parser.add_argument("--source-cash-report", type=Path, default=DEFAULT_SOURCE_CASH_REPORT)
    parser.add_argument("--data-quality-report", type=Path, default=DEFAULT_DATA_QUALITY_REPORT)
    parser.add_argument("--yhome-csv", type=Path, default=DEFAULT_YHOME_CSV)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--ignore-source-blockers", action="store_true")
    parser.add_argument(
        "--allow-downstream-balance-correction",
        action="store_true",
        help="Allow only the guarded canonical-source cash-row correction when structural source checks are clean.",
    )
    parser.add_argument("--create-missing-rows", action="store_true")
    parser.add_argument(
        "--apply-eco-net-dao-funds",
        dest="apply_eco_net_dao_funds",
        action="store_true",
        help="Update ECO Net DAO Funds only from a verified dated spendable-cash authority snapshot.",
    )
    parser.add_argument(
        "--apply-physical-eco-cash",
        dest="apply_eco_net_dao_funds",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    args = parser.parse_args()
    report = build_report(args)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, default=str, sort_keys=True) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "status": report["status"],
                "effective_mode": report.get("effective_mode"),
                "property_count": report.get("property_count"),
                "change_count": report.get("change_count"),
                "report": str(args.report),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
