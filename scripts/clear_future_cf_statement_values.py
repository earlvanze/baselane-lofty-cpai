#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import shutil
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CANONICAL_ROOT = Path("/home/digit/.openclaw/workspace")
ROOT = CANONICAL_ROOT if CANONICAL_ROOT.is_dir() else Path(__file__).resolve().parents[1]
DEFAULT_REAL_ESTATE_ROOT = Path("/mnt/c/Users/digit/Dropbox/Real Estate")
DEFAULT_RUNTIME_MAP = ROOT / "reports" / "baselane_financials_monthly_lofty_pm_runtime_map.json"
DEFAULT_PORTFOLIO_INDEX = ROOT.parent / "workspace-lofty-vp-comms" / "updates" / "2026-06-portfolio-update-index.csv"
DEFAULT_CF_WORKBOOK_MANIFEST = ROOT / "reports" / "baselane_cf_balance_sheet_cash_apply_report.json"
CANONICAL_STATEMENT_DIR = "07 - P&L & Owner Statements"
POLICY = "Revenue and Operating Expenses cells for future months must be blank until source GL data exists."


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def default_target_period(today: date | None = None) -> tuple[int, int]:
    today = today or date.today()
    return today.year, today.month


def is_canonical_cf_file(path: Path, *, include_archive: bool = False, include_conflicts: bool = False) -> bool:
    name = path.name.lower()
    if path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    if "cash flow statement" not in name:
        return False
    if not include_archive and "_archive" in {part.lower() for part in path.parts}:
        return False
    if not include_conflicts and ("conflicted copy" in name or " conflict" in name):
        return False
    return True


def iter_index_property_paths(runtime_map: Path = DEFAULT_RUNTIME_MAP, portfolio_index: Path = DEFAULT_PORTFOLIO_INDEX) -> list[Path]:
    property_paths: list[Path] = []
    if runtime_map.is_file():
        try:
            runtime = json.loads(runtime_map.read_text(encoding="utf-8"))
        except Exception:
            runtime = {}
        records = runtime.get("records") or runtime.get("properties") or []
        if isinstance(records, list):
            for record in records:
                if not isinstance(record, dict):
                    continue
                raw_path = record.get("property_path") or record.get("input_property_path")
                if raw_path:
                    property_paths.append(Path(str(raw_path)))
    if not property_paths and portfolio_index.is_file():
        try:
            with portfolio_index.open(newline="", encoding="utf-8-sig") as handle:
                for row in csv.DictReader(handle):
                    raw_path = row.get("property_path")
                    if raw_path:
                        property_paths.append(Path(raw_path))
        except Exception:
            return []
    return sorted(set(property_paths), key=lambda path: str(path).lower())


def discover_indexed_cf_files(real_estate_root: Path, *, include_conflicts: bool = False) -> list[Path]:
    property_paths = iter_index_property_paths()
    if not property_paths:
        return []
    paths: list[Path] = []
    for property_path in property_paths:
        try:
            property_path.relative_to(real_estate_root)
        except ValueError:
            continue
        for statement_root in (
            property_path / CANONICAL_STATEMENT_DIR,
            property_path / "Public" / CANONICAL_STATEMENT_DIR,
        ):
            if not statement_root.is_dir():
                continue
            for path in statement_root.rglob("*Cash Flow Statement*.xls*"):
                if is_canonical_cf_file(path, include_archive=False, include_conflicts=include_conflicts):
                    paths.append(path)
    return sorted(set(paths), key=lambda path: str(path).lower())


def discover_manifest_cf_files(
    real_estate_root: Path,
    *,
    manifest: Path = DEFAULT_CF_WORKBOOK_MANIFEST,
    include_archive: bool = False,
    include_conflicts: bool = False,
) -> list[Path]:
    if not manifest.is_file():
        return []
    try:
        data = json.loads(manifest.read_text(encoding="utf-8"))
    except Exception:
        return []
    paths: list[Path] = []
    for row in data.get("summaries") or []:
        if not isinstance(row, dict):
            continue
        raw_path = row.get("workbook")
        if not raw_path:
            continue
        path = Path(str(raw_path))
        try:
            path.relative_to(real_estate_root)
        except ValueError:
            continue
        if not path.is_file():
            continue
        if is_canonical_cf_file(path, include_archive=include_archive, include_conflicts=include_conflicts):
            paths.append(path)
    return sorted(set(paths), key=lambda path: str(path).lower())


def discover_all_cf_files(real_estate_root: Path, *, include_archive: bool = False, include_conflicts: bool = False) -> list[Path]:
    manifest_paths = discover_manifest_cf_files(
        real_estate_root,
        include_archive=include_archive,
        include_conflicts=include_conflicts,
    )
    if manifest_paths:
        return manifest_paths

    indexed_paths = discover_indexed_cf_files(real_estate_root, include_conflicts=include_conflicts)
    if indexed_paths:
        return indexed_paths

    statement_roots: list[Path] = []
    state_dirs = sorted(real_estate_root.iterdir() if real_estate_root.is_dir() else [], key=lambda path: path.name.lower())
    for state_dir in state_dirs:
        if not state_dir.is_dir():
            continue
        if state_dir.name == "_Archive":
            if not include_archive:
                continue
            for pattern in (
                f"*/Public/{CANONICAL_STATEMENT_DIR}",
                f"*/*/Public/{CANONICAL_STATEMENT_DIR}",
                f"*/*/*/Public/{CANONICAL_STATEMENT_DIR}",
                f"*/*/*/*/Public/{CANONICAL_STATEMENT_DIR}",
                f"*/*/*/*/*/Public/{CANONICAL_STATEMENT_DIR}",
            ):
                statement_roots.extend(path for path in state_dir.glob(pattern) if path.is_dir())
            continue
        property_dirs = sorted((path for path in state_dir.iterdir() if path.is_dir()), key=lambda path: path.name.lower())
        for property_dir in property_dirs:
            for candidate in (property_dir / CANONICAL_STATEMENT_DIR, property_dir / "Public" / CANONICAL_STATEMENT_DIR):
                if candidate.is_dir():
                    statement_roots.append(candidate)

    paths: list[Path] = []
    for statement_root in sorted(set(statement_roots), key=lambda path: str(path).lower()):
        for path in statement_root.rglob("*Cash Flow Statement*.xls*"):
            if is_canonical_cf_file(path, include_archive=include_archive, include_conflicts=include_conflicts):
                paths.append(path)
    return sorted(set(paths), key=lambda path: str(path).lower())


def target_columns(sheet: Any, start_month: int) -> list[int]:
    return [column for column in range(2, min(sheet.max_column, 13) + 1) if column - 1 >= start_month]


def normalize_label(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def should_clear_row(label: Any, active_section: str | None) -> tuple[bool, str | None]:
    normalized = normalize_label(label)
    if not normalized:
        return False, active_section
    if normalized in {"capital expenditures", "capex", "balance sheet"}:
        return False, None
    if normalized in {"revenue", "income", "operating revenue"}:
        return True, "revenue"
    if normalized in {"operating expenses", "expenses", "operating expense"}:
        return True, "operating_expenses"
    if active_section in {"revenue", "operating_expenses"}:
        return True, active_section
    return False, active_section


def clear_workbook_values(path: Path, year: int, start_month: int, *, apply: bool) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    try:
        sheet_name = str(year)
        if sheet_name not in workbook.sheetnames:
            return {"file": str(path), "status": "missing_year_sheet", "changed_cell_count": 0, "changed_rows_bounded": []}
        sheet = workbook[sheet_name]
        columns = target_columns(sheet, start_month)
        changed_by_row: dict[str, int] = defaultdict(int)
        active_section: str | None = None
        for row in range(1, sheet.max_row + 1):
            label = sheet.cell(row, 1).value
            clear_row, active_section = should_clear_row(label, active_section)
            if not clear_row:
                continue
            row_label = str(label or f"row {row}").strip()
            for column in columns:
                cell = sheet.cell(row, column)
                if cell.value is not None:
                    changed_by_row[row_label] += 1
                    if apply:
                        cell.value = None
        changed_cell_count = sum(changed_by_row.values())
        if apply and changed_cell_count:
            workbook.save(path)
        return {
            "file": str(path),
            "status": "ok",
            "changed_cell_count": changed_cell_count,
            "changed_rows_bounded": [
                {"label": label, "changed_cell_count": count}
                for label, count in sorted(changed_by_row.items(), key=lambda item: item[0].lower())[:25]
            ],
            "target_column_count": len(columns),
            "target_columns": [f"{sheet.cell(1, column).column_letter}1" for column in columns],
        }
    finally:
        workbook.close()


def backup_workbook(path: Path, real_estate_root: Path, backup_root: Path) -> tuple[Path | None, str | None]:
    try:
        relative = path.relative_to(real_estate_root)
    except ValueError:
        relative = Path(path.name)
    backup_path = backup_root / relative
    try:
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, backup_path)
        return backup_path, None
    except Exception as exc:
        return None, str(exc)


def build_report(
    real_estate_root: Path,
    year: int,
    start_month: int,
    *,
    apply: bool = False,
    backup_root: Path | None = None,
    include_archive: bool = False,
    include_conflicts: bool = False,
    as_of_date: date | None = None,
) -> dict[str, Any]:
    cf_files = discover_all_cf_files(real_estate_root, include_archive=include_archive, include_conflicts=include_conflicts)
    changed_workbooks: list[dict[str, Any]] = []
    unreadable_workbooks: list[dict[str, Any]] = []
    backup_failed_workbooks: list[dict[str, Any]] = []
    changed_cell_count = 0
    backup_count = 0
    backup_failed_count = 0

    for path in cf_files:
        try:
            audit_result = clear_workbook_values(path, year, start_month, apply=False)
            workbook_changed_count = int(audit_result.get("changed_cell_count") or 0)
            if apply and workbook_changed_count:
                backup_path: Path | None = None
                if backup_root is not None:
                    backup_path, backup_error = backup_workbook(path, real_estate_root, backup_root)
                    if backup_path is not None:
                        backup_count += 1
                    else:
                        backup_failed_count += 1
                        backup_failed_workbooks.append({"file": str(path), "error": backup_error})
                audit_result = clear_workbook_values(path, year, start_month, apply=True)
                if backup_path is not None:
                    audit_result["backup_path"] = str(backup_path)
            if workbook_changed_count:
                changed_workbooks.append(audit_result)
                changed_cell_count += workbook_changed_count
        except Exception as exc:
            unreadable_workbooks.append({"file": str(path), "error": str(exc)})

    issues: list[dict[str, str]] = []
    if not apply and changed_cell_count:
        issues.append({"code": "future_cf_values_present", "detail": str(changed_cell_count)})
    if unreadable_workbooks:
        issues.append({"code": "future_cf_workbooks_unreadable", "detail": str(len(unreadable_workbooks))})
    if backup_failed_count:
        issues.append({"code": "future_cf_backup_failed", "detail": str(backup_failed_count)})

    return {
        "job": "clear-future-cf-statement-values",
        "status": "review" if issues else "ok",
        "generated_at": utc_now(),
        "mode": "apply" if apply else "audit",
        "policy": POLICY,
        "real_estate_root": str(real_estate_root),
        "as_of_date": (as_of_date or date.today()).isoformat(),
        "year": year,
        "start_month": start_month,
        "include_archive": include_archive,
        "include_conflicts": include_conflicts,
        "backup_root": str(backup_root) if backup_root is not None else None,
        "backup_count": backup_count,
        "backup_failed_count": backup_failed_count,
        "backup_failed_workbooks_bounded": backup_failed_workbooks[:25],
        "cf_file_count": len(cf_files),
        "changed_workbook_count": len(changed_workbooks),
        "changed_cell_count": changed_cell_count,
        "changed_workbooks_bounded": changed_workbooks[:100],
        "unreadable_count": len(unreadable_workbooks),
        "unreadable_workbooks_bounded": unreadable_workbooks[:25],
        "issue_count": len(issues),
        "issues": issues,
    }


def parse_args() -> argparse.Namespace:
    default_year, default_month = default_target_period()
    parser = argparse.ArgumentParser(description=POLICY)
    parser.add_argument("--real-estate-root", type=Path, default=DEFAULT_REAL_ESTATE_ROOT)
    parser.add_argument("--year", type=int, default=default_year)
    parser.add_argument("--start-month", type=int, default=default_month)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--include-archive", action="store_true")
    parser.add_argument("--include-conflicts", action="store_true")
    parser.add_argument("--backup-root", type=Path, default=None)
    parser.add_argument("--report", type=Path, default=ROOT / "reports" / "future_cf_statement_values_clear_report.json")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    backup_root = args.backup_root
    if args.apply and backup_root is None:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup_root = ROOT / "reports" / "future_cf_statement_values_backups" / stamp
    report = build_report(
        args.real_estate_root,
        args.year,
        args.start_month,
        apply=args.apply,
        backup_root=backup_root,
        include_archive=args.include_archive,
        include_conflicts=args.include_conflicts,
    )
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"changed_cell_count": report["changed_cell_count"], "issue_count": report["issue_count"], "report": str(args.report), "status": report["status"]}, sort_keys=True))
    return 2 if report["status"] == "review" else 0


if __name__ == "__main__":
    raise SystemExit(main())
