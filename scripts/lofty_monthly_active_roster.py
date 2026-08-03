#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from lofty_property_paths import public_dir_for_property


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_POLICY = ROOT / "config" / "lofty_active_property_roster_policy.json"
DEFAULT_REPORT = ROOT / "reports" / "lofty_monthly_active_property_roster.json"
DEFAULT_INDEX = ROOT / "reports" / "lofty_monthly_active_reporting_index.csv"
INDEX_FIELDS = ("property_path", "managed_name", "draft_path", "status", "template", "notes")
WORD_ALIASES = {
    "avenue": "ave",
    "boulevard": "blvd",
    "circle": "cir",
    "drive": "dr",
    "east": "e",
    "lane": "ln",
    "north": "n",
    "place": "pl",
    "road": "rd",
    "south": "s",
    "street": "st",
    "west": "w",
}


def iso_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"JSON source must be an object: {path}")
    return data


def resolve_config_path(value: object, *, root: Path = ROOT) -> Path:
    path = Path(str(value or "")).expanduser()
    return path if path.is_absolute() else root / path


def first_existing(values: Iterable[object], *, root: Path = ROOT) -> Path | None:
    for value in values:
        path = resolve_config_path(value, root=root)
        if path.is_file():
            return path
    return None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def source_evidence(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path),
        "sha256": sha256_file(path),
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
    }


def normalize(value: object) -> str:
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"\blfty\d+\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = [WORD_ALIASES.get(token, token) for token in text.split()]
    return " ".join(tokens)


def street_key(value: object) -> str:
    return normalize(str(value or "").split(",", 1)[0])


def dao_key(value: object) -> str:
    return normalize(value)


def display_street(value: object) -> str:
    return str(value or "").split(",", 1)[0].strip()


def find_header(headers: list[object], expected: str) -> int | None:
    expected_key = normalize(expected)
    for index, value in enumerate(headers):
        if normalize(value) == expected_key:
            return index
    return None


def load_schedule(path: Path, worksheet: str, active_marker: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[worksheet] if worksheet in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows, ()))
    columns = {
        "portfolio": find_header(headers, "Portfolio (Internal Name)"),
        "address": find_header(headers, "Address"),
        "pm": find_header(headers, "PM / Sub-PM"),
        "on_lofty": find_header(headers, "On Lofty?"),
        "dao": find_header(headers, "DAO"),
        "current_status": find_header(headers, "Current Status (Occupied Units)"),
        "total_units": find_header(headers, "Total Units"),
    }
    missing = [name for name in ("address", "on_lofty", "dao") if columns[name] is None]
    if missing:
        raise ValueError(f"schedule is missing required columns: {', '.join(missing)}")

    def cell(values: tuple[object, ...], name: str) -> object:
        index = columns[name]
        return values[index] if index is not None and index < len(values) else None

    records: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows, 2):
        if str(cell(values, "on_lofty") or "").strip() != active_marker:
            continue
        address = str(cell(values, "address") or "").strip()
        dao = str(cell(values, "dao") or "").strip()
        records.append(
            {
                "source": "ownership_schedule",
                "source_row": row_number,
                "portfolio": str(cell(values, "portfolio") or "").strip() or None,
                "address": address,
                "address_key": street_key(address),
                "dao": dao,
                "dao_key": dao_key(dao),
                "pm": str(cell(values, "pm") or "").strip() or None,
                "current_status": str(cell(values, "current_status") or "").strip() or None,
                "total_units": str(cell(values, "total_units") or "").strip() or None,
            }
        )
    return records


def policy_records(policy: dict[str, Any], field: str, *, split_only: bool = False) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for value in policy.get(field) or []:
        payload = value if isinstance(value, dict) else {"address": value}
        if split_only and payload.get("split_exclude") is not True:
            continue
        address = str(payload.get("address") or payload.get("property_name") or "").strip()
        if address:
            records.append({**payload, "address": address, "address_key": street_key(address)})
    return records


def manager_properties(payload: dict[str, Any]) -> list[dict[str, Any]]:
    current: Any = payload
    for _ in range(4):
        if not isinstance(current, dict):
            break
        properties = current.get("properties")
        if isinstance(properties, list):
            return [item for item in properties if isinstance(item, dict)]
        if isinstance(current.get("response"), dict):
            current = current["response"]
            continue
        if isinstance(current.get("data"), dict):
            current = current["data"]
            continue
        break
    return []


def manager_address(record: dict[str, Any]) -> str:
    nested = record.get("property") if isinstance(record.get("property"), dict) else {}
    return str(record.get("address") or record.get("fullAddress") or nested.get("address") or "").strip()


def manager_status(record: dict[str, Any]) -> str:
    entry = record.get("plEntry") if isinstance(record.get("plEntry"), dict) else {}
    return str(entry.get("status") or record.get("status") or "").strip().lower()


def manager_id(record: dict[str, Any]) -> str | None:
    entry = record.get("plEntry") if isinstance(record.get("plEntry"), dict) else {}
    value = record.get("id") or record.get("propertyId") or entry.get("propertyId")
    return str(value).strip() if value else None


def load_legacy_index(path: Path | None) -> list[dict[str, str]]:
    if path is None:
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def legacy_row_key(row: dict[str, str]) -> str:
    property_path = Path(str(row.get("property_path") or ""))
    candidate = property_path.name
    if candidate.lower() == "public":
        candidate = property_path.parent.name
    candidate = re.sub(r"\s+Public$", "", candidate, flags=re.IGNORECASE)
    return street_key(row.get("managed_name") or candidate)


def load_property_map(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.is_file():
        return []
    payload = read_json(path)
    return [item for item in payload.get("properties") or [] if isinstance(item, dict)]


def property_map_key(record: dict[str, Any]) -> str:
    return street_key(record.get("full_address") or record.get("property_name"))


def pick_path(
    physical_records: list[dict[str, Any]],
    target_override: dict[str, Any] | None,
    path_overrides: dict[str, str],
    legacy_rows: list[dict[str, str]],
    property_map: list[dict[str, Any]],
) -> tuple[str, str, str | None]:
    if target_override and target_override.get("property_path"):
        return str(target_override["property_path"]), "reporting_target_override", None
    for physical in physical_records:
        override = path_overrides.get(str(physical.get("address_key") or ""))
        if override:
            return override, "property_path_override", None
    physical_keys = {str(record.get("address_key") or "") for record in physical_records}
    target_key = street_key((target_override or {}).get("managed_name"))
    wanted_keys = {key for key in [*physical_keys, target_key] if key}
    legacy_matches = [row for row in legacy_rows if legacy_row_key(row) in wanted_keys]
    if len(legacy_matches) == 1:
        return str(legacy_matches[0].get("property_path") or ""), "legacy_index_path", None
    if len(legacy_matches) > 1:
        return "", "ambiguous_legacy_index_path", ",".join(str(row.get("property_path") or "") for row in legacy_matches)
    map_matches = [record for record in property_map if property_map_key(record) in wanted_keys]
    if len(map_matches) == 1:
        update_path = Path(str(map_matches[0].get("updates_md") or ""))
        if update_path.name == "UPDATES.md" and len(update_path.parents) >= 3:
            return str(update_path.parents[2]), "property_update_map", None
    if len(map_matches) > 1:
        return "", "ambiguous_property_update_map", ",".join(str(record.get("property_name") or "") for record in map_matches)
    return "", "unresolved", None


def managed_name_for_target(
    physical_records: list[dict[str, Any]],
    target_override: dict[str, Any] | None,
    legacy_rows: list[dict[str, str]],
    property_map: list[dict[str, Any]],
) -> str:
    if target_override and target_override.get("managed_name"):
        return str(target_override["managed_name"]).strip()
    physical_keys = {str(record.get("address_key") or "") for record in physical_records}
    for row in legacy_rows:
        if legacy_row_key(row) in physical_keys and str(row.get("managed_name") or "").strip():
            return str(row["managed_name"]).strip()
    for record in property_map:
        if property_map_key(record) in physical_keys:
            return display_street(record.get("full_address") or record.get("property_name"))
    return display_street(physical_records[0].get("address"))


def build_roster(
    policy_path: Path = DEFAULT_POLICY,
    *,
    run_month: str | None = None,
    schedule_workbook: Path | None = None,
    manager_snapshot: Path | None = None,
    legacy_index: Path | None = None,
    property_map_path: Path | None = None,
    listing_policy_path: Path | None = None,
) -> dict[str, Any]:
    policy = read_json(policy_path)
    run_month = run_month or str(policy.get("effective_month") or "")
    schedule_config = policy.get("schedule") if isinstance(policy.get("schedule"), dict) else {}
    schedule_workbook = schedule_workbook or first_existing(schedule_config.get("workbook_candidates") or [])
    manager_snapshot = manager_snapshot or first_existing(policy.get("manager_snapshot_candidates") or [])
    legacy_index = legacy_index or first_existing(policy.get("legacy_index_candidates") or [])
    property_map_path = property_map_path or resolve_config_path(policy.get("property_map"))
    listing_policy_path = listing_policy_path or resolve_config_path(policy.get("listing_update_policy"))
    if schedule_workbook is None:
        raise FileNotFoundError("no configured ownership schedule is available")
    if manager_snapshot is None:
        raise FileNotFoundError("no configured Lofty manager snapshot is available")

    schedule_records = load_schedule(
        schedule_workbook,
        str(schedule_config.get("worksheet") or "Sheet1"),
        str(schedule_config.get("active_marker") or "Yes"),
    )
    listing_policy = read_json(listing_policy_path)
    sold_records = policy_records(listing_policy, "sold_ignore_listing_updates")
    split_exclusions = policy_records(listing_policy, "operational_ignore_listing_updates", split_only=True)
    hard_exclusions = [*sold_records, *split_exclusions]
    excluded: list[dict[str, Any]] = []
    active_physical: list[dict[str, Any]] = []
    for record in schedule_records:
        exclusion = next((item for item in hard_exclusions if item["address_key"] == record["address_key"]), None)
        if exclusion:
            excluded.append(
                {
                    **record,
                    "exclusion_source": "sold_policy" if exclusion in sold_records else "operational_split_exclusion",
                    "exclusion_reason": exclusion.get("reason"),
                }
            )
        else:
            active_physical.append(record)

    manager_payload = read_json(manager_snapshot)
    live_properties = manager_properties(manager_payload)
    live_by_key = {
        street_key(manager_address(record)): record
        for record in live_properties
        if street_key(manager_address(record))
    }
    active_keys = {str(record.get("address_key") or "") for record in active_physical}
    configured_additions: list[dict[str, Any]] = []
    addition_issues: list[str] = []
    for addition in policy.get("live_ready_additions") or []:
        if not isinstance(addition, dict):
            continue
        address = str(addition.get("address") or "").strip()
        key = street_key(address)
        live = live_by_key.get(key)
        if not live:
            addition_issues.append(f"configured_live_addition_missing:{address}")
            continue
        if manager_status(live) != "ready":
            addition_issues.append(f"configured_live_addition_not_ready:{address}:status={manager_status(live) or 'missing'}")
            continue
        if key in active_keys:
            continue
        record = {
            "source": "configured_live_ready_addition",
            "source_row": None,
            "portfolio": None,
            "address": address,
            "address_key": key,
            "dao": str(addition.get("dao") or "").strip(),
            "dao_key": dao_key(addition.get("dao")),
            "pm": None,
            "current_status": "live_ready",
            "total_units": None,
            "addition_reason": addition.get("reason"),
            "property_path_override": addition.get("property_path"),
        }
        active_physical.append(record)
        configured_additions.append(record)
        active_keys.add(key)

    path_overrides = {
        street_key(record.get("address")): str(record.get("property_path") or "")
        for record in policy.get("property_path_overrides") or []
        if isinstance(record, dict) and record.get("address") and record.get("property_path")
    }
    for addition in policy.get("live_ready_additions") or []:
        if isinstance(addition, dict) and addition.get("address") and addition.get("property_path"):
            path_overrides[street_key(addition["address"])] = str(addition["property_path"])
    target_overrides = {
        dao_key(record.get("dao")): record
        for record in policy.get("reporting_target_overrides") or []
        if isinstance(record, dict) and record.get("dao")
    }
    legacy_rows = load_legacy_index(legacy_index)
    property_map = load_property_map(property_map_path)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    issues: list[str] = [*addition_issues]
    for record in active_physical:
        key = str(record.get("dao_key") or "")
        if not key:
            issues.append(f"active_property_missing_dao:{record.get('address')}")
            key = f"missing-dao:{record.get('address_key')}"
        grouped[key].append(record)

    reporting_targets: list[dict[str, Any]] = []
    for key, physical_records in grouped.items():
        physical_records.sort(key=lambda item: str(item.get("address") or ""))
        target_override = target_overrides.get(key)
        managed_name = managed_name_for_target(physical_records, target_override, legacy_rows, property_map)
        property_path, path_source, path_detail = pick_path(
            physical_records,
            target_override,
            path_overrides,
            legacy_rows,
            property_map,
        )
        if not property_path:
            issues.append(f"reporting_target_path_unresolved:{managed_name}:{path_source}:{path_detail or ''}")
        path = Path(property_path) if property_path else Path("/")
        public_dir = public_dir_for_property(path) if property_path else path
        docs_dir = public_dir / "00 - README & Property Snapshot"
        draft_path = docs_dir / f"{run_month}-owner-update-checkin-draft.md"
        target_live = live_by_key.get(street_key(managed_name))
        if target_live is None:
            target_live = next((live_by_key.get(str(record.get("address_key") or "")) for record in physical_records if live_by_key.get(str(record.get("address_key") or ""))), None)
        target = {
            "reporting_target_key": key,
            "property_name": managed_name,
            "managed_name": managed_name,
            "dao": str(physical_records[0].get("dao") or "").strip(),
            "property_path": property_path,
            "property_path_source": path_source,
            "property_path_exists": bool(property_path and path.exists()),
            "physical_property_count": len(physical_records),
            "physical_addresses": [record["address"] for record in physical_records],
            "physical_address_keys": [record["address_key"] for record in physical_records],
            "updates_md": str(docs_dir / "UPDATES.md") if property_path else None,
            "financials_md": str(public_dir / "07 - P&L & Owner Statements" / "FINANCIALS.md") if property_path else None,
            "draft_path": str(draft_path) if property_path else None,
            "status": "existing" if property_path and draft_path.is_file() else "would_create",
            "template": "existing" if property_path and draft_path.is_file() else "checkin",
            "notes": f"authoritative active roster; {len(physical_records)} physical propert{'y' if len(physical_records) == 1 else 'ies'} in reporting target",
            "lofty_property_id": manager_id(target_live) if target_live else None,
            "lofty_manager_status": manager_status(target_live) if target_live else None,
            "lofty_live_mutation_available": bool(target_live and manager_status(target_live) == "ready"),
            "summary_scope": "active_reporting_target",
            "summary_scope_reason": "Included by the authoritative active physical-property roster for the close month.",
        }
        reporting_targets.append(target)

    active_physical.sort(key=lambda item: str(item.get("address") or ""))
    reporting_targets.sort(key=lambda item: normalize(item.get("managed_name")))
    duplicate_address_keys = sorted(
        key for key in {record["address_key"] for record in active_physical}
        if sum(record["address_key"] == key for record in active_physical) > 1
    )
    duplicate_target_names = sorted(
        name for name in {normalize(record["managed_name"]) for record in reporting_targets}
        if sum(normalize(record["managed_name"]) == name for record in reporting_targets) > 1
    )
    if duplicate_address_keys:
        issues.append(f"duplicate_active_physical_address_keys:{','.join(duplicate_address_keys)}")
    if duplicate_target_names:
        issues.append(f"duplicate_reporting_target_names:{','.join(duplicate_target_names)}")

    expected = policy.get("expected_counts") if isinstance(policy.get("expected_counts"), dict) else {}
    expected_physical = int(expected.get("active_physical_properties") or 0)
    expected_targets = int(expected.get("active_reporting_targets") or 0)
    if len(active_physical) != expected_physical:
        issues.append(f"active_physical_property_count_mismatch:{len(active_physical)}:expected={expected_physical}")
    if len(reporting_targets) != expected_targets:
        issues.append(f"active_reporting_target_count_mismatch:{len(reporting_targets)}:expected={expected_targets}")
    if any(not record.get("property_path_exists") for record in reporting_targets):
        missing = [str(record.get("managed_name")) for record in reporting_targets if not record.get("property_path_exists")]
        issues.append(f"reporting_target_paths_missing:{','.join(missing)}")

    sources = {
        "policy": source_evidence(policy_path),
        "ownership_schedule": source_evidence(schedule_workbook),
        "lofty_manager_snapshot": source_evidence(manager_snapshot),
        "legacy_index": source_evidence(legacy_index) if legacy_index else None,
        "property_map": source_evidence(property_map_path) if property_map_path and property_map_path.is_file() else None,
        "listing_update_policy": source_evidence(listing_policy_path),
    }
    live_ready_keys = {street_key(manager_address(record)) for record in live_properties if manager_status(record) == "ready"}
    unmatched_ready = sorted(key for key in live_ready_keys if key and key not in active_keys)
    return {
        "generated_at": iso_z(),
        "status": "ok" if not issues else "review",
        "run_month": run_month,
        "policy_version": policy.get("version"),
        "policy_authority": policy.get("authority"),
        "authoritative_active_property_count": len(active_physical),
        "authoritative_reporting_target_count": len(reporting_targets),
        "expected_active_property_count": expected_physical,
        "expected_reporting_target_count": expected_targets,
        "schedule_active_marker_record_count": len(schedule_records),
        "schedule_excluded_record_count": len(excluded),
        "configured_live_ready_addition_count": len(configured_additions),
        "lofty_manager_property_count": len(live_properties),
        "lofty_manager_ready_count": sum(manager_status(record) == "ready" for record in live_properties),
        "unmatched_live_ready_property_keys": unmatched_ready,
        "physical_property_count": len(active_physical),
        "reporting_target_count": len(reporting_targets),
        "issue_count": len(issues),
        "issues": issues,
        "source_evidence": sources,
        "excluded_physical_properties": excluded,
        "physical_properties": active_physical,
        "reporting_targets": reporting_targets,
        "records": reporting_targets,
    }


def write_index(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f"{path.name}.tmp")
    with temporary.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=INDEX_FIELDS)
        writer.writeheader()
        for record in report.get("reporting_targets") or []:
            writer.writerow({field: record.get(field) or "" for field in INDEX_FIELDS})
    temporary.replace(path)


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f"{path.name}.tmp")
    temporary.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    temporary.replace(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the authoritative monthly physical-property roster and grouped reporting index.")
    parser.add_argument("--policy", type=Path, default=DEFAULT_POLICY)
    parser.add_argument("--run-month")
    parser.add_argument("--schedule-workbook", type=Path)
    parser.add_argument("--manager-snapshot", type=Path)
    parser.add_argument("--legacy-index", type=Path)
    parser.add_argument("--property-map", type=Path)
    parser.add_argument("--listing-update-policy", type=Path)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--index-csv", type=Path, default=DEFAULT_INDEX)
    args = parser.parse_args(argv)
    report = build_roster(
        args.policy,
        run_month=args.run_month,
        schedule_workbook=args.schedule_workbook,
        manager_snapshot=args.manager_snapshot,
        legacy_index=args.legacy_index,
        property_map_path=args.property_map,
        listing_policy_path=args.listing_update_policy,
    )
    write_report(args.report, report)
    write_index(args.index_csv, report)
    print(
        f"status={report['status']} physical={report['authoritative_active_property_count']} "
        f"targets={report['authoritative_reporting_target_count']} issues={report['issue_count']} "
        f"report={args.report} index={args.index_csv}"
    )
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
