#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised operationally
    openpyxl = None


CONFLICT_THRESHOLD = 0.01
APPROVAL_SCOPE = "baselane_cf_conflict_resolution"
APPLICABLE_ACTIONS = {"clear_from_verified_void", "fill_from_gl", "overwrite", "review_accrual_in_baselane"}
DEFAULT_BLOCKED_ACTION_REASONS = {
    "cf_has_value_gl_empty": "CF has a manual/accrual value while Baselane GL is empty; fix Baselane tagging/accrual or explicitly review manually.",
}


def parse_number(value: Any) -> float | None:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except ValueError:
        return None


def amounts_match(left: Any, right: Any) -> bool:
    left_number = parse_number(left)
    right_number = parse_number(right)
    if left_number is None or right_number is None:
        return str(left or "").strip() == str(right or "").strip()
    return abs(left_number - right_number) <= CONFLICT_THRESHOLD


def property_identity(value: Any) -> str:
    """Compare the address portion only; source rows omit city/state inconsistently."""
    address = str(value or "").split(",", 1)[0].lower()
    return re.sub(r"[^a-z0-9]", "", address)


def conflict_id(row: dict[str, Any]) -> str:
    material = "|".join(
        str(row.get(key, "")).strip()
        for key in (
            "property",
            "file",
            "row",
            "label",
            "action",
            "cf_value",
            "gl_total",
            "verified_void_baselane_id",
            "verified_voided_amount",
        )
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()[:16]


def load_packet(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"Conflict packet must be a JSON object: {path}")
    return data


def load_verified_voids(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.is_file():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"Verified-void file must be a JSON object: {path}")
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        raise ValueError(f"Verified-void rows must be a list: {path}")
    return [row for row in rows if isinstance(row, dict)]


def load_transaction_index(path: Path | None) -> list[dict[str, str]]:
    if not path or not path.is_file():
        return []
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def voided_amount_from_notes(notes: Any) -> float | None:
    """Read the explicit legacy-accrual amount from a Baselane PM-void note."""
    text = str(notes or "")
    match = re.search(r"(?:Original\s+|Voided legacy manual PM row of\s+\$?)(-?\$?[\d,]+(?:\.\d{1,2})?)", text, re.IGNORECASE)
    if not match:
        return None
    return parse_number(match.group(1))


def verified_void_for_row(
    row: dict[str, Any],
    month: str | None,
    verified_voids: list[dict[str, Any]],
    transactions: list[dict[str, str]],
) -> dict[str, Any] | None:
    """Return exact current-source evidence for a deliberately voided CF PM row."""
    if row.get("action") != "cf_has_value_gl_empty" or not month:
        return None
    property_name = str(row.get("property") or "").strip()
    label = str(row.get("label") or "").strip()
    row_number = str(row.get("row") or "").strip()
    cf_value = parse_number(row.get("cf_value"))
    if cf_value is None or abs(cf_value) <= CONFLICT_THRESHOLD:
        return None
    for candidate in verified_voids:
        if (
            str(candidate.get("month") or "").strip() != month
            or str(candidate.get("property") or "").strip() != property_name
            or str(candidate.get("label") or "").strip() != label
            or str(candidate.get("row") or "").strip() != row_number
            or not amounts_match(candidate.get("original_amount"), cf_value)
        ):
            continue
        baselane_id = str(candidate.get("void_baselane_id") or "").strip()
        note_marker = str(candidate.get("required_note_marker") or "").strip()
        required_voided_amount = candidate.get("required_voided_amount")
        if not baselane_id or not note_marker or required_voided_amount is None:
            continue
        for transaction in transactions:
            if str(transaction.get("BaselaneId") or "").strip() != baselane_id:
                continue
            transaction_month = str(transaction.get("ISODate") or "")[:7]
            if transaction_month != month:
                continue
            if property_identity(transaction.get("Property")) != property_identity(property_name):
                continue
            category = str(transaction.get("Category") or "").strip()
            sub_category = str(transaction.get("Sub-category") or "").strip()
            transaction_type = str(transaction.get("Type") or "").strip()
            is_property_management = (
                (transaction_type == "Operating Expenses" and category == "Property Management")
                or (category == "Operating Expenses" and sub_category == "Property Management")
            )
            if not is_property_management:
                continue
            if not amounts_match(transaction.get("Amount"), 0):
                continue
            if note_marker not in str(transaction.get("Notes") or ""):
                continue
            source_voided_amount = voided_amount_from_notes(transaction.get("Notes"))
            if source_voided_amount is None or not amounts_match(source_voided_amount, required_voided_amount):
                continue
            return {
                "baselane_id": baselane_id,
                "note_marker": note_marker,
                "source_amount": transaction.get("Amount"),
                "voided_amount": source_voided_amount,
            }
    return None


def infer_month(packet: dict[str, Any], explicit_month: str | None) -> str | None:
    if explicit_month:
        return explicit_month
    audit_report = str(packet.get("audit_report") or "")
    stem = Path(audit_report).stem
    if stem.startswith("audit_"):
        candidate = stem.removeprefix("audit_")
        if len(candidate) == 7 and candidate[4] == "-":
            return candidate
    return None


def month_column(sheet: Any, month: str) -> int | None:
    year, month_number = [int(part) for part in month.split("-")]
    for header_row in (1, 4):
        for column in range(2, 14):
            value = sheet.cell(row=header_row, column=column).value
            if value is None:
                continue
            if hasattr(value, "year") and hasattr(value, "month"):
                if value.year == year and value.month == month_number:
                    return column
            text = str(value).strip()
            for fmt in ("%b-%y", "%B-%y", "%b %y", "%B %y", "%b-%Y", "%B-%Y", "%b %Y", "%B %Y", "%Y-%m"):
                try:
                    parsed = datetime.strptime(text, fmt)
                except ValueError:
                    continue
                if parsed.year == year and parsed.month == month_number:
                    return column
    return None


def normalize_packet_rows(
    packet: dict[str, Any],
    month: str | None = None,
    verified_voids: list[dict[str, Any]] | None = None,
    transactions: list[dict[str, str]] | None = None,
) -> list[dict[str, Any]]:
    normalized = []
    for row in packet.get("rows") or []:
        if not isinstance(row, dict):
            continue
        row_copy = dict(row)
        verified_void = verified_void_for_row(row_copy, month, verified_voids or [], transactions or [])
        if verified_void:
            row_copy["action"] = "clear_from_verified_void"
            row_copy["gl_total"] = "0"
            row_copy["verified_void_baselane_id"] = verified_void["baselane_id"]
            row_copy["verified_void_note_marker"] = verified_void["note_marker"]
            row_copy["verified_voided_amount"] = verified_void["voided_amount"]
        row_copy["id"] = conflict_id(row_copy)
        row_copy["applicable"] = row_copy.get("action") in APPLICABLE_ACTIONS
        row_copy["blocked_reason"] = "" if row_copy["applicable"] else DEFAULT_BLOCKED_ACTION_REASONS.get(
            row_copy.get("action"),
            "Conflict action is not eligible for automated workbook update.",
        )
        normalized.append(row_copy)
    return normalized


def approval_template(
    packet: dict[str, Any],
    month: str | None,
    verified_voids: list[dict[str, Any]] | None = None,
    transactions: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    rows = normalize_packet_rows(packet, month, verified_voids, transactions)
    return {
        "approval_scope": APPROVAL_SCOPE,
        "month": month,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "instructions": "Set approved=true only after human review. Automated apply requires exact id/action/file/row/label/current/new value match.",
        "approved": [
            {
                "id": row["id"],
                "approved": False,
                "property": row.get("property"),
                "file": row.get("file"),
                "row": row.get("row"),
                "label": row.get("label"),
                "action": row.get("action"),
                "current_value": row.get("cf_value"),
                "new_value": row.get("gl_total"),
                "verified_void_baselane_id": row.get("verified_void_baselane_id"),
                "verified_voided_amount": row.get("verified_voided_amount"),
            }
            for row in rows
            if row.get("applicable")
        ],
        "blocked": [
            {
                "id": row["id"],
                "property": row.get("property"),
                "file": row.get("file"),
                "row": row.get("row"),
                "label": row.get("label"),
                "action": row.get("action"),
                "reason": row.get("blocked_reason"),
            }
            for row in rows
            if not row.get("applicable")
        ],
    }


def load_approvals(path: Path | None, expected_month: str | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("approval_scope") != APPROVAL_SCOPE:
        raise ValueError(f"Approval file scope must be {APPROVAL_SCOPE}")
    approval_month = data.get("month")
    if expected_month and approval_month and approval_month != expected_month:
        raise ValueError(f"Approval month {approval_month} does not match run month {expected_month}")
    approvals = {}
    for entry in data.get("approved") or []:
        if isinstance(entry, dict) and entry.get("approved") is True:
            approvals[str(entry.get("id"))] = entry
    return approvals


def apply_or_plan_row(row: dict[str, Any], approval: dict[str, Any] | None, month: str | None, apply: bool) -> dict[str, Any]:
    result = {
        "id": row["id"],
        "property": row.get("property"),
        "file": row.get("file"),
        "row": row.get("row"),
        "label": row.get("label"),
        "action": row.get("action"),
        "cf_value": row.get("cf_value"),
        "gl_total": row.get("gl_total"),
        "approved": bool(approval),
        "status": "skipped",
    }
    if not row.get("applicable"):
        result["status"] = "blocked_action"
        result["reason"] = row.get("blocked_reason")
        return result
    if not approval:
        result["status"] = "needs_approval"
        return result
    for field in ("file", "row", "label", "action"):
        if str(approval.get(field, "")).strip() != str(row.get(field, "")).strip():
            result["status"] = "approval_mismatch"
            result["reason"] = f"Approval {field} does not match conflict packet."
            return result
    if not amounts_match(approval.get("current_value"), row.get("cf_value")) or not amounts_match(approval.get("new_value"), row.get("gl_total")):
        result["status"] = "approval_value_mismatch"
        result["reason"] = "Approval current/new values do not match conflict packet."
        return result
    if row.get("action") == "clear_from_verified_void" and not amounts_match(
        approval.get("verified_voided_amount"), row.get("verified_voided_amount")
    ):
        result["status"] = "approval_void_evidence_mismatch"
        result["reason"] = "Approval verified void amount does not match current-source evidence."
        return result
    if not month:
        result["status"] = "missing_month"
        result["reason"] = "Cannot locate workbook month column without --month or audit_YYYY-MM report name."
        return result
    if openpyxl is None:
        result["status"] = "openpyxl_missing"
        return result
    workbook_path = Path(str(row.get("file") or ""))
    if not workbook_path.exists():
        result["status"] = "workbook_missing"
        return result
    year = int(month.split("-")[0])
    workbook = openpyxl.load_workbook(workbook_path)
    try:
        if str(year) not in workbook.sheetnames:
            result["status"] = "year_sheet_missing"
            return result
        sheet = workbook[str(year)]
        column = month_column(sheet, month)
        if not column:
            result["status"] = "month_column_missing"
            return result
        row_number = int(str(row.get("row")))
        label = str(sheet.cell(row=row_number, column=1).value or "").strip()
        if label != str(row.get("label") or "").strip():
            result["status"] = "label_mismatch"
            result["actual_label"] = label
            return result
        cell = sheet.cell(row=row_number, column=column)
        current_value = cell.value
        if isinstance(current_value, str) and current_value.startswith("="):
            if row.get("action") != "review_accrual_in_baselane":
                result["status"] = "formula_preserved"
                return result
            if str(current_value).strip() != str(row.get("cf_value") or "").strip():
                result["status"] = "current_formula_mismatch"
                result["actual_current_value"] = current_value
                return result
        elif not amounts_match(current_value, row.get("cf_value")):
            if amounts_match(current_value, row.get("gl_total")):
                result["status"] = "already_current"
                return result
            result["status"] = "current_value_mismatch"
            result["actual_current_value"] = current_value
            return result
        new_value = parse_number(row.get("gl_total"))
        if new_value is None:
            result["status"] = "new_value_not_numeric"
            return result
        result["old_value"] = current_value
        result["new_value"] = round(new_value, 2)
        result["status"] = "would_update"
        if apply:
            cell.value = round(new_value, 2)
            workbook.save(workbook_path)
            result["status"] = "updated"
    finally:
        workbook.close()
    return result


def build_report(
    packet_path: Path,
    approval_path: Path | None,
    month: str | None,
    apply: bool,
    verified_voids: list[dict[str, Any]] | None = None,
    transactions: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    packet = load_packet(packet_path)
    resolved_month = infer_month(packet, month)
    rows = normalize_packet_rows(packet, resolved_month, verified_voids, transactions)
    approvals = load_approvals(approval_path, resolved_month)
    results = [
        apply_or_plan_row(row, approvals.get(row["id"]), resolved_month, apply)
        for row in rows
    ]
    status_counts = Counter(result["status"] for result in results)
    action_counts = Counter(row.get("action") for row in rows)
    applicable_rows = [row for row in rows if row.get("applicable")]
    approved_applicable = [result for result in results if result.get("approved") and result.get("action") in APPLICABLE_ACTIONS]
    return {
        "status": "review" if rows else "ok",
        "mode": "apply" if apply else "dry_run",
        "month": resolved_month,
        "packet": str(packet_path),
        "approval_file": str(approval_path) if approval_path else None,
        "conflict_count": len(rows),
        "applicable_count": len(applicable_rows),
        "blocked_count": len(rows) - len(applicable_rows),
        "approved_applicable_count": len(approved_applicable),
        "status_counts": dict(sorted(status_counts.items())),
        "action_counts": dict(sorted(action_counts.items())),
        "results": results,
    }


def write_markdown(report: dict[str, Any], path: Path) -> None:
    lines = [
        "# Baselane CF Conflict Resolution Plan",
        "",
        f"- Status: {report['status']}",
        f"- Mode: {report['mode']}",
        f"- Month: {report.get('month')}",
        f"- Conflicts: {report['conflict_count']}",
        f"- Applicable with approval: {report['applicable_count']}",
        f"- Blocked/manual-only: {report['blocked_count']}",
        f"- Approved applicable: {report['approved_applicable_count']}",
        "",
        "## Status Counts",
    ]
    for status, count in report["status_counts"].items():
        lines.append(f"- {status}: {count}")
    lines.extend(["", "## Action Counts"])
    for action, count in report["action_counts"].items():
        lines.append(f"- {action}: {count}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create/apply guarded approvals for Baselane CF conflict rows.")
    parser.add_argument("--packet", required=True, type=Path)
    parser.add_argument("--month", default=None)
    parser.add_argument("--approval-json", type=Path, default=None)
    parser.add_argument("--approval-template-out", type=Path, default=None)
    parser.add_argument("--verified-voids-json", type=Path, default=None)
    parser.add_argument("--source-transaction-index", type=Path, default=None)
    parser.add_argument("--report-json", required=True, type=Path)
    parser.add_argument("--report-markdown", required=True, type=Path)
    parser.add_argument("--apply", action="store_true", help="Write approved exact-match workbook cells. Defaults to dry-run.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    packet = load_packet(args.packet)
    month = infer_month(packet, args.month)
    verified_voids = load_verified_voids(args.verified_voids_json)
    transactions = load_transaction_index(args.source_transaction_index)
    if args.approval_template_out:
        template = approval_template(packet, month, verified_voids, transactions)
        args.approval_template_out.parent.mkdir(parents=True, exist_ok=True)
        args.approval_template_out.write_text(json.dumps(template, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    report = build_report(args.packet, args.approval_json, month, args.apply, verified_voids, transactions)
    args.report_json.parent.mkdir(parents=True, exist_ok=True)
    args.report_json.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    write_markdown(report, args.report_markdown)
    print(json.dumps({k: report[k] for k in ("status", "mode", "conflict_count", "applicable_count", "approved_applicable_count")}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
