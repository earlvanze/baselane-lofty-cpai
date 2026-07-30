#!/usr/bin/env python3
"""Remove repeated template mortgage balances from canonical CF workbooks."""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import tempfile
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from lofty_monthly_exclusions import DEFAULT_MANUAL_EXCLUDED_PROPERTIES
except ImportError:
    DEFAULT_MANUAL_EXCLUDED_PROPERTIES = (
        "3560 Saint Albans Rd",
        "1935 S Glen Rd",
        "402 N Wild Olive Ave",
    )

try:
    from coownership_mortgage_policy import (
        is_no_dao_mortgage_property,
        is_yhome_stolen_deed_mortgage_property,
    )
except ImportError:
    def is_no_dao_mortgage_property(value: Any) -> bool:
        normalized = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
        return any(key in normalized for key in ("86 madison ave", "88 madison ave", "90 madison ave", "724 3rd ave"))

    def is_yhome_stolen_deed_mortgage_property(value: Any) -> bool:
        return False


OWNER_STATEMENTS_DIR = "07 - P&L & Owner Statements"
MORTGAGE_BALANCE_LABEL = "mortgage principal balance"
TEMPLATE_SENTINEL = Decimal("1200000")
SENTINEL_TOLERANCE = Decimal("0.01")
REPEATED_SENTINEL_MINIMUM = 6
NO_MORTGAGE_STATES = {"IL", "OH", "TN"}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\blfty\d+\b", " ", text)
    text = re.sub(r"\bpublic\b", " ", text)
    replacements = {
        "avenue": "ave",
        "street": "st",
        "road": "rd",
        "lane": "ln",
        "north": "n",
        "south": "s",
        "east": "e",
        "west": "w",
    }
    for source, target in replacements.items():
        text = re.sub(rf"\b{source}\b", target, text)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def is_manually_excluded(path: Path) -> bool:
    haystack = normalize(str(path))
    return any(normalize(name) and normalize(name) in haystack for name in DEFAULT_MANUAL_EXCLUDED_PROPERTIES)


def is_canonical_cf_file(path: Path) -> bool:
    if path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    lowered_name = path.name.lower()
    if not lowered_name.startswith("cash flow statement"):
        return False
    if any(marker in lowered_name for marker in ("conflict", ".before-", ".backup", " backup", "-backup")):
        return False
    if any(part.casefold() == "_archive" for part in path.parts):
        return False
    statement_dir = path.parent
    if statement_dir.name.casefold() != OWNER_STATEMENTS_DIR.casefold():
        return False
    owner_dir = statement_dir.parent
    owner_name = owner_dir.name.casefold()
    return owner_name == "public" or owner_name.endswith(" public")


def discover_canonical_cf_files(real_estate_root: Path) -> list[Path]:
    if not real_estate_root.is_dir():
        return []
    find_result = subprocess.run(
        [
            "find",
            str(real_estate_root),
            "-type",
            "f",
            "-path",
            f"*/{OWNER_STATEMENTS_DIR}/Cash Flow Statement*.xlsx",
            "-print0",
        ],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    paths: list[Path] = []
    for raw_path in find_result.stdout.split(b"\0"):
        if not raw_path:
            continue
        candidate = Path(os.fsdecode(raw_path))
        if is_canonical_cf_file(candidate):
            paths.append(candidate)
    return sorted(paths, key=lambda path: str(path).casefold())


def parse_number(value: Any) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("$", "").replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def is_template_sentinel(value: Any) -> bool:
    number = parse_number(value)
    return number is not None and abs(number - TEMPLATE_SENTINEL) <= SENTINEL_TOLERANCE


def state_for_path(real_estate_root: Path, path: Path) -> str:
    try:
        relative = path.relative_to(real_estate_root)
    except ValueError:
        return ""
    return relative.parts[0] if relative.parts else ""


def zero_policy_applies(real_estate_root: Path, path: Path) -> bool:
    state = state_for_path(real_estate_root, path).upper()
    if state in NO_MORTGAGE_STATES:
        return not is_yhome_stolen_deed_mortgage_property(str(path))
    return is_no_dao_mortgage_property(str(path)) and not is_yhome_stolen_deed_mortgage_property(str(path))


def month_value(sheet: Any, column: int) -> str | None:
    value = sheet.cell(row=1, column=column).value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    text = str(value or "").strip()
    return text or None


def scan_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    sentinel_cells: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            for row_number in range(1, (sheet.max_row or 0) + 1):
                label = str(sheet.cell(row=row_number, column=1).value or "").strip().casefold()
                if label != MORTGAGE_BALANCE_LABEL:
                    continue
                for column in range(2, min(sheet.max_column or 0, 13) + 1):
                    value = sheet.cell(row=row_number, column=column).value
                    if is_template_sentinel(value):
                        sentinel_cells.append(
                            {
                                "sheet": str(sheet.title),
                                "cell": sheet.cell(row=row_number, column=column).coordinate,
                                "month": month_value(sheet, column),
                                "value": str(value),
                            }
                        )
    finally:
        workbook.close()
    repeated = len(sentinel_cells) >= REPEATED_SENTINEL_MINIMUM
    return {
        "sentinel_cells": sentinel_cells,
        "sentinel_cell_count": len(sentinel_cells),
        "repeated_sentinel": repeated,
    }


def scan_workbook_isolated(path: Path) -> dict[str, Any]:
    result = subprocess.run(
        [os.environ.get("PYTHON", "python3"), str(Path(__file__).resolve()), "--scan-file", str(path)],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    if result.returncode != 0:
        detail = result.stderr.strip() or f"isolated scan exited with return code {result.returncode}"
        raise RuntimeError(detail)
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"isolated scan returned invalid JSON: {result.stdout!r}") from exc


def save_workbook_atomically(path: Path, workbook: Any) -> None:
    suffix = path.suffix or ".xlsx"
    file_descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}.", suffix=suffix, dir=path.parent)
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def repair_workbook(path: Path, real_estate_root: Path, *, apply: bool, scan: dict[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {
        "file": str(path),
        "state": state_for_path(real_estate_root, path),
        "property_folder": path.parent.parent.name,
        "sentinel_cell_count": int(scan["sentinel_cell_count"]),
        "repeated_sentinel": bool(scan["repeated_sentinel"]),
        "candidate": bool(scan["repeated_sentinel"]),
        "changed": False,
        "changes": [],
    }
    if not scan["repeated_sentinel"]:
        record["remaining_sentinel_cell_count"] = 0
        return record
    if not apply:
        record["remaining_sentinel_cell_count"] = int(scan["sentinel_cell_count"])
        return record
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    changes: list[dict[str, Any]] = []
    replacement = 0 if zero_policy_applies(real_estate_root, path) else None
    try:
        for sheet in workbook.worksheets:
            for row_number in range(1, (sheet.max_row or 0) + 1):
                label = str(sheet.cell(row=row_number, column=1).value or "").strip().casefold()
                if label != MORTGAGE_BALANCE_LABEL:
                    continue
                for column in range(2, min(sheet.max_column or 0, 13) + 1):
                    cell = sheet.cell(row=row_number, column=column)
                    if not is_template_sentinel(cell.value):
                        continue
                    changes.append(
                        {
                            "sheet": str(sheet.title),
                            "cell": cell.coordinate,
                            "month": month_value(sheet, column),
                            "before": str(cell.value),
                            "after": replacement,
                        }
                    )
                    cell.value = replacement
        if changes:
            save_workbook_atomically(path, workbook)
    finally:
        workbook.close()
    record["changed"] = bool(changes)
    record["changes"] = changes
    record["replacement_policy"] = "zero_no_mortgage_policy" if replacement == 0 else "blank_until_authoritative_mortgage_source"
    after = scan_workbook(path)
    record["remaining_sentinel_cell_count"] = int(after["sentinel_cell_count"])
    return record


def build_report(real_estate_root: Path, *, apply: bool) -> dict[str, Any]:
    files = discover_canonical_cf_files(real_estate_root)
    records: list[dict[str, Any]] = []
    skipped_excluded_count = 0
    unreadable_count = 0
    for path in files:
        if is_manually_excluded(path):
            skipped_excluded_count += 1
            continue
        try:
            scan = scan_workbook_isolated(path)
            records.append(repair_workbook(path, real_estate_root, apply=apply, scan=scan))
        except Exception as exc:
            unreadable_count += 1
            records.append({"file": str(path), "status": "read_error", "error": str(exc), "candidate": False, "changed": False})
    candidate_records = [record for record in records if record.get("candidate")]
    changed_records = [record for record in records if record.get("changed")]
    remaining_sentinel_count = sum(int(record.get("remaining_sentinel_cell_count") or 0) for record in records)
    unresolved_candidates = [
        record for record in records if record.get("candidate") and int(record.get("remaining_sentinel_cell_count") or (record.get("sentinel_cell_count") if not apply else 0)) > 0
    ]
    status = "ok" if unreadable_count == 0 and not unresolved_candidates else "review"
    return {
        "generated_at": utc_now(),
        "status": status,
        "apply": apply,
        "real_estate_root": str(real_estate_root),
        "canonical_scope": "Public/07 - P&L & Owner Statements only; private siblings, archives, conflicts, and manual exclusions are not modified",
        "template_sentinel": float(TEMPLATE_SENTINEL),
        "repeated_sentinel_minimum": REPEATED_SENTINEL_MINIMUM,
        "file_count": len(files),
        "scanned_file_count": len(records),
        "skipped_excluded_count": skipped_excluded_count,
        "unreadable_count": unreadable_count,
        "candidate_workbook_count": len(candidate_records),
        "changed_workbook_count": len(changed_records),
        "sentinel_cell_count": sum(int(record.get("sentinel_cell_count") or 0) for record in records),
        "changed_cell_count": sum(len(record.get("changes") or []) for record in records),
        "remaining_sentinel_cell_count": remaining_sentinel_count,
        "records": records,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Guard canonical CF mortgage balances against repeated template sentinel values.")
    parser.add_argument("--real-estate-root", type=Path, default=Path("/mnt/c/Users/digit/Dropbox/Real Estate"))
    parser.add_argument("--report", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--scan-file", type=Path, help=argparse.SUPPRESS)
    args = parser.parse_args(argv)
    if args.scan_file is not None:
        print(json.dumps(scan_workbook(args.scan_file), sort_keys=True))
        return 0
    if args.report is None:
        parser.error("--report is required unless --scan-file is used")
    report = build_report(args.real_estate_root, apply=args.apply)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({key: report[key] for key in ("status", "apply", "file_count", "candidate_workbook_count", "changed_cell_count", "remaining_sentinel_cell_count")}, sort_keys=True))
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
