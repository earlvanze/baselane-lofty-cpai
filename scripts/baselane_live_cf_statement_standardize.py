#!/usr/bin/env python3
"""Standardize live Lofty CF statements from per-property source files.

This is the targeted bridge for the Lofty PM monthly pipeline:
- scope is the current live Lofty financial capture records
- monthly P&L rows come from each property's split ECO GL CSV
- the complete property-split ECO GL Column E balance remains an internal
  accounting control; it is distinct from ECO Net DAO Funds (spendable),
  physical bank cash, monthly cash flow, and transfer amounts
- selected balance-sheet/tokenomics cells come from live Lofty fields

Applying more than one month at a time is intentionally blocked. The live
Lofty capture contains the current token snapshot, not historical monthly
snapshots, so reusing it for prior columns would silently corrupt history.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import os
import re
import shutil
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

try:
    from lofty_monthly_exclusions import match_exclusion_guard, monthly_exclusion_guards
except ImportError:  # pragma: no cover - workspace script imports are available in production
    match_exclusion_guard = None
    monthly_exclusion_guards = None


ROOT = Path(
    os.environ.get("OPENCLAW_WORKSPACE_ROOT")
    or os.environ.get("OPENCLAW_WORKSPACE")
    or Path(__file__).resolve().parents[1]
).absolute()
UPDATE_CF_SCRIPT = ROOT / "skills" / "baselane-financials" / "scripts" / "update_cf_statements.py"
DEFAULT_LIVE_CAPTURE = ROOT / "reports/baselane_financials_monthly_live_financial_capture.json"
DEFAULT_LIVE_ROWS = ROOT / "tmp/utilities-six-figure-current/get-manager-properties.current2.json"
DEFAULT_REPORT = ROOT / "reports/baselane_live_cf_statement_standardize_report.json"
DEFAULT_FINANCE_TRUTH_REFRESH = ROOT / "reports/baselane_monthly_finance_truth_refresh.json"
DEFAULT_SOURCE_CASH_REPORT = ROOT / "reports/baselane_daily_source_cash_balance_report.json"
DEFAULT_SOURCE_FIX_PLAN = ROOT / "reports/baselane_ecogl_source_fix_plan.json"
DEFAULT_UNTAGGED_REVIEW_REPORT = ROOT / "reports/baselane_cf_untagged_review_packet.json"
DEFAULT_YHOME_CSV = ROOT / "reports/yhome_transition_reconciliation.csv"
SOURCE_CASH_REPORT_MAX_AGE_HOURS = 36
UNTAGGED_REVIEW_REPORT_MAX_AGE_HOURS = 36
DEFAULT_MASTER_GL = Path(
    os.environ.get("BASELANE_LEDGER_PATH")
    or "/mnt/c/Users/digit/Dropbox/Projects/assetrail/ECO Systems General Ledger.csv"
)


def load_update_cf_module():
    spec = importlib.util.spec_from_file_location("update_cf_statements", UPDATE_CF_SCRIPT)
    if not spec or not spec.loader:
        raise RuntimeError(f"Cannot load {UPDATE_CF_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def iso_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {"status": "missing", "path": str(path)}
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else {"status": "unreadable", "path": str(path)}


def finance_truth_apply_gate(path: Path) -> dict[str, Any]:
    """Allow workbook writes only after a verified Baselane truth refresh."""
    if not path.is_file():
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_report_missing",
            "report": str(path),
        }
    payload = read_json(path)
    if payload.get("status") not in {"ok", "ok_dry_run"}:
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_not_ok",
            "report": str(path),
            "source_status": payload.get("status"),
        }
    if payload.get("auth_blocked") is True or payload.get("cdp_blocked") is True:
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_auth_or_cdp_blocked",
            "report": str(path),
            "auth_blocked": payload.get("auth_blocked") is True,
            "cdp_blocked": payload.get("cdp_blocked") is True,
        }
    if payload.get("downstream_generation_allowed") is not True:
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_downstream_generation_not_allowed",
            "report": str(path),
        }
    return {
        "allowed": True,
        "status": "ok",
        "reason": "verified_finance_truth_refresh",
        "report": str(path),
        "source_generated_at": payload.get("generated_at"),
    }


def balance_sheet_source_apply_gate(path: Path) -> dict[str, Any]:
    """Allow source-row repairs from verified local truth, even if live auth is stale."""
    if not path.is_file():
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_report_missing",
            "report": str(path),
        }
    payload = read_json(path)
    if payload.get("status") not in {"ok", "ok_dry_run"}:
        if (
            payload.get("status") == "failed"
            and payload.get("failed_step") == "baselane_auth_preflight"
            and payload.get("auth_blocked") is True
            and payload.get("cdp_blocked") is True
            and payload.get("safe_to_retry_after_manual_auth") is True
        ):
            return {
                "allowed": True,
                "status": "ok",
                "reason": "verified_local_source_row_repair_auth_independent",
                "report": str(path),
                "source_generated_at": payload.get("generated_at"),
                "live_auth_blocked": True,
            }
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_not_ok",
            "report": str(path),
            "source_status": payload.get("status"),
        }
    if payload.get("auth_blocked") is True or payload.get("cdp_blocked") is True:
        return {
            "allowed": False,
            "status": "blocked",
            "reason": "finance_truth_refresh_auth_or_cdp_blocked",
            "report": str(path),
        }
    return {
        "allowed": True,
        "status": "ok",
        "reason": "verified_local_finance_truth_refresh_for_source_row_only",
        "report": str(path),
        "source_generated_at": payload.get("generated_at"),
    }


def source_cash_report_entries(path: Path) -> dict[str, Any]:
    """Load the bounded daily-audit manifest used to select canonical split GLs."""
    payload = read_json(path)
    manifest_fields = (
        "source_cash_balance_mode",
        "violation_count",
        "missing_row_count",
        "missing_month_column_count",
        "unreadable_count",
        "noncanonical_source_count",
        "blocking_no_match_count",
        "split_scope_missing_property_count",
        "apply_blocked_by_raw_no_dao_mortgage_guard",
    )
    entries: dict[str, dict[str, Any]] = {}
    for item in payload.get("checked_workbooks_bounded") or []:
        if not isinstance(item, dict):
            continue
        property_name = str(item.get("property") or "").strip()
        raw_sources = item.get("source_cash_sources")
        source_strings = (
            [str(source) for source in raw_sources if str(source).strip()]
            if isinstance(raw_sources, list)
            else [str(item.get("source_cash_source") or "")]
        )
        sources = [Path(source) for source in source_strings]
        if property_name and sources and all(source.is_file() for source in sources):
            entries[normalize_name(property_name)] = {**item, "source_cash_sources": source_strings}
    return {
        "report": str(path),
        "status": payload.get("status", "missing"),
        "generated_at": payload.get("generated_at"),
        "entry_count": len(entries),
        **{field: payload.get(field, 0) for field in manifest_fields},
        "entries": entries,
    }


def source_cash_report_entry(property_name: str, entries: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    """Match a live capture name to the audit's canonical GL property name."""
    exact = entries.get(normalize_name(property_name))
    if exact:
        return exact
    property_key = normalize_name(property_name)
    property_tokens = set(property_key.split())
    property_number = next((token for token in property_key.split() if token.isdigit()), None)
    best: tuple[int, int, str, dict[str, Any]] | None = None
    for source_key, entry in entries.items():
        source_tokens = set(source_key.split())
        source_number = next((token for token in source_key.split() if token.isdigit()), None)
        if property_number and source_number and property_number != source_number:
            continue
        shared = len(property_tokens & source_tokens)
        if shared < min(3, len(property_tokens), len(source_tokens)):
            continue
        ranked = (shared, -len(source_tokens), source_key, entry)
        if best is None or ranked[:3] > best[:3]:
            best = ranked
    return best[3] if best else None


def source_cash_apply_gate(
    records: list[dict[str, Any]],
    manifest: dict[str, Any],
    *,
    allow_workbook_repair: bool = False,
) -> dict[str, Any]:
    """Require fresh canonical split-GL coverage before any workbook write."""
    status = str(manifest.get("status") or "missing")
    violation_count = int(manifest.get("violation_count") or 0)
    source_quality_blockers = {
        "missing_row_count": int(manifest.get("missing_row_count") or 0),
        "missing_month_column_count": int(manifest.get("missing_month_column_count") or 0),
        "unreadable_count": int(manifest.get("unreadable_count") or 0),
        "noncanonical_source_count": int(manifest.get("noncanonical_source_count") or 0),
        "blocking_no_match_count": int(manifest.get("blocking_no_match_count") or 0),
        "split_scope_missing_property_count": int(manifest.get("split_scope_missing_property_count") or 0),
        "raw_no_dao_mortgage_guard": int(bool(manifest.get("apply_blocked_by_raw_no_dao_mortgage_guard"))),
    }
    repairable_or_out_of_scope = {
        "missing_row_count",
        "missing_month_column_count",
        "blocking_no_match_count",
        "split_scope_missing_property_count",
    }
    effective_blockers = {
        key: value
        for key, value in source_quality_blockers.items()
        if not (allow_workbook_repair and key in repairable_or_out_of_scope)
    }
    if status not in {"ok", "review"} or any(effective_blockers.values()):
        return {
            "allowed": False,
            "reason": "source_cash_report_not_usable",
            "status": status,
            "violation_count": violation_count,
            **source_quality_blockers,
        }
    generated_at = str(manifest.get("generated_at") or "")
    try:
        generated = datetime.fromisoformat(generated_at.replace("Z", "+00:00"))
        age_hours = (datetime.now(timezone.utc) - generated).total_seconds() / 3600
    except ValueError:
        return {"allowed": False, "reason": "source_cash_report_timestamp_invalid", "status": status}
    if age_hours > SOURCE_CASH_REPORT_MAX_AGE_HOURS:
        return {
            "allowed": False,
            "reason": "source_cash_report_stale",
            "status": status,
            "age_hours": round(age_hours, 2),
            "max_age_hours": SOURCE_CASH_REPORT_MAX_AGE_HOURS,
        }
    entries = manifest.get("entries") or {}
    missing: list[str] = []
    noncanonical: list[str] = []
    for record in records:
        property_name = str(record.get("property_name") or Path(str(record.get("input_property_path"))).name)
        if "package" in normalize_name(property_name):
            continue
        entry = source_cash_report_entry(property_name, entries)
        if not entry:
            missing.append(property_name)
        elif entry.get("source_cash_source_mode") not in {
            "canonical_property_split_gl",
            "canonical_aggregate_property_split_gl",
        }:
            noncanonical.append(property_name)
    if missing or noncanonical:
        return {
            "allowed": False,
            "reason": "source_cash_report_incomplete_canonical_coverage",
            "status": status,
            "age_hours": round(age_hours, 2),
            "missing_properties": missing,
            "noncanonical_properties": noncanonical,
        }
    return {
        "allowed": True,
        "reason": (
            "fresh_canonical_split_gl_coverage_with_downstream_balance_violations"
            if violation_count
            else "fresh_canonical_split_gl_coverage"
        ),
        "status": status,
        "violation_count": violation_count,
        "age_hours": round(age_hours, 2),
        "entry_count": len(entries),
        "allow_workbook_repair": allow_workbook_repair,
        "ignored_downstream_or_out_of_scope_counts": (
            {
                key: source_quality_blockers[key]
                for key in sorted(repairable_or_out_of_scope)
                if source_quality_blockers[key]
            }
            if allow_workbook_repair
            else {}
        ),
    }


def unresolved_source_fix_entries(path: Path) -> dict[str, Any]:
    """Group unresolved upstream source defects for property-scoped workbook holds."""
    payload = read_json(path)
    resolved_statuses = {
        "already applied",
        "cleared",
        "not applicable",
        "resolved",
        "verified fixed",
    }
    entries: dict[str, dict[str, Any]] = {}
    for action in payload.get("actions") or []:
        if not isinstance(action, dict):
            continue
        automation_status = normalize_name(str(action.get("automation_status") or ""))
        action_status = normalize_name(str(action.get("status") or ""))
        if automation_status in resolved_statuses or action_status in resolved_statuses:
            continue
        property_name = str(action.get("property") or action.get("property_name") or "").strip()
        if not property_name:
            continue
        key = normalize_name(property_name)
        entry = entries.setdefault(
            key,
            {
                "property": property_name,
                "actions": [],
            },
        )
        entry["actions"].append(action)
    return {
        "status": payload.get("status") or "missing",
        "generated_at": payload.get("generated_at"),
        "action_count": sum(len(entry["actions"]) for entry in entries.values()),
        "property_count": len(entries),
        "entries": entries,
        "path": str(path),
    }


def source_fix_entry(property_name: str, manifest: dict[str, Any]) -> dict[str, Any] | None:
    return source_cash_report_entry(property_name, manifest.get("entries") or {})


def untagged_review_apply_gate(path: Path | None) -> dict[str, Any]:
    """Require all fallback rows to be explicitly tagged before writes."""
    if path is None:
        return {"allowed": True, "reason": "untagged_review_gate_not_configured"}
    if not path.is_file():
        return {
            "allowed": False,
            "reason": "untagged_review_report_missing",
            "report": str(path),
        }
    payload = read_json(path)
    status = str(payload.get("status") or "missing")
    raw_review_required_count = int(payload.get("review_required_count") or 0)
    raw_untagged_row_count = int(payload.get("untagged_row_count") or 0)
    review_required_count = int(payload.get("effective_review_required_count", raw_review_required_count) or 0)
    untagged_row_count = int(payload.get("effective_untagged_row_count", raw_untagged_row_count) or 0)
    generated_at = str(payload.get("generated_at") or "")
    try:
        generated = datetime.fromisoformat(generated_at.replace("Z", "+00:00"))
        age_hours = (datetime.now(timezone.utc) - generated).total_seconds() / 3600
    except ValueError:
        return {
            "allowed": False,
            "reason": "untagged_review_report_timestamp_invalid",
            "report": str(path),
        }
    if age_hours > UNTAGGED_REVIEW_REPORT_MAX_AGE_HOURS:
        return {
            "allowed": False,
            "reason": "untagged_review_report_stale",
            "report": str(path),
            "age_hours": round(age_hours, 2),
            "max_age_hours": UNTAGGED_REVIEW_REPORT_MAX_AGE_HOURS,
        }
    if status != "ok" or review_required_count or untagged_row_count:
        return {
            "allowed": False,
            "reason": "baselane_untagged_rows_require_explicit_tagging",
            "report": str(path),
            "status": status,
            "age_hours": round(age_hours, 2),
            "untagged_row_count": untagged_row_count,
            "review_required_count": review_required_count,
            "raw_untagged_row_count": raw_untagged_row_count,
            "raw_review_required_count": raw_review_required_count,
        }
    return {
        "allowed": True,
        "reason": "all_source_rows_explicitly_tagged",
        "report": str(path),
        "status": status,
        "age_hours": round(age_hours, 2),
        "untagged_row_count": 0,
        "review_required_count": 0,
        "raw_untagged_row_count": raw_untagged_row_count,
        "raw_review_required_count": raw_review_required_count,
    }


def parse_months(months: list[str]) -> list[tuple[int, int]]:
    parsed = []
    for month in months:
        year_text, month_text = month.split("-", 1)
        parsed.append((int(year_text), int(month_text)))
    return parsed


def live_capture_records(path: Path) -> list[dict[str, Any]]:
    payload = read_json(path)
    records = payload.get("records")
    if not isinstance(records, list):
        return []
    return [record for record in records if isinstance(record, dict) and record.get("input_property_path")]


def filter_excluded_live_capture_records(
    records: list[dict[str, Any]],
    yhome_csv: Path | None = DEFAULT_YHOME_CSV,
    runtime_map: Path | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Keep sold/offboarded properties out of CF source selection and writes."""
    guards = []
    if match_exclusion_guard and monthly_exclusion_guards:
        guards, _yhome_guard, _manual_exclusions = monthly_exclusion_guards(yhome_csv)

    runtime_exclusions: dict[str, dict[str, Any]] = {}
    if runtime_map and runtime_map.is_file():
        payload = read_json(runtime_map)
        runtime_records = payload.get("records") if isinstance(payload, dict) else None
        for row in runtime_records if isinstance(runtime_records, list) else []:
            if not isinstance(row, dict):
                continue
            status = str(row.get("status") or row.get("index_status") or "").strip().lower()
            index_status = str(row.get("index_status") or "").strip().lower()
            if not (
                status.startswith(("excluded_", "skipped_", "omitted_"))
                or index_status.startswith(("excluded_", "skipped_", "omitted_"))
            ):
                continue
            for value in (
                row.get("lofty_property_id"),
                row.get("property_id"),
                row.get("input_property_path"),
                row.get("property_path"),
            ):
                key = str(value or "").strip()
                if key:
                    runtime_exclusions[key] = row

    included: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for record in records:
        property_path = Path(str(record.get("input_property_path") or record.get("property_path") or ""))
        runtime_match = next(
            (
                runtime_exclusions.get(str(value or "").strip())
                for value in (
                    record.get("lofty_property_id"),
                    record.get("property_id"),
                    record.get("input_property_path"),
                    record.get("property_path"),
                )
                if str(value or "").strip() in runtime_exclusions
            ),
            None,
        )
        match = match_exclusion_guard(property_path, guards) if match_exclusion_guard else None
        if runtime_match:
            excluded.append(
                {
                    "property": str(record.get("property_name") or property_path.name),
                    "property_path": str(property_path),
                    "source": "lofty_pm_runtime_map",
                    "exclude_reason": str(
                        runtime_match.get("exclude_reason")
                        or runtime_match.get("status")
                        or runtime_match.get("index_status")
                    ),
                }
            )
            continue
        if match:
            excluded.append(
                {
                    "property": str(record.get("property_name") or property_path.name),
                    "property_path": str(property_path),
                    "source": match.get("source"),
                    "exclude_reason": match.get("exclude_reason"),
                }
            )
        else:
            included.append(record)
    return included, excluded


def flatten_live_rows(payload: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    stack = [payload]
    while stack:
        item = stack.pop()
        if isinstance(item, list):
            if item and all(isinstance(child, dict) for child in item) and any(
                "address" in child or "utilities" in child for child in item
            ):
                rows.extend(child for child in item if isinstance(child, dict))
            else:
                stack.extend(item)
        elif isinstance(item, dict):
            stack.extend(item.values())
    return rows


def live_rows_by_id(path: Path) -> dict[str, dict[str, Any]]:
    payload = read_json(path)
    rows = {}
    for row in flatten_live_rows(payload):
        property_id = str(row.get("id") or row.get("property_id") or "").strip()
        if property_id:
            rows[property_id] = row
    return rows


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\bpublic\b", " ", text)
    replacements = {
        "avenue": "ave",
        "street": "st",
        "road": "rd",
        "lane": "ln",
        "drive": "dr",
        "place": "pl",
        "circle": "cir",
        "north": "n",
        "south": "s",
        "east": "e",
        "west": "w",
    }
    for source, target in replacements.items():
        text = re.sub(rf"\b{source}\b", target, text)
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def matching_public_sibling(property_path: Path) -> Path | None:
    if property_path.name.lower().endswith(" public"):
        return property_path
    parent = property_path.parent
    if not parent.is_dir():
        return None
    source_key = normalize_name(property_path.name)
    source_tokens = set(source_key.split())
    best: tuple[int, str, Path] | None = None
    for candidate in parent.glob("* Public"):
        if not candidate.is_dir():
            continue
        candidate_key = normalize_name(candidate.name)
        candidate_tokens = set(candidate_key.split())
        if not source_tokens or not candidate_tokens:
            continue
        source_number = next((token for token in source_key.split() if token.isdigit()), None)
        candidate_number = next((token for token in candidate_key.split() if token.isdigit()), None)
        shared = len(source_tokens & candidate_tokens)
        if source_number and candidate_number and source_number != candidate_number:
            continue
        if shared < min(len(source_tokens), 3):
            continue
        score = -shared
        ranked = (score, str(candidate).lower(), candidate)
        if best is None or ranked < best:
            best = ranked
    return best[2] if best else None


def source_paths(property_path: Path) -> list[Path]:
    paths = [property_path]
    sibling = matching_public_sibling(property_path)
    if sibling and sibling not in paths:
        paths.insert(0, sibling)
    nested_public = property_path / "Public"
    if nested_public.is_dir() and nested_public not in paths:
        paths.insert(0, nested_public)
    return paths


def owner_statement_dirs(property_path: Path) -> list[Path]:
    dirs: list[Path] = []
    for source_path in source_paths(property_path):
        owner_dir = source_path / "07 - P&L & Owner Statements"
        if owner_dir.is_dir() and owner_dir not in dirs:
            dirs.append(owner_dir)
    return dirs


def csv_row_count(path: Path) -> int:
    try:
        with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
            return max(sum(1 for _ in handle) - 1, 0)
    except OSError:
        return 0


def component_ledger_key(path: Path) -> str:
    stem = re.sub(r"^ECO Systems General Ledger\s*-\s*", "", path.stem, flags=re.I)
    stem = re.sub(r"\s*-\s*Ohio 3 Property Package.*$", "", stem, flags=re.I)
    return normalize_name(stem)


def split_ledger_paths(property_path: Path, property_name: str) -> list[Path]:
    for owner_dir in owner_statement_dirs(property_path):
        candidates = sorted(owner_dir.glob("ECO Systems General Ledger*.csv"))
        if not candidates:
            continue
        if "package" not in normalize_name(property_name) or len(candidates) == 1:
            selected = max(
                candidates,
                key=lambda path: (csv_row_count(path), path.stat().st_mtime, path.name.lower()),
            )
            return [selected]
        best_by_component: dict[str, Path] = {}
        for candidate in candidates:
            key = component_ledger_key(candidate)
            current = best_by_component.get(key)
            if current is None:
                best_by_component[key] = candidate
                continue
            candidate_rank = (csv_row_count(candidate), candidate.stat().st_mtime, len(candidate.name))
            current_rank = (csv_row_count(current), current.stat().st_mtime, len(current.name))
            if candidate_rank > current_rank:
                best_by_component[key] = candidate
        return sorted(best_by_component.values(), key=lambda path: str(path).lower())
    return []


def transaction_csv_fallback_path(property_path: Path) -> Path | None:
    for owner_dir in owner_statement_dirs(property_path):
        candidates = [
            path
            for path in owner_dir.glob("Transactions*.csv")
            if path.is_file() and "conflict" not in path.name.lower()
        ]
        if candidates:
            return max(candidates, key=lambda path: (csv_row_count(path), path.stat().st_mtime, path.name))
    return None


def cf_workbook_path(property_path: Path) -> Path | None:
    candidates: list[Path] = []
    for owner_dir in owner_statement_dirs(property_path):
        candidates.extend(
            path
            for path in owner_dir.glob("Cash Flow Statement*.xlsx")
            if not any(
                marker in path.name.lower()
                for marker in (
                    "conflict",
                    "conflicted copy",
                    ".before-",
                    ".backup",
                    " backup",
                    "-backup",
                )
            )
        )
    if not candidates:
        return None

    def schema_priority(path: Path) -> tuple[int, int, str]:
        has_eco_gl_row = False
        has_eco_cash_row = False
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
            for sheet in workbook.worksheets:
                labels = {
                    str(sheet.cell(row=row, column=1).value or "").strip()
                    for row in range(1, min(sheet.max_row or 0, 120) + 1)
                }
                has_eco_gl_row = has_eco_gl_row or "ECO General Ledger (ECO GL Column E Total)" in labels
                has_eco_cash_row = has_eco_cash_row or update_cf_label in labels
                if has_eco_gl_row and has_eco_cash_row:
                    break
            workbook.close()
        except Exception:
            return (2, len(path.name), str(path).lower())
        return (0 if has_eco_gl_row and has_eco_cash_row else 1, len(path.name), str(path).lower())

    update_cf_label = "ECO Operating Cash"
    return sorted(candidates, key=schema_priority)[0]


def create_clean_cf_workbook(update_cf: Any, property_path: Path, property_name: str, year: int, apply: bool) -> dict[str, Any]:
    owner_dirs = owner_statement_dirs(property_path)
    target_dir = owner_dirs[0] if owner_dirs else property_path / "Public" / "07 - P&L & Owner Statements"
    target_path = target_dir / update_cf.safe_cf_filename(property_name)
    result = {
        "target_path": str(target_path),
        "template_path": str(update_cf.TEMPLATE_PATH) if update_cf.TEMPLATE_PATH.exists() else None,
        "status": "exists" if target_path.exists() else "would_create",
    }
    if target_path.exists() or not apply:
        return result
    if not update_cf.TEMPLATE_PATH.exists():
        result["status"] = "blocked"
        result["reason"] = "template_missing"
        return result
    target_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(update_cf.TEMPLATE_PATH, target_path)
    workbook = openpyxl.load_workbook(target_path)
    try:
        result["cleared_template_cell_count"] = update_cf.clear_template_workbook_data(workbook, year)
        no_mortgage_clear = update_cf.clear_no_mortgage_debt_rows(workbook, target_path)
        result["no_mortgage_debt_state"] = no_mortgage_clear["state"]
        result["no_mortgage_debt_checked"] = no_mortgage_clear["checked"]
        result["no_mortgage_debt_cleared_cell_count"] = no_mortgage_clear["cleared_cell_count"]
        workbook.save(target_path)
    finally:
        workbook.close()
    result["status"] = "created"
    return result


def load_split_transactions(update_cf: Any, ledger_paths: list[Path]) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []
    for ledger_path in ledger_paths:
        rows = update_cf.load_gl_data(ledger_path)
        for row in rows:
            row["_source_ledger"] = str(ledger_path)
        transactions.extend(rows)
    return transactions


def source_cash_mode_for_month(year: int, month: int, today: date | None = None) -> str:
    current_date = today or datetime.now(timezone.utc).date()
    return "as_of_month_end" if (year, month) < (current_date.year, current_date.month) else "full_column_e"


def source_cash_mode_for_manifest(
    year: int,
    month: int,
    source_cash_manifest: dict[str, Any] | None = None,
) -> str:
    """Honor a clean daily source-cash manifest when it declares cash semantics.

    The daily source-cash audit is the authoritative reconciliation for the
    same workbook cells.  Its declared mode must win over the standalone
    historical default so a later standardization pass cannot undo a verified
    source-cash normalization.
    """
    declared_mode = str((source_cash_manifest or {}).get("source_cash_balance_mode") or "")
    if declared_mode in {"full_column_e", "as_of_month_end"}:
        return declared_mode
    return source_cash_mode_for_month(year, month)


def month_end_source_summary(
    transactions: list[dict[str, Any]], year: int, month: int, source_cash_mode: str
) -> dict[str, Any]:
    return {
        "month": f"{year}-{month:02d}",
        **load_update_cf_module().source_cash_balance_for_month(
            transactions, year, month, source_cash_mode
        ),
    }


def apply_month_scope_gate(
    months: list[tuple[int, int]],
    apply: bool,
    p_and_l_only: bool = False,
    balance_sheet_source_only: bool = False,
) -> dict[str, Any]:
    """Require one statement month for tokenomics writes, not P&L-only writes."""
    if not apply or p_and_l_only or balance_sheet_source_only or len(months) <= 1:
        return {
            "allowed": True,
            "reason": "source_only_or_p_and_l_only_or_single_month_or_read_only",
            "month_count": len(months),
            "p_and_l_only": p_and_l_only,
            "balance_sheet_source_only": balance_sheet_source_only,
        }
    return {
        "allowed": False,
        "reason": "historical_tokenomics_requires_single_month_apply",
        "month_count": len(months),
        "requested_months": [f"{year:04d}-{month:02d}" for year, month in months],
    }


def p_and_l_row_scope(update_cf: Any) -> list[str]:
    """Return rows that may be sourced from month-specific ECO GL data.

    This deliberately excludes tokenomics and generic balance-sheet rows. ECO
    The canonical ECO General Ledger row is included using month-end history
    for closed columns and the full current balance for the active column.
    Physical ECO Operating Cash is a separate mapped-bank field and is never
    sourced from the property GL. No-mortgage debt rows are included so the
    policy can clear stale template debt without touching mortgage tokenomics.
    """
    rows = set(update_cf.ROW_TO_GL_CATEGORIES)
    rows.update(update_cf.FORMULA_ROWS)
    rows.update(update_cf.ECO_GL_NET_CASH_BALANCE_LABELS)
    rows.update(update_cf.NO_MORTGAGE_DEBT_ROW_LABELS)
    return sorted(rows)


def balance_sheet_source_row_scope(update_cf: Any) -> list[str]:
    """Return only full-Column-E internal accounting-control labels."""
    return [
        update_cf.ECO_GL_NET_CASH_BALANCE_LABEL,
        *update_cf.ECO_GL_NET_CASH_BALANCE_LEGACY_LABELS,
    ]


def parse_money(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text or text in {"-", "—"}:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def token_float(live_row: dict[str, Any]) -> tuple[float | None, str | None]:
    for key in ("currentTokenFloat", "numIssued", "num_issued", "issuedTokens", "issued_tokens", "num_sold", "tokens"):
        value = parse_money(live_row.get(key))
        if value is not None and value > 0:
            return value, key
    return None, None


def total_tokens(live_row: dict[str, Any]) -> tuple[float | None, str | None]:
    for key in ("tokens", "number_of_tokens", "numberOfTokens", "num_sold"):
        value = parse_money(live_row.get(key))
        if value is not None and value > 0:
            return value, key
    return None, None


def token_price(live_row: dict[str, Any]) -> tuple[float, str]:
    for key in ("initialTokenPrice", "originalTokenPrice", "offeringTokenPrice", "token_price"):
        value = parse_money(live_row.get(key))
        if value is not None and value > 0:
            return value, key
    return 50.0, "fixed_lofty_token_price"


def find_row(sheet: Any, label: str) -> int | None:
    for row_number in range(1, (sheet.max_row or 0) + 1):
        if str(sheet.cell(row=row_number, column=1).value or "").strip() == label:
            return row_number
    return None


def set_cell(sheet: Any, label: str, column: int, value: float | int | None, changes: list[dict[str, Any]]) -> None:
    row = find_row(sheet, label)
    if row is None or value is None:
        return
    cell = sheet.cell(row=row, column=column)
    old_value = cell.value
    old_numeric = parse_money(old_value)
    if old_numeric is not None and abs(round(old_numeric - float(value), 2)) <= 0.01 and not str(old_value).startswith("="):
        return
    cell.value = value
    changes.append({"label": label, "cell": cell.coordinate, "old_value": old_value, "new_value": value})


def standardize_tokenomics(
    update_cf: Any,
    workbook_path: Path,
    *,
    year: int,
    month: int,
    live_row: dict[str, Any] | None,
    apply: bool,
) -> list[dict[str, Any]]:
    if not live_row:
        return []
    workbook = openpyxl.load_workbook(workbook_path)
    try:
        sheet = update_cf.get_year_sheet(workbook, year)
        if sheet is None:
            return []
        column = update_cf.get_month_column(sheet, year, month)
        if column is None:
            return []
        floating, floating_source = token_float(live_row)
        tokens, tokens_source = total_tokens(live_row)
        price, price_source = token_price(live_row)
        current_loan = parse_money(live_row.get("current_loan")) or 0.0
        if update_cf.no_mortgage_debt_policy_applies(workbook_path):
            current_loan = 0.0
        total_market_cap = parse_money(live_row.get("total_investment"))
        if total_market_cap is None and tokens is not None:
            total_market_cap = round(tokens * price, 2)
        escrowed = None
        if tokens is not None and floating is not None:
            escrowed = max(round(tokens - floating, 8), 0.0)
        estimated_equity = None
        if total_market_cap is not None:
            estimated_equity = round(total_market_cap - current_loan, 2)
        changes: list[dict[str, Any]] = []
        set_cell(sheet, "Mortgage Principal Balance", column, current_loan, changes)
        set_cell(sheet, "Total Market Cap (excl. unapproved EARLDAO Interest)", column, total_market_cap, changes)
        set_cell(sheet, "Total Market Cap", column, total_market_cap, changes)
        set_cell(sheet, "Estimated Equity", column, estimated_equity, changes)
        set_cell(sheet, "Escrowed Tokens", column, escrowed, changes)
        set_cell(sheet, "Floating Shares", column, floating, changes)
        set_cell(sheet, "Total Shares", column, tokens, changes)
        set_cell(sheet, "Last Price Per Share", column, price, changes)
        if changes and apply:
            workbook.save(workbook_path)
        for change in changes:
            change["floating_token_source"] = floating_source
            change["total_token_source"] = tokens_source
            change["token_price_source"] = price_source
        return changes
    finally:
        workbook.close()


def build_report(args: argparse.Namespace) -> dict[str, Any]:
    update_cf = load_update_cf_module()
    months = parse_months(args.month)
    p_and_l_only = bool(args.p_and_l_only or args.apply_p_and_l_only)
    balance_sheet_source_only = bool(
        getattr(args, "balance_sheet_source_only", False)
        or getattr(args, "apply_balance_sheet_source_only", False)
    )
    write_requested = bool(
        args.apply
        or args.apply_p_and_l_only
        or getattr(args, "apply_balance_sheet_source_only", False)
    )
    month_scope_gate = apply_month_scope_gate(
        months,
        write_requested,
        p_and_l_only,
        balance_sheet_source_only,
    )
    mode_conflict = p_and_l_only and balance_sheet_source_only
    apply_gate = finance_truth_apply_gate(args.finance_truth_refresh)
    source_only_apply_gate = balance_sheet_source_apply_gate(args.finance_truth_refresh)
    write_apply_gate = source_only_apply_gate if balance_sheet_source_only else apply_gate
    all_live_records = live_capture_records(args.live_capture)
    records, excluded_live_records = filter_excluded_live_capture_records(
        all_live_records, args.yhome_csv, args.runtime_map
    )
    source_cash_manifest = source_cash_report_entries(args.source_cash_report)
    source_cash_gate = source_cash_apply_gate(
        records,
        source_cash_manifest,
        allow_workbook_repair=balance_sheet_source_only,
    )
    source_fix_manifest = unresolved_source_fix_entries(args.source_fix_plan)
    untagged_review_gate = untagged_review_apply_gate(args.untagged_review_report)
    if write_requested and not balance_sheet_source_only and source_cash_gate["allowed"] and not untagged_review_gate["allowed"]:
        source_cash_gate = {
            **source_cash_gate,
            "allowed": False,
            "reason": untagged_review_gate["reason"],
            "untagged_review": untagged_review_gate,
        }
    if mode_conflict:
        return {
            "status": "blocked",
            "generated_at": iso_z(),
            "apply": False,
            "apply_requested": write_requested,
            "apply_mode": "invalid_conflicting_modes",
            "p_and_l_only": p_and_l_only,
            "balance_sheet_source_only": balance_sheet_source_only,
            "write_blocked": True,
            "write_block_reason": "source_only_and_p_and_l_modes_are_mutually_exclusive",
            "finance_truth_refresh": apply_gate,
            "source_cash_report": {key: value for key, value in source_cash_manifest.items() if key != "entries"},
            "source_cash_apply_gate": source_cash_gate,
            "untagged_review_apply_gate": untagged_review_gate,
            "live_capture": str(args.live_capture),
            "live_rows": str(args.live_rows),
            "excluded_live_property_count": len(excluded_live_records),
            "excluded_live_properties": excluded_live_records,
            "months": args.month,
            "property_count": 0,
            "changed_property_count": 0,
            "missing_source_count": 0,
            "action_counts": {},
            "records": [],
        }
    if write_requested and not month_scope_gate["allowed"]:
        return {
            "status": "blocked",
            "generated_at": iso_z(),
            "apply": False,
            "apply_requested": write_requested,
            "apply_mode": "balance_sheet_source_only" if balance_sheet_source_only else ("p_and_l_only" if p_and_l_only else "full"),
            "p_and_l_only": p_and_l_only,
            "balance_sheet_source_only": balance_sheet_source_only,
            "write_blocked": True,
            "write_block_reason": month_scope_gate["reason"],
            "month_scope_gate": month_scope_gate,
            "finance_truth_refresh": apply_gate,
            "source_cash_report": {key: value for key, value in source_cash_manifest.items() if key != "entries"},
            "source_cash_apply_gate": source_cash_gate,
            "untagged_review_apply_gate": untagged_review_gate,
            "live_capture": str(args.live_capture),
            "live_rows": str(args.live_rows),
            "excluded_live_property_count": len(excluded_live_records),
            "excluded_live_properties": excluded_live_records,
            "months": args.month,
            "property_count": 0,
            "changed_property_count": 0,
            "missing_source_count": 0,
            "action_counts": {},
            "records": [],
        }
    if write_requested and not write_apply_gate["allowed"]:
        return {
            "status": "blocked",
            "generated_at": iso_z(),
            "apply": False,
            "apply_requested": write_requested,
            "apply_mode": "balance_sheet_source_only" if balance_sheet_source_only else ("p_and_l_only" if p_and_l_only else "full"),
            "p_and_l_only": p_and_l_only,
            "balance_sheet_source_only": balance_sheet_source_only,
            "write_blocked": True,
            "write_block_reason": write_apply_gate["reason"],
            "finance_truth_refresh": apply_gate,
            "source_only_apply_gate": source_only_apply_gate,
            "source_cash_report": {key: value for key, value in source_cash_manifest.items() if key != "entries"},
            "source_cash_apply_gate": source_cash_gate,
            "untagged_review_apply_gate": untagged_review_gate,
            "live_capture": str(args.live_capture),
            "live_rows": str(args.live_rows),
            "excluded_live_property_count": len(excluded_live_records),
            "excluded_live_properties": excluded_live_records,
            "months": args.month,
            "property_count": 0,
            "changed_property_count": 0,
            "missing_source_count": 0,
            "action_counts": {},
            "records": [],
        }
    if write_requested and not source_cash_gate["allowed"]:
        return {
            "status": "blocked",
            "generated_at": iso_z(),
            "apply": False,
            "apply_requested": write_requested,
            "apply_mode": "balance_sheet_source_only" if balance_sheet_source_only else ("p_and_l_only" if p_and_l_only else "full"),
            "p_and_l_only": p_and_l_only,
            "balance_sheet_source_only": balance_sheet_source_only,
            "write_blocked": True,
            "write_block_reason": source_cash_gate["reason"],
            "finance_truth_refresh": apply_gate,
            "source_cash_report": {key: value for key, value in source_cash_manifest.items() if key != "entries"},
            "source_cash_apply_gate": source_cash_gate,
            "untagged_review_apply_gate": untagged_review_gate,
            "live_capture": str(args.live_capture),
            "live_rows": str(args.live_rows),
            "excluded_live_property_count": len(excluded_live_records),
            "excluded_live_properties": excluded_live_records,
            "months": args.month,
            "property_count": 0,
            "changed_property_count": 0,
            "missing_source_count": 0,
            "action_counts": {},
            "records": [],
        }
    live_by_id = live_rows_by_id(args.live_rows)
    source_cash_entries = source_cash_manifest["entries"]
    row_scope = (
        balance_sheet_source_row_scope(update_cf)
        if balance_sheet_source_only
        else p_and_l_row_scope(update_cf)
        if p_and_l_only
        else None
    )
    all_records = []
    action_counts: Counter[str] = Counter()
    source_cash_mode_counts: Counter[str] = Counter()
    master_transactions: list[dict[str, Any]] | None = None
    master_properties: set[str] | None = None
    for record in records:
        property_path = Path(str(record.get("input_property_path")))
        property_name = str(record.get("property_name") or property_path.name)
        property_id = str(record.get("lofty_property_id") or "")
        source_fix_hold = source_fix_entry(property_name, source_fix_manifest)
        property_write_requested = write_requested and source_fix_hold is None
        source_entry = source_cash_report_entry(property_name, source_cash_entries)
        if source_entry:
            raw_sources = source_entry.get("source_cash_sources")
            ledger_paths = (
                [Path(str(source)) for source in raw_sources]
                if isinstance(raw_sources, list) and raw_sources
                else [Path(str(source_entry["source_cash_source"]))]
            )
            source_selection = "daily_source_cash_report"
        elif "package" in normalize_name(property_name):
            ledger_paths = split_ledger_paths(property_path, property_name)
            source_selection = "local_verified_package_components"
            if not ledger_paths:
                result = {
                    "property": property_name,
                    "property_id": property_id,
                    "property_path": str(property_path),
                    "source_selection": "missing_canonical_source",
                    "source_mode": "missing_canonical_package_components",
                    "status": "missing_source",
                    "missing_source_reason": "canonical_package_component_gl_not_verified",
                    "workbook": str(cf_workbook_path(property_path) or "") or None,
                    "months": [],
                }
                all_records.append(result)
                continue
        else:
            result = {
                "property": property_name,
                "property_id": property_id,
                "property_path": str(property_path),
                "source_selection": "missing_canonical_source",
                "source_mode": "missing_canonical_property_split_eco_gl",
                "status": "missing_source",
                "missing_source_reason": "canonical_property_split_eco_gl_not_verified",
                "workbook": str(cf_workbook_path(property_path) or "") or None,
                "months": [],
            }
            all_records.append(result)
            continue
        workbook_path = cf_workbook_path(property_path)
        workbook_create: dict[str, Any] | None = None
        if workbook_path is None and ledger_paths:
            workbook_create = create_clean_cf_workbook(
                update_cf,
                property_path,
                property_name,
                months[0][0],
                property_write_requested,
            )
            created_target = Path(str(workbook_create.get("target_path") or ""))
            if created_target.is_file():
                workbook_path = created_target
        result = {
            "property": property_name,
            "property_id": property_id,
            "property_path": str(property_path),
            "ledger": str(ledger_paths[0]) if len(ledger_paths) == 1 else None,
            "ledgers": [str(path) for path in ledger_paths],
            "source_selection": source_selection,
            "workbook": str(workbook_path) if workbook_path else None,
            "months": [],
        }
        if workbook_create is not None:
            result["workbook_create"] = workbook_create
        if source_fix_hold is not None:
            result["source_quality_hold"] = {
                "property": source_fix_hold.get("property"),
                "action_count": len(source_fix_hold.get("actions") or []),
                "action_types": sorted(
                    {
                        str(action.get("action_type") or "unknown")
                        for action in source_fix_hold.get("actions") or []
                    }
                ),
                "automation_statuses": sorted(
                    {
                        str(action.get("automation_status") or "unknown")
                        for action in source_fix_hold.get("actions") or []
                    }
                ),
            }
        if workbook_path is None:
            result["status"] = "missing_source"
            all_records.append(result)
            continue
        source_mode = "property_split_eco_gl"
        matched_master_property = None
        transactions = load_split_transactions(update_cf, ledger_paths)
        accounting_month = max(months) if months else None
        transactions = update_cf.filter_through_date(
            transactions, args.cutoff_date, accounting_month=accounting_month
        )
        if len(ledger_paths) > 1:
            source_mode = "aggregate_property_split_eco_gl"
        result["status"] = "held_source_quality" if source_fix_hold is not None else "ok"
        result["source_mode"] = source_mode
        result["matched_master_property"] = matched_master_property
        if source_entry:
            result["source_cash_report_entry"] = {
                "property": source_entry.get("property"),
                "source_cash_source": source_entry.get("source_cash_source"),
                "source_cash_sources": source_entry.get("source_cash_sources"),
                "source_cash_source_mode": source_entry.get("source_cash_source_mode"),
            }
        for year, month in months:
            month_data = update_cf.filter_by_month(transactions, year, month)
            source_cash_mode = source_cash_mode_for_manifest(
                year, month, source_cash_manifest if source_entry else None
            )
            source_cash_mode_counts[source_cash_mode] += 1
            changes = update_cf.update_xlsx(
                workbook_path,
                property_name,
                month_data,
                year,
                month,
                dry_run=not property_write_requested,
                source_cash_data=transactions,
                source_cash_only=balance_sheet_source_only,
                only_rows=row_scope,
                source_cash_mode=source_cash_mode,
            )
            for change in changes:
                change.setdefault("property", property_name)
                change.setdefault("month", f"{year}-{month:02d}")
            tokenomics_changes = []
            if not p_and_l_only and not balance_sheet_source_only:
                tokenomics_changes = standardize_tokenomics(
                    update_cf,
                    workbook_path,
                    year=year,
                    month=month,
                    live_row=live_by_id.get(property_id),
                    apply=args.apply and source_fix_hold is None,
                )
                for change in tokenomics_changes:
                    change.setdefault("property", property_name)
                    change.setdefault("month", f"{year}-{month:02d}")
            for change in changes:
                action_counts[str(change.get("action") or "unknown")] += 1
            action_counts["tokenomics_balance_sheet"] += len(tokenomics_changes)
            result["months"].append(
                {
                    "month": f"{year}-{month:02d}",
                    "source_month_row_count": len(month_data),
                    "eco_cash": month_end_source_summary(
                        transactions, year, month, source_cash_mode
                    ),
                    "change_count": len(changes),
                    "tokenomics_change_count": len(tokenomics_changes),
                    "changes": changes[:50],
                    "tokenomics_changes": tokenomics_changes[:50],
                }
            )
        all_records.append(result)
    changed_property_count = sum(
        1
        for record in all_records
        if any(month.get("change_count") or month.get("tokenomics_change_count") for month in record.get("months", []))
    )
    missing_count = sum(1 for record in all_records if record.get("status") == "missing_source")
    source_quality_hold_count = sum(
        1 for record in all_records if record.get("status") == "held_source_quality"
    )
    historical_review_required = bool(
        not write_requested and not p_and_l_only and len(months) > 1 and changed_property_count > 0
    )
    p_and_l_review_required = bool(
        not write_requested and p_and_l_only and changed_property_count > 0
    )
    return {
        "status": (
            "review"
            if missing_count != 0
            or source_quality_hold_count != 0
            or not source_cash_gate["allowed"]
            or historical_review_required
            or p_and_l_review_required
            else "ok"
        ),
        "generated_at": iso_z(),
        "apply": bool(write_requested),
        "apply_requested": bool(write_requested),
        "apply_mode": "balance_sheet_source_only" if balance_sheet_source_only else ("p_and_l_only" if p_and_l_only else "full"),
        "p_and_l_only": p_and_l_only,
        "balance_sheet_source_only": balance_sheet_source_only,
        "tokenomics_skipped": p_and_l_only or balance_sheet_source_only,
        "p_and_l_review_required": p_and_l_review_required,
        "p_and_l_review_reason": (
            "month_specific_eco_gl_rows_differ_from_current_workbooks"
            if p_and_l_review_required
            else None
        ),
        "write_blocked": False,
        "historical_review_required": historical_review_required,
        "historical_review_reason": (
            "multi_month_read_only_detected_changes_require_authoritative_month_specific_tokenomics"
            if historical_review_required
            else None
        ),
        "finance_truth_refresh": apply_gate,
        "source_only_apply_gate": source_only_apply_gate,
        "source_cash_report": {
            key: value for key, value in source_cash_manifest.items() if key != "entries"
        },
        "source_cash_apply_gate": source_cash_gate,
        "source_fix_plan": {
            key: value for key, value in source_fix_manifest.items() if key != "entries"
        },
        "source_quality_hold_count": source_quality_hold_count,
        "source_cash_balance_policy": {
            "current_balance": "full_property_split_eco_gl_column_e_across_all_rows",
            "closed_month_balance": "property_split_eco_gl_column_e_through_month_end",
            "current_mode": "full_column_e",
            "closed_month_mode": "as_of_month_end",
            "mode_counts": dict(sorted(source_cash_mode_counts.items())),
        },
        "reporting_cutoff_date": args.cutoff_date.isoformat() if args.cutoff_date else None,
        "untagged_review_apply_gate": untagged_review_gate,
        "month_scope_gate": month_scope_gate,
        "live_capture": str(args.live_capture),
        "live_rows": str(args.live_rows),
        "excluded_live_property_count": len(excluded_live_records),
        "excluded_live_properties": excluded_live_records,
        "months": args.month,
        "property_count": len(all_records),
        "changed_property_count": changed_property_count,
        "missing_source_count": missing_count,
        "action_counts": dict(sorted(action_counts.items())),
        "source_cash_change_count": action_counts.get("set_source_cash_balance", 0),
        "source_only_scope": balance_sheet_source_row_scope(update_cf) if balance_sheet_source_only else [],
        "records": all_records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Standardize live Lofty CF statements from split ECO GL files.")
    parser.add_argument("--live-capture", type=Path, default=DEFAULT_LIVE_CAPTURE)
    parser.add_argument("--live-rows", type=Path, default=DEFAULT_LIVE_ROWS)
    parser.add_argument("--master-gl", type=Path, default=DEFAULT_MASTER_GL)
    parser.add_argument("--month", action="append", required=True, help="Month to standardize, YYYY-MM. Repeatable.")
    parser.add_argument(
        "--cutoff-date",
        type=date.fromisoformat,
        default=None,
        help="Include split-ledger transactions dated on or before YYYY-MM-DD.",
    )
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--finance-truth-refresh", type=Path, default=DEFAULT_FINANCE_TRUTH_REFRESH)
    parser.add_argument("--source-cash-report", type=Path, default=DEFAULT_SOURCE_CASH_REPORT)
    parser.add_argument("--source-fix-plan", type=Path, default=DEFAULT_SOURCE_FIX_PLAN)
    parser.add_argument("--untagged-review-report", type=Path, default=DEFAULT_UNTAGGED_REVIEW_REPORT)
    parser.add_argument("--yhome-csv", type=Path, default=DEFAULT_YHOME_CSV)
    parser.add_argument(
        "--runtime-map",
        type=Path,
        default=None,
        help="Optional monthly Lofty runtime map whose terminal exclusions are removed from source selection.",
    )
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--p-and-l-only",
        action="store_true",
        help="Audit only month-specific ECO GL rows and the full Column E accounting control; never inspect or apply tokenomics.",
    )
    parser.add_argument(
        "--apply-p-and-l-only",
        action="store_true",
        help="Apply month-specific ECO GL rows and the full Column E accounting control only; never apply tokenomics.",
    )
    parser.add_argument(
        "--balance-sheet-source-only",
        action="store_true",
        help="Audit only the canonical full-Column-E internal accounting-control row; never touch custody cash, P&L, debt, or tokenomics.",
    )
    parser.add_argument(
        "--apply-balance-sheet-source-only",
        action="store_true",
        help="Apply only the verified full-Column-E internal accounting-control row; all other workbook rows remain untouched.",
    )
    args = parser.parse_args()
    if args.apply_p_and_l_only and not args.p_and_l_only:
        args.p_and_l_only = True
    if args.apply_balance_sheet_source_only and not args.balance_sheet_source_only:
        args.balance_sheet_source_only = True
    report = build_report(args)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, default=str, sort_keys=True) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "status": report["status"],
                "apply": report["apply"],
                "property_count": report["property_count"],
                "changed_property_count": report["changed_property_count"],
                "missing_source_count": report["missing_source_count"],
                "report": str(args.report),
            },
            indent=2,
            sort_keys=True,
        )
    )
    if report["status"] == "blocked":
        return 3
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
