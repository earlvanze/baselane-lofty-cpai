#!/usr/bin/env python3
"""Idempotently add verified/accrued Gipson mortgage equity to 804's workbook.

The workbook contains Excel pivot caches and GETPIVOTDATA formulas.  Editing it
with a generic XLSX writer can discard those structures, so this utility makes
small, explicit OOXML changes and preserves every unrelated ZIP member byte for
byte.  It updates the visible capital detail, pivot cache, cached summary values,
filter range, and workbook recalculation flags.
"""

from __future__ import annotations

import argparse
import html
import json
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    "/mnt/c/Users/digit/Dropbox/Real Estate/CO/804 S Quitman St, Denver, CO 80219/"
    "Financials/Capital Accounts.xlsx"
)


@dataclass(frozen=True)
class Contribution:
    contribution_date: date
    description: str
    amount: Decimal
    note: str


CONTRIBUTIONS = (
    Contribution(
        date(2026, 6, 1),
        "June 2026 Mortgage Payment",
        Decimal("2700.00"),
        "Additional equity contribution: mortgage advanced by Nathaniel outside Baselane; "
        "premium/conversion tracked by Gipson and pending later application.",
    ),
    Contribution(
        date(2026, 7, 1),
        "July 2026 Mortgage Payment",
        Decimal("2700.00"),
        "Additional equity contribution: mortgage advanced by Nathaniel outside Baselane; "
        "premium/conversion tracked by Gipson and pending later application.",
    ),
    Contribution(
        date(2026, 8, 1),
        "August 2026 Mortgage Payment (Accrued)",
        Decimal("2700.00"),
        "Accrued additional equity contribution for August mortgage; premium/conversion "
        "tracked by Gipson and pending later application.",
    ),
)


def excel_serial(value: date) -> int:
    return (value - date(1899, 12, 30)).days


def money(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.01")), "f")


def number(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _replace_cell_value(xml: str, cell_ref: str, value: Decimal) -> str:
    pattern = re.compile(
        rf'(<c\s+r="{re.escape(cell_ref)}"[^>]*>(?:(?!</c>).)*?<v>)([^<]*)(</v>(?:(?!</c>).)*?</c>)',
        re.DOTALL,
    )
    updated, count = pattern.subn(rf"\g<1>{number(value)}\g<3>", xml, count=1)
    if count != 1:
        raise RuntimeError(f"Could not update cached value for {cell_ref}")
    return updated


def _new_cells(row_number: int, contribution: Contribution) -> str:
    who = html.escape("Nathaniel Gipson")
    desc = html.escape(contribution.description)
    note = html.escape(contribution.note)
    return (
        f'<c r="A{row_number}" s="2"><v>{excel_serial(contribution.contribution_date)}</v></c>'
        f'<c r="B{row_number}" t="inlineStr"><is><t>{who}</t></is></c>'
        f'<c r="C{row_number}" t="inlineStr"><is><t>{desc}</t></is></c>'
        f'<c r="D{row_number}" s="15"><v>{number(contribution.amount)}</v></c>'
        f'<c r="E{row_number}" t="inlineStr"><is><t>{note}</t></is></c>'
    )


def _insert_sheet_cells(xml: str, row_number: int, contribution: Contribution) -> str:
    cells = _new_cells(row_number, contribution)
    existing_row = re.compile(rf'(<row\s+r="{row_number}"[^>]*>)')
    if existing_row.search(xml):
        return existing_row.sub(rf"\g<1>{cells}", xml, count=1)

    later_row = re.search(rf'<row\s+r="([0-9]+)"[^>]*>', xml)
    candidates = [
        (int(match.group(1)), match.start())
        for match in re.finditer(r'<row\s+r="([0-9]+)"[^>]*>', xml)
        if int(match.group(1)) > row_number
    ]
    insertion_point = min(candidates)[1] if candidates else xml.index("</sheetData>")
    new_row = f'<row r="{row_number}" spans="1:9">{cells}</row>'
    return xml[:insertion_point] + new_row + xml[insertion_point:]


def _pivot_record(contribution: Contribution) -> str:
    return (
        "<r>"
        f'<d v="{contribution.contribution_date.isoformat()}T00:00:00"/>'
        '<x v="1"/>'  # Nathaniel Gipson in cache field 1 shared-items list
        f'<s v="{html.escape(contribution.description, quote=True)}"/>'
        f'<n v="{number(contribution.amount)}"/>'
        f'<s v="{html.escape(contribution.note, quote=True)}"/>'
        "</r>"
    )


def _read_state(path: Path) -> tuple[dict[str, int], dict[str, Decimal], Decimal]:
    formulas = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    ws_formula = formulas["Sheet1"]
    ws_value = values["Sheet1"]

    descriptions: dict[str, int] = {}
    for row in range(2, ws_formula.max_row + 1):
        if ws_formula.cell(row, 2).value == "Nathaniel Gipson":
            description = ws_formula.cell(row, 3).value
            if description:
                descriptions[str(description)] = row

    summary: dict[str, Decimal] = {}
    for row in range(13, 25):
        member = ws_value.cell(row, 8).value
        amount = ws_value.cell(row, 9).value
        if member is not None and amount is not None:
            summary[str(member)] = Decimal(str(amount))

    current_nathaniel = sum(
        Decimal(str(ws_value.cell(row, 4).value))
        for row in range(2, ws_value.max_row + 1)
        if ws_value.cell(row, 2).value == "Nathaniel Gipson"
        and ws_value.cell(row, 4).value is not None
    )
    return descriptions, summary, current_nathaniel


def build_plan(path: Path) -> dict[str, object]:
    descriptions, summary, current_nathaniel = _read_state(path)
    missing = [item for item in CONTRIBUTIONS if item.description not in descriptions]
    target_nathaniel = current_nathaniel + sum((item.amount for item in missing), Decimal("0"))
    summary["Nathaniel Gipson"] = target_nathaniel
    grand_total = sum(summary.values(), Decimal("0"))
    member_equity = grand_total - summary["Rocket Mortgage"] - summary["Lofty Holding 1039 Mt Vernon Road DAO LLC"]
    return {
        "workbook": str(path),
        "missing": [
            {
                "date": item.contribution_date.isoformat(),
                "description": item.description,
                "amount": money(item.amount),
                "note": item.note,
            }
            for item in missing
        ],
        "current_nathaniel_capital": money(current_nathaniel),
        "target_nathaniel_capital": money(target_nathaniel),
        "target_grand_total": money(grand_total),
        "target_member_equity": money(member_equity),
        "summary": {key: money(value) for key, value in summary.items()},
    }


def apply_plan(path: Path, plan: dict[str, object]) -> Path | None:
    missing_descriptions = {item["description"] for item in plan["missing"]}  # type: ignore[index]
    missing = [item for item in CONTRIBUTIONS if item.description in missing_descriptions]
    if not missing:
        return None

    descriptions, summary, _ = _read_state(path)
    used_rows = set(descriptions.values())
    next_row = 62
    row_by_description: dict[str, int] = {}
    for item in missing:
        while next_row in used_rows:
            next_row += 1
        row_by_description[item.description] = next_row
        used_rows.add(next_row)
        next_row += 1

    summary = {key: Decimal(value) for key, value in plan["summary"].items()}  # type: ignore[union-attr]
    grand_total = Decimal(str(plan["target_grand_total"]))
    member_equity = Decimal(str(plan["target_member_equity"]))

    with zipfile.ZipFile(path, "r") as source:
        parts = {info.filename: source.read(info.filename) for info in source.infolist()}
        infos = source.infolist()

    sheet = parts["xl/worksheets/sheet2.xml"].decode("utf-8")
    for item in missing:
        sheet = _insert_sheet_cells(sheet, row_by_description[item.description], item)

    final_row = max(row_by_description.values())
    sheet = re.sub(r'(<autoFilter\s+ref=")A1:E[0-9]+', rf'\g<1>A1:E{final_row}', sheet, count=1)

    member_rows = {
        "Rocket Mortgage": 13,
        "Nathaniel Gipson": 14,
        "Wesley Babcock": 15,
        "Lofty Holding 1039 Mt Vernon Road DAO LLC": 16,
        "Ian Haber": 17,
        "Earl Vanze Co": 18,
        "NARWALL Holdings, LLC": 19,
        "Thomas A. Austin": 20,
        "EVCO Holdings LLC (Earl)": 21,
        "Daniel Murrey": 22,
        "Brandon McArthur": 23,
        "Kyle McArthur": 24,
    }
    for member, row in member_rows.items():
        sheet = _replace_cell_value(sheet, f"I{row}", summary[member])
        j_value = Decimal("0") if member in {"Rocket Mortgage", "Lofty Holding 1039 Mt Vernon Road DAO LLC"} else summary[member] / member_equity
        k_value = summary[member] / grand_total
        sheet = _replace_cell_value(sheet, f"J{row}", j_value)
        sheet = _replace_cell_value(sheet, f"K{row}", k_value)
    sheet = _replace_cell_value(sheet, "I25", grand_total)
    sheet = _replace_cell_value(sheet, "J25", Decimal("1"))
    sheet = _replace_cell_value(sheet, "K25", Decimal("1"))
    parts["xl/worksheets/sheet2.xml"] = sheet.encode("utf-8")

    workbook = parts["xl/workbook.xml"].decode("utf-8")
    workbook = re.sub(
        r'(name="_xlnm\._FilterDatabase"[^>]*>Sheet1!\$A\$1:\$E\$)[0-9]+',
        rf'\g<1>{final_row}',
        workbook,
        count=1,
    )
    workbook = re.sub(
        r'<calcPr\s+([^>]*)/>',
        lambda match: '<calcPr ' + re.sub(r'\s+(?:fullCalcOnLoad|forceFullCalc)="[^"]*"', '', match.group(1)).strip() + ' fullCalcOnLoad="1" forceFullCalc="1"/>',
        workbook,
        count=1,
    )
    parts["xl/workbook.xml"] = workbook.encode("utf-8")

    cache_definition = parts["xl/pivotCache/pivotCacheDefinition1.xml"].decode("utf-8")
    cache_definition = re.sub(
        r'(<pivotCacheDefinition\s+)(?![^>]*refreshOnLoad=)',
        r'\1refreshOnLoad="1" ',
        cache_definition,
        count=1,
    )
    cache_definition = re.sub(
        r'maxDate="[^"]+"',
        f'maxDate="{max(item.contribution_date for item in CONTRIBUTIONS).isoformat()}T00:00:00"',
        cache_definition,
        count=1,
    )
    old_count_match = re.search(r'recordCount="([0-9]+)"', cache_definition)
    if not old_count_match:
        raise RuntimeError("pivot cache recordCount was not found")
    old_count = int(old_count_match.group(1))
    cache_definition = cache_definition.replace(
        f'recordCount="{old_count}"', f'recordCount="{old_count + len(missing)}"', 1
    )
    parts["xl/pivotCache/pivotCacheDefinition1.xml"] = cache_definition.encode("utf-8")

    cache_records = parts["xl/pivotCache/pivotCacheRecords1.xml"].decode("utf-8")
    blank_marker = '<r><m/><x v="13"/><m/><m/><m/></r>'
    if blank_marker not in cache_records:
        raise RuntimeError("pivot cache blank-record insertion point was not found")
    added_records = "".join(_pivot_record(item) for item in missing)
    cache_records = cache_records.replace(blank_marker, added_records + blank_marker, 1)
    cache_records = re.sub(
        r'(<pivotCacheRecords[^>]*\scount=")([0-9]+)(")',
        lambda match: f'{match.group(1)}{int(match.group(2)) + len(missing)}{match.group(3)}',
        cache_records,
        count=1,
    )
    parts["xl/pivotCache/pivotCacheRecords1.xml"] = cache_records.encode("utf-8")

    backup_dir = path.parent / "Capital Accounts Cleanup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    backup = backup_dir / f"Capital Accounts.backup-gipson-equity-{stamp}.xlsx"
    shutil.copy2(path, backup)

    fd, temp_name = tempfile.mkstemp(prefix="capital-accounts-", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with zipfile.ZipFile(temp_path, "w") as output:
            for info in infos:
                output.writestr(info, parts[info.filename])
        with zipfile.ZipFile(temp_path, "r") as check:
            bad = check.testzip()
            if bad:
                raise RuntimeError(f"updated workbook ZIP failed integrity check at {bad}")
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return backup


def verify(path: Path) -> dict[str, object]:
    plan = build_plan(path)
    formulas = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    ws_formula = formulas["Sheet1"]
    ws_value = values["Sheet1"]
    details = []
    for item in CONTRIBUTIONS:
        matches = [
            row
            for row in range(2, ws_formula.max_row + 1)
            if ws_formula.cell(row, 2).value == "Nathaniel Gipson"
            and ws_formula.cell(row, 3).value == item.description
        ]
        details.append(
            {
                "description": item.description,
                "rows": matches,
                "amounts": [ws_value.cell(row, 4).value for row in matches],
                "notes": [ws_formula.cell(row, 5).value for row in matches],
            }
        )
    with zipfile.ZipFile(path, "r") as archive:
        cache = archive.read("xl/pivotCache/pivotCacheDefinition1.xml").decode("utf-8")
        records = archive.read("xl/pivotCache/pivotCacheRecords1.xml").decode("utf-8")
    return {
        "status": "ok" if not plan["missing"] and all(len(item["rows"]) == 1 for item in details) else "error",
        "workbook": str(path),
        "details": details,
        "nathaniel_capital": ws_value["I14"].value,
        "grand_total": ws_value["I25"].value,
        "pivot_refresh_on_load": 'refreshOnLoad="1"' in cache,
        "pivot_record_count": int(re.search(r'recordCount="([0-9]+)"', cache).group(1)),
        "pivot_new_record_count": sum(records.count(item.description) for item in CONTRIBUTIONS),
        "remaining_missing": plan["missing"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--verify", action="store_true")
    args = parser.parse_args()

    if args.verify:
        result = verify(args.workbook)
        print(json.dumps(result, indent=2, default=str))
        return 0 if result["status"] == "ok" else 1

    plan = build_plan(args.workbook)
    print(json.dumps({"mode": "apply" if args.apply else "preview", **plan}, indent=2))
    if args.apply:
        backup = apply_plan(args.workbook, plan)
        print(json.dumps({"backup": str(backup) if backup else None, "verification": verify(args.workbook)}, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
