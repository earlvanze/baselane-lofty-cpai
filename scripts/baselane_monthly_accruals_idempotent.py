#!/usr/bin/env python3
"""
Monthly idempotent accrual generator for Baselane GL.

Runs as part of the daily sync cron. Checks which accrual entries already
exist in the GL for the target month and creates any that are missing.

Idempotency: Each accrual entry is tagged with a stable note marker:
  AOPS-{GROUP}-ACCRUAL|{kind}|{property}|{YYYY-MM}|{amount} | {description}

The script reads the GL CSV, finds existing markers for the target month,
and only generates entries for missing (property, kind, month) combinations.

Usage:
  python3 baselane_monthly_accruals_idempotent.py --gl-csv "ECO Systems General Ledger.csv"
  python3 baselane_monthly_accruals_idempotent.py --gl-csv "..." --month 2026-06 --apply
  python3 baselane_monthly_accruals_idempotent.py --gl-csv "..." --start-month 2025-07 --end-month 2026-07 --property "326-332 S Alcott" --kind taxes --kind insurance --apply
  python3 baselane_monthly_accruals_idempotent.py --self-test
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from functools import lru_cache
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from baselane_ledger_revenue_policy import is_categoryless_known_rent_revenue
from coownership_reserve_policy import (
    LOCAL_FINANCIALS_ONLY_PROPERTIES,
    POLICY_EFFECTIVE_MONTH,
    POLICY_PROPERTIES,
    RETAINED_KIND,
    RETAINED_PREFIX,
    calculate_replenishment,
    canonical_property as canonical_reserve_property,
    load_lofty_reserves,
    live_lofty_reserve_required_properties,
    properties_from_lofty_response,
)
from coownership_mortgage_policy import is_no_dao_mortgage_property

# ── Accrual templates ─────────────────────────────────────────────────────────
# Each template defines: property, kind, monthly_amount, category, sub_category,
# description_template, merchant_template, source_prefix, group
#
# Fixed amounts (dao, legal, insurance, taxes, mortgage_interest) are known.
# PM fees are computed from gross rent in the GL for the target month.

CASH_BASIS_INSURANCE_STATES = {"OH", "IL", "TN"}
PM_DAO_KIND = "pm_dao"
PM_ECO_KIND = "pm_eco"
DAO_ECO_KIND = "dao_eco"
DAO_LLC_ADMIN_ANNUAL_CHARGE = Decimal("750.00")
DAO_LLC_ADMIN_MONTHLY_ACCRUAL = Decimal("62.50")
DAO_LLC_ADMIN_ECO_LOFTY_PAYABLE = Decimal("200.00")
DAO_LLC_ADMIN_ECO_FILING_COST_REFERENCE = Decimal("125.00")
DAO_LLC_ADMIN_ECO_REVENUE_NOTE = (
    "ECO Systems LLC DAO registration/admin fee revenue; matched to the "
    "DAO-side expense. The DAO owes the full $750.00 annual charge, recognized "
    "at $62.50 per month. ECO's separate $200.00 annual payable to Lofty and "
    "ECO's actual filing/vendor costs (historical reference $125.00) are ECO "
    "expenses and never reduce the DAO's fee or payable. Extraordinary back-filing "
    "costs require their own evidence and classification. Accounting/manual "
    "accrual only, no bank transfer."
)
# PM accrual treatment is transaction-based, not geography-based. Hemlane can
# remit net rent in any state; direct deposits can likewise reach Baselane in
# any state. Only actual Hemlane-originated net rent is excluded from the
# manual DAO/ECO PM accrual basis.
PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS: set[tuple[str, str]] = set()


def default_lofty_reserve_snapshot() -> Path:
    """Return the live-workspace reserve snapshot without resolving legacy links."""
    return Path(__file__).absolute().parents[1] / "reports" / "lofty-pm-current" / "get-manager-properties.full-response.json"


def default_retained_capital_approved_exceptions() -> Path:
    return Path(__file__).absolute().parents[1] / "config" / "baselane_retained_capital_approved_exceptions.json"


def load_retained_capital_approved_exceptions(path: Path) -> list[dict[str, Any]]:
    """Load narrowly scoped, evidence-backed retained-capital reporting exceptions."""
    if not path.is_file():
        raise ValueError(f"retained-capital approved-exceptions file is missing: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"retained-capital approved-exceptions file is unreadable: {path}: {exc}") from exc
    records = payload.get("exceptions") if isinstance(payload, dict) else None
    if not isinstance(records, list):
        raise ValueError("retained-capital approved-exceptions payload requires an exceptions list")
    validated: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            raise ValueError(f"retained-capital exception {index} must be an object")
        property_name = canonical_reserve_property(record.get("property"))
        month = str(record.get("month") or "")
        approval_id = str(record.get("approval_id") or "").strip()
        reason = str(record.get("reason") or "").strip()
        try:
            amount = Decimal(str(record.get("amount") or "0"))
        except Exception as exc:
            raise ValueError(f"retained-capital exception {index} has an invalid amount") from exc
        key = (property_name or "", month)
        if not property_name or not re.fullmatch(r"20\d\d-(0[1-9]|1[0-2])", month):
            raise ValueError(f"retained-capital exception {index} requires a canonical property and YYYY-MM month")
        if amount <= 0 or not approval_id or not reason:
            raise ValueError(f"retained-capital exception {index} requires a positive amount, approval_id, and reason")
        if key in seen:
            raise ValueError(f"duplicate retained-capital exception for {property_name} {month}")
        seen.add(key)
        validated.append({
            "property": property_name,
            "month": month,
            "amount": float(amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "approval_id": approval_id,
            "reason": reason,
        })
    return validated


# Madison's 25% total PM burden includes direct co-host charges paid from the
# DAO account.  These reviewed, unreimbursed amounts reduce the ECO portion of
# the PM fee for the service month.  Use an explicit override because older
# Baselane descriptions do not consistently include the "3%" wording used by
# the deterministic row detector below.
MADISON_UNREIMBURSED_COHOST_PM_OVERRIDES: dict[tuple[str, str], float] = {
    ("88 Madison Ave", "2025-07"): 231.42,
    ("88 Madison Ave", "2026-05"): 40.00,
    ("90 Madison Ave", "2025-07"): 244.08,
}


def is_pm_accrual_kind(kind: str) -> bool:
    """Return whether a marker kind is one side of a PM fee accrual."""
    return kind in {"pm", PM_DAO_KIND, PM_ECO_KIND}


def normalized_accrual_kind(kind: str) -> str:
    """Map PM accrual sides to the user-facing PM coverage bucket."""
    if is_pm_accrual_kind(kind):
        return "pm"
    return "dao" if kind == DAO_ECO_KIND else kind

ACCRUAL_TEMPLATES: list[dict[str, Any]] = [
    # ── PNL properties (84/86/88/90 Madison, 724 3rd Ave, 9 Country Club) ──
    {"property": "84 Madison Ave", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "86 Madison Ave", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "88 Madison Ave", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "90 Madison Ave", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "724 3rd Ave", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "9 Country Club Ln N", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},

    # 9 Country Club - insurance, taxes, mortgage_interest
    {"property": "9 Country Club Ln N", "kind": "insurance", "amount": 1552.91, "category": "Rental Dwelling", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "Insurance Accrual | {property} | {month_label}", "merchant": "Insurance Accrual | {property} | {month_label}"},
    {"property": "9 Country Club Ln N", "kind": "taxes", "amount": 3225.01, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    {"property": "9 Country Club Ln N", "kind": "mortgage_interest", "amount": 2875.00, "category": "Mortgage Interest Payments", "sub_category": "", "source_prefix": "AOPS-PNL-ACCRUAL", "description": "Mortgage Interest Accrual | {property} | {month_label}", "merchant": "Mortgage Interest Accrual | {property} | {month_label}"},

    # ── OH/IL properties (AOPS-OHIL-ACCRUAL) ──
    # DAO LLC fee: $750/year ÷ 12 = $62.50/month for each property
    # Taxes are property-specific annual values ÷ 12.
    # OH/IL/TN insurance is cash-basis only via OSC Risk Secure transactions.

    # 10724 Gooding Ave
    {"property": "10724 Gooding Ave, Cleveland, OH 44108", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    # 12028 Wade Park Ave
    {"property": "12028 Wade Park Ave, Cleveland, OH 44106", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    # 1278 E 187th St
    {"property": "1278 E 187th St, Cleveland, OH 44110", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "1278 E 187th St, Cleveland, OH 44110", "kind": "taxes", "amount": 1755.94, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 1315 E 114th St
    {"property": "1315 E 114th St, Cleveland, OH 44106", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    # 13806 Coit Rd
    {"property": "13806 Coit Rd, Cleveland, OH 44110", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    # 1456 W 85th St
    {"property": "1456 W 85th St, Cleveland, OH 44102", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "1456 W 85th St, Cleveland, OH 44102", "kind": "taxes", "amount": 4799.60, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 15555 Millard Ave
    {"property": "15555 Millard Ave, Markham, IL 60428", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "15555 Millard Ave, Markham, IL 60428", "kind": "taxes", "amount": 8571.14, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 2337 Greenvale Rd
    {"property": "2337 Greenvale Rd, Cleveland, OH 44121", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "2337 Greenvale Rd, Cleveland, OH 44121", "kind": "taxes", "amount": 1385.77, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 25 Circle Dr
    {"property": "25 Circle Dr, Dixmoor, IL 60426", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "25 Circle Dr, Dixmoor, IL 60426", "kind": "taxes", "amount": 3114.59, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 26931 Shoreview Ave
    {"property": "26931 Shoreview Ave, Euclid, OH 44132", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "26931 Shoreview Ave, Euclid, OH 44132", "kind": "taxes", "amount": 4313.44, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 3024 W 103rd St
    {"property": "3024 W 103rd St, Cleveland, OH 44111", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    # 428 Cross St
    {"property": "428 Cross St, Akron, OH 44311", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "428 Cross St, Akron, OH 44311", "kind": "taxes", "amount": 4432.47, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 4318 Clybourne Ave
    {"property": "4318 Clybourne Ave, Cleveland, OH 44109", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "4318 Clybourne Ave, Cleveland, OH 44109", "kind": "taxes", "amount": 5547.20, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 5541 S Peoria St
    {"property": "5541 S Peoria St, Chicago, IL 60621", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "5541 S Peoria St, Chicago, IL 60621", "kind": "taxes", "amount": 2628.00, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 566 Nash St
    {"property": "566 Nash St, Akron, OH 44306", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "566 Nash St, Akron, OH 44306", "kind": "taxes", "amount": 2038.36, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 7542 & 7656 S Colfax Ave
    {"property": "7542 and 7656 S Colfax Ave, Chicago, IL 60649", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "7542 and 7656 S Colfax Ave, Chicago, IL 60649", "kind": "taxes", "amount": 1214.71, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 783 Leonard St
    {"property": "783 Leonard St, Akron, OH 44307", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "783 Leonard St, Akron, OH 44307", "kind": "taxes", "amount": 113.87, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 8143 S Sangamon St
    {"property": "8143 S Sangamon St, Chicago, IL 60620", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "8143 S Sangamon St, Chicago, IL 60620", "kind": "taxes", "amount": 4552.72, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 8708 Willard Ave
    {"property": "8708 Willard Ave, Cleveland, OH 44102", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "8708 Willard Ave, Cleveland, OH 44102", "kind": "taxes", "amount": 1842.94, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 917 Pawnee Ave
    {"property": "917 Pawnee Ave, Memphis, TN 38109", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "917 Pawnee Ave, Memphis, TN 38109", "kind": "taxes", "amount": 98.38, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 9634 S Green St
    {"property": "9634 S Green St, Chicago, IL 60643", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "9634 S Green St, Chicago, IL 60643", "kind": "taxes", "amount": 2843.56, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    # 9919 S Oglesby Ave
    {"property": "9919 S Oglesby Ave, Chicago, IL 60617", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-OHIL-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},

    # ── PAU properties (22164 Umland Cir, 326-332 S Alcott) ──
    {"property": "22164 Umland Cir, Jenner, CA 95450", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "22164 Umland Cir, Jenner, CA 95450", "kind": "insurance", "amount": 367.25, "category": "Rental Dwelling", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "Insurance Accrual | {property} | {month_label}", "merchant": "Insurance Accrual | {property} | {month_label}"},
    {"property": "22164 Umland Cir, Jenner, CA 95450", "kind": "taxes", "amount": 1020.18, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
    {"property": "326-332 S Alcott St, Denver, CO 80219", "kind": "dao", "amount": 62.50, "category": "Legal Fees", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "DAO LLC Fee Accrual | {property} | {month_label}", "merchant": "DAO LLC Fee Accrual | {property} | {month_label}"},
    {"property": "326-332 S Alcott St, Denver, CO 80219", "kind": "insurance", "amount": 256.37, "category": "Rental Dwelling", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "Insurance Accrual | {property} | {month_label}", "merchant": "Insurance Accrual | {property} | {month_label}"},
    {"property": "326-332 S Alcott St, Denver, CO 80219", "kind": "taxes", "amount": 374.19, "category": "City, State, & Local Taxes", "sub_category": "", "source_prefix": "AOPS-PAU-ACCRUAL", "description": "Tax Accrual | {property} | {month_label}", "merchant": "Tax Accrual | {property} | {month_label}"},
]

# PM fee properties: (property, pm_rate, source_prefix)
# PM fees are computed as rate × gross rent for the target month
PM_FEE_PROPERTIES = [
    ("10724 Gooding Ave, Cleveland, OH 44108", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("1278 E 187th St, Cleveland, OH 44110", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("1456 W 85th St, Cleveland, OH 44102", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("25 Circle Dr, Dixmoor, IL 60426", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("3024 W 103rd St, Cleveland, OH 44111", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("428 Cross St, Akron, OH 44311", 0.15, "AOPS-OHIL-ACCRUAL"),
    ("5541 S Peoria St, Chicago, IL 60621", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("566 Nash St, Akron, OH 44306", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("724 3rd Ave", 0.06, "AOPS-PNL-ACCRUAL"),
    ("84 Madison Ave", 0.15, "AOPS-PNL-ACCRUAL"),
    ("86 Madison Ave", 0.25, "AOPS-PNL-ACCRUAL"),
    ("88 Madison Ave", 0.25, "AOPS-PNL-ACCRUAL"),
    ("90 Madison Ave", 0.25, "AOPS-PNL-ACCRUAL"),
    ("9 Country Club Ln N", 0.20, "AOPS-PNL-ACCRUAL"),
    ("8708 Willard Ave, Cleveland, OH 44102", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("8143 S Sangamon St, Chicago, IL 60620", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("917 Pawnee Ave, Memphis, TN 38109", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("9634 S Green St, Chicago, IL 60643", 0.10, "AOPS-OHIL-ACCRUAL"),
    ("22164 Umland Cir, Jenner, CA 95450", 0.15, "AOPS-PAU-ACCRUAL"),
    ("326-332 S Alcott St, Denver, CO 80219", 0.10, "AOPS-PAU-ACCRUAL"),
    ("85-104 Alawa Pl", 0.25, "AOPS-PM-FEE"),
]

PM_FEE_COMPONENTS: dict[str, list[dict[str, Any]]] = {}
HEMLANE_PM_FEE_BASIS: dict[str, dict[str, Any]] = {}
PM_RATE_HISTORY: dict[str, list[tuple[str, float]]] = {
    # Madison's total PM burden is always 25%. Separately booked Jazmine 3%
    # co-host charges are deducted from ECO's component, never from this rate.
    "86 Madison Ave": [("0000-01", 0.25)],
    "88 Madison Ave": [("0000-01", 0.25)],
    "90 Madison Ave": [("0000-01", 0.25)],
    "724 3rd Ave": [("0000-01", 0.05), ("2026-05", 0.06)],
}
SCHEDULE_DAO_MONTHLY_AMOUNT = float(DAO_LLC_ADMIN_MONTHLY_ACCRUAL)
SCHEDULE_SOURCE_PREFIX = "AOPS-MONTHLY-ACCRUAL"
MANUAL_EXCLUDED_SCHEDULE_PROPERTIES = (
    "3560 Saint Albans Rd",
    "1935 S Glen Rd",
    "5401 Odom Ave",
    "1236 W 7th St",
    # Coolwood belongs to a separate accounting workspace. Its presence in
    # the shared PM schedule must not create Baselane accruals in this ledger.
    "1 Coolwood Drive",
)


def default_listing_update_policy_path() -> Path:
    candidates = [
        Path(os.environ["LOFTY_LISTING_UPDATE_POLICY"]) if os.environ.get("LOFTY_LISTING_UPDATE_POLICY") else None,
        Path(os.environ["WORKSPACE_ROOT"]) / "config" / "lofty_listing_update_policy.json"
        if os.environ.get("WORKSPACE_ROOT")
        else None,
        Path.cwd() / "config" / "lofty_listing_update_policy.json",
        Path(__file__).resolve().parents[1] / "config" / "lofty_listing_update_policy.json",
        Path.home() / ".openclaw" / "workspace" / "config" / "lofty_listing_update_policy.json",
        Path("/home/digit/.openclaw/workspace/config/lofty_listing_update_policy.json"),
    ]
    for candidate in candidates:
        if candidate and candidate.is_file():
            return candidate
    return Path(__file__).resolve().parents[1] / "config" / "lofty_listing_update_policy.json"

# Properties with known Baselane property name variations
PROPERTY_ALIASES = {
    "10724 Gooding Ave, Cleveland, OH 44108": ["10724 Gooding Ave"],
    "12028 Wade Park Ave, Cleveland, OH 44106": ["12028 Wade Park Ave"],
    "326-332 S Alcott St, Denver, CO 80219": ["326 South Alcott Street"],
    "7542 and 7656 S Colfax Ave, Chicago, IL 60649": ["7542 & 7656 S Colfax Ave"],
    "8143 S Sangamon St, Chicago, IL 60620": [" 8143 S Sangamon St."],
    "1315 E 114th St, Cleveland, OH 44106": ["1315 E 114th St"],
    "13806 Coit Rd, Cleveland, OH 44110": ["13806 Coit Rd"],
    "1456 W 85th St, Cleveland, OH 44102": ["1456 W 85th St."],
    "15555 Millard Ave, Markham, IL 60428": ["15555 Millard Ave"],
    "2337 Greenvale Rd, Cleveland, OH 44121": ["2337 Greenvale Rd"],
    "26931 Shoreview Ave, Euclid, OH 44132": ["26931 Shoreview Ave"],
    "3024 W 103rd St, Cleveland, OH 44111": ["3024 W 103rd St"],
    "4318 Clybourne Ave, Cleveland, OH 44109": ["4318 Clybourne Ave"],
    "5541 S Peoria St, Chicago, IL 60621": ["5541 S Peoria St"],
    "566 Nash St, Akron, OH 44306": ["566 Nash St"],
    "783 Leonard St, Akron, OH 44307": ["783 Leonard St"],
    "428 Cross St, Akron, OH 44311": ["428 Cross St."],
    "1278 E 187th St, Cleveland, OH 44110": ["1278 E 187th St"],
    "25 Circle Dr, Dixmoor, IL 60426": ["25 Circle Dr"],
    "22164 Umland Cir, Jenner, CA 95450": ["22164 Umland Circle"],
    "8708 Willard Ave, Cleveland, OH 44102": ["8708 Willard Ave"],
    "917 Pawnee Ave, Memphis, TN 38109": ["917 Pawnee Ave"],
    "9634 S Green St, Chicago, IL 60643": ["9634 S Green St"],
    "9919 S Oglesby Ave, Chicago, IL 60617": ["9919 S Oglesby Ave"],
    "85-104 Alawa Pl": ["85-104 Alawa Pl, Waianae, HI 96792"],
}

# Launch dates: no accruals before these dates
PROPERTY_LAUNCH_DATES: dict[str, str] = {
    "84 Madison Ave": "2025-08-25",
    "86 Madison Ave": "2024-12-06",
    "88 Madison Ave": "2024-01-29",
    "90 Madison Ave": "2024-05-14",
    "724 3rd Ave": "2024-04-24",
    "85-104 Alawa Pl": "2025-03-14",
    "9 Country Club Ln N": "2025-08-15",
}

# NY/HI co-ownership accounting begins in the month before the first token sale.
PROPERTY_ACCRUAL_START_MONTHS: dict[str, str] = {
    "84 Madison Ave": "2025-07",
    "86 Madison Ave": "2024-11",
    "88 Madison Ave": "2023-12",
    "90 Madison Ave": "2024-04",
    "724 3rd Ave": "2024-03",
    "85-104 Alawa Pl": "2025-02",
    "9 Country Club Ln N": "2025-07",
}

# Property-level accounting can begin before a specific obligation becomes the
# DAO's responsibility. 9CC mortgage interest starts with September 2025.
PROPERTY_KIND_ACCRUAL_START_MONTHS: dict[tuple[str, str], str] = {
    ("9 Country Club Ln N", "mortgage_interest"): "2025-09",
}

RETAINED_CAPITAL_RULE = {
    "kind": RETAINED_KIND,
    "effective_month": POLICY_EFFECTIVE_MONTH,
    "source_prefix": RETAINED_PREFIX,
}
RETAINED_CAPITAL_PROPERTIES = tuple(POLICY_PROPERTIES)

MONTH_NAME_RE = re.compile(
    r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(20\d\d)\b",
    re.IGNORECASE,
)
NUMERIC_SERVICE_MONTH_RE = re.compile(r"\b(0?[1-9]|1[0-2])/(20\d\d)\b")

ACCRUAL_AMOUNT_OVERRIDES: list[dict[str, Any]] = [
    {
        "property": "9 Country Club Ln N",
        "kind": "insurance",
        "start_month": "2026-04",
        "amount": 918.00,
        "reason": "New York Central premium changed to $918/month beginning April 2026",
    },
    {
        "property": "326-332 S Alcott St, Denver, CO 80219",
        "kind": "taxes",
        "start_month": "2026-01",
        "amount": 333.49,
        "reason": "Alcott 2026 tax accrual adjusted to $333.49/month",
    },
]

ANNUAL_REFERENCE_FIXED_ACCRUALS: set[tuple[str, str]] = {
    ("1278 E 187th St, Cleveland, OH 44110", "taxes"),
    ("1456 W 85th St, Cleveland, OH 44102", "taxes"),
    ("15555 Millard Ave, Markham, IL 60428", "taxes"),
    ("2337 Greenvale Rd, Cleveland, OH 44121", "taxes"),
    ("25 Circle Dr, Dixmoor, IL 60426", "taxes"),
    ("26931 Shoreview Ave, Euclid, OH 44132", "taxes"),
    ("428 Cross St, Akron, OH 44311", "taxes"),
    ("4318 Clybourne Ave, Cleveland, OH 44109", "taxes"),
    ("5541 S Peoria St, Chicago, IL 60621", "taxes"),
    ("566 Nash St, Akron, OH 44306", "taxes"),
    ("8143 S Sangamon St, Chicago, IL 60620", "taxes"),
    ("8708 Willard Ave, Cleveland, OH 44102", "taxes"),
    ("9634 S Green St, Chicago, IL 60643", "taxes"),
}

NO_FIXED_ACCRUAL_TEMPLATE_REQUIRED: dict[str, str] = {
    "6914 Polonia Ave, Cleveland, OH 44105": "sold property; no synthetic tax, insurance, PM, or DAO LLC fee accruals",
    "49 Bannbury Ln, Palm Coast, FL 32137": "existing ledger activity has direct tax/license, management fee, mortgage interest, and rent rows; fixed monthly accrual amounts not established",
    "804 S Quitman St, Denver, CO 80219": "rehab/development property with direct expense and financing rows; fixed monthly accrual amounts not established",
    "9902 Garfield Ave, Cleveland, Ohio 44108": "ledger shows direct insurance/legal/loan/sale activity; fixed monthly accrual amounts not established",
    "122 Florida Park Dr, Palm Coast, FL 32137": "ledger shows direct operating and license rows; fixed monthly accrual amounts not established",
    "27 Pillar Ln, Palm Coast, FL 32164": "short-term-rental operating property with direct expense and mortgage rows; fixed monthly accrual amounts not established",
    "Earl DAO": "portfolio/entity row, not a property-level accrual template target",
    "918 Frederick Blvd, Akron, Ohio 44320": "ledger shows direct rent, insurance, tax, management, and utility rows; fixed monthly accrual amounts not established",
    "3178 W 41st St, Cleveland, Ohio 44109": "ledger shows direct rent, insurance, tax, and management rows; fixed monthly accrual amounts not established",
    "7411 Elton Ave, Cleveland, Ohio 44102": "ledger shows direct rent, insurance, tax, management, and legal rows; fixed monthly accrual amounts not established",
    "656 E 126th St, Cleveland, Ohio 44108": "ledger shows direct insurance, transfer, permit, and rent rows; fixed monthly accrual amounts not established",
    "1935 S Glen Rd, Shelby, MI 49455": "no Baselane ledger rows found for fixed monthly accrual inference",
    "3850 W 17th St, Cleveland, Ohio 44109": "ledger shows direct insurance, transfer, repair, rent, and sale rows; fixed monthly accrual amounts not established",
    "3139 West Blvd, Cleveland, OH 44111": "no Baselane ledger rows found for fixed monthly accrual inference",
    "1845 W 48th St, Cleveland, Ohio 44102": "no Baselane ledger rows found for fixed monthly accrual inference",
    "Ohio 3 Property Package, Akron, Ohio 44117": "package row without a single fixed property-level accrual template target",
    "1432 Sara Ave, Akron, Ohio 44305": "ledger shows direct rent, insurance, loan, management, and operating rows; fixed monthly accrual amounts not established",
    "254 Bowmanville St, Akron, Ohio 44305": "ledger shows direct transfer, loan, insurance, repair, and revenue rows; fixed monthly accrual amounts not established",
    "614 E 97th St, Cleveland, Ohio 44108": "no property-specific fixed monthly accrual amounts established from ledger",
    "5401 Odom Ave, Fort Worth, TX 76114": "ledger shows direct operating, insurance, transfer, and loan rows; fixed monthly accrual amounts not established",
    "4183 E 146th St, Cleveland, OH 44128": "ledger shows direct owner distribution, loan interest, and management rows; fixed monthly accrual amounts not established",
    "1236 W 7th St, Davenport, IA 52802": "ledger shows direct transfer, tax, repair, and landscaping rows; fixed monthly accrual amounts not established",
    "1090 Diagonal Rd, Akron, Ohio 44320": "ledger shows direct loan and rent rows; fixed monthly accrual amounts not established",
}


def iso_z() -> str:
    return dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def is_zero_amount_text(value: object) -> bool:
    return str(value or "").replace(",", "").strip() in {"", "0", "0.0", "0.00", "-0", "-0.0", "-0.00"}


def has_nonzero_accrual_rows(rows: list[dict[str, str]]) -> bool:
    """Return whether a dry run still has accounting work to apply."""
    return any(not is_zero_amount_text(row.get("Amount")) for row in rows)


def default_target_month() -> str:
    """Default target month is the previous month."""
    today = dt.date.today()
    first = today.replace(day=1)
    prev = first - dt.timedelta(days=1)
    return f"{prev.year:04d}-{prev.month:02d}"


def month_label(month: str) -> str:
    """Convert 2026-06 to 'June 2026'."""
    year, m = [int(x) for x in month.split("-")]
    return dt.date(year, m, 1).strftime("%B %Y")


def month_end_date(month: str) -> str:
    """Last day of the month as 'Month DD, YYYY'."""
    year, m = [int(x) for x in month.split("-")]
    if m == 12:
        last = dt.date(year, 12, 31)
    else:
        last = dt.date(year, m + 1, 1) - dt.timedelta(days=1)
    return last.strftime("%B %d, %Y")


def accrual_posting_date(month: str) -> str:
    """Standard monthly accrual posting date as 'Month 28, YYYY'."""
    year, m = [int(x) for x in month.split("-")]
    return dt.date(year, m, 28).strftime("%B %d, %Y")


def iter_months(start_month: str, end_month: str) -> list[str]:
    """Inclusive list of YYYY-MM months."""
    start_year, start_m = [int(x) for x in start_month.split("-")]
    end_year, end_m = [int(x) for x in end_month.split("-")]
    current = dt.date(start_year, start_m, 1)
    end = dt.date(end_year, end_m, 1)
    if current > end:
        raise ValueError("--start-month cannot be after --end-month")
    months = []
    while current <= end:
        months.append(f"{current.year:04d}-{current.month:02d}")
        if current.month == 12:
            current = dt.date(current.year + 1, 1, 1)
        else:
            current = dt.date(current.year, current.month + 1, 1)
    return months


def stable_id(prefix: str, kind: str, property_name: str, month: str, amount: float) -> str:
    material = f"{prefix}|{kind}|{property_name}|{month}|{amount:.2f}"
    return hashlib.sha256(material.encode("utf-8")).hexdigest()[:16]


def effective_accrual_amount(property_name: str, kind: str, target_month: str, template_amount: float) -> float:
    amount = float(template_amount)
    for override in ACCRUAL_AMOUNT_OVERRIDES:
        if override["property"] != property_name:
            continue
        if override["kind"] != kind:
            continue
        if target_month >= override["start_month"]:
            amount = float(override["amount"])
    return amount


def fixed_accrual_reference_amount(property_name: str, kind: str, target_month: str, template_amount: float) -> float:
    return effective_accrual_amount(property_name, kind, target_month, template_amount)


def fixed_accrual_posting_amount(property_name: str, kind: str, target_month: str, template_amount: float) -> float:
    reference_amount = Decimal(str(fixed_accrual_reference_amount(property_name, kind, target_month, template_amount)))
    if (property_name, kind) in ANNUAL_REFERENCE_FIXED_ACCRUALS:
        return round_money(reference_amount / Decimal("12"))
    return round_money(reference_amount)


def effective_pm_rate(property_name: str, target_month: str, default_rate: float) -> float:
    rate = float(default_rate)
    for start_month, historical_rate in PM_RATE_HISTORY.get(property_name, []):
        if target_month >= start_month:
            rate = float(historical_rate)
    return rate


def round_money(value: Decimal | float | str) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def previous_month(month: str) -> str:
    year, month_number = [int(part) for part in month.split("-")]
    first = dt.date(year, month_number, 1)
    prev = first - dt.timedelta(days=1)
    return f"{prev.year:04d}-{prev.month:02d}"


def retained_capital_amount(
    rows: list[dict[str, str]],
    target_month: str,
    property_name: str,
    lofty_reserves: dict[str, float],
    approved_exceptions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    canonical = canonical_reserve_property(property_name)
    if canonical is None:
        raise ValueError(f"property is not covered by the co-ownership reserve policy: {property_name}")
    if canonical not in lofty_reserves:
        return {
            "property": canonical,
            "target_month": target_month,
            "basis_month": target_month,
            "status": "missing_lofty_operating_reserve",
            "amount": 0.0,
        }
    result = calculate_replenishment(
        rows, canonical, target_month, lofty_reserves[canonical], approved_exceptions
    )
    return {
        **result,
        "target_month": target_month,
        "basis_month": target_month,
        "status": "ok",
    }


def retained_capital_note(property_name: str, target_month: str, amount: float, retained_summary: dict[str, Any]) -> str:
    approved_exception = retained_summary.get("approved_exception")
    if isinstance(approved_exception, dict):
        return (
            f"{RETAINED_CAPITAL_RULE['source_prefix']}|{RETAINED_CAPITAL_RULE['kind']}|{property_name}|{target_month}|{amount:.2f} | "
            f"Approved retained-earnings reporting exception {approved_exception.get('approval_id')}: "
            f"{approved_exception.get('reason')} Outstanding cash reserve settlement requirement after "
            f"${retained_summary.get('cash_settled_amount', 0):.2f} of verified same-month cash settlement."
        )
    return (
        f"{RETAINED_CAPITAL_RULE['source_prefix']}|{RETAINED_CAPITAL_RULE['kind']}|{property_name}|{target_month}|{amount:.2f} | "
        f"OR Replenishment per March 2026 governance vote: {retained_summary['replenishment_rate'] * 100:g}% of "
        f"{retained_summary['basis_month']} NOI ${retained_summary['noi']:.2f}. Reserve basis ${retained_summary['combined_reserve_basis']:.2f} "
        f"= ECO GL net of accruals ${retained_summary['eco_gl_net_of_accruals']:.2f} + Lofty OR "
        f"${retained_summary['lofty_operating_reserve']:.2f}. Outstanding cash reserve settlement requirement "
        f"after ${retained_summary.get('cash_settled_amount', 0):.2f} of verified same-month cash settlement."
    )


def pm_fee_rule_summary(property_name: str, target_month: str, default_rate: float, components: list[dict[str, Any]]) -> str:
    if property_name == "9 Country Club Ln N":
        return "20% total: non-Evolve gross x 20%; Evolve 90%-net remittance / 9 for ECO's remaining 10% of gross"
    source_components = components or [{"rate": default_rate, "deduction": 0.0}]
    parts = []
    for item in source_components:
        part = f"{effective_pm_rate(property_name, target_month, float(item['rate'])) * 100:g}% of direct gross rent"
        deduction = float(item.get("deduction") or 0)
        if deduction:
            part += f" less ${deduction:.2f} contractual DAO credit"
        parts.append(part)
    summary = "; ".join(parts)
    if property_name in {"86 Madison Ave", "88 Madison Ave", "90 Madison Ave"}:
        summary += "; less any separately booked Jazmine 3% co-host PM fee"
    hemlane_basis = HEMLANE_PM_FEE_BASIS.get(property_name)
    if hemlane_basis and hemlane_basis.get("month") == target_month:
        summary += (
            f"; Hemlane live request basis rent ${float(hemlane_basis.get('rent_request_amount') or 0):.2f}, "
            f"PM request ${float(hemlane_basis.get('pm_fee_request_amount') or 0):.2f}"
        )
    return summary


def parse_marker(notes: str) -> dict[str, str] | None:
    """Parse an AOPS accrual marker from the Notes field."""
    # Format: AOPS-{GROUP}-ACCRUAL|{kind}|{property}|{YYYY-MM}|{amount} | {description}
    # The older HI PM prefix did not include ``-ACCRUAL``.  It originally used
    # a direct one-sided format, while paired rows use an explicit PM kind.
    pm_fee_match = re.match(
        r"(AOPS-PM-FEE)\|(?:(pm_(?:dao|eco))\|)?([^|]+)\|(\d{4}-\d{2})\|([\d.]+)",
        notes.strip(),
    )
    if pm_fee_match:
        return {
            "prefix": pm_fee_match.group(1),
            "kind": pm_fee_match.group(2) or "pm",
            "property": pm_fee_match.group(3),
            "month": pm_fee_match.group(4),
            "amount": pm_fee_match.group(5),
        }
    match = re.match(
        r"(AOPS-[A-Z]+-ACCRUAL)\|([^|]+)\|([^|]+)\|(\d{4}-\d{2})\|([\d.]+)",
        notes.strip(),
    )
    if not match:
        return None
    prefix = match.group(1)
    kind = match.group(2)
    if prefix == "AOPS-PNL-ACCRUAL" and kind == "legal":
        kind = "dao"
    return {
        "prefix": prefix,
        "kind": kind,
        "property": match.group(3),
        "month": match.group(4),
        "amount": match.group(5),
    }


def parse_pm_fee_marker(notes: str) -> dict[str, str] | None:
    """Parse direct PM fee markers such as AOPS-PM-FEE|90 Madison Ave|2026-05|1053.05."""
    match = re.search(
        r"AOPS-PM-FEE\|([^|]+)\|(\d{4}-\d{2})\|([\d.]+)",
        notes.strip(),
    )
    if not match:
        return None
    return {
        "prefix": "AOPS-PM-FEE",
        "kind": "pm",
        "property": match.group(1),
        "month": match.group(2),
        "amount": match.group(3),
    }


def previous_month(target_month: str) -> str:
    year, month = [int(part) for part in target_month.split("-")]
    if month == 1:
        return f"{year - 1:04d}-12"
    return f"{year:04d}-{month - 1:02d}"


def normalize_gl_row(row: dict[str | None, object], fieldnames: list[str]) -> dict[str, str]:
    return {field: "" if row.get(field) is None else str(row.get(field)) for field in fieldnames}


def read_gl(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        return fieldnames, [normalize_gl_row(row, fieldnames) for row in reader]


def row_text(row: dict[str, str], fields: list[str] | None = None) -> str:
    keys = fields or ["Account", "Merchant", "Description", "Category", "Sub-category", "Property", "Unit", "Notes"]
    return " ".join(str(row.get(key) or "") for key in keys)


def row_category_names(row: dict[str, str]) -> set[str]:
    return {
        str(row.get(field) or "").strip().lower()
        for field in ("Category", "Sub-category")
        if str(row.get(field) or "").strip()
    }


def parse_row_date(row: dict[str, str]) -> dt.date | None:
    value = str(row.get("Date") or "").strip()
    for date_format in ("%Y-%m-%d", "%B %d, %Y"):
        try:
            return dt.datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return None


def parse_reporting_cutoff(value: str | None) -> dt.date | None:
    if not value:
        return None
    try:
        return dt.date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"reporting cutoff date must be YYYY-MM-DD, got {value!r}"
        ) from exc


def rows_through_reporting_cutoff(
    rows: list[dict[str, str]],
    cutoff: dt.date | None,
    target_months: list[str],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Exclude later ordinary transactions while retaining target-month AOPS rows."""
    if cutoff is None:
        return list(rows), []

    target_month_set = set(target_months)
    included: list[dict[str, str]] = []
    excluded: list[dict[str, str]] = []
    for row in rows:
        row_date = parse_row_date(row)
        if row_date is None or row_date <= cutoff:
            included.append(row)
            continue

        notes = str(row.get("Notes") or "").strip()
        marker = parse_marker(notes) or parse_pm_fee_marker(notes)
        row_month = f"{row_date.year:04d}-{row_date.month:02d}"
        if (
            notes.upper().startswith("AOPS-")
            and row_month in target_month_set
            and (marker is None or marker["month"] == row_month)
        ):
            included.append(row)
            continue
        excluded.append(row)
    return included, excluded


def is_rent_revenue(row: dict[str, str], amount: float) -> bool:
    if amount <= 0:
        return False
    txn_type = str(row.get("Type") or "").strip().lower()
    category = str(row.get("Category") or "").strip().lower()
    sub_category = str(row.get("Sub-category") or "").strip().lower()
    if "rent" in category or "rent" in sub_category:
        return txn_type == "revenue" or category in {"rents", "short term rents", "long term rents"} or sub_category in {"rents", "short term rents", "long term rents"}
    if txn_type == "revenue" and category not in {"fees & other revenue", "interest income"}:
        return True
    if category or sub_category:
        return False
    merchant = str(row.get("Merchant") or "").strip().lower()
    description = str(row.get("Description") or "").strip().lower()
    notes = str(row.get("Notes") or "").strip().lower()
    if (
        notes.startswith(("aligned clearing detail import", "aligned/evernest clearing detail import"))
        and "| rent or tenant receipt |" in notes
    ):
        return True
    if "rent voucher" in merchant or "rent voucher" in description or "cbrap" in merchant or "cbrap" in description:
        return True
    return is_categoryless_known_rent_revenue(row, amount)


def is_hemlane_net_rent(row: dict[str, str]) -> bool:
    """Return whether a Baselane rent receipt was remitted net by Hemlane."""
    return "hemlane" in row_text(row, ["Merchant", "Description", "Notes"]).lower()


def row_matches_property(row: dict[str, str], property_name: str) -> bool:
    aliases = [property_name] + PROPERTY_ALIASES.get(property_name, [])
    explicit_property = str(row.get("Property") or "").strip().lower()
    if explicit_property:
        return any(alias.lower() in explicit_property or explicit_property in alias.lower() for alias in aliases)
    haystack = row_text(row).lower()
    return any(alias.lower() in haystack for alias in aliases)


def hemlane_net_rent_amount(rows: list[dict[str, str]], property_name: str, target_month: str) -> float:
    total = Decimal("0")
    for row in rows:
        row_date = parse_row_date(row)
        if row_date is None or f"{row_date.year:04d}-{row_date.month:02d}" != target_month:
            continue
        try:
            amount = float(str(row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        if row_matches_property(row, property_name) and is_rent_revenue(row, amount) and is_hemlane_net_rent(row):
            total += Decimal(str(amount))
    return round_money(total)


def hemlane_tx_matches_property(tx: dict[str, Any], property_name: str) -> bool:
    tx_property = normalize_schedule_address(f"{tx.get('property') or ''} {tx.get('property_address') or ''}")
    if not tx_property:
        return False
    aliases = [property_name] + PROPERTY_ALIASES.get(property_name, [])
    for alias in aliases:
        normalized_alias = normalize_schedule_address(alias)
        if normalized_alias and (normalized_alias in tx_property or tx_property in normalized_alias):
            return True
    return False


def hemlane_tx_month(tx: dict[str, Any]) -> str:
    for field in ("due_date", "transaction_date", "posted_at"):
        raw = str(tx.get(field) or "").strip()
        if not raw:
            continue
        if re.match(r"^\d{4}-\d{2}", raw):
            return raw[:7]
        try:
            parsed = dt.datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return f"{parsed.year:04d}-{parsed.month:02d}"
        except ValueError:
            continue
    return ""


def decimal_amount(value: Any) -> Decimal:
    try:
        return Decimal(str(value or "0").replace(",", "")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def hemlane_category_text_is_rent(text: str) -> bool:
    normalized = normalize_property_key(text)
    if "non rental" in normalized:
        return False
    tokens = set(normalized.split())
    return "rent" in tokens or "rents" in tokens or "rental income" in normalized


def load_hemlane_pm_fee_basis(path: Path | None, target_month: str) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    report: dict[str, Any] = {
        "path": str(path) if path else None,
        "status": "not_configured",
        "month": target_month,
        "property_count": 0,
    }
    if not path:
        return {}, report
    if not path.is_file():
        report["status"] = "missing"
        return {}, report
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        report.update({"status": "unreadable", "error": str(exc)})
        return {}, report
    if payload.get("status") != "ok":
        report.update({"status": "source_not_ok", "source_status": payload.get("status")})
        return {}, report
    transactions = [tx for tx in payload.get("transactions") or [] if isinstance(tx, dict)]
    basis: dict[str, dict[str, Any]] = {}
    for property_name, rate, _prefix in PM_FEE_PROPERTIES:
        rent_amount = Decimal("0.00")
        rent_success_amount = Decimal("0.00")
        rent_success_evidence_count = 0
        pm_fee_amount = Decimal("0.00")
        rent_ids: list[str] = []
        pm_fee_ids: list[str] = []
        for tx in transactions:
            if hemlane_tx_month(tx) != target_month or not hemlane_tx_matches_property(tx, property_name):
                continue
            category_text = f"{tx.get('payment_category') or ''} {tx.get('payment_subcategory') or ''}".lower()
            request_amount = decimal_amount(tx.get("request_amount") if tx.get("request_amount") is not None else tx.get("amount"))
            if request_amount <= 0:
                continue
            if hemlane_category_text_is_rent(category_text):
                rent_amount += request_amount
                if "success_amount" in tx:
                    rent_success_evidence_count += 1
                    rent_success_amount += decimal_amount(tx.get("success_amount"))
                if tx.get("id"):
                    rent_ids.append(str(tx["id"]))
            elif "management" in category_text and "fee" in category_text:
                pm_fee_amount += request_amount
                if tx.get("id"):
                    pm_fee_ids.append(str(tx["id"]))
        if rent_amount <= 0 and pm_fee_amount <= 0:
            continue
        amount = pm_fee_amount if pm_fee_amount > 0 else rent_amount * Decimal(str(rate))
        basis[property_name] = {
            "property": property_name,
            "month": target_month,
            "rent_request_amount": round_money(rent_amount),
            "rent_success_amount": round_money(rent_success_amount),
            "rent_success_evidence_count": rent_success_evidence_count,
            "pm_fee_request_amount": round_money(pm_fee_amount),
            "computed_pm_fee_amount": round_money(amount),
            "rent_transaction_ids": rent_ids,
            "pm_fee_transaction_ids": pm_fee_ids,
            "source": str(path),
            "rule": "Hemlane live requested rent/PM fee evidence; accrues PM obligation only and does not create rent cash revenue.",
        }
    report.update(
        {
            "status": "ok",
            "transaction_count": len(transactions),
            "property_count": len(basis),
            "properties": basis,
        }
    )
    return basis, report


def reporting_property_name(property_name: str) -> str:
    if "," not in property_name:
        return property_name
    aliases = PROPERTY_ALIASES.get(property_name, [])
    return aliases[0] if aliases else property_name


def normalize_property_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def normalize_schedule_address(value: str) -> str:
    text = normalize_property_key(value)
    replacements = {
        "avenue": "ave", "street": "st", "road": "rd", "lane": "ln",
        "place": "pl", "drive": "dr", "north": "n", "south": "s",
        "east": "e", "west": "w", "saint": "st",
    }
    for source, target in replacements.items():
        text = re.sub(rf"\b{source}\b", target, text)
    return text


def property_state(value: str) -> str:
    match = re.search(r",\s*([A-Z]{2})(?:\s+\d{5})?\s*$", value.strip())
    if match:
        return match.group(1)
    lowered = value.lower()
    if re.search(r"\bohio\b", lowered):
        return "OH"
    if re.search(r"\billinois\b", lowered):
        return "IL"
    if re.search(r"\btennessee\b", lowered):
        return "TN"
    if "cleveland" in lowered or "akron" in lowered or "euclid" in lowered:
        return "OH"
    if "chicago" in lowered or "markham" in lowered or "dixmoor" in lowered:
        return "IL"
    if "memphis" in lowered:
        return "TN"
    # Several actively managed STR DAO names intentionally omit their city and
    # state in Baselane.  Keep the accounting policy deterministic for them.
    if value.strip() in {"724 3rd Ave", "84 Madison Ave", "86 Madison Ave", "88 Madison Ave", "90 Madison Ave", "9 Country Club Ln N"}:
        return "NY"
    if value.strip() == "85-104 Alawa Pl":
        return "HI"
    return ""


def is_cash_basis_insurance_accrual(property_name: str, kind: str) -> bool:
    return kind == "insurance" and property_state(property_name) in CASH_BASIS_INSURANCE_STATES


def is_no_dao_mortgage_escrow_accrual(property_name: str, kind: str) -> bool:
    """Manual tax/insurance accruals are invalid where the servicer pays escrow."""
    return kind in {"taxes", "insurance"} and is_no_dao_mortgage_property(property_name)


def pm_manual_accrual_required(property_name: str, target_month: str) -> bool:
    """Whether a PM fee needs a Baselane manual DAO/ECO accrual pair."""
    return True


@lru_cache(maxsize=1)
def sold_listing_update_policy_excluded_schedule_names() -> tuple[str, ...]:
    path = default_listing_update_policy_path()
    if not path.is_file():
        return ()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return ()
    values = payload.get("sold_ignore_listing_updates") if isinstance(payload, dict) else None
    if not isinstance(values, list):
        return ()
    names: list[str] = []
    for value in values:
        raw_value = value if isinstance(value, dict) else {}
        full_address = str(raw_value.get("address") or raw_value.get("property_name") or value or "").strip()
        property_name = full_address.split(",", 1)[0].strip()
        if property_name:
            names.append(property_name)
        if full_address and full_address != property_name:
            names.append(full_address)
    return tuple(dict.fromkeys(names))


def schedule_address_is_excluded(address: str) -> bool:
    normalized_address = normalize_schedule_address(address)
    excluded_names = (*MANUAL_EXCLUDED_SCHEDULE_PROPERTIES, *sold_listing_update_policy_excluded_schedule_names())
    for name in excluded_names:
        normalized_name = normalize_schedule_address(name)
        if normalized_name and (normalized_name in normalized_address or normalized_address in normalized_name):
            return True
    return False


def canonical_accrual_property_name(value: str) -> str:
    normalized = normalize_schedule_address(value)
    candidates = [str(item["property"]) for item in ACCRUAL_TEMPLATES]
    candidates.extend(str(item[0]) for item in PM_FEE_PROPERTIES)
    matches = {candidate for candidate in candidates if normalize_schedule_address(candidate) == normalized}
    return next(iter(matches)) if len(matches) == 1 else value


def default_pm_rate_schedule() -> Path | None:
    configured = str(os.environ.get("BASELANE_PM_RATE_SCHEDULE") or "").strip()
    candidates = [
        Path(configured) if configured else None,
        Path("/mnt/c/Users/digit/Dropbox/Real Estate/Lofty PM/Earl Co - Property Management Portfolio (Bookkeeping).xlsx"),
        Path("/data/Dropbox/Real Estate/Lofty PM/Earl Co - Property Management Portfolio (Bookkeeping).xlsx"),
    ]
    return next((path for path in candidates if path is not None and path.is_file()), None)


def parse_pm_rate(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        rate = float(value)
    else:
        match = re.match(r"\s*(\d+(?:\.\d+)?)\s*%", str(value or ""))
        if not match:
            return None
        rate = float(match.group(1)) / 100.0
    return rate if 0 < rate < 1 else None


def parse_pm_deduction(value: Any, occupied_units: Any) -> float:
    text = str(value or "")
    # “Less Hemlane fee” is an internal ECO profitability note in the source
    # workbook. It is not a contractual DAO credit.
    if "hemlane fee" in text.lower():
        return 0.0
    per_unit = re.search(r"\$?([\d.]+)\s*/\s*unit\s*/\s*month", text, re.IGNORECASE)
    if per_unit:
        try:
            match = re.match(r"\s*(\d+)", str(occupied_units or ""))
            units = int(match.group(1)) if match else 0
        except (TypeError, ValueError):
            units = 0
        return round(float(per_unit.group(1)) * units, 2)
    monthly = re.search(r"\$?([\d.]+)\s*/\s*month", text, re.IGNORECASE)
    return round(float(monthly.group(1)), 2) if monthly else 0.0


def schedule_row_is_in_scope(row: dict[str, Any]) -> bool:
    address = str(row.get("address") or "")
    if schedule_address_is_excluded(address):
        return False
    if str(row.get("dao") or "").strip() in {"", "-"}:
        return False
    lofty = normalize_property_key(str(row.get("lofty") or ""))
    if lofty not in {"yes", "intend to launch", "limited private listing exclusive for accredited investors"}:
        return False
    status = normalize_property_key(str(row.get("status") or ""))
    pm = normalize_property_key(str(row.get("pm") or ""))
    return status != "sold" and pm != "sold" and lofty != "sold"


def schedule_property_match(address: str, candidates: list[str]) -> str | None:
    address_key = normalize_schedule_address(address)
    address_numbers = set(re.findall(r"\b\d+\b", address_key))
    address_tokens = set(address_key.split()) - {"oh", "il", "ny", "ca", "co", "fl", "hi", "tn", "tx", "ia", "mi", "mo"}
    matches: list[tuple[int, str]] = []
    for candidate in candidates:
        candidate_key = normalize_schedule_address(candidate)
        candidate_numbers = set(re.findall(r"\b\d+\b", candidate_key))
        if address_numbers and not address_numbers.intersection(candidate_numbers):
            continue
        shared = address_tokens.intersection(candidate_key.split())
        if len(shared) < 2:
            continue
        score = len(shared) * 10 + (50 if address_key in candidate_key or candidate_key in address_key else 0)
        matches.append((score, candidate))
    matches.sort(key=lambda item: (item[0], len(item[1])), reverse=True)
    return matches[0][1] if matches else None


def load_pm_schedule_rows(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, ())
        headers = {normalize_property_key(str(value or "")): index for index, value in enumerate(header_row)}
        address_column = headers.get("address")
        rate_column = headers.get("pm fee of gross rents")
        if rate_column is None:
            rate_column = headers.get("pm sub pm")
        if address_column is None or rate_column is None:
            raise ValueError(f"PM schedule is missing Address or PM Fee columns: {path}")
        schedule_rows: list[dict[str, Any]] = []
        for row in rows:
            address = str(row[address_column] or "").strip() if len(row) > address_column else ""
            if not address:
                continue
            value = row[rate_column] if len(row) > rate_column else None
            schedule_rows.append({
                "address": address,
                "dao": row[headers.get("dao")] if headers.get("dao") is not None and len(row) > headers["dao"] else None,
                "pm": row[headers.get("pm sub pm")] if headers.get("pm sub pm") is not None and len(row) > headers["pm sub pm"] else None,
                "lofty": row[headers.get("on lofty")] if headers.get("on lofty") is not None and len(row) > headers["on lofty"] else None,
                "status": row[headers.get("current status occupied units")] if headers.get("current status occupied units") is not None and len(row) > headers["current status occupied units"] else None,
                "total_units": row[headers.get("total units")] if headers.get("total units") is not None and len(row) > headers["total units"] else None,
                "rate": parse_pm_rate(value),
                "deduction": parse_pm_deduction(value, row[headers.get("current status occupied units")] if headers.get("current status occupied units") is not None and len(row) > headers["current status occupied units"] else None),
                "raw_rate": value,
            })
    finally:
        workbook.close()
    return schedule_rows


def pm_fee_properties_from_schedule(
    path: Path,
    base: list[tuple[str, float, str]] | None = None,
    gl_rows: list[dict[str, str]] | None = None,
) -> tuple[list[tuple[str, float, str]], dict[str, Any]]:
    global ACCRUAL_TEMPLATES, PM_FEE_COMPONENTS

    properties = list(base or PM_FEE_PROPERTIES)
    schedule_rows = load_pm_schedule_rows(path)
    if base is not None and not any(str(item.get("dao") or "").strip() for item in schedule_rows):
        resolved = []
        matched = {}
        for property_name, fallback_rate, prefix in properties:
            canonical = schedule_property_match(property_name, [str(item["address"]) for item in schedule_rows])
            schedule_row = next((item for item in schedule_rows if item["address"] == canonical), None)
            rate = float(schedule_row["rate"]) if schedule_row and schedule_row.get("rate") is not None else fallback_rate
            resolved.append((property_name, rate, prefix))
            if schedule_row and schedule_row.get("rate") is not None:
                matched[property_name] = rate
        return resolved, {
            "path": str(path),
            "status": "ok" if matched else "review",
            "matched_property_count": len(matched),
            "matched_rates": dict(sorted(matched.items())),
        }
    ledger_names = sorted({str(row.get("Property") or "").strip() for row in gl_rows or [] if str(row.get("Property") or "").strip()})
    known_names = [str(item["property"]) for item in ACCRUAL_TEMPLATES] + [item[0] for item in properties]
    candidates = list(dict.fromkeys([*known_names, *ledger_names]))

    mapped_rows: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []
    excluded_canonical: set[str] = set()
    for schedule_row in schedule_rows:
        canonical = schedule_property_match(str(schedule_row["address"]), candidates)
        if not schedule_row_is_in_scope(schedule_row):
            excluded_rows.append({"address": schedule_row["address"], "status": schedule_row.get("status"), "canonical": canonical})
            if canonical:
                excluded_canonical.add(canonical)
            continue
        if not canonical:
            canonical = str(schedule_row["address"]).split(",", 1)[0].strip()
            unmatched_rows.append({"address": schedule_row["address"], "dao": schedule_row.get("dao"), "fallback_canonical": canonical})
        schedule_row = {**schedule_row, "canonical": canonical}
        mapped_rows.append(schedule_row)
        aliases = PROPERTY_ALIASES.setdefault(canonical, [])
        if schedule_row["address"] not in aliases:
            aliases.append(str(schedule_row["address"]))

    if excluded_canonical:
        ACCRUAL_TEMPLATES = [
            template
            for template in ACCRUAL_TEMPLATES
            if str(template["property"]) not in excluded_canonical
        ]

    prefixes = {name: prefix for name, _rate, prefix in properties}
    components: dict[str, list[dict[str, Any]]] = defaultdict(list)
    matched: dict[str, float] = {}
    fallback_rate_matches: dict[str, float] = {}
    for schedule_row in mapped_rows:
        rate = schedule_row.get("rate")
        if rate is None:
            continue
        canonical = str(schedule_row["canonical"])
        components[canonical].append({
            "address": schedule_row["address"],
            "basis_token": re.match(r"\d+", normalize_schedule_address(str(schedule_row["address"]))).group(0),
            "rate": float(rate),
            "deduction": float(schedule_row.get("deduction") or 0),
            "raw_rate": schedule_row.get("raw_rate"),
        })
        matched[canonical] = float(rate)
    if not matched:
        mapped_canonical = {str(item["canonical"]) for item in mapped_rows}
        for property_name, fallback_rate, _prefix in properties:
            if property_name in excluded_canonical:
                continue
            if mapped_canonical and property_name not in mapped_canonical:
                continue
            components[property_name].append({
                "address": property_name,
                "basis_token": None,
                "rate": float(fallback_rate),
                "deduction": 0.0,
                "raw_rate": "fallback_existing_rate_schedule_missing_rate_column",
            })
            fallback_rate_matches[property_name] = float(fallback_rate)
    PM_FEE_COMPONENTS = dict(components)
    resolved = [
        (canonical, float(items[0]["rate"]), prefixes.get(canonical, SCHEDULE_SOURCE_PREFIX))
        for canonical, items in sorted(components.items())
    ]

    existing_dao_by_name = {
        str(item["property"]): item
        for item in ACCRUAL_TEMPLATES
        if str(item.get("kind")) == "dao"
    }
    dao_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for schedule_row in mapped_rows:
        dao_groups[normalize_property_key(str(schedule_row.get("dao") or ""))].append(schedule_row)
    added_dao: list[str] = []
    for dao_rows in dao_groups.values():
        canonical = next((str(item["canonical"]) for item in dao_rows if str(item["canonical"]) in existing_dao_by_name), str(dao_rows[0]["canonical"]))
        if canonical in existing_dao_by_name:
            continue
        ACCRUAL_TEMPLATES.append({
            "property": canonical,
            "kind": "dao",
            "amount": SCHEDULE_DAO_MONTHLY_AMOUNT,
            "category": "Legal Fees",
            "sub_category": "",
            "source_prefix": SCHEDULE_SOURCE_PREFIX,
            "description": "DAO LLC Fee Accrual | {property} | {month_label}",
            "merchant": "DAO LLC Fee Accrual | {property} | {month_label}",
        })
        added_dao.append(canonical)

    return resolved, {
        "path": str(path),
        "status": "ok" if mapped_rows else "review",
        "schedule_property_count": len(schedule_rows),
        "eligible_property_count": len(mapped_rows),
        "excluded_property_count": len(excluded_rows),
        "source_only_fallback_property_count": len(unmatched_rows),
        "source_only_fallback_properties": unmatched_rows,
        "matched_property_count": len(matched),
        "matched_rates": dict(sorted(matched.items())),
        "fallback_rate_property_count": len(fallback_rate_matches),
        "fallback_rates": dict(sorted(fallback_rate_matches.items())),
        "added_dao_template_count": len(added_dao),
        "added_dao_templates": sorted(added_dao),
        "component_property_count": sum(len(items) > 1 for items in components.values()),
    }


def template_property_keys() -> set[str]:
    names = {str(item["property"]) for item in ACCRUAL_TEMPLATES}
    names.update(str(item[0]) for item in PM_FEE_PROPERTIES)
    for canonical, aliases in PROPERTY_ALIASES.items():
        names.add(canonical)
        names.update(aliases)
    return {normalize_property_key(name) for name in names if normalize_property_key(name)}


def no_fixed_accrual_policy_by_key() -> dict[str, str]:
    return {normalize_property_key(name): reason for name, reason in NO_FIXED_ACCRUAL_TEMPLATE_REQUIRED.items()}


def load_active_property_map(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    records = data.get("properties") if isinstance(data, dict) else data
    if not isinstance(records, list):
        raise ValueError(f"{path} did not contain a properties list")
    active: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        if str(record.get("status") or "").strip().lower() in {"inactive", "sold", "archived"}:
            continue
        property_name = str(record.get("property_name") or "").strip()
        full_address = str(record.get("full_address") or "").strip()
        if not property_name and not full_address:
            continue
        active.append(
            {
                "property_name": property_name,
                "full_address": full_address,
                "assetUnit": record.get("assetUnit"),
                "lofty_property_id": record.get("lofty_property_id"),
            }
        )
    return active


def find_active_properties_without_accrual_templates(active_properties: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys = template_property_keys()
    policy = no_fixed_accrual_policy_by_key()
    missing: list[dict[str, Any]] = []
    for record in active_properties:
        candidates = [str(record.get("property_name") or ""), str(record.get("full_address") or "")]
        candidate_keys = {normalize_property_key(value) for value in candidates if normalize_property_key(value)}
        if any(key in keys for key in candidate_keys):
            continue
        if any(key in policy for key in candidate_keys):
            continue
        missing.append(record)
    return missing


def active_properties_without_fixed_accrual_requirement(active_properties: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys = template_property_keys()
    policy = no_fixed_accrual_policy_by_key()
    records: list[dict[str, Any]] = []
    for record in active_properties:
        candidates = [str(record.get("property_name") or ""), str(record.get("full_address") or "")]
        candidate_keys = {normalize_property_key(value) for value in candidates if normalize_property_key(value)}
        if any(key in keys for key in candidate_keys):
            continue
        reason = next((policy[key] for key in candidate_keys if key in policy), "")
        if reason:
            records.append({**record, "reason": reason})
    return records


def infer_service_month(row: dict[str, str]) -> str:
    text = row_text(row, ["Merchant", "Description", "Notes"])
    match = MONTH_NAME_RE.search(text)
    if match:
        service_date = dt.datetime.strptime(
            f"{match.group(1).title()} {match.group(2)}",
            "%B %Y",
        )
        return f"{service_date.year:04d}-{service_date.month:02d}"
    match = NUMERIC_SERVICE_MONTH_RE.search(text)
    if match:
        return f"{int(match.group(2)):04d}-{int(match.group(1)):02d}"
    row_date = parse_row_date(row)
    if row_date is None:
        return ""
    return f"{row_date.year:04d}-{row_date.month:02d}"


def row_day(row: dict[str, str]) -> int | None:
    row_date = parse_row_date(row)
    return row_date.day if row_date else None


def actual_scruggs_kind(row: dict[str, str]) -> str | None:
    text = row_text(row, ["Merchant", "Description", "Category", "Sub-category", "Notes"]).lower()
    if "scruggs investments" not in text:
        return None
    if "insurance" in text or "rental dwelling" in text:
        return "insurance"
    if "tax" in text or "taxes" in text or "property taxes" in text:
        return "taxes"
    return None


def find_existing_accruals(rows: list[dict[str, str]], target_month: str) -> set[str]:
    """Return set of 'property|kind' that already have AOPS accruals for target_month."""
    existing = set()
    for row in rows:
        notes = str(row.get("Notes") or "")
        marker = parse_marker(notes)
        if marker and marker["month"] == target_month:
            # Historical direct AOPS-PM-FEE rows dated on day one were stale
            # prior-period imports, not current-month coverage.  Explicit
            # paired PM markers remain valid regardless of posting day.
            if marker["prefix"] == "AOPS-PM-FEE" and marker["kind"] == "pm" and row_day(row) == 1:
                continue
            property_name = canonical_accrual_property_name(marker["property"])
            existing.add(f"{property_name}|{normalized_accrual_kind(marker['kind'])}")
            continue
        pm_marker = parse_pm_fee_marker(notes)
        if (
            pm_marker
            and pm_marker["month"] == target_month
            and row_day(row) != 1
        ):
            property_name = canonical_accrual_property_name(pm_marker["property"])
            existing.add(f"{property_name}|pm")
    return existing


def pm_accrual_sides(rows: list[dict[str, str]], target_month: str) -> dict[str, set[str]]:
    """Return PM accrual sides already posted, preserving legacy one-sided rows.

    Pre-pair ``pm`` and ``AOPS-PM-FEE`` rows are treated as legacy DAO-side
    postings.  They are not automatically backfilled as ECO revenue: that
    requires a reviewed migration so historical cash can first be reclassified.
    """
    sides: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        marker = parse_marker(str(row.get("Notes") or ""))
        if marker and marker["month"] == target_month:
            property_name = canonical_accrual_property_name(marker["property"])
            if marker["kind"] == PM_DAO_KIND:
                sides[property_name].add("dao")
            elif marker["kind"] == PM_ECO_KIND:
                sides[property_name].add("eco")
            elif marker["kind"] == "pm":
                sides[property_name].update({"dao", "legacy"})
            continue
        marker = parse_pm_fee_marker(str(row.get("Notes") or ""))
        if marker and marker["month"] == target_month and row_day(row) != 1:
            property_name = canonical_accrual_property_name(marker["property"])
            sides[property_name].update({"dao", "legacy"})
    return sides


def dao_fee_accrual_sides(rows: list[dict[str, str]], target_month: str) -> dict[str, set[str]]:
    """Return fixed DAO service-fee sides already posted.

    Legacy ``dao``/``legal`` markers are the DAO expense side. ``dao_eco`` is
    ECO's matching fee revenue. The underlying filing-vendor debits are
    separate ECO expenses and never count as either side.
    """
    sides: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        marker = parse_marker(str(row.get("Notes") or ""))
        if not marker or marker["month"] != target_month:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        if marker["kind"] == "dao":
            sides[property_name].add("dao")
        elif marker["kind"] == DAO_ECO_KIND:
            sides[property_name].add("eco")
    return sides


def expected_fixed_accrual_keys(kind_filters: set[str] | None = None, property_filters: list[str] | None = None) -> set[str]:
    keys: set[str] = set()
    for template in ACCRUAL_TEMPLATES:
        property_name = str(template["property"])
        kind = str(template["kind"])
        if schedule_address_is_excluded(property_name):
            continue
        if is_cash_basis_insurance_accrual(property_name, kind):
            continue
        if is_no_dao_mortgage_escrow_accrual(property_name, kind):
            continue
        if kind_filters and kind not in kind_filters:
            continue
        if property_filters and not any(normalize_property_key(item) in normalize_property_key(property_name) or normalize_property_key(property_name) in normalize_property_key(item) for item in property_filters):
            continue
        keys.add(f"{property_name}|{kind}")
    for property_name, _rate, _prefix in PM_FEE_PROPERTIES:
        if schedule_address_is_excluded(property_name):
            continue
        if kind_filters and "pm" not in kind_filters:
            continue
        if property_filters and not any(normalize_property_key(item) in normalize_property_key(property_name) or normalize_property_key(property_name) in normalize_property_key(item) for item in property_filters):
            continue
        keys.add(f"{property_name}|pm")
    return keys


def retained_capital_expected_keys(
    rows: list[dict[str, str]],
    target_month: str,
    lofty_reserves: dict[str, float],
    kind_filters: set[str] | None = None,
    property_filters: list[str] | None = None,
    approved_exceptions: list[dict[str, Any]] | None = None,
) -> set[str]:
    kind = str(RETAINED_CAPITAL_RULE["kind"])
    keys: set[str] = set()
    for property_name in RETAINED_CAPITAL_PROPERTIES:
        if not template_in_scope(property_name, kind, property_filters, kind_filters):
            continue
        if retained_capital_amount(rows, target_month, property_name, lofty_reserves, approved_exceptions)["amount"] <= 0:
            continue
        keys.add(f"{property_name}|{kind}")
    return keys


def coverage_by_kind(keys: set[str]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for key in keys:
        if "|" not in key:
            continue
        counts[key.rsplit("|", 1)[1]] += 1
    return dict(sorted(counts.items()))


def coverage_key_details(keys: set[str], target_month: str | None = None) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for key in sorted(keys):
        if "|" not in key:
            continue
        property_name, kind = key.rsplit("|", 1)
        detail = {
            "property": property_name,
            "kind": kind,
            "key": key,
        }
        if target_month:
            detail["month"] = target_month
        details.append(detail)
    return details


def find_existing_coverage(rows: list[dict[str, str]], target_month: str) -> set[str]:
    """Return property/kind coverage from AOPS accruals or actual Scruggs escrow transfers."""
    existing = find_existing_accruals(rows, target_month)
    for row in rows:
        marker = parse_pm_fee_marker(str(row.get("Notes") or ""))
        if not marker or marker["month"] != target_month:
            continue
        if row_day(row) == 1:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        existing.add(f"{property_name}|pm")
    for template in ACCRUAL_TEMPLATES:
        kind = template["kind"]
        if kind not in {"taxes", "insurance"}:
            continue
        property_name = template["property"]
        if is_cash_basis_insurance_accrual(str(property_name), str(kind)):
            continue
        if is_no_dao_mortgage_escrow_accrual(str(property_name), str(kind)):
            continue
        key = f"{property_name}|{kind}"
        if key in existing:
            continue
        for row in rows:
            if not row_matches_property(row, property_name):
                continue
            if actual_scruggs_kind(row) != kind:
                continue
            if infer_service_month(row) == target_month:
                existing.add(key)
                break
    insurance_key = "9 Country Club Ln N|insurance"
    if insurance_key not in existing:
        for row in rows:
            if not row_matches_property(row, "9 Country Club Ln N"):
                continue
            if infer_service_month(row) != target_month:
                continue
            if "rental dwelling" not in row_category_names(row):
                continue
            if parse_marker(str(row.get("Notes") or "")):
                continue
            try:
                amount = float(str(row.get("Amount") or "0").replace(",", ""))
            except ValueError:
                continue
            if amount < 0:
                existing.add(insurance_key)
                break
    return existing


def find_duplicate_accrual_markers(
    rows: list[dict[str, str]],
    target_month: str,
    property_filters: list[str] | None = None,
    kind_filters: set[str] | None = None,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row_index, row in enumerate(rows):
        notes = str(row.get("Notes") or "")
        marker = parse_marker(notes) or parse_pm_fee_marker(notes)
        if not marker or marker["month"] != target_month:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        kind = marker.get("kind") or "pm"
        if not template_in_scope(property_name, kind, property_filters, kind_filters):
            continue
        key = f"{property_name}|{kind}|{target_month}"
        grouped[key].append({
            "row_index": row_index,
            "date": row.get("Date", ""),
            "amount": row.get("Amount", ""),
            "description": row.get("Description", ""),
            "notes": notes,
        })
    return [
        {
            "key": key,
            "property": key.rsplit("|", 2)[0],
            "kind": key.rsplit("|", 2)[1],
            "month": target_month,
            "row_count": len(marker_rows),
            "rows": [{field: value for field, value in item.items() if field != "row_index"} for item in marker_rows],
        }
        for key, marker_rows in sorted(grouped.items())
        if len(marker_rows) > 1
    ]


def find_blocked_first_day_pm_fees(
    rows: list[dict[str, str]],
    target_month: str,
    property_filters: list[str] | None = None,
    kind_filters: set[str] | None = None,
) -> list[dict[str, Any]]:
    """Return first-day direct PM fee rows that block month-end PM accrual generation."""
    if kind_filters and "pm" not in kind_filters:
        return []

    blocked: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        marker = parse_pm_fee_marker(str(row.get("Notes") or ""))
        if not marker or marker["month"] != target_month:
            continue
        if row_day(row) != 1:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        if not template_in_scope(property_name, "pm", property_filters, kind_filters):
            continue
        blocked.append({
            "row_index": index,
            "property": property_name,
            "kind": "pm",
            "month": target_month,
            "date": row.get("Date", ""),
            "amount": row.get("Amount", ""),
            "category": row.get("Category", ""),
            "description": row.get("Description", ""),
            "notes": row.get("Notes", ""),
            "reason": "first_day_pm_fee_source_row_blocks_month_end_pm_accrual",
        })
    return blocked


def find_amount_mismatches(
    rows: list[dict[str, str]],
    target_month: str,
    property_filters: list[str] | None = None,
    kind_filters: set[str] | None = None,
    lofty_reserves: dict[str, float] | None = None,
    approved_exceptions: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    lofty_reserves = lofty_reserves or {}
    templates = {(item["property"], item["kind"]): item for item in ACCRUAL_TEMPLATES}
    pm_expected = compute_pm_fees(rows, target_month)
    pm_settlement_expected = compute_pm_settlements(
        rows, target_month, expected_fees=pm_expected
    )
    hemlane_net_basis_cache: dict[str, float] = {}
    mismatches = []
    for index, row in enumerate(rows):
        notes = str(row.get("Notes") or "")
        marker = parse_marker(notes) or parse_pm_fee_marker(notes)
        if not marker or marker["month"] != target_month:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        kind = marker["kind"]
        template = templates.get((property_name, "dao" if kind == DAO_ECO_KIND else kind))
        retained_summary = None
        if property_name not in hemlane_net_basis_cache:
            hemlane_net_basis_cache[property_name] = hemlane_net_rent_amount(
                rows, property_name, target_month
            )
        hemlane_net_basis = hemlane_net_basis_cache[property_name]
        void_hemlane_direct_split = False
        if is_no_dao_mortgage_escrow_accrual(property_name, kind):
            # Mortgage-servicer tax and insurance escrow is the source of
            # truth for these properties. A manual accrual would double-count.
            expected = 0.0
        elif is_pm_accrual_kind(kind):
            expected = pm_expected.get(property_name)
        elif kind == "pm_settlement":
            expected = pm_settlement_expected.get(property_name)
        elif template:
            expected = fixed_accrual_posting_amount(
                property_name,
                "dao" if kind == DAO_ECO_KIND else kind,
                target_month,
                float(template["amount"]),
            )
        elif kind == RETAINED_CAPITAL_RULE["kind"] and canonical_reserve_property(property_name):
            retained_summary = retained_capital_amount(
                rows, target_month, property_name, lofty_reserves, approved_exceptions
            )
            expected = retained_summary["amount"] if retained_summary.get("status") == "ok" else None
        else:
            expected = None
        if (
            expected is None
            and hemlane_net_basis > 0
            and (is_pm_accrual_kind(kind) or kind == "pm_settlement")
        ):
            # Transaction-level evidence shows that this property-month's rent
            # arrived from Hemlane net of the PM split, with no direct gross
            # rent remaining as a manual-accrual basis.
            expected = 0.0
            void_hemlane_direct_split = True
        if expected is None:
            continue
        if not template_in_scope(property_name, kind, property_filters, kind_filters):
            continue
        current_marker_amount = float(marker["amount"])
        try:
            current_row_amount = abs(float(str(row.get("Amount") or "0").replace(",", "")))
        except ValueError:
            current_row_amount = current_marker_amount
        legacy_dao_label = kind == "dao" and "|legal|" in notes
        stale_pm_rule = False
        stale_retained_rule = False
        if is_pm_accrual_kind(kind):
            if void_hemlane_direct_split:
                # A zero-dollar void marker has no accounting or live mutation
                # effect. Older explanatory wording must not block a close.
                stale_pm_rule = (
                    round(current_row_amount, 2) != 0.0
                    or round(expected, 2) != 0.0
                ) and "transaction-level Hemlane rent evidence" not in notes
            else:
                default_rate = next((item[1] for item in PM_FEE_PROPERTIES if item[0] == property_name), 0.0)
                expected_rule = pm_fee_rule_summary(
                    property_name,
                    target_month,
                    default_rate,
                    PM_FEE_COMPONENTS.get(property_name) or [],
                )
                stale_pm_rule = expected_rule not in notes
        elif kind == RETAINED_CAPITAL_RULE["kind"] and canonical_reserve_property(property_name):
            expected_label = f"OR Replenishment | {property_name} | {month_label(target_month)}"
            expected_note = retained_capital_note(property_name, target_month, expected, retained_summary)
            description = str(row.get("Description") or "")
            classification_ok = (
                str(row.get("Type") or "") == "Transfers & Other"
                and str(row.get("Category") or "") == "Owner Contributions/Distributions"
            ) or (
                str(row.get("Type") or "") == "Manual"
                and str(row.get("Category") or "") == "Transfers & Other"
                and str(row.get("Sub-category") or "") == "Owner Contributions/Distributions"
            )
            canonical_note_prefix = (
                f"{RETAINED_CAPITAL_RULE['source_prefix']}|{RETAINED_CAPITAL_RULE['kind']}|"
                f"{property_name}|{target_month}|{expected:.2f} | "
            )
            policy_snapshot_ok = (
                notes.startswith(canonical_note_prefix)
                and "Outstanding cash reserve settlement requirement" in notes
            )
            stale_retained_rule = any(
                (
                    str(row.get("Merchant") or "") != expected_label,
                    # Baselane's export leaves Description blank for manual
                    # transactions even when merchantName is canonical.
                    description not in {"", expected_label},
                    not classification_ok,
                    not policy_snapshot_ok,
                )
            )
            if retained_summary.get("approved_exception"):
                # The preserved historical note identifies the approved source
                # basis; do not replace it with a current raw-export formula.
                stale_retained_rule = False
        if (
            round(current_row_amount, 2) == round(expected, 2)
            and not stale_pm_rule
            and not stale_retained_rule
        ):
            continue
        mismatches.append({
            "row_index": index,
            "property": property_name,
            "kind": kind,
            "month": target_month,
            "current_marker_amount": round(current_marker_amount, 2),
            "current_row_amount": round(current_row_amount, 2),
            "expected_amount": round(expected, 2),
            "legacy_dao_label": legacy_dao_label,
            "stale_pm_rule": stale_pm_rule,
            "stale_retained_rule": stale_retained_rule,
            "void_hemlane_direct_split": void_hemlane_direct_split,
        })
    return mismatches


def apply_amount_mismatch_updates(
    rows: list[dict[str, str]],
    mismatches: list[dict[str, Any]],
    lofty_reserves: dict[str, float] | None = None,
    approved_exceptions: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    lofty_reserves = lofty_reserves or {}
    applied = []
    for mismatch in mismatches:
        row = rows[mismatch["row_index"]]
        expected = float(mismatch["expected_amount"])
        notes = str(row.get("Notes") or "")
        if mismatch.get("void_hemlane_direct_split"):
            row["Amount"] = "0.00"
        else:
            row["Amount"] = f"{expected:.2f}" if mismatch["kind"] in {"pm_settlement", PM_ECO_KIND, DAO_ECO_KIND} else f"-{expected:.2f}"
        row["Notes"] = re.sub(r"(\|)\d+(?:\.\d+)?(\s+\|)", rf"\g<1>{expected:.2f}\g<2>", notes, count=1)
        if mismatch.get("void_hemlane_direct_split"):
            property_name = mismatch["property"]
            original_amount = float(mismatch["current_row_amount"])
            label = f"Hemlane Direct-Split PM Void | {property_name} | {month_label(mismatch['month'])}"
            row["Description"] = label
            row["Merchant"] = label
            row["Notes"] = (
                f"{parse_marker(notes)['prefix']}|{mismatch['kind']}|{property_name}|{mismatch['month']}|0.00 | "
                f"Voided legacy manual PM row of ${original_amount:.2f}: transaction-level Hemlane rent evidence "
                "shows that the PM fee was withheld/remitted before net rent reached Baselane, so a manual accrual "
                "or settlement would double-count it. "
                "Accounting correction only; no cash movement."
            )
        elif mismatch.get("legacy_dao_label"):
            row["Notes"] = row["Notes"].replace("|legal|", "|dao|", 1)
            label = f"DAO LLC Fee Accrual | {mismatch['property']} | {month_label(mismatch['month'])}"
            row["Description"] = label
            row["Merchant"] = label
        if mismatch.get("void_hemlane_direct_split"):
            pass
        elif mismatch["kind"] == DAO_ECO_KIND:
            property_name = mismatch["property"]
            template = next(
                item
                for item in ACCRUAL_TEMPLATES
                if item["property"] == property_name and item["kind"] == "dao"
            )
            prefix = template["source_prefix"]
            label = f"ECO Systems LLC DAO Registration Fee Revenue | {property_name} | {month_label(mismatch['month'])}"
            row["Description"] = label
            row["Merchant"] = label
            row["Type"] = "Revenue"
            row["Category"] = "Fees & Other Revenue"
            row["Sub-category"] = ""
            row["Property"] = "Mining, Sales, Consulting, and PM"
            row["Notes"] = (
                f"{prefix}|{DAO_ECO_KIND}|{property_name}|{mismatch['month']}|{expected:.2f} | "
                f"{DAO_LLC_ADMIN_ECO_REVENUE_NOTE}"
            )
        elif is_pm_accrual_kind(mismatch["kind"]):
            property_name = mismatch["property"]
            prefix = next((item[2] for item in PM_FEE_PROPERTIES if item[0] == property_name), SCHEDULE_SOURCE_PREFIX)
            components = PM_FEE_COMPONENTS.get(property_name) or []
            default_rate = next((item[1] for item in PM_FEE_PROPERTIES if item[0] == property_name), 0.0)
            component_summary = pm_fee_rule_summary(property_name, mismatch["month"], default_rate, components)
            accrual_id = f"{prefix}|pm|{property_name}|{mismatch['month']}|{expected:.2f}"
            if mismatch["kind"] == PM_ECO_KIND:
                label = f"ECO Systems LLC PM Fee Revenue | {property_name} | {month_label(mismatch['month'])}"
                row["Description"] = label
                row["Merchant"] = label
                row["Type"] = "Revenue"
                row["Category"] = "Fees & Other Revenue"
                row["Sub-category"] = "Property Management"
                row["Notes"] = f"{prefix}|{PM_ECO_KIND}|{property_name}|{mismatch['month']}|{expected:.2f} | Accrual ID {accrual_id}. ECO Systems LLC PM fee revenue for {month_label(mismatch['month'])} ({component_summary}); matched to the DAO-side expense. Accounting/manual accrual only, no bank transfer."
            else:
                label = f"PM Fee Accrual | {property_name} | {month_label(mismatch['month'])}"
                row["Description"] = label
                row["Merchant"] = label
                row["Type"] = "Operating Expenses"
                row["Category"] = "Property Management"
                row["Sub-category"] = ""
                row["Notes"] = f"{prefix}|{PM_DAO_KIND}|{property_name}|{mismatch['month']}|{expected:.2f} | Accrual ID {accrual_id}. DAO-side PM expense on Baselane gross rent for {month_label(mismatch['month'])} ({component_summary}); matched to ECO revenue. Accounting/manual accrual only, no bank transfer."
        elif mismatch["kind"] == "pm_settlement":
            property_name = mismatch["property"]
            label = f"PM Fee Settlement | {property_name} | {month_label(mismatch['month'])}"
            row["Description"] = label
            row["Merchant"] = label
            row["Notes"] = f"AOPS-PNL-ACCRUAL|pm_settlement|{property_name}|{mismatch['month']}|{expected:.2f} | Reverses the paid portion of the monthly PM fee accrual so the later ECO payment is not double-counted."
        elif mismatch["kind"] == RETAINED_CAPITAL_RULE["kind"] and canonical_reserve_property(mismatch["property"]):
            property_name = mismatch["property"]
            label = f"OR Replenishment | {property_name} | {month_label(mismatch['month'])}"
            retained_summary = retained_capital_amount(
                rows, mismatch["month"], property_name, lofty_reserves, approved_exceptions
            )
            row["Description"] = label
            row["Merchant"] = label
            row["Type"] = "Transfers & Other"
            row["Category"] = "Owner Contributions/Distributions"
            row["Sub-category"] = ""
            row["Notes"] = retained_capital_note(property_name, mismatch["month"], expected, retained_summary)
        elif (mismatch["property"], mismatch["kind"]) in ANNUAL_REFERENCE_FIXED_ACCRUALS:
            annual_reference = expected * 12
            row["Notes"] = (
                f"{parse_marker(notes)['prefix']}|{mismatch['kind']}|{mismatch['property']}|{mismatch['month']}|{expected:.2f} | "
                f"Monthly accrual from annual reference ${annual_reference:.2f}. "
                "Accounting/manual accrual only, no bank transfer."
            )
        applied.append({
            "property": mismatch["property"],
            "kind": mismatch["kind"],
            "month": mismatch["month"],
            "old_amount": mismatch["current_row_amount"],
            "new_amount": round(expected, 2),
            "description": row.get("Description", ""),
        })
    return applied


def template_in_scope(
    property_name: str,
    kind: str,
    property_filters: list[str] | None,
    kind_filters: set[str] | None,
) -> bool:
    normalized_kind = normalized_accrual_kind(kind)
    if kind_filters and kind not in kind_filters and normalized_kind not in kind_filters:
        return False
    if not property_filters:
        return True
    aliases = [property_name] + PROPERTY_ALIASES.get(property_name, [])
    normalized_aliases = [alias.lower() for alias in aliases]
    for property_filter in property_filters:
        needle = property_filter.lower()
        if any(needle in alias or alias in needle for alias in normalized_aliases):
            return True
    return False


def gross_revenue_by_property(rows: list[dict[str, str]], target_month: str) -> dict[str, float]:
    rent_by_property: dict[str, float] = defaultdict(float)
    for row in rows:
        row_date = parse_row_date(row)
        if row_date is None:
            continue
        row_month = f"{row_date.year:04d}-{row_date.month:02d}"
        if row_month != target_month:
            continue

        amount_str = str(row.get("Amount") or "0").replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            continue

        if is_rent_revenue(row, amount):
            property_name = str(row.get("Property") or "").strip()
            rent_by_property[property_name] += amount
    return dict(rent_by_property)


def property_gross_revenue(rows: list[dict[str, str]], property_name: str, target_month: str) -> float:
    total = 0.0
    for row in rows:
        row_date = parse_row_date(row)
        try:
            amount = float(str(row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        if row_date is None:
            continue
        if f"{row_date.year:04d}-{row_date.month:02d}" != target_month:
            continue
        if is_rent_revenue(row, amount) and row_matches_property(row, property_name):
            total += amount
    return round(total, 2)


def separately_booked_jazmine_pm_fee(rows: list[dict[str, str]], property_name: str, target_month: str) -> float:
    if property_name not in {"86 Madison Ave", "88 Madison Ave", "90 Madison Ave"}:
        return 0.0
    total = 0.0
    for row in rows:
        if not row_matches_property(row, property_name) or infer_service_month(row) != target_month:
            continue
        notes = str(row.get("Notes") or "")
        if parse_marker(notes) or parse_pm_fee_marker(notes):
            continue
        text = row_text(row, ["Merchant", "Description", "Notes"]).lower()
        if "jazmine" not in text:
            continue
        if not ("3 percent" in text or "3%" in text):
            continue
        if not any(token in text for token in ("co-host", "cohost", "guest specialist", "pm fee")):
            continue
        try:
            amount = float(str(row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        if amount < 0:
            total += abs(amount)
    reviewed_override = MADISON_UNREIMBURSED_COHOST_PM_OVERRIDES.get((property_name, target_month))
    if reviewed_override is not None:
        return reviewed_override
    return round(total, 2)


def separately_booked_eco_pm_fee_payment(rows: list[dict[str, str]], property_name: str, target_month: str) -> float:
    total = 0.0
    for row in rows:
        if not row_matches_property(row, property_name) or infer_service_month(row) != target_month:
            continue
        notes = str(row.get("Notes") or "")
        if parse_marker(notes) or parse_pm_fee_marker(notes):
            continue
        categories = row_category_names(row)
        text = row_text(row, ["Merchant", "Description", "Notes"]).lower()
        is_pm_expense = bool({"property management", "management fees"} & categories)
        is_explicit_pm_transfer = (
            "transfers between accounts" in categories
            and any(token in text for token in ("pm fee", "property management fee", "management fees"))
        )
        if not (is_pm_expense or is_explicit_pm_transfer):
            continue
        if "eco systems" not in text and "pm fee" not in text and "property management fee" not in text and "management fees" not in text:
            continue
        try:
            amount = float(str(row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        if amount < 0:
            total += abs(amount)
    return round(total, 2)


def separately_booked_eco_pm_fee_payment_date(rows: list[dict[str, str]], property_name: str, target_month: str) -> str:
    payment_dates = []
    for row in rows:
        if not row_matches_property(row, property_name) or infer_service_month(row) != target_month:
            continue
        notes = str(row.get("Notes") or "")
        if parse_marker(notes) or parse_pm_fee_marker(notes):
            continue
        categories = row_category_names(row)
        text = row_text(row, ["Merchant", "Description", "Notes"]).lower()
        is_pm_expense = bool({"property management", "management fees"} & categories)
        is_explicit_pm_transfer = (
            "transfers between accounts" in categories
            and any(token in text for token in ("pm fee", "property management fee", "management fees"))
        )
        if not (is_pm_expense or is_explicit_pm_transfer):
            continue
        if "eco systems" not in text and "pm fee" not in text and "property management fee" not in text and "management fees" not in text:
            continue
        try:
            amount = float(str(row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        row_date = parse_row_date(row)
        if amount < 0 and row_date is not None:
            payment_dates.append(row_date)
    return max(payment_dates).strftime("%B %d, %Y") if payment_dates else accrual_posting_date(target_month)


def madison_jazmine_pm_fee_deductions(rows: list[dict[str, str]], target_month: str) -> dict[str, float]:
    return {
        property_name: amount
        for property_name in ("86 Madison Ave", "88 Madison Ave", "90 Madison Ave")
        if (amount := separately_booked_jazmine_pm_fee(rows, property_name, target_month)) > 0
    }


def compute_pm_fees(rows: list[dict[str, str]], target_month: str) -> dict[str, float]:
    """Compute PM fees as rate × gross rent for each PM property in the target month."""
    pm_fees: dict[str, float] = {}
    for property_name, rate, _ in PM_FEE_PROPERTIES:
        components = PM_FEE_COMPONENTS.get(property_name) or [{"rate": rate, "deduction": 0.0, "basis_token": None}]
        property_rows = []
        for row in rows:
            row_date = parse_row_date(row)
            try:
                row_amount = float(str(row.get("Amount") or "0").replace(",", ""))
            except ValueError:
                continue
            if row_date is None:
                continue
            if f"{row_date.year:04d}-{row_date.month:02d}" != target_month:
                continue
            if row_matches_property(row, property_name) and is_rent_revenue(row, row_amount):
                property_rows.append(row)
        if property_name == "9 Country Club Ln N":
            evolve_net = sum(
                (Decimal(str(row.get("Amount") or "0").replace(",", ""))
                for row in property_rows
                if "evolve" in row_text(row, ["Merchant", "Description", "Notes"]).lower()),
                Decimal("0"),
            )
            non_evolve_gross = sum(
                (Decimal(str(row.get("Amount") or "0").replace(",", ""))
                for row in property_rows
                if "evolve" not in row_text(row, ["Merchant", "Description", "Notes"]).lower()),
                Decimal("0"),
            )
            if evolve_net > 0 or non_evolve_gross > 0:
                # Evolve already retains its 10% share before remitting 90% net.
                # ECO's remaining 10% of gross is therefore net / 9.
                pm_fees[property_name] = round_money(
                    evolve_net / Decimal("9") + non_evolve_gross * Decimal("0.20")
                )
            continue
        amount = Decimal("0")
        has_revenue_basis = False
        for component in components:
            basis_token = str(component.get("basis_token") or "") if len(components) > 1 else ""
            component_rows = [
                row
                for row in property_rows
                if (not basis_token or basis_token in row_text(row, ["Merchant", "Description", "Notes", "Unit"]).lower())
                and not is_hemlane_net_rent(row)
            ]
            gross_rent = sum(
                (Decimal(str(row.get("Amount") or "0").replace(",", "")) for row in component_rows),
                Decimal("0"),
            )
            if gross_rent <= 0:
                continue
            has_revenue_basis = True
            component_rate = effective_pm_rate(property_name, target_month, float(component["rate"]))
            amount += max(
                Decimal("0"),
                gross_rent * Decimal(str(component_rate)) - Decimal(str(component.get("deduction") or 0)),
            )
        if has_revenue_basis:
            amount = max(
                Decimal("0"),
                amount - Decimal(str(separately_booked_jazmine_pm_fee(rows, property_name, target_month))),
            )
            pm_fees[property_name] = round_money(amount)
            continue
        hemlane_basis = HEMLANE_PM_FEE_BASIS.get(property_name)
        if (
            (property_name, target_month) in PM_DIRECT_SPLIT_UNPAID_EXCEPTIONS
            and hemlane_basis
            and hemlane_basis.get("month") == target_month
        ):
            pm_amount = Decimal(str(hemlane_basis.get("computed_pm_fee_amount") or "0"))
            if pm_amount > 0:
                pm_fees[property_name] = round_money(pm_amount)
    return pm_fees


def compute_pm_settlements(
    rows: list[dict[str, str]],
    target_month: str,
    expected_fees: dict[str, float] | None = None,
) -> dict[str, float]:
    expected_fees = expected_fees if expected_fees is not None else compute_pm_fees(rows, target_month)
    settlements = {}
    for property_name, expected_fee in expected_fees.items():
        paid = separately_booked_eco_pm_fee_payment(rows, property_name, target_month)
        if paid > 0 and expected_fee > 0:
            settlements[property_name] = round(min(paid, expected_fee), 2)
    return settlements


def aligned_net_transfer_basis_gap_evidence(rows: list[dict[str, str]], property_name: str, target_month: str) -> dict[str, Any] | None:
    def row_amount(row: dict[str, str]) -> Decimal:
        try:
            return Decimal(str(row.get("Amount") or "0").replace(",", ""))
        except Exception:
            return Decimal("0")

    matches = [
        row
        for row in rows
        if (row_date := parse_row_date(row)) is not None
        and f"{row_date.year:04d}-{row_date.month:02d}" == target_month
        and row_matches_property(row, property_name)
        and str(row.get("Type") or "").strip() == "Transfers & Other"
        and str(row.get("Category") or "").strip() == "Transfers Between Accounts"
        and "aligned" in " ".join(str(row.get(key) or "") for key in ("Merchant", "Description", "Notes")).lower()
        and row_amount(row) > 0
    ]
    if not matches:
        return None
    amount = sum((row_amount(row) for row in matches), Decimal("0"))
    return {
        "aligned_net_transfer_count": len(matches),
        "aligned_net_transfer_amount": float(round_money(amount)),
        "aligned_net_transfer_dates": sorted({str(row.get("Date") or "") for row in matches if str(row.get("Date") or "")}),
        "aligned_net_transfer_merchants": sorted({str(row.get("Merchant") or "") for row in matches if str(row.get("Merchant") or "")})[:5],
    }


def pm_fee_basis_gaps(
    rows: list[dict[str, str]],
    target_month: str,
    existing: set[str],
    property_filters: list[str] | None = None,
    kind_filters: set[str] | None = None,
) -> list[dict[str, Any]]:
    if kind_filters and "pm" not in kind_filters:
        return []
    pm_fees = compute_pm_fees(rows, target_month)
    gaps: list[dict[str, Any]] = []
    for property_name, rate, _prefix in PM_FEE_PROPERTIES:
        if not pm_manual_accrual_required(property_name, target_month):
            continue
        if not template_in_scope(property_name, "pm", property_filters, kind_filters):
            continue
        if is_before_launch(property_name, target_month):
            continue

        key = f"{property_name}|pm"
        if key in existing or property_name in pm_fees:
            continue
        if hemlane_net_rent_amount(rows, property_name, target_month) > 0:
            # Hemlane remits these receipts after withholding its PM split.
            # The positive bank receipt is evidence of rent activity, while a
            # second manual PM accrual would duplicate the withheld fee.
            continue
        hemlane_basis = HEMLANE_PM_FEE_BASIS.get(property_name) or {}
        rent_ids = hemlane_basis.get("rent_transaction_ids") or []
        if (
            hemlane_basis.get("month") == target_month
            and float(hemlane_basis.get("rent_request_amount") or 0) > 0
            and int(hemlane_basis.get("rent_success_evidence_count") or 0) == len(rent_ids)
            and len(rent_ids) > 0
            and float(hemlane_basis.get("rent_success_amount") or 0) == 0
        ):
            # The full Baselane ledger has no rent basis and Hemlane explicitly
            # reports zero successful collection for every scheduled rent row.
            # A zero PM marker is deterministic; carrying prior rent forward
            # would create an expense and ECO revenue that never occurred.
            continue
        previous = previous_month(target_month)
        previous_gross_rent = property_gross_revenue(rows, property_name, previous)
        if previous_gross_rent <= 1.0:
            continue
        aligned_evidence = aligned_net_transfer_basis_gap_evidence(rows, property_name, target_month)
        reason = "no_current_month_revenue_basis_for_pm_fee"
        review_action = (
            "Verify whether current-month rent was truly zero or whether Baselane/APG/Evernest rent rows "
            "are missing or tagged to the wrong property before waiving the PM accrual."
        )
        if aligned_evidence:
            reason = "aligned_owner_statement_detail_missing_for_pm_fee"
            review_action = (
                "Aligned net cash hit Baselane for the month, but Aligned owner-statement detail has not "
                "been imported for the same month. Import or attach the Aligned detail before waiving PM accrual."
            )
        gaps.append({
            "property": property_name,
            "kind": "pm",
            "key": key,
            "month": target_month,
            "rate": rate,
            "current_month_gross_rent": 0.0,
            "previous_month": previous,
            "previous_month_gross_rent": previous_gross_rent,
            "reason": reason,
            "review_action": review_action,
            **(aligned_evidence or {}),
        })
    return gaps


def accrual_gap_action_queue(
    missing_fixed_coverage: list[dict[str, Any]],
    pm_basis_gaps: list[dict[str, Any]],
    active_without_fixed_requirement: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    queue: list[dict[str, Any]] = []
    pm_gap_keys = {str(item.get("key") or "") for item in pm_basis_gaps}
    for item in pm_basis_gaps:
        queue.append(
            {
                "property": item.get("property"),
                "kind": item.get("kind"),
                "month": item.get("month"),
                "severity": "blocks_monthly_close",
                "action": "verify_missing_rent_or_approve_zero_pm",
                "reason": item.get("reason"),
                "review_action": item.get("review_action"),
                "current_month_gross_rent": item.get("current_month_gross_rent"),
                "previous_month": item.get("previous_month"),
                "previous_month_gross_rent": item.get("previous_month_gross_rent"),
                "aligned_net_transfer_count": item.get("aligned_net_transfer_count"),
                "aligned_net_transfer_amount": item.get("aligned_net_transfer_amount"),
                "aligned_net_transfer_dates": item.get("aligned_net_transfer_dates"),
                "aligned_net_transfer_merchants": item.get("aligned_net_transfer_merchants"),
                "key": item.get("key"),
            }
        )
    for item in missing_fixed_coverage:
        if str(item.get("key") or "") in pm_gap_keys:
            continue
        queue.append(
            {
                "property": item.get("property"),
                "kind": item.get("kind"),
                "month": item.get("month"),
                "severity": "blocks_monthly_close",
                "action": "generate_or_verify_missing_fixed_accrual",
                "reason": "expected_fixed_accrual_not_covered",
                "review_action": "Rerun monthly accrual generation/apply for this property/kind, then verify the AOPS marker appears exactly once in the ECO GL.",
                "key": item.get("key"),
            }
        )
    for item in active_without_fixed_requirement:
        queue.append(
            {
                "property": item.get("full_address") or item.get("property_name"),
                "kind": "fixed_accrual_template",
                "month": None,
                "severity": "audit_only",
                "action": "document_or_add_fixed_accrual_requirement",
                "reason": item.get("reason"),
                "review_action": "Confirm whether this active DAO should have recurring tax, insurance, PM, or DAO LLC fee accrual templates; add templates when recurring obligations exist.",
                "lofty_property_id": item.get("lofty_property_id"),
                "assetUnit": item.get("assetUnit"),
            }
        )
    return queue


def blocking_gap_actions(queue: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [item for item in queue if item.get("severity") == "blocks_monthly_close"]


def markdown_money(value: object) -> str:
    try:
        return f"${float(value or 0):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def build_review_markdown(report: dict[str, Any]) -> str:
    queue = report.get("gap_action_queue") if isinstance(report.get("gap_action_queue"), list) else []
    blocking = [item for item in queue if isinstance(item, dict) and item.get("severity") == "blocks_monthly_close"]
    audit_only = [item for item in queue if isinstance(item, dict) and item.get("severity") == "audit_only"]
    scaffold = report.get("gap_approval_scaffold") if isinstance(report.get("gap_approval_scaffold"), dict) else {}
    review_csv = report.get("gap_approval_review_csv") if isinstance(report.get("gap_approval_review_csv"), dict) else {}
    import_commands = report.get("gap_approval_import_commands") if isinstance(report.get("gap_approval_import_commands"), dict) else {}
    lines = [
        "# Monthly Accrual Completeness Review",
        "",
        f"Generated at: {report.get('generated_at')}",
        f"Target month: `{report.get('target_month')}`",
        f"Status: `{report.get('status')}`",
        "",
        "## Gate Summary",
        "",
        f"- Missing fixed accrual coverage: `{report.get('missing_fixed_accrual_coverage_count')}`",
        f"- PM fee basis gaps: `{report.get('pm_fee_basis_gap_count')}`",
        f"- Blocking actions: `{report.get('blocking_gap_action_count')}`",
        f"- Audit-only template questions: `{len(audit_only)}`",
        "",
        "## Blocking Actions",
        "",
    ]
    if not blocking:
        lines.append("- None")
    else:
        lines.extend(
            [
                "| Property | Kind | Key | Action | Current Rent | Previous Rent | Current PM | Previous PM | Evidence Digest | Review |",
                "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- | --- |",
            ]
        )
        for item in blocking:
            evidence_digest = item.get("evidence_digest") or gap_approval_evidence_digest(item)
            lines.append(
                "| "
                + " | ".join(
                    [
                        str(item.get("property") or "").replace("|", "/"),
                        str(item.get("kind") or ""),
                        f"`{item.get('key')}`",
                        str(item.get("action") or ""),
                        markdown_money(item.get("current_month_gross_rent")),
                        markdown_money(item.get("previous_month_gross_rent")),
                        markdown_money(item.get("current_implied_pm_fee_amount")),
                        markdown_money(item.get("previous_month_implied_pm_fee_amount")),
                        f"`{str(evidence_digest)}`",
                        str(item.get("review_action") or "").replace("|", "/"),
                    ]
                )
                + " |"
            )
    lines.extend(
        [
            "",
            "## Approval File Schema",
            "",
            "Use `--gap-approvals config/baselane_monthly_accrual_gap_approvals.json` for reviewed non-mutating decisions.",
            "Only `approve_zero_pm_fee` is currently accepted, and only for `verify_missing_rent_or_approve_zero_pm` rows.",
            "Each approval must keep the exact `evidence_digest` from the table/scaffold; stale digests are rejected.",
            "Reviewed approvals must include a substantive `note` documenting the rent-source review performed.",
            "",
            "```json",
            '{\"approvals\":[{\"key\":\"PROPERTY|pm\",\"decision\":\"approve_zero_pm_fee\",\"reviewed\":true,\"reviewed_at\":\"2026-07-13T00:00:00Z\",\"evidence_digest\":\"sha256\",\"note\":\"Checked Baselane/APG/Evernest; no June rent rows found or expected.\"}]}',
            "```",
            "",
            "## Approval Artifacts",
            "",
            f"- Scaffold JSON: `{scaffold.get('path') or 'not written'}`",
            f"- Review CSV: `{review_csv.get('path') or 'not written'}`",
            f"- Guarded import script: `{import_commands.get('path') or 'not written'}`",
            "- Import command: `bash reports/baselane_monthly_accrual_gap_approvals_import.requires-explicit-approval.sh`",
            "- The import script reruns the accrual audit and refuses to clear monthly close while any approval, coverage, or PM-basis blocker remains.",
        ]
    )
    lines.extend(["", "## Audit-Only Template Questions", ""])
    if not audit_only:
        lines.append("- None")
    else:
        for item in audit_only[:50]:
            lines.append(
                f"- `{item.get('property')}` — {item.get('reason')} Action: {item.get('review_action')}"
            )
    return "\n".join(lines).rstrip() + "\n"


def write_preflight_failure_report(
    report_path: Path | None,
    review_markdown_path: Path | None,
    *,
    target_month: str,
    code: str,
    message: str,
    **details: Any,
) -> dict[str, Any]:
    report: dict[str, Any] = {
        "generated_at": iso_z(),
        "target_month": target_month,
        "months": [target_month] if target_month else [],
        "month_count": 1 if target_month else 0,
        "status": "blocked",
        "preflight_failure": {
            "code": code,
            "message": message,
            **details,
        },
        "missing_fixed_accrual_coverage_count": None,
        "pm_fee_basis_gap_count": None,
        "unapproved_pm_fee_basis_gap_count": None,
        "blocking_gap_action_count": None,
        "amount_mismatch_count": None,
        "blocked_first_day_pm_fee_count": None,
        "duplicate_accrual_marker_count": None,
        "gap_action_queue": [],
        "missing_accruals": [],
        "appended": False,
        "mode": "preflight_blocked",
    }
    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    if review_markdown_path:
        review_markdown_path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            "# Monthly Accrual Completeness Review",
            "",
            f"Generated at: {report['generated_at']}",
            f"Target month: `{target_month}`",
            "Status: `blocked`",
            "",
            f"- Preflight failure: `{code}`",
            f"- Message: {message}",
        ]
        missing_reserves = details.get("missing_lofty_reserves")
        diagnostics = details.get("missing_lofty_reserve_diagnostics")
        if isinstance(missing_reserves, list) and missing_reserves:
            lines.append("- Missing live reserve properties: " + ", ".join(f"`{item}`" for item in missing_reserves))
        if isinstance(diagnostics, list):
            for item in diagnostics:
                if not isinstance(item, dict):
                    continue
                property_name = str(item.get("property") or "unknown property")
                status = str(item.get("status") or "unknown")
                action = str(item.get("required_action") or "")
                lines.append(f"- `{property_name}`: `{status}`. {action}".rstrip())
        review_markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report


def load_gap_approvals(path: Path | None) -> dict[str, Any]:
    if not path:
        return {"status": "not_configured", "path": None, "approvals": []}
    if not path.is_file():
        return {"status": "missing", "path": str(path), "approvals": []}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return {"status": "unreadable", "path": str(path), "approvals": [], "error": f"{type(exc).__name__}: {exc}"}
    approvals = payload.get("approvals") if isinstance(payload, dict) else []
    return {
        "status": "loaded" if isinstance(approvals, list) else "invalid",
        "path": str(path),
        "approvals": approvals if isinstance(approvals, list) else [],
    }



def gap_approval_evidence_payload(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "key": item.get("key"),
        "property": item.get("property"),
        "kind": item.get("kind"),
        "month": item.get("month"),
        "action": item.get("action"),
        "current_month_gross_rent": round(float(item.get("current_month_gross_rent") or 0), 2),
        "previous_month": item.get("previous_month"),
        "previous_month_gross_rent": round(float(item.get("previous_month_gross_rent") or 0), 2),
    }


def gap_approval_evidence_digest(item: dict[str, Any]) -> str:
    payload = gap_approval_evidence_payload(item)
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def implied_pm_fee_amount(item: dict[str, Any]) -> float:
    return round(float(item.get("current_month_gross_rent") or 0) * 0.10, 2)

def apply_gap_approvals(queue: list[dict[str, Any]], approval_payload: dict[str, Any]) -> dict[str, Any]:
    approvals = approval_payload.get("approvals") if isinstance(approval_payload.get("approvals"), list) else []
    approval_by_key = {
        str(item.get("key") or ""): item
        for item in approvals
        if isinstance(item, dict) and str(item.get("key") or "")
    }
    issues: list[dict[str, Any]] = []
    approved_count = 0
    stale_keys = set(approval_by_key)
    for item in queue:
        key = str(item.get("key") or "")
        if key in stale_keys:
            stale_keys.remove(key)
        approval = approval_by_key.get(key)
        if not approval:
            continue
        decision = str(approval.get("decision") or "")
        if approval.get("reviewed") is not True:
            issues.append({"key": key, "code": "approval_not_reviewed"})
            continue
        if not approval.get("reviewed_at"):
            issues.append({"key": key, "code": "approval_missing_reviewed_at"})
            continue
        if len(str(approval.get("note") or "").strip()) < 20:
            issues.append({"key": key, "code": "approval_missing_substantive_note"})
            continue
        if item.get("action") == "verify_missing_rent_or_approve_zero_pm":
            if decision != "approve_zero_pm_fee":
                issues.append({"key": key, "code": "invalid_pm_gap_decision", "decision": decision})
                continue
            expected_digest = gap_approval_evidence_digest(item)
            if approval.get("evidence_digest") != expected_digest:
                issues.append(
                    {
                        "key": key,
                        "code": "approval_evidence_digest_mismatch",
                        "expected_digest": expected_digest,
                        "approval_digest": approval.get("evidence_digest"),
                    }
                )
                continue
            item["approval_decision"] = decision
            item["approval_reviewed_at"] = approval.get("reviewed_at")
            item["approval_note"] = approval.get("note")
            item["severity"] = "reviewed_nonblocking"
            item["review_action"] = "Approved as zero PM-fee basis for this month; no synthetic PM accrual will be generated."
            approved_count += 1
            continue
        issues.append({"key": key, "code": "approval_not_supported_for_action", "action": item.get("action")})
    ignored_stale_approval_keys = sorted(stale_keys)
    return {
        "status": "ok" if not issues else "review",
        "path": approval_payload.get("path"),
        "loaded_status": approval_payload.get("status"),
        "approval_count": len(approvals),
        "approved_count": approved_count,
        "ignored_stale_approval_count": len(ignored_stale_approval_keys),
        "ignored_stale_approval_keys": ignored_stale_approval_keys[:200],
        "issue_count": len(issues),
        "issues": issues[:200],
    }


def build_gap_approval_scaffold(queue: list[dict[str, Any]], existing_payload: dict[str, Any] | None = None) -> dict[str, Any]:
    existing_approvals = (
        existing_payload.get("approvals")
        if isinstance(existing_payload, dict) and isinstance(existing_payload.get("approvals"), list)
        else []
    )
    existing_by_key = {
        str(item.get("key") or ""): item
        for item in existing_approvals
        if isinstance(item, dict) and str(item.get("key") or "")
    }
    approvals: list[dict[str, Any]] = []
    for item in queue:
        if item.get("action") != "verify_missing_rent_or_approve_zero_pm":
            continue
        key = str(item.get("key") or "")
        if not key:
            continue
        existing = existing_by_key.get(key)
        evidence_payload = gap_approval_evidence_payload(item)
        evidence_digest = gap_approval_evidence_digest(item)
        current_implied_pm_fee_amount = implied_pm_fee_amount(item)
        if existing and existing.get("reviewed") is True and existing.get("evidence_digest") == evidence_digest:
            approvals.append(existing)
            continue
        approvals.append(
            {
                "key": key,
                "property": item.get("property"),
                "kind": item.get("kind"),
                "month": item.get("month"),
                "decision": "approve_zero_pm_fee",
                "reviewed": False,
                "reviewed_at": None,
                "note": (
                    "Set reviewed=true only after confirming current-month rent is truly zero "
                    "and no Baselane/APG/Evernest rent rows are missing or mis-tagged."
                ),
                "required_review_evidence": [
                    "Baselane current-month GL/rent rows checked for this property and aliases",
                    "APG/Evernest/current PM portal or owner statement checked when applicable",
                    "No current-month rent was expected or collected, or missing/mistagged rent rows were corrected upstream",
                ],
                "review_note_template": (
                    "Checked Baselane/APG/Evernest for {month}; no rent rows were missing or mistagged because ..."
                ),
                "approval_effect": "waive_pm_accrual_for_month_only_no_cash_transfer",
                "evidence_digest": evidence_digest,
                "evidence": evidence_payload,
                "current_month_gross_rent": item.get("current_month_gross_rent"),
                "current_implied_pm_fee_amount": current_implied_pm_fee_amount,
                "previous_month": item.get("previous_month"),
                "previous_month_gross_rent": item.get("previous_month_gross_rent"),
                "previous_month_implied_pm_fee_amount": round(float(item.get("previous_month_gross_rent") or 0) * 0.10, 2),
            }
        )
    return {
        "status": "review_required" if approvals else "ok",
        "generated_at": iso_z(),
        "policy": "Human review scaffold only. Reviewed approvals are non-mutating and only waive zero-rent PM fee accrual generation for the listed month/key. Reviewed approvals must retain the current evidence_digest or they are rejected as stale.",
        "approval_count": len(approvals),
        "approvals": approvals,
    }


def write_gap_approval_scaffold(path: Path | None, queue: list[dict[str, Any]]) -> dict[str, Any]:
    if not path:
        return {"status": "not_configured", "path": None, "written": False}
    existing = load_gap_approvals(path) if path.exists() else {}
    payload = build_gap_approval_scaffold(queue, existing)
    existing_loaded_status = existing.get("status") if isinstance(existing, dict) else None
    existing_recovered = bool(path.exists() and existing_loaded_status not in {None, "loaded"})
    existing_reviewed_count = sum(
        1
        for item in (existing.get("approvals") if isinstance(existing.get("approvals"), list) else [])
        if isinstance(item, dict) and item.get("reviewed") is True
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {
        "status": "written",
        "path": str(path),
        "written": True,
        "approval_count": payload["approval_count"],
        "existing_reviewed_count": existing_reviewed_count,
        "existing_loaded_status": existing_loaded_status,
        "existing_recovered": existing_recovered,
        "policy": payload["policy"],
    }


def write_gap_approval_review_csv(path: Path | None, scaffold_payload: dict[str, Any]) -> dict[str, Any]:
    if not path:
        return {"status": "not_configured", "path": None, "written": False}
    approvals = scaffold_payload.get("approvals") if isinstance(scaffold_payload.get("approvals"), list) else []
    fieldnames = [
        "key",
        "property",
        "kind",
        "month",
        "decision",
        "reviewed",
        "reviewed_at",
        "note",
        "evidence_digest",
        "current_month_gross_rent",
        "previous_month_gross_rent",
        "approval_effect",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for approval in approvals:
            if not isinstance(approval, dict):
                continue
            writer.writerow({key: approval.get(key) for key in fieldnames})
    return {"status": "written", "path": str(path), "written": True, "approval_count": len(approvals)}


def truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "reviewed", "approved"}


def import_gap_approval_review_csv(csv_path: Path | None, approvals_path: Path | None, min_note_length: int = 20) -> dict[str, Any]:
    if not csv_path:
        return {"status": "not_configured", "csv": None, "approvals": str(approvals_path) if approvals_path else None, "imported_count": 0}
    if not approvals_path:
        return {"status": "missing_approvals_path", "csv": str(csv_path), "approvals": None, "imported_count": 0}
    if not csv_path.is_file():
        return {"status": "missing_csv", "csv": str(csv_path), "approvals": str(approvals_path), "imported_count": 0}
    payload = load_gap_approvals(approvals_path)
    approvals = payload.get("approvals") if isinstance(payload.get("approvals"), list) else []
    by_key = {str(item.get("key") or ""): item for item in approvals if isinstance(item, dict)}
    issues: list[dict[str, Any]] = []
    imported_count = 0
    skipped_count = 0
    with csv_path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), start=2):
            key = str(row.get("key") or "").strip()
            if not key:
                issues.append({"row": row_number, "code": "missing_key"})
                continue
            approval = by_key.get(key)
            if not approval:
                issues.append({"row": row_number, "key": key, "code": "key_not_in_current_scaffold"})
                continue
            if not truthy(row.get("reviewed")):
                skipped_count += 1
                continue
            decision = str(row.get("decision") or "").strip()
            reviewed_at = str(row.get("reviewed_at") or "").strip()
            note = str(row.get("note") or "").strip()
            evidence_digest = str(row.get("evidence_digest") or "").strip()
            if decision != "approve_zero_pm_fee":
                issues.append({"row": row_number, "key": key, "code": "invalid_decision", "decision": decision})
                continue
            if not reviewed_at:
                issues.append({"row": row_number, "key": key, "code": "missing_reviewed_at"})
                continue
            if len(note) < min_note_length:
                issues.append({"row": row_number, "key": key, "code": "review_note_too_short", "min_note_length": min_note_length})
                continue
            if evidence_digest != str(approval.get("evidence_digest") or ""):
                issues.append({"row": row_number, "key": key, "code": "evidence_digest_mismatch"})
                continue
            approval["reviewed"] = True
            approval["reviewed_at"] = reviewed_at
            approval["note"] = note
            approval["decision"] = decision
            imported_count += 1
    output = {
        **({"path": str(approvals_path)} if payload.get("path") else {}),
        "status": "review_required",
        "generated_at": iso_z(),
        "policy": "Human review scaffold only. Reviewed approvals are non-mutating and only waive zero-rent PM fee accrual generation for the listed month/key. Reviewed approvals must retain the current evidence_digest or they are rejected as stale.",
        "approval_count": len(approvals),
        "approvals": approvals,
        "csv_import": {
            "status": "ok" if not issues else "review",
            "csv": str(csv_path),
            "imported_count": imported_count,
            "skipped_unreviewed_count": skipped_count,
            "issue_count": len(issues),
            "issues": issues[:200],
        },
    }
    approvals_path.parent.mkdir(parents=True, exist_ok=True)
    approvals_path.write_text(json.dumps(output, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return output["csv_import"]


def write_gap_approval_import_commands(path: Path | None) -> dict[str, Any]:
    if not path:
        return {"status": "not_configured", "path": None, "written": False}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "#!/usr/bin/env bash",
                "set -euo pipefail",
                "# Requires explicit human approval after completing the accrual gap approval CSV.",
                "# Imports reviewed=true zero-PM approvals, reruns accrual audit, and refuses dirty validation.",
                'echo "[monthly-accruals] importing reviewed gap approval CSV and rerunning audit"',
                "BASELANE_MONTHLY_ACCRUALS_IMPORT_GAP_APPROVAL_CSV=1 "
                "BASELANE_MONTHLY_ACCRUALS_APPLY=0 bash scripts/baselane_monthly_accruals_cron.sh",
                'STATUS="$(python3 - <<\'PY\'',
                "import json",
                "payload=json.load(open('reports/baselane_monthly_accruals_202606.json', encoding='utf-8'))",
                "status=payload.get('status') or ''",
                "print(status)",
                "if status != 'ok':",
                "    print('missing_fixed_accrual_coverage_count=' + str(payload.get('missing_fixed_accrual_coverage_count') or 0))",
                "    print('pm_fee_basis_gap_count=' + str(payload.get('pm_fee_basis_gap_count') or 0))",
                "    print('unapproved_pm_fee_basis_gap_count=' + str(payload.get('unapproved_pm_fee_basis_gap_count') or 0))",
                "    print('blocking_gap_action_count=' + str(payload.get('blocking_gap_action_count') or 0))",
                "    approval_import=(payload.get('gap_approval_import') or {})",
                "    print('gap_approval_import_status=' + str(approval_import.get('status') or ''))",
                "    for issue in approval_import.get('issues') or []:",
                "        print('gap_approval_import_issue=' + str(issue))",
                "    for action in payload.get('gap_action_queue') or []:",
                "        print('gap_action=' + str(action.get('key') or '') + '|' + str(action.get('reason') or ''))",
                "PY",
                ')"',
                'if [ "${STATUS%%$\'\\n\'*}" != "ok" ]; then',
                '  echo "[monthly-accruals] accrual report status is $STATUS; refusing monthly close" >&2',
                "  exit 1",
                "fi",
                "",
            ]
        ),
        encoding="utf-8",
    )
    path.chmod(0o600)
    return {"status": "written", "path": str(path), "written": True}


def is_before_launch(property_name: str, target_month: str) -> bool:
    accrual_start = PROPERTY_ACCRUAL_START_MONTHS.get(property_name)
    if accrual_start:
        return target_month < accrual_start
    launch_str = PROPERTY_LAUNCH_DATES.get(property_name)
    if not launch_str:
        return False
    launch_date = dt.date.fromisoformat(launch_str)
    target_year, target_m = [int(x) for x in target_month.split("-")]
    target_date = dt.date(target_year, target_m, 1)
    return target_date < launch_date


def generate_missing_accruals(
    rows: list[dict[str, str]],
    target_month: str,
    fieldnames: list[str],
    lofty_reserves: dict[str, float] | None = None,
    property_filters: list[str] | None = None,
    kind_filters: set[str] | None = None,
    approved_exceptions: list[dict[str, Any]] | None = None,
) -> list[dict[str, str]]:
    """Generate GL rows for missing accrual entries."""
    existing = find_existing_coverage(rows, target_month)
    blocked_pm_keys = {
        f"{item['property']}|pm"
        for item in find_blocked_first_day_pm_fees(rows, target_month, property_filters, kind_filters)
    }
    pm_fees = compute_pm_fees(rows, target_month)
    existing_pm_sides = pm_accrual_sides(rows, target_month)
    existing_dao_fee_sides = dao_fee_accrual_sides(rows, target_month)
    end_date = accrual_posting_date(target_month)
    m_label = month_label(target_month)

    new_rows: list[dict[str, str]] = []

    lofty_reserves = lofty_reserves or {}

    # Complete every existing DAO registration/admin fee expense with ECO's
    # matching revenue side, including historical or active properties that do
    # not yet have a fixed-accrual template. This pass is deliberately driven
    # only by an existing negative DAO-side marker; it never invents a new fee
    # obligation. ``--kind dao_eco`` can therefore be used for a safe,
    # counterpart-only historical backfill.
    dao_expense_markers: dict[str, tuple[dict[str, str], float]] = {}
    for source_row in rows:
        marker = parse_marker(str(source_row.get("Notes") or ""))
        if not marker or marker["month"] != target_month or marker["kind"] != "dao":
            continue
        try:
            source_amount = float(str(source_row.get("Amount") or "0").replace(",", ""))
        except ValueError:
            continue
        if source_amount >= 0:
            continue
        property_name = canonical_accrual_property_name(marker["property"])
        # Older annual-reference DAO rows correctly posted a monthly -$62.50
        # amount while retaining $750.00 in the marker as the annual basis.
        # The ECO counterpart must mirror the actual posted row, not the
        # marker's historical reference amount.
        dao_expense_markers.setdefault(property_name, (marker, abs(source_amount)))

    for property_name, (marker, source_amount) in sorted(dao_expense_markers.items()):
        sides = existing_dao_fee_sides.get(property_name, set())
        if "eco" in sides or not template_in_scope(
            property_name,
            DAO_ECO_KIND,
            property_filters,
            kind_filters,
        ):
            continue
        amount = round(source_amount, 2)
        label = f"ECO Systems LLC DAO Registration Fee Revenue | {property_name} | {m_label}"
        row = {field: "" for field in fieldnames}
        row.update({
            "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
            "Date": end_date,
            "Merchant": label,
            "Description": label,
            "Amount": f"{amount:.2f}",
            "Type": "Revenue",
            "Category": "Fees & Other Revenue",
            "Sub-category": "",
            "Property": "Mining, Sales, Consulting, and PM",
            "Unit": "",
            "Notes": (
                f"{marker['prefix']}|{DAO_ECO_KIND}|{property_name}|{target_month}|{amount:.2f} | "
                f"{DAO_LLC_ADMIN_ECO_REVENUE_NOTE}"
            ),
        })
        new_rows.append(row)
        existing_dao_fee_sides[property_name].add("eco")

    retained_kind = str(RETAINED_CAPITAL_RULE["kind"])
    for retained_property in RETAINED_CAPITAL_PROPERTIES:
        retained_summary = retained_capital_amount(
            rows, target_month, retained_property, lofty_reserves, approved_exceptions
        )
        retained_key = f"{retained_property}|{retained_kind}"
        if (
            retained_summary["amount"] <= 0
            or retained_key in existing
            or not template_in_scope(retained_property, retained_kind, property_filters, kind_filters)
        ):
            continue
        amount = float(retained_summary["amount"])
        label = f"OR Replenishment | {retained_property} | {m_label}"
        row = {field: "" for field in fieldnames}
        row.update({
            "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
            "Date": end_date,
            "Merchant": label,
            "Description": label,
            "Amount": f"-{amount:.2f}",
            "Type": "Transfers & Other",
            "Category": "Owner Contributions/Distributions",
            "Sub-category": "",
            "Property": reporting_property_name(retained_property),
            "Unit": "",
            "Notes": retained_capital_note(retained_property, target_month, amount, retained_summary),
        })
        new_rows.append(row)

    # Fixed-amount accruals
    for template in ACCRUAL_TEMPLATES:
        property_name = template["property"]
        kind = template["kind"]

        if schedule_address_is_excluded(str(property_name)):
            continue
        if is_cash_basis_insurance_accrual(str(property_name), str(kind)):
            continue
        if is_no_dao_mortgage_escrow_accrual(str(property_name), str(kind)):
            continue

        if not template_in_scope(property_name, kind, property_filters, kind_filters):
            continue

        if is_before_launch(property_name, target_month):
            continue

        kind_start = PROPERTY_KIND_ACCRUAL_START_MONTHS.get((property_name, kind))
        if kind_start and target_month < kind_start:
            continue

        reference_amount = fixed_accrual_reference_amount(property_name, kind, target_month, float(template["amount"]))
        amount = fixed_accrual_posting_amount(property_name, kind, target_month, float(template["amount"]))
        prefix = template["source_prefix"]
        description = template["description"].format(property=property_name, month_label=m_label)
        merchant = template["merchant"].format(property=property_name, month_label=m_label)
        if (property_name, kind) in ANNUAL_REFERENCE_FIXED_ACCRUALS:
            accrual_note = (
                f"{prefix}|{kind}|{property_name}|{target_month}|{amount:.2f} | "
                f"Monthly accrual from annual reference ${reference_amount:.2f}. "
                "Accounting/manual accrual only, no bank transfer."
            )
        else:
            accrual_note = f"{prefix}|{kind}|{property_name}|{target_month}|{amount:.2f} | Monthly accrual entry. Accounting/manual accrual only, no bank transfer."

        key = f"{property_name}|{kind}"
        sides = existing_dao_fee_sides.get(property_name, set()) if kind == "dao" else set()
        if kind != "dao" and key in existing:
            continue

        if kind != "dao" or "dao" not in sides:
            row = {field: "" for field in fieldnames}
            row.update({
                "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
                "Date": end_date,
                "Merchant": merchant,
                "Description": description,
                "Amount": f"-{amount:.2f}",
                "Type": "Operating Expenses",
                "Category": template["category"],
                "Sub-category": template.get("sub_category", ""),
                "Property": reporting_property_name(property_name),
                "Unit": "",
                "Notes": accrual_note,
            })
            new_rows.append(row)

        if kind == "dao" and "eco" not in sides:
            label = f"ECO Systems LLC DAO Registration Fee Revenue | {property_name} | {m_label}"
            row = {field: "" for field in fieldnames}
            row.update({
                "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
                "Date": end_date,
                "Merchant": label,
                "Description": label,
                "Amount": f"{amount:.2f}",
                "Type": "Revenue",
                "Category": "Fees & Other Revenue",
                "Sub-category": "",
                "Property": "Mining, Sales, Consulting, and PM",
                "Unit": "",
                "Notes": (
                    f"{prefix}|{DAO_ECO_KIND}|{property_name}|{target_month}|{amount:.2f} | "
                    f"{DAO_LLC_ADMIN_ECO_REVENUE_NOTE}"
                ),
            })
            new_rows.append(row)

    # PM fee accruals (computed from gross rent)
    for property_name, rate, prefix in PM_FEE_PROPERTIES:
        if not pm_manual_accrual_required(property_name, target_month):
            continue
        if not template_in_scope(property_name, "pm", property_filters, kind_filters):
            continue

        if is_before_launch(property_name, target_month):
            continue

        key = f"{property_name}|pm"
        sides = existing_pm_sides.get(property_name, set())
        # Do not silently backfill an ECO revenue row for a historical,
        # one-sided PM accrual. Historical cash needs reviewed reclassification
        # first. Explicit pm_dao rows, however, may be safely completed after a
        # partial prior run.
        if "legacy" in sides:
            continue
        if key in blocked_pm_keys:
            continue

        components = PM_FEE_COMPONENTS.get(property_name) or [{"rate": rate, "deduction": 0.0}]
        component_summary = pm_fee_rule_summary(property_name, target_month, rate, components)
        pm_amount = pm_fees.get(property_name, 0)
        accrual_id = f"{prefix}|pm|{property_name}|{target_month}|{pm_amount:.2f}"
        if pm_amount <= 0:
            if "dao" not in sides:
                label = f"PM Fee Accrual | {property_name} | {m_label}"
                row = {field: "" for field in fieldnames}
                row.update({
                    "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
                    "Date": end_date,
                    "Merchant": label,
                    "Description": label,
                    "Amount": "0.00",
                    "Type": "Operating Expenses",
                    "Category": "Property Management",
                    "Sub-category": "",
                    "Property": reporting_property_name(property_name),
                    "Unit": "",
                    "Notes": f"{prefix}|{PM_DAO_KIND}|{property_name}|{target_month}|0.00 | Accrual ID {accrual_id}. Zero-dollar coverage marker only; skipped by live Baselane apply.",
                })
                new_rows.append(row)
            continue

        if "dao" not in sides:
            label = f"PM Fee Accrual | {property_name} | {m_label}"
            row = {field: "" for field in fieldnames}
            row.update({
                "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
                "Date": end_date,
                "Merchant": label,
                "Description": label,
                "Amount": f"-{pm_amount:.2f}",
                "Type": "Operating Expenses",
                "Category": "Property Management",
                "Sub-category": "",
                "Property": reporting_property_name(property_name),
                "Unit": "",
                "Notes": f"{prefix}|{PM_DAO_KIND}|{property_name}|{target_month}|{pm_amount:.2f} | Accrual ID {accrual_id}. DAO-side PM expense on Baselane gross rent for {m_label} ({component_summary}); matched to ECO revenue. Accounting/manual accrual only, no bank transfer.",
            })
            new_rows.append(row)

        if "eco" not in sides:
            label = f"ECO Systems LLC PM Fee Revenue | {property_name} | {m_label}"
            row = {field: "" for field in fieldnames}
            row.update({
                "Account": "ECO Systems, LLC-ECO Systems Accrual Overlay",
                "Date": end_date,
                "Merchant": label,
                "Description": label,
                "Amount": f"{pm_amount:.2f}",
                "Type": "Revenue",
                "Category": "Fees & Other Revenue",
                "Sub-category": "Property Management",
                "Property": reporting_property_name(property_name),
                "Unit": "",
                "Notes": f"{prefix}|{PM_ECO_KIND}|{property_name}|{target_month}|{pm_amount:.2f} | Accrual ID {accrual_id}. ECO Systems LLC PM fee revenue for {m_label} ({component_summary}); matched to the DAO-side expense. Accounting/manual accrual only, no bank transfer.",
            })
            new_rows.append(row)

    return new_rows


def ledger_sha256(gl_path: Path) -> str:
    """Return a content fingerprint for optimistic GL write protection."""
    return hashlib.sha256(gl_path.read_bytes()).hexdigest()


def require_unchanged_ledger(gl_path: Path, expected_sha256: str) -> None:
    actual_sha256 = ledger_sha256(gl_path)
    if actual_sha256 != expected_sha256:
        raise RuntimeError(
            "GL changed during accrual run; refusing to write "
            f"(expected {expected_sha256}, found {actual_sha256})"
        )


def append_rows_to_gl(
    gl_path: Path,
    new_rows: list[dict[str, str]],
    fieldnames: list[str],
    expected_sha256: str | None = None,
) -> None:
    """Append new rows to the GL CSV."""
    if expected_sha256:
        require_unchanged_ledger(gl_path, expected_sha256)
    backup_gl(gl_path)
    with gl_path.open("a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        for row in new_rows:
            writer.writerow(row)


def backup_gl(gl_path: Path) -> Path:
    backup_path = gl_path.with_name(f"{gl_path.name}.bak-{dt.datetime.now(dt.UTC).strftime('%Y%m%dT%H%M%SZ')}")
    backup_path.write_bytes(gl_path.read_bytes())
    return backup_path


def write_gl(
    gl_path: Path,
    rows: list[dict[str, str]],
    fieldnames: list[str],
    expected_sha256: str | None = None,
) -> None:
    if expected_sha256:
        require_unchanged_ledger(gl_path, expected_sha256)
    backup_gl(gl_path)
    tmp_path = gl_path.with_suffix(gl_path.suffix + ".tmp")
    with tmp_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    tmp_path.replace(gl_path)


def self_test() -> int:
    """Run self-tests."""
    # Test marker parsing
    marker = parse_marker("AOPS-PNL-ACCRUAL|legal|88 Madison Ave|2026-05|62.50 | Backfill")
    assert marker is not None, "marker parsing failed"
    assert marker["kind"] == "dao"
    assert marker["month"] == "2026-05"
    assert marker["property"] == "88 Madison Ave"

    # Test month_label
    assert month_label("2026-06") == "June 2026"
    assert month_label("2025-12") == "December 2025"

    # Test month_end_date
    assert month_end_date("2026-06") == "June 30, 2026"
    assert month_end_date("2026-02") == "February 28, 2026"
    assert month_end_date("2024-02") == "February 29, 2024"  # leap year
    assert accrual_posting_date("2026-06") == "June 28, 2026"
    assert iter_months("2026-05", "2026-07") == ["2026-05", "2026-06", "2026-07"]
    assert effective_accrual_amount("326-332 S Alcott St, Denver, CO 80219", "taxes", "2025-12", 374.19) == 374.19
    assert effective_accrual_amount("326-332 S Alcott St, Denver, CO 80219", "taxes", "2026-01", 374.19) == 333.49
    assert fixed_accrual_posting_amount("5541 S Peoria St, Chicago, IL 60621", "taxes", "2026-06", 2628.00) == 219.00
    assert fixed_accrual_posting_amount("9 Country Club Ln N", "taxes", "2026-06", 3225.01) == 3225.01

    # Test launch date filtering
    assert is_before_launch("9 Country Club Ln N", "2025-06") is True
    assert is_before_launch("9 Country Club Ln N", "2025-07") is False
    assert PROPERTY_KIND_ACCRUAL_START_MONTHS[("9 Country Club Ln N", "mortgage_interest")] == "2025-09"
    assert is_before_launch("88 Madison Ave", "2025-01") is False

    # Test idempotency: existing markers prevent duplicates
    rows = [{
        "Notes": "AOPS-PNL-ACCRUAL|legal|88 Madison Ave|2026-06|62.50 | test",
        "Date": "June 28, 2026",
        "Property": "88 Madison Ave",
        "Amount": "-62.50",
        "Category": "Legal Fees",
        "Type": "Operating Expenses",
    }]
    existing = find_existing_accruals(rows, "2026-06")
    assert "88 Madison Ave|dao" in existing
    assert "88 Madison Ave|insurance" not in existing

    # Legacy DAO markers sometimes carry the $750 annual reference even
    # though the posted monthly expense is -$62.50. The ECO revenue side must
    # mirror the posted monthly amount.
    annual_reference_dao_rows = [{
        "Notes": (
            "AOPS-OHIL-ACCRUAL|dao|25 Circle Dr, Dixmoor, IL 60426|"
            "2026-01|750.00 | Annual DAO LLC/admin fee amortized monthly"
        ),
        "Date": "January 28, 2026",
        "Property": "25 Circle Dr",
        "Amount": "-62.50",
        "Category": "Legal Fees",
        "Type": "Operating Expenses",
    }]
    dao_counterparts = generate_missing_accruals(
        annual_reference_dao_rows,
        "2026-01",
        ["Account", "Date", "Merchant", "Description", "Amount", "Type", "Category", "Sub-category", "Property", "Unit", "Notes"],
        property_filters=["25 Circle Dr"],
        kind_filters={DAO_ECO_KIND},
    )
    assert len(dao_counterparts) == 1
    assert dao_counterparts[0]["Amount"] == "62.50"
    assert "|dao_eco|25 Circle Dr, Dixmoor, IL 60426|2026-01|62.50" in dao_counterparts[0]["Notes"]

    pm_coverage_rows = [
        {
            "Notes": "AOPS-PM-FEE|90 Madison Ave|2026-06|420.00 | month-end PM fee",
            "Date": "June 28, 2026",
            "Category": "Property Management",
        },
        {
            "Notes": "AOPS-PM-FEE|88 Madison Ave|2026-06|300.00 | stale first-day PM fee",
            "Date": "June 01, 2026",
            "Category": "Property Management",
        },
    ]
    pm_coverage = find_existing_coverage(pm_coverage_rows, "2026-06")
    assert "90 Madison Ave|pm" in pm_coverage
    assert "88 Madison Ave|pm" not in pm_coverage

    # Test actual Scruggs escrow coverage prevents synthetic tax/insurance duplicates
    scruggs_rows = [
        {
            "Merchant": "Scruggs Investments | TRANSFER_OUT",
            "Description": "Scruggs Investments | TRANSFER_OUT",
            "Category": "Rental Dwelling",
            "Property": "326 South Alcott Street",
            "Notes": "Insurance Escrow July 2025",
            "Date": "August 10, 2025",
        },
        {
            "Merchant": "Scruggs Investments | TRANSFER_OUT",
            "Description": "Scruggs Investments | TRANSFER_OUT",
            "Category": "Property Taxes",
            "Property": "326 South Alcott Street",
            "Notes": "Taxes Escrow July 2025",
            "Date": "August 10, 2025",
        },
    ]
    coverage = find_existing_coverage(scruggs_rows, "2025-07")
    assert "326-332 S Alcott St, Denver, CO 80219|insurance" in coverage
    assert "326-332 S Alcott St, Denver, CO 80219|taxes" in coverage

    # Test PM fee computation
    rent_rows = [
        {"Date": "June 15, 2026", "Property": "88 Madison Ave", "Amount": "1000.00", "Category": "Short Term Rents", "Type": "Revenue"},
        {"Date": "June 20, 2026", "Property": "88 Madison Ave", "Amount": "500.00", "Category": "Short Term Rents", "Type": "Revenue"},
    ]
    pm_fees = compute_pm_fees(rent_rows, "2026-06")
    assert pm_fees.get("88 Madison Ave") == 375.0  # 25% of 1500
    assert separately_booked_jazmine_pm_fee([], "88 Madison Ave", "2025-07") == 231.42
    assert separately_booked_jazmine_pm_fee([], "88 Madison Ave", "2026-05") == 40.0
    assert separately_booked_jazmine_pm_fee([], "90 Madison Ave", "2025-07") == 244.08

    # Master GL exports place the detailed category in Sub-category.
    pm_payment_rows = [{
        "Date": "June 30, 2026",
        "Property": "9 Country Club Ln N",
        "Amount": "-125.00",
        "Category": "Operating Expenses",
        "Sub-category": "Property Management",
        "Merchant": "ECO Systems LLC",
        "Notes": "June 2026 PM fee",
    }]
    assert separately_booked_eco_pm_fee_payment(
        pm_payment_rows, "9 Country Club Ln N", "2026-06"
    ) == 125.0

    print("self-test ok")
    return 0


def main(argv: list[str] | None = None) -> int:
    global PM_FEE_PROPERTIES

    parser = argparse.ArgumentParser(description="Monthly idempotent accrual generator for Baselane GL.")
    parser.add_argument("--gl-csv", type=Path, required=False, help="Path to ECO Systems General Ledger CSV")
    parser.add_argument("--month", type=str, default=None, help="Target month YYYY-MM (default: previous month)")
    parser.add_argument("--start-month", type=str, default=None, help="Inclusive start month YYYY-MM for range backfill")
    parser.add_argument("--end-month", type=str, default=None, help="Inclusive end month YYYY-MM for range backfill")
    parser.add_argument(
        "--reporting-cutoff-date",
        default=os.environ.get("BASELANE_REPORTING_CUTOFF_DATE"),
        help="Exclude ordinary transactions after YYYY-MM-DD while retaining same-month AOPS synthetic rows",
    )
    parser.add_argument("--property", dest="property_filters", action="append", default=None, help="Limit to property template/address substring; may be repeated")
    parser.add_argument("--kind", dest="kind_filters", action="append", default=None, help="Limit to accrual kind, e.g. taxes, insurance, dao, pm; may be repeated")
    parser.add_argument("--active-property-map", type=Path, default=None, help="JSON property map used to require accrual template coverage for active DAOs")
    parser.add_argument("--pm-rate-schedule", type=Path, default=default_pm_rate_schedule(), help="Bookkeeping workbook containing Address and PM Fee (%% of Gross Rents) columns")
    parser.add_argument("--hemlane-live-transactions", type=Path, default=None, help="Live Hemlane transaction report used only as PM-fee accrual basis when Baselane rent cash is not present")
    parser.add_argument(
        "--lofty-reserve-snapshot",
        type=Path,
        default=default_lofty_reserve_snapshot(),
        help="Fresh get-manager-properties response containing curr_maintenance_reserve for co-ownership OR policy",
    )
    parser.add_argument(
        "--retained-capital-approved-exceptions",
        type=Path,
        default=default_retained_capital_approved_exceptions(),
        help="Exact property/month approved retained-capital reporting exceptions; malformed input blocks the run",
    )
    parser.add_argument("--update-amount-mismatches", action="store_true", help="With --apply, update existing accrual rows whose marker amount differs from the effective template amount")
    parser.add_argument("--apply", action="store_true", help="Append missing accruals to the GL CSV")
    parser.add_argument("--report", type=Path, default=None, help="Write JSON report to this path")
    parser.add_argument("--review-markdown", type=Path, default=None, help="Write human-review markdown for accrual gap actions")
    parser.add_argument("--gap-approvals", type=Path, default=None, help="Reviewed JSON approvals for non-mutating accrual gap decisions")
    parser.add_argument("--gap-approval-scaffold", type=Path, default=None, help="Write/update JSON scaffold for reviewed accrual gap approvals")
    parser.add_argument("--gap-approval-review-csv", type=Path, default=None, help="Write compact CSV for reviewing accrual gap approvals")
    parser.add_argument("--gap-approval-import-commands", type=Path, default=None, help="Write explicit approval command for importing reviewed gap approval CSV")
    parser.add_argument("--import-gap-approval-csv", action="store_true", help="Import reviewed gap approval CSV into --gap-approvals before validation")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args(argv)

    if args.self_test:
        return self_test()

    try:
        retained_capital_approved_exceptions = load_retained_capital_approved_exceptions(
            args.retained_capital_approved_exceptions
        )
    except ValueError as exc:
        parser.error(str(exc))

    if not args.gl_csv:
        parser.error("--gl-csv is required (unless --self-test)")

    if args.month and (args.start_month or args.end_month):
        parser.error("--month cannot be combined with --start-month/--end-month")
    if bool(args.start_month) != bool(args.end_month):
        parser.error("--start-month and --end-month must be provided together")

    gl_source_sha256 = ledger_sha256(args.gl_csv)
    fieldnames, rows = read_gl(args.gl_csv)
    if not args.lofty_reserve_snapshot.is_file():
        write_preflight_failure_report(
            args.report,
            args.review_markdown,
            target_month=args.month or args.start_month or default_target_month(),
            code="lofty_reserve_snapshot_missing",
            message=f"--lofty-reserve-snapshot not found: {args.lofty_reserve_snapshot}",
            lofty_reserve_snapshot=str(args.lofty_reserve_snapshot),
        )
        parser.error(f"--lofty-reserve-snapshot not found: {args.lofty_reserve_snapshot}")
    lofty_reserve_payload = json.loads(args.lofty_reserve_snapshot.read_text(encoding="utf-8"))
    lofty_reserves = load_lofty_reserves(args.lofty_reserve_snapshot)
    live_reserve_required_properties = live_lofty_reserve_required_properties()
    missing_lofty_reserves = sorted(set(live_reserve_required_properties) - set(lofty_reserves))
    local_financials_only_without_lofty_reserve = sorted(
        set(LOCAL_FINANCIALS_ONLY_PROPERTIES) - set(lofty_reserves)
    )
    if missing_lofty_reserves:
        live_manager_properties = properties_from_lofty_response(lofty_reserve_payload)
        missing_diagnostics = [
            {
                "property": property_name,
                "status": "absent_from_live_manager_roster",
                "required_action": (
                    "Restore the active property to Lofty manager access or capture a fresh authoritative "
                    "get-manager-properties reserve snapshot before monthly accruals can continue."
                ),
            }
            for property_name in missing_lofty_reserves
        ]
        write_preflight_failure_report(
            args.report,
            args.review_markdown,
            target_month=args.month or args.start_month or default_target_month(),
            code="lofty_reserve_properties_missing",
            message=(
                "--lofty-reserve-snapshot is missing co-ownership properties: "
                + ", ".join(missing_lofty_reserves)
            ),
            lofty_reserve_snapshot=str(args.lofty_reserve_snapshot),
            missing_lofty_reserves=missing_lofty_reserves,
            lofty_reserve_snapshot_property_count=len(live_manager_properties),
            available_lofty_reserve_properties=sorted(lofty_reserves),
            missing_lofty_reserve_diagnostics=missing_diagnostics,
        )
        parser.error(
            "--lofty-reserve-snapshot is missing co-ownership properties: "
            + ", ".join(missing_lofty_reserves)
        )
    months = iter_months(args.start_month, args.end_month) if args.start_month else [args.month or default_target_month()]
    target_month = months[0]
    try:
        reporting_cutoff = parse_reporting_cutoff(args.reporting_cutoff_date)
    except ValueError as exc:
        parser.error(str(exc))
    working_rows, post_cutoff_excluded_rows = rows_through_reporting_cutoff(
        rows, reporting_cutoff, months
    )

    pm_rate_schedule = {
        "path": str(args.pm_rate_schedule) if args.pm_rate_schedule else None,
        "status": "review",
        "matched_property_count": 0,
        "matched_rates": {},
    }
    if args.pm_rate_schedule:
        if not args.pm_rate_schedule.is_file():
            parser.error(f"--pm-rate-schedule not found: {args.pm_rate_schedule}")
        PM_FEE_PROPERTIES, pm_rate_schedule = pm_fee_properties_from_schedule(
            args.pm_rate_schedule, gl_rows=working_rows
        )
    global HEMLANE_PM_FEE_BASIS
    hemlane_pm_fee_basis_report = {
        "path": str(args.hemlane_live_transactions) if args.hemlane_live_transactions else None,
        "status": "not_configured",
    }
    if args.hemlane_live_transactions:
        HEMLANE_PM_FEE_BASIS, hemlane_pm_fee_basis_report = load_hemlane_pm_fee_basis(args.hemlane_live_transactions, target_month)
    else:
        HEMLANE_PM_FEE_BASIS = {}
    kind_filters = {kind.strip() for kind in args.kind_filters or [] if kind.strip()} or None
    active_properties = load_active_property_map(args.active_property_map)
    active_without_templates = find_active_properties_without_accrual_templates(active_properties)
    active_without_fixed_requirement = active_properties_without_fixed_accrual_requirement(active_properties)
    month_reports = []
    all_new_rows: list[dict[str, str]] = []
    all_mismatches: list[dict[str, Any]] = []
    all_updated_amount_mismatches: list[dict[str, Any]] = []
    all_blocked_first_day_pm_fees: list[dict[str, Any]] = []

    for month in months:
        existing = find_existing_coverage(working_rows, month)
        pm_fees = compute_pm_fees(working_rows, month)
        jazmine_pm_fee_deductions = madison_jazmine_pm_fee_deductions(working_rows, month)
        existing.update(f"{property_name}|pm" for property_name, amount in pm_fees.items() if amount <= 0)
        pm_basis_gaps = pm_fee_basis_gaps(
            working_rows,
            month,
            existing,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
        )
        blocking_pm_keys = {str(item["key"]) for item in pm_basis_gaps}
        for property_name, _rate, _prefix in PM_FEE_PROPERTIES:
            if not pm_manual_accrual_required(property_name, month):
                continue
            key = f"{property_name}|pm"
            if key not in blocking_pm_keys and property_name not in pm_fees:
                existing.add(key)
        expected_fixed_keys = expected_fixed_accrual_keys(kind_filters, args.property_filters)
        expected_fixed_keys |= retained_capital_expected_keys(
            working_rows,
            month,
            lofty_reserves,
            kind_filters,
            args.property_filters,
            retained_capital_approved_exceptions,
        )
        covered_fixed_keys = existing & expected_fixed_keys
        missing_fixed_keys = expected_fixed_keys - existing
        blocked_first_day_pm_fees = find_blocked_first_day_pm_fees(
            working_rows,
            month,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
        )
        all_blocked_first_day_pm_fees.extend(blocked_first_day_pm_fees)
        duplicate_accrual_markers = find_duplicate_accrual_markers(
            working_rows,
            month,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
        )
        amount_mismatches = find_amount_mismatches(
            working_rows,
            month,
            lofty_reserves=lofty_reserves,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
            approved_exceptions=retained_capital_approved_exceptions,
        )
        if args.apply and args.update_amount_mismatches:
            all_updated_amount_mismatches.extend(
                apply_amount_mismatch_updates(
                    working_rows, amount_mismatches, lofty_reserves, retained_capital_approved_exceptions
                )
            )
            amount_mismatches = find_amount_mismatches(
                working_rows,
                month,
                lofty_reserves=lofty_reserves,
                property_filters=args.property_filters,
                kind_filters=kind_filters,
                approved_exceptions=retained_capital_approved_exceptions,
            )
        new_rows = generate_missing_accruals(
            working_rows,
            month,
            fieldnames,
            lofty_reserves=lofty_reserves,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
            approved_exceptions=retained_capital_approved_exceptions,
        )
        all_new_rows.extend(new_rows)
        working_rows.extend(new_rows)
        reportable_new_rows = [row for row in new_rows if not is_zero_amount_text(row.get("Amount"))]
        amount_mismatches = find_amount_mismatches(
            working_rows,
            month,
            lofty_reserves=lofty_reserves,
            property_filters=args.property_filters,
            kind_filters=kind_filters,
            approved_exceptions=retained_capital_approved_exceptions,
        )
        if args.apply and args.update_amount_mismatches:
            all_updated_amount_mismatches.extend(
                apply_amount_mismatch_updates(
                    working_rows, amount_mismatches, lofty_reserves, retained_capital_approved_exceptions
                )
            )
            amount_mismatches = find_amount_mismatches(
                working_rows,
                month,
                lofty_reserves=lofty_reserves,
                property_filters=args.property_filters,
                kind_filters=kind_filters,
                approved_exceptions=retained_capital_approved_exceptions,
            )
        all_mismatches.extend(amount_mismatches)
        post_existing = find_existing_coverage(working_rows, month)
        covered_fixed_keys = post_existing & expected_fixed_keys
        missing_fixed_keys = expected_fixed_keys - post_existing
        month_reports.append({
            "target_month": month,
            "month_label": month_label(month),
            "existing_coverage_count": len(post_existing),
            "expected_fixed_accrual_coverage_count": len(expected_fixed_keys),
            "covered_fixed_accrual_coverage_count": len(covered_fixed_keys),
            "missing_fixed_accrual_coverage_count": len(missing_fixed_keys),
            "expected_fixed_accrual_coverage_by_kind": coverage_by_kind(expected_fixed_keys),
            "covered_fixed_accrual_coverage_by_kind": coverage_by_kind(covered_fixed_keys),
            "missing_fixed_accrual_coverage_by_kind": coverage_by_kind(missing_fixed_keys),
            "missing_fixed_accrual_coverage": coverage_key_details(missing_fixed_keys, month),
            "pm_fee_basis_gap_count": len(pm_basis_gaps),
            "pm_fee_basis_gaps": pm_basis_gaps,
            "pm_fees_computed": pm_fees,
            "or_replenishment": [
                retained_capital_amount(
                    working_rows, month, property_name, lofty_reserves, retained_capital_approved_exceptions
                )
                for property_name in RETAINED_CAPITAL_PROPERTIES
                if template_in_scope(property_name, RETAINED_CAPITAL_RULE["kind"], args.property_filters, kind_filters)
            ],
            "madison_jazmine_pm_fee_deductions": jazmine_pm_fee_deductions,
            "madison_jazmine_pm_fee_deduction_total": round(sum(jazmine_pm_fee_deductions.values()), 2),
            "blocked_first_day_pm_fee_count": len(blocked_first_day_pm_fees),
            "blocked_first_day_pm_fees": [
                {
                    key: value
                    for key, value in item.items()
                    if key != "row_index"
                }
                for item in blocked_first_day_pm_fees
            ],
            "duplicate_accrual_marker_count": len(duplicate_accrual_markers),
            "duplicate_accrual_markers": duplicate_accrual_markers,
            "missing_accruals": [
                {
                    "property": row["Property"],
                    "date": row["Date"],
                    "amount": row["Amount"],
                    "category": row["Category"],
                    "description": row["Description"],
                    "notes": row["Notes"],
                }
                for row in reportable_new_rows
            ],
            "missing_count": len(reportable_new_rows),
            "amount_mismatch_count": len(amount_mismatches),
            "amount_mismatches": [
                {
                    key: value
                    for key, value in mismatch.items()
                    if key != "row_index"
                }
                for mismatch in amount_mismatches
            ],
        })

    expected_fixed_coverage_count = sum(item["expected_fixed_accrual_coverage_count"] for item in month_reports)
    covered_fixed_coverage_count = sum(item["covered_fixed_accrual_coverage_count"] for item in month_reports)
    missing_fixed_coverage_count = sum(item["missing_fixed_accrual_coverage_count"] for item in month_reports)
    missing_fixed_coverage = [
        detail
        for item in month_reports
        for detail in item.get("missing_fixed_accrual_coverage", [])
    ]
    all_pm_fee_basis_gaps = [
        detail
        for item in month_reports
        for detail in item.get("pm_fee_basis_gaps", [])
    ]
    all_duplicate_accrual_markers = [
        detail
        for item in month_reports
        for detail in item.get("duplicate_accrual_markers", [])
    ]
    gap_action_queue = accrual_gap_action_queue(
        missing_fixed_coverage,
        all_pm_fee_basis_gaps,
        active_without_fixed_requirement,
    )
    if args.import_gap_approval_csv:
        gap_approval_csv_import = import_gap_approval_review_csv(args.gap_approval_review_csv, args.gap_approvals)
    else:
        gap_approval_csv_import = None
    gap_approvals = load_gap_approvals(args.gap_approvals)
    gap_approval_validation = apply_gap_approvals(gap_action_queue, gap_approvals)
    gap_approval_scaffold = write_gap_approval_scaffold(args.gap_approval_scaffold, gap_action_queue)
    gap_approval_scaffold_payload = load_gap_approvals(args.gap_approval_scaffold)
    gap_approval_review_csv = write_gap_approval_review_csv(args.gap_approval_review_csv, gap_approval_scaffold_payload)
    gap_approval_import_commands = write_gap_approval_import_commands(args.gap_approval_import_commands)
    effective_blocking_gap_actions = blocking_gap_actions(gap_action_queue)
    blocking_gap_action_count = len(effective_blocking_gap_actions)
    unapproved_pm_fee_basis_gaps = [
        item
        for item in gap_action_queue
        if item.get("action") == "verify_missing_rent_or_approve_zero_pm"
        and item.get("severity") == "blocks_monthly_close"
    ]

    report = {
        "generated_at": iso_z(),
        "target_month": target_month,
        "months": months,
        "month_count": len(months),
        "month_label": month_label(target_month) if len(months) == 1 else None,
        "gl_csv": str(args.gl_csv),
        "gl_source_sha256": gl_source_sha256,
        "reporting_cutoff_date": reporting_cutoff.isoformat() if reporting_cutoff else None,
        "post_cutoff_excluded_row_count": len(post_cutoff_excluded_rows),
        "post_cutoff_excluded_rows": [
            {
                "date": row.get("Date", ""),
                "property": row.get("Property", ""),
                "amount": row.get("Amount", ""),
                "merchant": row.get("Merchant", ""),
                "description": row.get("Description", ""),
            }
            for row in post_cutoff_excluded_rows
        ],
        "retained_capital_approved_exceptions": {
            "path": str(args.retained_capital_approved_exceptions),
            "sha256": ledger_sha256(args.retained_capital_approved_exceptions),
            "count": len(retained_capital_approved_exceptions),
            "keys": [
                f"{item['property']}|{item['month']}|{item['approval_id']}"
                for item in retained_capital_approved_exceptions
            ],
        },
        "lofty_reserve_snapshot": str(args.lofty_reserve_snapshot),
        "lofty_reserve_property_count": len(lofty_reserves),
        "live_lofty_reserve_required_properties": list(live_reserve_required_properties),
        "local_financials_only_properties": list(LOCAL_FINANCIALS_ONLY_PROPERTIES),
        "local_financials_only_without_lofty_reserve": local_financials_only_without_lofty_reserve,
        "pm_rate_schedule": pm_rate_schedule,
        "hemlane_pm_fee_basis": hemlane_pm_fee_basis_report,
        "mode": "apply" if args.apply else "dry_run",
        "property_filters": args.property_filters or [],
        "kind_filters": sorted(kind_filters or []),
        "active_property_map": str(args.active_property_map) if args.active_property_map else None,
        "active_property_count": len(active_properties),
        "active_without_accrual_template_count": len(active_without_templates),
        "active_without_accrual_templates": active_without_templates,
        "active_without_fixed_accrual_requirement_count": len(active_without_fixed_requirement),
        "active_without_fixed_accrual_requirement": active_without_fixed_requirement,
        "existing_accrual_count": month_reports[0]["existing_coverage_count"] if len(month_reports) == 1 else None,
        "existing_coverage_count": sum(item["existing_coverage_count"] for item in month_reports),
        "expected_fixed_accrual_coverage_count": expected_fixed_coverage_count,
        "covered_fixed_accrual_coverage_count": covered_fixed_coverage_count,
        "missing_fixed_accrual_coverage_count": missing_fixed_coverage_count,
        "expected_fixed_accrual_coverage_by_kind": dict(sum((Counter(item["expected_fixed_accrual_coverage_by_kind"]) for item in month_reports), Counter())),
        "covered_fixed_accrual_coverage_by_kind": dict(sum((Counter(item["covered_fixed_accrual_coverage_by_kind"]) for item in month_reports), Counter())),
        "missing_fixed_accrual_coverage_by_kind": coverage_by_kind(
            {
                str(detail.get("key") or "")
                for detail in missing_fixed_coverage
                if str(detail.get("key") or "")
            }
        ),
        "missing_fixed_accrual_coverage": missing_fixed_coverage,
        "pm_fee_basis_gap_count": len(all_pm_fee_basis_gaps),
        "unapproved_pm_fee_basis_gap_count": len(unapproved_pm_fee_basis_gaps),
        "pm_fee_basis_gaps": all_pm_fee_basis_gaps,
        "gap_action_queue_count": len(gap_action_queue),
        "blocking_gap_action_count": blocking_gap_action_count,
        "gap_action_queue": gap_action_queue,
        "gap_approvals": gap_approval_validation,
        "gap_approval_csv_import": gap_approval_csv_import,
        "gap_approval_scaffold": gap_approval_scaffold,
        "gap_approval_review_csv": gap_approval_review_csv,
        "gap_approval_import_commands": gap_approval_import_commands,
        "pm_fees_computed": month_reports[0]["pm_fees_computed"] if len(month_reports) == 1 else {},
        "month_reports": month_reports,
        "missing_accruals": [
            {
                "property": row["Property"],
                "date": row["Date"],
                "amount": row["Amount"],
                "category": row["Category"],
                "description": row["Description"],
                "notes": row["Notes"],
            }
            for row in all_new_rows
            if not is_zero_amount_text(row.get("Amount"))
        ],
        "missing_count": sum(1 for row in all_new_rows if not is_zero_amount_text(row.get("Amount"))),
        "blocked_first_day_pm_fee_count": len(all_blocked_first_day_pm_fees),
        "blocked_first_day_pm_fees": [
            {
                key: value
                for key, value in item.items()
                if key != "row_index"
            }
            for item in all_blocked_first_day_pm_fees
        ],
        "duplicate_accrual_marker_count": len(all_duplicate_accrual_markers),
        "duplicate_accrual_markers": all_duplicate_accrual_markers,
        "amount_mismatch_count": len(all_mismatches),
        "amount_mismatches": [
            {
                key: value
                for key, value in mismatch.items()
                if key != "row_index"
            }
            for mismatch in all_mismatches
        ],
        "status": (
            "review"
            if active_without_templates
            or pm_rate_schedule.get("status") != "ok"
            or has_nonzero_accrual_rows(all_new_rows)
            or blocking_gap_action_count
            or gap_approval_validation.get("status") != "ok"
            or all_mismatches
            or all_blocked_first_day_pm_fees
            or all_duplicate_accrual_markers
            else "ok"
        ),
    }

    if args.apply and (all_new_rows or all_updated_amount_mismatches):
        try:
            if args.update_amount_mismatches:
                write_gl(
                    args.gl_csv,
                    rows + all_new_rows,
                    fieldnames,
                    expected_sha256=gl_source_sha256,
                )
            else:
                append_rows_to_gl(args.gl_csv, all_new_rows, fieldnames, expected_sha256=gl_source_sha256)
        except RuntimeError as exc:
            report["status"] = "blocked"
            report["preflight_failure"] = {
                "code": "gl_changed_during_run",
                "message": str(exc),
                "expected_gl_sha256": gl_source_sha256,
                "actual_gl_sha256": ledger_sha256(args.gl_csv),
            }
            report["appended"] = False
            report["appended_count"] = 0
            report["updated_amount_mismatch_count"] = 0
        else:
            report["appended"] = True
            report["appended_count"] = len(all_new_rows)
            report["updated_amount_mismatch_count"] = len(all_updated_amount_mismatches) if args.update_amount_mismatches else 0
            report["updated_amount_mismatches"] = all_updated_amount_mismatches
    elif args.apply:
        report["appended"] = True
        report["appended_count"] = 0
        report["updated_amount_mismatch_count"] = 0
    else:
        report["appended"] = False
        report["updated_amount_mismatch_count"] = 0

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    if args.review_markdown:
        args.review_markdown.parent.mkdir(parents=True, exist_ok=True)
        args.review_markdown.write_text(build_review_markdown(report), encoding="utf-8")

    print(json.dumps({
        "status": report["status"],
        "target_month": report["target_month"],
        "months": report["months"],
        "existing": report["existing_accrual_count"],
        "existing_coverage": report["existing_coverage_count"],
        "missing": report["missing_count"],
        "amount_mismatches": report["amount_mismatch_count"],
        "active_without_accrual_templates": report["active_without_accrual_template_count"],
        "appended": report.get("appended_count", 0) if args.apply else None,
        "updated": report.get("updated_amount_mismatch_count", 0) if args.apply else None,
    }, indent=2))
    if (report.get("preflight_failure") or {}).get("code") == "gl_changed_during_run":
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
