#!/usr/bin/env python3
"""Synchronize four sold-property June tax cells after verified Baselane voids."""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def workspace_root() -> Path:
    for candidate in (Path.cwd(), Path(__file__).absolute().parent.parent):
        if (candidate / "reports").is_dir() and (candidate / "scripts").is_dir():
            return candidate
    raise RuntimeError("OpenClaw workspace root not found")


ROOT = workspace_root()
SOURCE_REPORT = ROOT / "reports" / "baselane_sold_property_june_tax_accrual_removal.json"
DEFAULT_REPORT = ROOT / "reports" / "baselane_sold_property_june_tax_cf_sync.json"
APPLY_ENV = "SOLD_PROPERTY_JUNE_TAX_CF_SYNC_APPLY"
TARGETS = (
    ("1315 E 114th St, Cleveland, OH 44106", 313980093, -2531.09,
     Path("/mnt/c/Users/digit/Dropbox/Real Estate/OH/1315 E 114th St, Cleveland, OH 44106/Public/07 - P&L & Owner Statements/Cash Flow Statement - 1315 E 114th St.xlsx")),
    ("13806 Coit Rd, Cleveland, OH 44110", 313980094, -2359.60,
     Path("/mnt/c/Users/digit/Dropbox/Real Estate/OH/13806 Coit Rd, Cleveland, OH 44110/Public/07 - P&L & Owner Statements/Cash Flow Statement - 13806 Coit Rd, Cleveland, OH 44110.xlsx")),
    ("3024 W 103rd St, Cleveland, OH 44111", 313980142, -2736.84,
     Path("/mnt/c/Users/digit/Dropbox/Real Estate/OH/3024 W. 103rd Street, Cleveland, OH 44111/Public/07 - P&L & Owner Statements/Cash Flow Statement - 3024 W. 103rd Street, Cleveland, OH 44111.xlsx")),
    ("9919 S Oglesby Ave, Chicago, IL 60617", 313980244, -1253.80,
     Path("/mnt/c/Users/digit/Dropbox/Real Estate/IL/9919 S Oglesby Ave, Chicago, IL 60617/Public/07 - P&L & Owner Statements/Cash Flow Statement - 9919 S Oglesby Ave Chicago IL, 60617.xlsx")),
)


def june_column(sheet: object) -> int:
    for column in range(2, 14):
        value = sheet.cell(row=1, column=column).value
        if getattr(value, "year", None) == 2026 and getattr(value, "month", None) == 6:
            return column
    raise RuntimeError("June 2026 column not found")


def approved_void_ids(source_report: Path) -> set[int]:
    payload = json.loads(source_report.read_text(encoding="utf-8"))
    if payload.get("status") != "applied":
        raise RuntimeError("Baselane sold-property tax-accrual removal is not applied")
    return {int(row["id"]) for row in payload.get("applied") or [] if row.get("isDeleted") is True}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--source-report", type=Path, default=SOURCE_REPORT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    required_ids = {target[1] for target in TARGETS}
    if approved_void_ids(args.source_report) != required_ids:
        raise RuntimeError("Baselane removal report does not prove the exact four expected voids")
    apply_allowed = args.apply and os.environ.get(APPLY_ENV) == "1"
    results = []
    for property_name, void_id, expected_old, workbook_path in TARGETS:
        workbook = openpyxl.load_workbook(workbook_path)
        try:
            sheet = workbook["2026"]
            column = june_column(sheet)
            cell = sheet.cell(row=14, column=column)
            label = str(sheet.cell(row=14, column=1).value or "").strip()
            old_value = cell.value
            if label != "Taxes":
                raise RuntimeError(f"{property_name}: expected Taxes at row 14, found {label!r}")
            if old_value not in (expected_old, 0) and round(float(old_value), 2) != round(expected_old, 2):
                raise RuntimeError(f"{property_name}: expected {expected_old}, found {old_value}")
            status = "already_current" if old_value == 0 else "would_update"
            if apply_allowed and old_value != 0:
                cell.value = 0
                workbook.save(workbook_path)
                status = "updated"
            results.append({
                "property": property_name,
                "void_baselane_id": void_id,
                "workbook": str(workbook_path),
                "row": 14,
                "old_value": old_value,
                "new_value": 0,
                "status": status,
            })
        finally:
            workbook.close()
    payload = {
        "status": "applied" if apply_allowed else "dry_run",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_report": str(args.source_report),
        "apply_requested": args.apply,
        "apply_allowed": apply_allowed,
        "results": results,
    }
    args.report.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"status": payload["status"], "results": results}, indent=2))


if __name__ == "__main__":
    main()
