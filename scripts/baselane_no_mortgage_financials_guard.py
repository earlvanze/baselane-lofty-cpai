#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from coownership_mortgage_policy import (
    is_no_dao_mortgage_property,
    is_yhome_stolen_deed_mortgage_property,
)


OWNER_STATEMENTS_DIR = Path("Public") / "07 - P&L & Owner Statements"
FINANCIALS_FILENAME = "FINANCIALS.md"
CASH_FLOW_STATEMENT_GLOB = "Cash Flow Statement*.xlsx"
NO_MORTGAGE_STATES = {"IL", "OH", "TN"}
NO_MORTGAGE_SCAN_STATES = NO_MORTGAGE_STATES | {"HI", "NY"}
NO_MORTGAGE_ROW_LABELS = {
    "Mortgage Principal Balance",
    "Mortgage Principal Payments",
    "Mortgage Interest Payments",
}
AUTO_BLOCK_START = "<!-- AUTO:CF_PRIOR_MONTH_START -->"
AUTO_BLOCK_END = "<!-- AUTO:CF_PRIOR_MONTH_END -->"
ZERO_MONEY = "$0.00"
MARKDOWN_FINANCIALS_NAMES = {"financials.md"}


def iso_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def money_is_zero(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return abs(float(value)) < 0.005
    text = str(value).strip().replace("$", "").replace(",", "")
    if text in {"", "-", "—", "n/a", "N/A"}:
        return True
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return abs(float(text)) < 0.005
    except ValueError:
        return str(value).strip() == ZERO_MONEY


def markdown_financials_paths(real_estate_root: Path, states: set[str]) -> list[Path]:
    paths: list[Path] = []
    for state in sorted(states):
        state_dir = real_estate_root / state
        if not state_dir.is_dir():
            continue
        for property_dir in sorted(path for path in state_dir.iterdir() if path.is_dir()):
            for statement_dir in (
                property_dir / OWNER_STATEMENTS_DIR,
                property_dir / "07 - P&L & Owner Statements",
                property_dir / "00 - README & Property Snapshot",
                property_dir / "Public" / "00 - README & Property Snapshot",
            ):
                if not statement_dir.is_dir():
                    continue
                for path in sorted(statement_dir.glob("*.md")):
                    lowered = path.name.casefold()
                    if any(marker in lowered for marker in (".backup", ".superseded")):
                        continue
                    if (
                        lowered in MARKDOWN_FINANCIALS_NAMES
                        or "financials-approved" in lowered
                        or "financial summary" in lowered
                        or lowered.startswith("financials -")
                    ):
                        paths.append(path)
    return sorted(set(paths))


def cash_flow_workbook_paths(real_estate_root: Path, states: set[str]) -> list[Path]:
    paths: list[Path] = []
    for state in sorted(states):
        state_dir = real_estate_root / state
        if not state_dir.is_dir():
            continue
        for path in sorted(
            list(state_dir.glob(f"*/{OWNER_STATEMENTS_DIR}/{CASH_FLOW_STATEMENT_GLOB}"))
            + list(state_dir.glob(f"*/07 - P&L & Owner Statements/{CASH_FLOW_STATEMENT_GLOB}"))
        ):
            if not path.name.startswith("~$"):
                paths.append(path)
    return sorted(set(paths))


def financials_paths(real_estate_root: Path, states: set[str]) -> list[Path]:
    return [
        path
        for path in markdown_financials_paths(real_estate_root, states) + cash_flow_workbook_paths(real_estate_root, states)
        if path_is_no_mortgage_policy_path(real_estate_root, path)
    ]


def patch_table_row(line: str) -> tuple[str, dict[str, Any] | None]:
    if not line.lstrip().startswith("|"):
        return line, None
    columns = [part.strip() for part in line.strip().strip("|").split("|")]
    if len(columns) < 2:
        return line, None
    label = columns[0]
    value = columns[1]
    if label not in NO_MORTGAGE_ROW_LABELS or money_is_zero(value):
        return line, None
    patched_columns = columns[:]
    patched_columns[1] = ZERO_MONEY
    patched_line = "| " + " | ".join(patched_columns) + " |"
    return patched_line, {"label": label, "before": value, "after": ZERO_MONEY}


def patch_auto_blocks(text: str) -> tuple[str, list[dict[str, Any]]]:
    changes: list[dict[str, Any]] = []
    lines = text.splitlines()
    patched_lines: list[str] = []
    in_auto_block = False
    for line_no, line in enumerate(lines, start=1):
        if AUTO_BLOCK_START in line:
            in_auto_block = True
            patched_lines.append(line)
            continue
        if AUTO_BLOCK_END in line:
            in_auto_block = False
            patched_lines.append(line)
            continue
        if in_auto_block:
            patched_line, change = patch_table_row(line)
            if change:
                change["line"] = line_no
                changes.append(change)
            patched_lines.append(patched_line)
        else:
            patched_lines.append(line)
    trailing_newline = "\n" if text.endswith("\n") else ""
    return "\n".join(patched_lines) + trailing_newline, changes


def find_nonzero_rows(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    in_auto_block = False
    for line_no, line in enumerate(text.splitlines(), start=1):
        if AUTO_BLOCK_START in line:
            in_auto_block = True
            continue
        if AUTO_BLOCK_END in line:
            in_auto_block = False
            continue
        if not in_auto_block:
            continue
        if not line.lstrip().startswith("|"):
            continue
        columns = [part.strip() for part in line.strip().strip("|").split("|")]
        if len(columns) < 2:
            continue
        label = columns[0]
        value = columns[1]
        if label in NO_MORTGAGE_ROW_LABELS and not money_is_zero(value):
            rows.append({"line": line_no, "label": label, "value": value})
    return rows


def find_nonzero_markdown_rows(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_no, line in enumerate(text.splitlines(), start=1):
        if not line.lstrip().startswith("|"):
            continue
        columns = [part.strip() for part in line.strip().strip("|").split("|")]
        if len(columns) < 2:
            continue
        label = columns[0]
        value = columns[1]
        if label in NO_MORTGAGE_ROW_LABELS and not money_is_zero(value):
            rows.append({"line": line_no, "label": label, "value": value})
    return rows


def state_for_path(real_estate_root: Path, path: Path) -> str:
    try:
        relative = path.relative_to(real_estate_root)
    except ValueError:
        return ""
    return relative.parts[0] if relative.parts else ""


def path_is_no_mortgage_policy_path(real_estate_root: Path, path: Path) -> bool:
    state = state_for_path(real_estate_root, path)
    if state in NO_MORTGAGE_STATES and not is_yhome_stolen_deed_mortgage_property(str(path)):
        return True
    candidates = [path.stem, *path.parts]
    return any(
        is_no_dao_mortgage_property(candidate) and not is_yhome_stolen_deed_mortgage_property(candidate)
        for candidate in candidates
    )


def patch_markdown_rows(text: str) -> tuple[str, list[dict[str, Any]]]:
    changes: list[dict[str, Any]] = []
    lines = text.splitlines()
    patched_lines: list[str] = []
    for line_no, line in enumerate(lines, start=1):
        patched_line, change = patch_table_row(line)
        if change:
            change["line"] = line_no
            changes.append(change)
        patched_lines.append(patched_line)
    trailing_newline = "\n" if text.endswith("\n") else ""
    return "\n".join(patched_lines) + trailing_newline, changes


def process_markdown_file(path: Path, real_estate_root: Path, apply: bool) -> dict[str, Any]:
    before = path.read_text(encoding="utf-8")
    full_scan = path.name.casefold() != FINANCIALS_FILENAME.casefold()
    patched, changes = patch_markdown_rows(before) if full_scan else patch_auto_blocks(before)
    if apply and patched != before:
        path.write_text(patched, encoding="utf-8")
    after = patched if apply else before
    remaining = find_nonzero_markdown_rows(after) if full_scan else find_nonzero_rows(after)
    return {
        "path": str(path),
        "artifact_type": "financials_markdown",
        "state": state_for_path(real_estate_root, path),
        "changed": patched != before,
        "change_count": len(changes),
        "changes": changes,
        "remaining_nonzero_count": len(remaining),
        "remaining_nonzero_rows": remaining,
        "before_sha256": sha256_text(before),
        "after_sha256": sha256_text(patched if apply else before),
    }


def nonzero_workbook_mortgage_cells(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            label_cell = next((cell for cell in row if cell.value in NO_MORTGAGE_ROW_LABELS), None)
            if not label_cell:
                continue
            for cell in row:
                if cell.column == label_cell.column:
                    continue
                if money_is_zero(cell.value):
                    continue
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "cell": cell.coordinate,
                        "label": label_cell.value,
                        "value": str(cell.value),
                    }
                )
    return rows


def patch_workbook_mortgage_cells(workbook: Any) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            label_cell = next((cell for cell in row if cell.value in NO_MORTGAGE_ROW_LABELS), None)
            if not label_cell:
                continue
            for cell in row:
                if cell.column == label_cell.column or money_is_zero(cell.value):
                    continue
                changes.append(
                    {
                        "sheet": worksheet.title,
                        "cell": cell.coordinate,
                        "label": label_cell.value,
                        "before": str(cell.value),
                        "after": 0,
                    }
                )
                cell.value = 0
    return changes


def process_workbook_file(path: Path, real_estate_root: Path, apply: bool) -> dict[str, Any]:
    before_sha256 = sha256_file(path)
    workbook = load_workbook(path)
    before_nonzero = nonzero_workbook_mortgage_cells(workbook)
    if apply:
        changes = patch_workbook_mortgage_cells(workbook)
        if changes:
            workbook.save(path)
        remaining = nonzero_workbook_mortgage_cells(workbook)
    else:
        changes = [
            {
                "sheet": row["sheet"],
                "cell": row["cell"],
                "label": row["label"],
                "before": row["value"],
                "after": 0,
            }
            for row in before_nonzero
        ]
        remaining = before_nonzero
    after_sha256 = sha256_file(path) if apply and changes else before_sha256
    return {
        "path": str(path),
        "artifact_type": "cash_flow_workbook",
        "state": state_for_path(real_estate_root, path),
        "changed": bool(changes),
        "change_count": len(changes),
        "changes": changes,
        "remaining_nonzero_count": len(remaining),
        "remaining_nonzero_rows": remaining,
        "before_sha256": before_sha256,
        "after_sha256": after_sha256,
    }


def process_file(path: Path, real_estate_root: Path, apply: bool) -> dict[str, Any]:
    try:
        if path.suffix.lower() == ".xlsx":
            return process_workbook_file(path, real_estate_root, apply)
        return process_markdown_file(path, real_estate_root, apply)
    except Exception as exc:  # noqa: BLE001
        return {
            "path": str(path),
            "artifact_type": "cash_flow_workbook" if path.suffix.lower() == ".xlsx" else "financials_markdown",
            "state": state_for_path(real_estate_root, path),
            "status": "read_error",
            "changed": False,
            "change_count": 0,
            "changes": [],
            "remaining_nonzero_count": 0,
            "remaining_nonzero_rows": [],
            "before_sha256": None,
            "after_sha256": None,
            "error_type": exc.__class__.__name__,
            "error": str(exc),
        }


def build_report(real_estate_root: Path, states: set[str], apply: bool) -> dict[str, Any]:
    paths = financials_paths(real_estate_root, states)
    records = [process_file(path, real_estate_root, apply) for path in paths]
    changed_records = [record for record in records if record["changed"]]
    remaining_records = [record for record in records if record["remaining_nonzero_count"]]
    read_error_records = [record for record in records if record.get("status") == "read_error"]
    artifact_type_counts: dict[str, int] = {}
    for record in records:
        artifact_type = str(record.get("artifact_type") or "unknown")
        artifact_type_counts[artifact_type] = artifact_type_counts.get(artifact_type, 0) + 1
    return {
        "generated_at": iso_z(),
        "status": "review" if remaining_records or read_error_records else "ok",
        "apply": apply,
        "real_estate_root": str(real_estate_root),
        "states": sorted(states),
        "file_count": len(records),
        "artifact_type_counts": artifact_type_counts,
        "changed_file_count": len(changed_records),
        "read_error_count": len(read_error_records),
        "change_count": sum(int(record["change_count"]) for record in records),
        "remaining_nonzero_count": sum(int(record["remaining_nonzero_count"]) for record in records),
        "records": records,
    }


def parse_states(value: str) -> set[str]:
    states = {part.strip().upper() for part in re.split(r"[,\s]+", value) if part.strip()}
    invalid = states - NO_MORTGAGE_SCAN_STATES
    if invalid:
        raise argparse.ArgumentTypeError(f"unsupported no-mortgage states: {', '.join(sorted(invalid))}")
    return states or set(NO_MORTGAGE_SCAN_STATES)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Enforce no-mortgage public FINANCIALS.md/CF workbook rows for no-DAO-mortgage properties.")
    parser.add_argument("--real-estate-root", type=Path, default=Path("/mnt/c/Users/digit/Dropbox/Real Estate"))
    parser.add_argument("--states", type=parse_states, default=set(NO_MORTGAGE_SCAN_STATES))
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args(argv)

    report = build_report(args.real_estate_root, args.states, args.apply)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({key: report[key] for key in ["status", "apply", "file_count", "changed_file_count", "change_count", "remaining_nonzero_count"]}, sort_keys=True))
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
