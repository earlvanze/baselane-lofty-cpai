#!/usr/bin/env python3
"""Sync Sara's June 2026 ECO cash rows from the canonical split ledger."""

from __future__ import annotations

import argparse
import csv
import json
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


WORKBOOK = Path("/home/digit/Dropbox/Real Estate/OH/1432 Sara Ave, Akron, Ohio 44305/Public/07 - P&L & Owner Statements/Cash Flow Statement - 1432 Sara Ave - 1432 Sara Ave, Akron, Ohio 44305.xlsx")
LEDGER = Path("/home/digit/Dropbox/Real Estate/OH/1432 Sara Ave, Akron, Ohio 44305/Public/07 - P&L & Owner Statements/ECO Systems General Ledger - 1432 Sara Ave..csv")
REPORT = Path.cwd() / "reports/baselane_sara_june_cash_balance_sync.json"
APPLY_ENV = "SARA_JUNE_CASH_BALANCE_SYNC_APPLY"


def june_balance() -> float:
    cutoff = datetime(2026, 6, 30, 23, 59, 59)
    total = 0.0
    with LEDGER.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if datetime.strptime(row["Date"], "%B %d, %Y") <= cutoff:
                total += float(row["Amount"])
    return round(total, 2)


def june_column(sheet: object) -> int:
    for header_row in (1, 4):
        for column in range(2, 14):
            value = sheet.cell(row=header_row, column=column).value
            if getattr(value, "year", None) == 2026 and getattr(value, "month", None) == 6:
                return column
            if isinstance(value, str) and value.strip().lower() in {"jun-26", "june-26", "jun 2026", "june 2026"}:
                return column
    raise RuntimeError("June 2026 column not found")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--report", type=Path, default=REPORT)
    args = parser.parse_args()
    expected = june_balance()
    if expected != 1796.51:
        raise RuntimeError(f"unexpected Sara June ledger balance: {expected}")
    apply_allowed = args.apply and os.environ.get(APPLY_ENV) == "1"
    workbook = openpyxl.load_workbook(WORKBOOK)
    try:
        sheet = workbook["2026"]
        column = june_column(sheet)
        results = []
        for row, label in ((31, "ECO Operating Cash"), (56, "ECO General Ledger (ECO GL Column E Total)")):
            if str(sheet.cell(row=row, column=1).value or "").strip() != label:
                raise RuntimeError(f"expected {label!r} at row {row}")
            cell = sheet.cell(row=row, column=column)
            old_value = cell.value
            if round(float(old_value), 2) not in {-1188.49, expected}:
                raise RuntimeError(f"unexpected {label} value: {old_value}")
            status = "already_current" if round(float(old_value), 2) == expected else "would_update"
            if apply_allowed and status == "would_update":
                cell.value = expected
                status = "updated"
            results.append({"row": row, "cell": cell.coordinate, "label": label, "old_value": old_value, "new_value": expected, "status": status})
        if apply_allowed:
            workbook.save(WORKBOOK)
    finally:
        workbook.close()
    payload = {"status": "applied" if apply_allowed else "dry_run", "generated_at": datetime.now(timezone.utc).isoformat(), "workbook": str(WORKBOOK), "ledger": str(LEDGER), "as_of": "2026-06-30", "expected": expected, "results": results}
    args.report.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
