#!/usr/bin/env python3
"""
Update canonical Cash Flow Statement xlsx files from Baselane GL data.

Reads: ECO Systems General Ledger CSV
Updates: Per-property CF xlsx files in Dropbox Real Estate directories
Flags: Conflicts where CF values don't match GL totals (threshold: $0.01)

Architecture:
  - Baselane GL is the single source of truth
  - CF statements are overwritten from GL each month
  - Formula cells in xlsx are preserved (never overwritten)
  - Any CF value that differs from GL by > $0.01 is flagged as a conflict
  - Accruals (insurance, taxes, PM fees) must exist in Baselane; this script
    surfaces mismatches so they can be reconciled

Usage:
  python update_cf_statements.py [--month 2026-05] [--dry-run] [--property "84 Madison"]
  python update_cf_statements.py --audit   # Just audit, don't write
  python update_cf_statements.py --create-missing  # Create CF xlsx for properties without one
"""

import argparse
import ast
import copy
import csv
import json
import operator
import os
import re
import shutil
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from collections import defaultdict


try:
    import openpyxl
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    sys.exit(1)

# ── Paths ──
WORKSPACE_ROOT = Path(__file__).resolve().parents[3]


def first_existing_path(*candidates):
    for candidate in candidates:
        path = Path(candidate).expanduser()
        if path.exists():
            return path
    return Path(candidates[0])


DROPBOX_ROOT = first_existing_path(
    os.environ.get("DROPBOX_ROOT") or "/mnt/c/Users/digit/Dropbox",
    "/home/digit/Dropbox",
    "/data/Dropbox",
)
GL_PATH = first_existing_path(
    os.environ.get("BASELANE_LEDGER_PATH") or DROPBOX_ROOT / "Projects/assetrail/ECO Systems General Ledger.csv",
    DROPBOX_ROOT / "Projects/assetrail/ECO Systems General Ledger.csv",
    "/data/Dropbox/Projects/assetrail/ECO Systems General Ledger.csv",
)
REAL_ESTATE_BASE = first_existing_path(
    os.environ.get("REAL_ESTATE_BASE") or DROPBOX_ROOT / "Real Estate",
    DROPBOX_ROOT / "Real Estate",
    "/data/Dropbox/Real Estate",
)
OUTPUT_DIR = WORKSPACE_ROOT / "reports" / "cf_statement_sync"
CONFLICT_THRESHOLD = 0.01  # $0.01
OWNER_STATEMENTS_DIR = "07 - P&L & Owner Statements"
LEGACY_PUBLIC_DIR_PART = "Public"
LEGACY_FINANCE_DIR_PART = "Fin" + "ancials"
US_STATE_DIRS = {
    "AL", "AR", "AZ", "CA", "CO", "CT", "FL", "GA", "HI", "IA",
    "IL", "IN", "KY", "MA", "MD", "MI", "MO", "NC", "NJ", "NY",
    "OH", "OR", "PA", "SC", "TN", "TX", "UT", "VA", "WA", "WI",
}

# Template xlsx for creating new CF statements
TEMPLATE_PATH = first_existing_path(
    REAL_ESTATE_BASE / "NY/84 Madison Ave Public/07 - P&L & Owner Statements/Cash Flow Statement - 84 Madison Avenue, Albany, NY 12202.xlsx",
    "/data/Dropbox/Real Estate/NY/84 Madison Ave Public/07 - P&L & Owner Statements/Cash Flow Statement - 84 Madison Avenue, Albany, NY 12202.xlsx",
)

# Properties to skip entirely (no longer managed)
SKIP_PROPERTIES = {
    "82 madison ave albany",
    "1935 s glen rd shelby",
}

EXCLUDED_PROPERTY_REPORTS = (
    "baselane_financials_monthly_lofty_pm_publish.json",
    "baselane_financials_monthly_guarded_apply.json",
    "baselane_monthly_owner_email_send_guard.json",
)

# ── CF xlsx row labels → GL category mapping ──
ROW_TO_GL_CATEGORIES = {
    # Revenue
    "Repairs Reimbursement": ["Repairs Reimbursement"],
    "Fees & Other Revenue": ["Fees & Other Revenue", "Other"],
    "Rents": ["Rents"],
    "Rental Income": ["Rents"],
    "Leasing⁺ Packages": ["Management Fees"],
    "Lockbox Fee": ["Repairs"],

    # Operating Expenses
    "Advertising": ["Advertising"],
    "Auto & Travel": ["Auto & Travel"],
    "Cleaning & Maintenance": ["Cleaning & Maintenance"],
    "Insurance": ["Insurance"],
    "Legal & Other Professional Fees": ["Legal & Other Professional Fees"],
    "Other Operating Expenses": ["Other Operating Expenses"],
    "Non-Property Expense": ["Non-Property Expense"],
    "Property Management fee": ["Management Fees", "Property Management"],
    "Property Management Fee": ["Management Fees", "Property Management"],
    "Software Subscriptions": ["Software Subscriptions"],
    "Repairs": ["Repairs"],
    "Repairs/Supplies": ["Repairs", "Supplies"],
    "Supplies": ["Supplies"],
    "Taxes": ["Taxes"],
    "Utilities": ["Utilities"],

    # Investing
    "Capital Expenditures": ["Capital Expenditures"],
    "Down Payments": ["Down Payments"],
    "Sale Proceeds": ["Sale Proceeds"],

    # Financing
    "Mortgage Principal Payments": ["Mortgage Principal Payments"],
    "Mortgage Interest Payments": ["Mortgage Interest Payments"],
    "Mortgage Interest-Only Payments": ["Mortgage Interest Payments"],
    "Mortgage Payments (P&I)": ["Mortgage Payments"],
    "Mortgages & Loans": ["Mortgage Payments"],
    "Other Loan Payments (P&I)": ["Other Loan Payments", "Other Loan Payments (P&I)"],
    "Other Loan Principal Payments": ["Other Loan Principal Payments"],
    "Other Loan Interest Payments": ["Other Loan Interest Payments"],
    "Short-Term Loan Principal Payments": ["Other Loan Principal Payments"],
    "Short-Term Loan Interest Payments": ["Other Loan Interest Payments"],
    "Owner Contributions/Distributions": ["Owner Contributions/Distributions"],
    "Owner Distributions": ["Owner Contributions/Distributions"],
    "Sent to Lofty (Distributions)": ["Owner Contributions/Distributions"],
    "Retained Earnings": ["Retained Earnings"],
    "Inter-Account Transfers": ["Transfers Between Accounts"],
    "Credit Card Payments": ["Credit Card Payments"],
    "Escrow Payments": ["Escrow Payments"],
    "Security Deposits": ["Security Deposits"],
    "Deposits": ["Security Deposits"],
    "Contributions & Distributions": ["Owner Contributions/Distributions"],
    "Solar Lease Payments": ["Solar Lease Payments", "Other Loan Payments", "Other Loan Payments (P&I)"],
    "Solar Loan Payments": ["Solar Lease Payments", "Other Loan Payments", "Other Loan Payments (P&I)"],
}

ECO_PHYSICAL_BANK_CASH_LABEL = "ECO Net DAO Funds"
ECO_PHYSICAL_BANK_CASH_LEGACY_LABELS = (
    "ECO Operating Cash",
)
ECO_PHYSICAL_BANK_CASH_LABELS = (
    ECO_PHYSICAL_BANK_CASH_LABEL,
    *ECO_PHYSICAL_BANK_CASH_LEGACY_LABELS,
)
ECO_GL_NET_CASH_BALANCE_LABEL = "ECO General Ledger (ECO GL Column E Total)"
ECO_GL_NET_CASH_BALANCE_LEGACY_LABELS = (
    "ECO Operating Cash (ECO GL Column E Total)",
    "ECO GL Net Cash Balance (excl. EARLDAO Interest)",
)
ECO_GL_NET_CASH_BALANCE_LABELS = (
    ECO_GL_NET_CASH_BALANCE_LABEL,
    *ECO_GL_NET_CASH_BALANCE_LEGACY_LABELS,
)
MORTGAGE_PRINCIPAL_PAYMENTS_LABEL = "Mortgage Principal Payments"
MORTGAGE_INTEREST_PAYMENTS_LABEL = "Mortgage Interest Payments"
MORTGAGE_PRINCIPAL_BALANCE_LABEL = "Mortgage Principal Balance"
CITADEL_TEXT_RE = re.compile(r"\b(CITADEL|ACRA|LOANSPHERE|LOANDEPOT|FREEDOM|NEWREZ|SHELLPOIN(?:T)?|MORTGAGE\s+SERV)\b", re.I)

# Balance sheet rows that are not monthly-category sourced.
BALANCE_SHEET_ROWS = {
    "Lofty Operating Reserve (OR) Balance",
    "Operating Reserve (OR) Balance",
    *ECO_PHYSICAL_BANK_CASH_LABELS,
    *ECO_GL_NET_CASH_BALANCE_LABELS,
    "Total DAO Cash Balance",
    "Mortgage Principal Balance",
    "Other Liabilities",
    "Other Liabilities (excl. EARLDAO Interest)",
    "Total Market Cap",
    "Total Market Cap (excl. unapproved EARLDAO Interest)",
    "Estimated Equity",
    "Escrowed Tokens",
    "Shares to Buyback and Burn",
    "Floating Shares",
    "New OR Balance",
    "Owner Contributions/Distributions",
    "Sent to Lofty (Distributions)",
}

NO_MORTGAGE_STATES = {"IL", "OH", "TN"}
NO_DAO_MORTGAGE_PROPERTY_KEYS = {
    "85 104 alawa pl",
    "86 madison ave",
    "88 madison ave",
    "90 madison ave",
    "724 3rd ave",
}
NO_MORTGAGE_DEBT_ROW_LABELS = {
    "Mortgage Principal Payments",
    "Mortgage Interest Payments",
    "Mortgage Interest-Only Payments",
    "Mortgage Principal Balance",
}

# Rows that are always formulas and should never be overwritten
FORMULA_ROWS = {
    "Total",
    "Net Operating Income (NOI)",
    "Net Operating Cashflow (CF)",
    "Percentage Change",
}

# Rows that are section headers - skip
SKIP_ROWS = {
    "Revenue",
    "Operating Expenses",
    "Balance Sheet",
    "Tokenomics",
    "Retained Earnings (50%)",
}

RETAINED_EARNINGS_RECONCILIATION_EXEMPTIONS = {
    "22164 umland circle jenner",
}


def retained_earnings_reconciliation_exempt(property_name):
    return any(
        normalized_property_is_match(property_name, exempt_property)
        for exempt_property in RETAINED_EARNINGS_RECONCILIATION_EXEMPTIONS
    )


def path_contains_parts(path, expected_parts):
    """Return true when path contains expected path parts contiguously."""
    parts = [part.lower() for part in path.parts]
    expected = [part.lower() for part in expected_parts]
    if not expected:
        return False
    for index in range(0, len(parts) - len(expected) + 1):
        if parts[index:index + len(expected)] == expected:
            return True
    return False


def is_legacy_public_finance_path(path):
    return LEGACY_FINANCE_DIR_PART.lower() in [part.lower() for part in Path(path).parts]


def is_canonical_owner_statement_path(path):
    return OWNER_STATEMENTS_DIR in path.parts


def is_direct_property_owner_statement_dir(path, prop_dir):
    path = Path(path)
    prop_dir = Path(prop_dir)
    return path.parent == prop_dir or (path.parent.name == "Public" and path.parent.parent == prop_dir)


def property_state_from_path(path):
    parts = list(Path(path).parts)
    lower_parts = [part.lower() for part in parts]
    for index, part in enumerate(lower_parts):
        if part == "real estate" and index + 1 < len(parts):
            return parts[index + 1].upper()
    for part in parts:
        upper_part = part.upper()
        if upper_part in NO_MORTGAGE_STATES:
            return upper_part
    return None


def no_mortgage_debt_policy_applies(path):
    if property_state_from_path(path) in NO_MORTGAGE_STATES:
        return True
    normalized_parts = [normalize_property_name(part) for part in Path(path).parts]
    normalized_filename = normalize_property_name(property_name_from_cf_file(path))
    return any(
        key == normalized_filename
        or key in normalized_filename
        or any(key == part or key in part for part in normalized_parts)
        for key in NO_DAO_MORTGAGE_PROPERTY_KEYS
    )


def is_90_madison_cf_path(path):
    normalized_parts = [normalize_property_name(part) for part in Path(path).parts]
    normalized_filename = normalize_property_name(property_name_from_cf_file(path))
    key = "90 madison ave"
    return (
        key == normalized_filename
        or key in normalized_filename
        or any(key == part or key in part for part in normalized_parts)
    )


def no_mortgage_debt_row_is_exempt(path, label):
    """Allow only 90 Madison's exact GL-backed NOI principal curtailments.

    90 Madison remains a no-DAO-mortgage property: ordinary principal,
    interest, and the outstanding principal balance belong to ECO.  Its
    separately approved NOI principal curtailments are nevertheless DAO
    expenses and share the Mortgage Principal Payments CF row.  Baselane's
    deterministic AOPS-90-CURTAILMENT rows are the sole source for that row.
    """
    if str(label or "").strip() != "Mortgage Principal Payments":
        return False
    return is_90_madison_cf_path(path)


def ensure_madison_90_principal_payments_row(sheet, xlsx_path, dry_run=True):
    """Reuse a legacy zero interest-only row when 2024 lacks a principal row.

    The 2024 workbook predates the approved-curtailment presentation and has
    only a zero Mortgage Interest-Only Payments row.  Renaming that otherwise
    unused row avoids inserting rows and destabilizing historical formulas.
    """
    if not is_90_madison_cf_path(xlsx_path):
        return None
    labels = {
        str(sheet.cell(row=row_number, column=1).value or "").strip(): row_number
        for row_number in range(2, sheet.max_row + 1)
    }
    if "Mortgage Principal Payments" in labels:
        return None
    row_number = labels.get("Mortgage Interest-Only Payments")
    if not row_number:
        return None
    if not dry_run:
        sheet.cell(row=row_number, column=1).value = "Mortgage Principal Payments"
    return {
        "row": row_number,
        "label": "Mortgage Principal Payments",
        "action": "rename_legacy_90_madison_principal_row",
        "old_value": "Mortgage Interest-Only Payments",
        "new_value": "Mortgage Principal Payments",
    }


def cf_candidate_priority(path):
    """Prefer direct canonical CF files, then canonical nested statement files."""
    if path.parent.name == OWNER_STATEMENTS_DIR:
        return 0
    return 1


ADDRESS_TOKEN_ALIASES = {
    "avenue": "ave",
    "street": "st",
    "road": "rd",
    "drive": "dr",
    "lane": "ln",
    "place": "pl",
    "circle": "cir",
}


def normalized_address_alias(value):
    """Normalize street suffix variants before scoring duplicate CF workbooks."""
    normalized = normalize_property_name(str(value or ""))
    return " ".join(ADDRESS_TOKEN_ALIASES.get(token, token) for token in normalized.split())


def property_scope_matches_key(property_scope, property_key):
    """Match a requested property scope against a canonical folder, including street aliases."""
    if normalized_property_is_match(property_key, property_scope):
        return True
    return normalized_property_is_match(
        normalized_address_alias(property_key),
        normalized_address_alias(property_scope),
    )


def cf_candidate_priority_for_property(path, prop_dir_name):
    """Prefer the canonical CF workbook whose filename best matches its property folder."""
    filename = normalized_address_alias(property_name_from_cf_file(path))
    folder = normalized_address_alias(prop_dir_name)
    filename_tokens = property_tokens(filename)
    folder_tokens = property_tokens(folder)
    shared_tokens = filename_tokens & folder_tokens
    exact_match = bool(filename and folder and filename == folder)
    contained_match = bool(filename and folder and (filename in folder or folder in filename))
    has_city_context = bool(folder_tokens and len(shared_tokens) >= min(len(folder_tokens), 3))
    return (
        cf_candidate_priority(path),
        0 if exact_match else 1,
        0 if contained_match else 1,
        0 if has_city_context else 1,
        -len(shared_tokens),
        str(path).lower(),
    )


def cf_workbook_schema_priority(path):
    """Prefer DAO/ECO-aware CF workbooks over generic OR templates when duplicates exist."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return (3, "unreadable")
    try:
        labels = set()
        for sheet_name in ("2026", "2025"):
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row_number in range(1, min(sheet.max_row or 0, 80) + 1):
                value = sheet.cell(row=row_number, column=1).value
                if value:
                    labels.add(str(value).strip())
            break
        has_eco_cash = ECO_GL_NET_CASH_BALANCE_LABEL in labels
        has_lofty_distribution = "Sent to Lofty (Distributions)" in labels
        has_generic_owner_contribution = "Owner Contributions/Distributions" in labels
        if has_eco_cash and has_lofty_distribution:
            return (0, "dao_eco_template")
        if has_generic_owner_contribution:
            return (2, "generic_owner_contribution_template")
        if has_eco_cash:
            return (1, "eco_template")
        return (2, "unknown_template")
    finally:
        wb.close()


def is_top_level_public_property_dir(path):
    return normalize_property_name(Path(path).name).endswith(" public")


def owner_statement_dir_has_cf_workbook(path):
    owner_dir = Path(path) / OWNER_STATEMENTS_DIR
    if not owner_dir.is_dir():
        return False
    return any(
        child.is_file()
        and child.name.startswith("Cash Flow Statement")
        and child.suffix.lower() == ".xlsx"
        and "conflict" not in child.name.lower()
        for child in owner_dir.iterdir()
    )


def matching_public_sibling_key(key, public_sibling_keys):
    for public_key in public_sibling_keys:
        if normalized_property_is_match(key, public_key) or normalized_property_is_match(public_key, key):
            return public_key
        key_number = leading_address_number(key)
        public_number = leading_address_number(public_key)
        if not key_number or key_number != public_number:
            continue
        shared_non_numeric_tokens = {
            token
            for token in property_tokens(key) & property_tokens(public_key)
            if not token.isdigit()
        }
        if shared_non_numeric_tokens:
            return public_key
    return None


def discover_cf_files(include_metadata=False, excluded_properties=None, property_scope=None):
    """Scan Real Estate directories for CF xlsx files."""
    cf_files = {}
    candidates = defaultdict(list)
    skipped = defaultdict(list)
    duplicate_candidates = {}
    owner_statement_dirs = {}
    excluded_owner_statement_dirs = {}
    public_shadowed_property_dirs = {}
    property_dir_names = {}
    excluded_properties = set(excluded_properties or [])
    property_scope_key = normalize_property_name(property_scope or "")
    property_scope_filtered = []
    real_estate = REAL_ESTATE_BASE
    if not real_estate.exists():
        for candidate in (
            Path("/home/digit/Dropbox/Real Estate"),
            Path("/mnt/c/Users/digit/Dropbox/Real Estate"),
            Path("/data/Dropbox/Real Estate"),
        ):
            if candidate.exists():
                real_estate = candidate
                break

    for state_dir in sorted(real_estate.iterdir()):
        if not state_dir.is_dir():
            continue
        if state_dir.name.upper() not in US_STATE_DIRS:
            continue
        if state_dir.name.startswith("_"):
            continue
        if state_dir.name in ("Lofty PM", "Archived Property Analyses", "Resources", "Yhome", "Other", "Archive", "Archives"):
            continue
        state_property_dirs = sorted(path for path in state_dir.iterdir() if path.is_dir())
        public_sibling_dirs = {
            normalize_property_name(path.name): path
            for path in state_property_dirs
            if is_top_level_public_property_dir(path) and owner_statement_dir_has_cf_workbook(path)
        }
        for prop_dir in state_property_dirs:
            if not prop_dir.is_dir():
                continue
            if prop_dir.name.startswith("_"):
                continue
            if prop_dir.name in {LEGACY_PUBLIC_DIR_PART, "reports"}:
                continue
            key = normalize_property_name(prop_dir.name)
            if not is_top_level_public_property_dir(prop_dir):
                public_sibling_key = matching_public_sibling_key(key, public_sibling_dirs)
                if public_sibling_key:
                    public_shadowed_property_dirs[key] = {
                        "path": str(prop_dir),
                        "shadowed_by": str(public_sibling_dirs[public_sibling_key]),
                        "shadowed_by_key": public_sibling_key,
                    }
                    continue
            if property_scope_key and not property_scope_matches_key(property_scope_key, key):
                property_scope_filtered.append(key)
                continue
            property_dir_names[key] = prop_dir.name
            if is_excluded_property_key(key, excluded_properties):
                excluded_owner_statement_dirs[key] = str(prop_dir)
                continue
            for root, dirs, files in os.walk(prop_dir):
                root_path = Path(root)
                if (
                    root_path.name == OWNER_STATEMENTS_DIR
                    and is_canonical_owner_statement_path(root_path)
                    and is_direct_property_owner_statement_dir(root_path, prop_dir)
                    and not is_legacy_public_finance_path(root_path)
                ):
                    owner_statement_dirs[key] = str(root_path)
                for f in files:
                    if f.startswith("Cash Flow Statement") and f.endswith(".xlsx"):
                        if "conflicted copy" in f.lower() or "conflict" in f.lower():
                            continue
                        full_path = Path(root) / f
                        if is_legacy_public_finance_path(full_path):
                            skipped[key].append({
                                "path": str(full_path),
                                "reason": "legacy_public_finance_dir_ignored",
                            })
                            continue
                        if not is_canonical_owner_statement_path(full_path):
                            skipped[key].append({
                                "path": str(full_path),
                                "reason": "noncanonical_owner_statement_dir_ignored",
                            })
                            continue
                        candidates[key].append(full_path)

    for key, paths in candidates.items():
        prop_dir_name = property_dir_names.get(key) or key
        schema_priorities = {}
        if len(paths) > 1:
            schema_priorities = {path: cf_workbook_schema_priority(path) for path in paths}
        ranked_paths = sorted(
            paths,
            key=lambda path: (
                schema_priorities.get(path, (0, "not_needed"))[0],
                cf_candidate_priority_for_property(path, prop_dir_name),
            ),
        )
        cf_files[key] = ranked_paths[0]
        if len(ranked_paths) > 1:
            duplicate_candidates[key] = {
                "selected": str(ranked_paths[0]),
                "ignored": [str(path) for path in ranked_paths[1:]],
                "candidate_count": len(ranked_paths),
                "schema_priorities": {str(path): schema_priorities.get(path) for path in ranked_paths},
            }

    if not include_metadata:
        return cf_files
    duplicate_template_mismatch_candidates = duplicate_template_mismatches(duplicate_candidates)
    metadata = {
        "skipped": dict(skipped),
        "canonical_property_count": len(cf_files),
        "skipped_property_count": len(skipped),
        "canonical_owner_statement_dirs": owner_statement_dirs,
        "excluded_owner_statement_dirs": excluded_owner_statement_dirs,
        "excluded_property_count": len(excluded_owner_statement_dirs),
        "public_shadowed_property_dir_count": len(public_shadowed_property_dirs),
        "public_shadowed_property_dirs": public_shadowed_property_dirs,
        "duplicate_candidates": duplicate_candidates,
        "duplicate_candidate_property_count": len(duplicate_candidates),
        "duplicate_template_mismatch_property_count": len(duplicate_template_mismatch_candidates),
        "duplicate_template_mismatch_candidates": duplicate_template_mismatch_candidates,
        "property_scope": property_scope,
        "property_scope_filter_enabled": bool(property_scope_key),
        "property_scope_filtered_count": len(property_scope_filtered),
    }
    return cf_files, metadata


def schema_label(schema_priority):
    if isinstance(schema_priority, (list, tuple)) and len(schema_priority) > 1:
        return schema_priority[1]
    return None


def duplicate_template_mismatches(duplicate_candidates):
    """Return duplicate CF groups where candidates use multiple template schemas."""
    mismatches = []
    for key, record in sorted((duplicate_candidates or {}).items()):
        schema_priorities = record.get("schema_priorities") or {}
        labels = {
            schema_label(priority)
            for priority in schema_priorities.values()
            if schema_label(priority)
        }
        if len(labels) <= 1:
            continue
        selected = record.get("selected")
        mismatches.append({
            "property_key": key,
            "selected": selected,
            "selected_schema_priority": schema_priorities.get(selected),
            "candidate_count": record.get("candidate_count"),
            "schema_labels": sorted(labels),
            "schema_priorities": schema_priorities,
            "ignored": record.get("ignored") or [],
        })
    return mismatches


def duplicate_owner_statement_dir_has_matching_cf(key, cf_files):
    """Return true when a short/public owner statement folder duplicates a CF-backed folder."""
    key_norm = normalize_property_name(key)
    key_number = leading_address_number(key_norm)
    key_tokens = property_tokens(key_norm)
    for cf_key in cf_files:
        cf_norm = normalize_property_name(cf_key)
        if normalized_property_is_match(key_norm, cf_norm) or normalized_property_is_match(cf_norm, key_norm):
            return True
        cf_number = leading_address_number(cf_norm)
        if key_number and cf_number == key_number and key_tokens & property_tokens(cf_norm):
            if key_norm.endswith(" public") or " public " in f" {key_norm} ":
                return True
            if cf_norm.endswith(" public") or " public " in f" {cf_norm} ":
                shared_non_numeric_tokens = {
                    token
                    for token in key_tokens & property_tokens(cf_norm)
                    if not token.isdigit()
                }
                if shared_non_numeric_tokens:
                    return True
    return False


def missing_owner_statement_suppression_reason(key, cf_files, gl_properties=None):
    if duplicate_owner_statement_dir_has_matching_cf(key, cf_files):
        return "duplicate_owner_statement_dir_has_matching_cf"
    if gl_properties is not None and not match_gl_property(key, gl_properties):
        return "owner_statement_dir_not_in_baselane_gl_scope"
    return None


def build_discovery_review_rows(cf_files, discovery_metadata, gl_properties=None):
    """Create audit rows for properties where only ignored CF candidates exist."""
    rows = []
    skipped = discovery_metadata.get("skipped") or {}
    suppressed_missing = []
    for key in sorted(skipped):
        if key in cf_files:
            continue
        suppression_reason = missing_owner_statement_suppression_reason(key, cf_files, gl_properties)
        if suppression_reason:
            suppressed_missing.append({"property": key, "reason": suppression_reason})
            continue
        candidates_for_property = skipped[key]
        summary = {
            "file": candidates_for_property[0].get("path") if candidates_for_property else None,
            "property": key,
            "rows": 0,
            "matches": 0,
            "conflicts": 0,
            "missing_in_gl": 0,
            "formulas": 0,
            "balance_sheet_skipped": 0,
            "untagged_gl_rows": 0,
        }
        rows.append({
            "summary": summary,
            "property": key,
            "error": "Canonical CF workbook missing; ignored noncanonical or legacy candidates",
            "ignored_candidates": candidates_for_property[:10],
            "conflicts": [],
        })
    owner_statement_dirs = discovery_metadata.get("canonical_owner_statement_dirs") or {}
    for key in sorted(owner_statement_dirs):
        if key in cf_files or key in skipped:
            continue
        suppression_reason = missing_owner_statement_suppression_reason(key, cf_files, gl_properties)
        if suppression_reason:
            suppressed_missing.append({
                "property": key,
                "path": owner_statement_dirs[key],
                "reason": suppression_reason,
            })
            continue
        rows.append({
            "summary": {
                "file": owner_statement_dirs[key],
                "property": key,
                "rows": 0,
                "matches": 0,
                "conflicts": 0,
                "missing_in_gl": 0,
                "formulas": 0,
                "balance_sheet_skipped": 0,
                "untagged_gl_rows": 0,
            },
            "property": key,
            "error": "Canonical owner statement folder exists but has no Cash Flow Statement workbook",
            "ignored_candidates": [],
            "conflicts": [],
        })
    discovery_metadata["suppressed_missing_owner_statement_dirs"] = suppressed_missing
    discovery_metadata["suppressed_missing_owner_statement_dir_count"] = len(suppressed_missing)
    return rows


def discovery_report(discovery_metadata, cf_files, year, month, gl_properties=None):
    skipped = discovery_metadata.get("skipped") or {}
    owner_statement_dirs = discovery_metadata.get("canonical_owner_statement_dirs") or {}
    excluded_owner_statement_dirs = discovery_metadata.get("excluded_owner_statement_dirs") or {}
    duplicate_candidates = discovery_metadata.get("duplicate_candidates") or {}
    duplicate_template_mismatch_candidates = duplicate_template_mismatches(duplicate_candidates)
    property_scope_exclusion_overrides = discovery_metadata.get("property_scope_exclusion_overrides") or []
    property_scope = discovery_metadata.get("property_scope")
    skipped_by_reason = defaultdict(int)
    missing_canonical_properties = []
    for key, candidates_for_property in skipped.items():
        for candidate in candidates_for_property:
            skipped_by_reason[candidate.get("reason") or "unknown"] += 1
        if key not in cf_files:
            if missing_owner_statement_suppression_reason(key, cf_files, gl_properties):
                continue
            missing_canonical_properties.append(key)
    missing_from_owner_statement_dirs = [
        key
        for key in sorted(owner_statement_dirs)
        if key not in cf_files
        and key not in skipped
        and not missing_owner_statement_suppression_reason(key, cf_files, gl_properties)
    ]
    missing_canonical_total = sorted(set(missing_canonical_properties) | set(missing_from_owner_statement_dirs))
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "month": f"{year}-{month:02d}",
        "status": "review" if missing_canonical_total else "ok",
        "property_scope": property_scope,
        "property_scope_filter_enabled": bool(discovery_metadata.get("property_scope_filter_enabled")),
        "property_scope_filtered_count": discovery_metadata.get("property_scope_filtered_count") or 0,
        "canonical_property_count": len(cf_files),
        "canonical_owner_statement_dir_count": len(owner_statement_dirs),
        "excluded_property_count": len(excluded_owner_statement_dirs),
        "excluded_properties": sorted(excluded_owner_statement_dirs)[:100],
        "property_scope_exclusion_override_count": len(property_scope_exclusion_overrides),
        "property_scope_exclusion_overrides": property_scope_exclusion_overrides[:100],
        "ignored_candidate_count": sum(len(paths) for paths in skipped.values()),
        "ignored_legacy_candidate_count": skipped_by_reason.get("legacy_public_finance_dir_ignored", 0),
        "ignored_noncanonical_candidate_count": skipped_by_reason.get("noncanonical_owner_statement_dir_ignored", 0),
        "duplicate_candidate_property_count": len(duplicate_candidates),
        "duplicate_candidates": duplicate_candidates,
        "duplicate_template_mismatch_property_count": len(duplicate_template_mismatch_candidates),
        "duplicate_template_mismatch_candidates": duplicate_template_mismatch_candidates,
        "suppressed_missing_owner_statement_dir_count": discovery_metadata.get("suppressed_missing_owner_statement_dir_count") or 0,
        "suppressed_missing_owner_statement_dirs": (discovery_metadata.get("suppressed_missing_owner_statement_dirs") or [])[:100],
        "missing_canonical_from_ignored_count": len(missing_canonical_properties),
        "missing_canonical_from_ignored_properties": missing_canonical_properties[:100],
        "missing_canonical_from_owner_statement_dir_count": len(missing_from_owner_statement_dirs),
        "missing_canonical_from_owner_statement_dir_properties": missing_from_owner_statement_dirs[:100],
        "missing_canonical_total_count": len(missing_canonical_total),
        "missing_canonical_total_properties": missing_canonical_total[:100],
        "skipped_by_reason": dict(sorted(skipped_by_reason.items())),
    }


def normalize_property_name(name):
    """Normalize property name for matching."""
    name = re.sub(r',?\s+(AL|AR|CA|CO|FL|GA|HI|IA|IL|MI|MO|NY|OH|SC|TN|TX|UT|WA)\s+\d{5}', '', name)
    name = re.sub(r',?\s+(AL|AR|CA|CO|FL|GA|HI|IA|IL|MI|MO|NY|OH|SC|TN|TX|UT|WA)\s*$', '', name)
    name = name.strip().lower()
    name = re.sub(r"\bcir\b", "circle", name)
    name = re.sub(r'[^\w\s]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name


PROPERTY_GL_ALIASES = {
    "326332 s alcott st denver": {"326 south alcott street"},
    "326332 s alcott st public": {"326 south alcott street"},
    "3139 west blvd cleveland oh 44111": {"3139 west blvd"},
    "3139 west blvd cleveland": {"3139 west blvd"},
}


def normalized_property_is_match(candidate, reference):
    candidate_norm = normalize_property_name(str(candidate or ""))
    reference_norm = normalize_property_name(str(reference or ""))
    if not candidate_norm or not reference_norm:
        return False
    if candidate_norm == reference_norm:
        return True
    candidate_number = leading_address_number(candidate_norm)
    reference_number = leading_address_number(reference_norm)
    if not candidate_number or candidate_number != reference_number:
        return False
    shared = property_tokens(candidate_norm) & property_tokens(reference_norm)
    if len(shared) < 2:
        return False
    return (
        candidate_norm in reference_norm
        or reference_norm in candidate_norm
        or len(shared) >= 3
    )


def load_excluded_property_names(root=None):
    """Load sold/closed/manual excluded property names from policy and guarded monthly reports."""
    root = Path(root or os.environ.get("ROOT") or WORKSPACE_ROOT)
    reports = root / "reports"
    policy = root / "config" / "lofty_listing_update_policy.json"
    excluded = set()
    if policy.is_file():
        try:
            payload = json.loads(policy.read_text(encoding="utf-8"))
        except Exception:
            payload = {}
        for field in ("sold_ignore_listing_updates", "operational_ignore_listing_updates"):
            for item in payload.get(field) or []:
                name = item.get("address") if isinstance(item, dict) else item
                normalized = normalize_property_name(str(name or ""))
                if normalized:
                    excluded.add(normalized)
    for report_name in EXCLUDED_PROPERTY_REPORTS:
        path = reports / report_name
        if not path.is_file():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        for field in (
            "excluded_property_names",
            "excluded_total_property_names",
            "manual_excluded_property_names",
            "skipped_closed_property_names",
        ):
            for name in payload.get(field) or []:
                normalized = normalize_property_name(str(name))
                if normalized:
                    excluded.add(normalized)
        guard = payload.get("yhome_transition_guard") if isinstance(payload.get("yhome_transition_guard"), dict) else {}
        for name in guard.get("excluded_property_names") or []:
            normalized = normalize_property_name(str(name))
            if normalized:
                excluded.add(normalized)
    return excluded


def is_excluded_property_key(key, excluded_property_names):
    return any(normalized_property_is_match(key, excluded) for excluded in excluded_property_names or [])


def property_scope_matches_exclusion(property_scope, excluded_property_name):
    """Let a deliberate --property run process that one property even if it is excluded globally."""
    if not property_scope:
        return False
    scope_key = normalize_property_name(property_scope)
    excluded_key = normalize_property_name(excluded_property_name)
    return normalized_property_is_match(scope_key, excluded_key) or normalized_property_is_match(excluded_key, scope_key)


PROPERTY_MATCH_STOPWORDS = {
    "ave", "avenue", "blvd", "ct", "dr", "drive",
    "east", "e", "heights", "lane", "ln", "north", "n", "ohio", "park",
    "place", "pl", "public", "rd", "road", "south", "s", "st", "street",
    "the", "west", "w",
}


def property_tokens(normalized_name):
    return {
        token
        for token in normalized_name.split()
        if len(token) > 1 and token not in PROPERTY_MATCH_STOPWORDS
    }


def leading_address_number(normalized_name):
    match = re.match(r"^(\d+)\b", normalized_name)
    return match.group(1) if match else None


def match_gl_property(cf_name, gl_properties):
    """Match a CF workbook property key to a GL property without loose numeric substrings."""
    cf_norm = normalize_property_name(cf_name)
    cf_aliases = PROPERTY_GL_ALIASES.get(cf_norm, set())
    cf_number = leading_address_number(cf_norm)
    cf_tokens = property_tokens(cf_norm)
    best_prop = None
    best_score = 0.0

    for gl_prop in sorted(gl_properties):
        gl_norm = normalize_property_name(gl_prop)
        if gl_norm == cf_norm:
            return gl_prop
        if gl_norm in cf_aliases:
            return gl_prop

        gl_number = leading_address_number(gl_norm)
        gl_tokens = property_tokens(gl_norm)
        shared = cf_tokens & gl_tokens
        token_union = cf_tokens | gl_tokens
        score = 0.0

        if cf_number and gl_number and cf_number == gl_number and shared:
            score = 0.75 + (len(shared) / max(len(token_union), 1))
        elif len(cf_norm) >= 8 and len(gl_norm) >= 8 and (cf_norm in gl_norm or gl_norm in cf_norm):
            score = 0.70 + (len(shared) / max(len(token_union), 1))
        elif cf_number and gl_number and cf_number == gl_number:
            score = 0.55
        elif shared and len(shared) >= 2:
            score = len(shared) / max(len(token_union), 1)

        if score > best_score:
            best_prop = gl_prop
            best_score = score

    return best_prop if best_score >= 0.70 else None


def property_name_from_cf_file(path):
    """Extract a property label from a Cash Flow Statement workbook filename."""
    stem = Path(path).stem
    stem = re.sub(r"^Cash\s+Flow\s+Statement\s*[-–—]\s*", "", stem, flags=re.IGNORECASE).strip()
    return stem or Path(path).stem


# ── GL category sets ──

CF_OPERATING_INFLOW = {
    "Rents", "Fees & Other Revenue", "Interest Received",
}

CF_OPERATING_OUTFLOW = {
    "Repairs", "Cleaning & Maintenance", "Supplies",
    "Repairs Reimbursement", "Software Subscriptions",
    "Utilities", "Insurance", "Taxes",
    "Management Fees", "Property Management", "Legal & Other Professional Fees",
    "Advertising", "Auto & Travel",
    "Other Operating Expenses", "Non-Property Expense",
    "Mortgage Interest Payments", "Other Loan Interest Payments",
    "Escrow Payments", "Solar Lease Payments",
}

CF_INVESTING = {
    "Capital Expenditures", "Down Payments", "Sale Proceeds",
}

CF_FINANCING = {
    "Transfers Between Accounts", "Credit Card Payments",
    "Owner Contributions/Distributions",
    "Security Deposits",
    "Mortgage Payments", "Other Loan Payments",
    "Mortgage Principal Payments", "Other Loan Principal Payments",
    "Mortgage Disbursements Received", "Other Loan Disbursements Received",
}

CF_SUBCATEGORY_LABELS = {
    "Rents": "Rents",
    "Fees & Other Revenue": "Fees & Other Revenue",
    "Interest Received": "Interest Received",
    "Repairs": "Repairs",
    "Repairs Reimbursement": "Repairs Reimbursement",
    "Cleaning & Maintenance": "Cleaning & Maintenance",
    "Supplies": "Supplies",
    "Software Subscriptions": "Software Subscriptions",
    "Utilities": "Utilities",
    "Insurance": "Insurance",
    "Taxes": "Taxes",
    "Management Fees": "Management Fees",
    "Property Management": "Management Fees",
    "Legal & Other Professional Fees": "Legal & Other Professional Fees",
    "Advertising": "Advertising",
    "Auto & Travel": "Auto & Travel",
    "Other Operating Expenses": "Other Operating Expenses",
    "Non-Property Expense": "Non-Property Expense",
    "Mortgage Interest Payments": "Mortgage Interest Payments",
    "Other Loan Interest Payments": "Other Loan Interest Payments",
    "Solar Lease Payments": "Solar Lease Payments",
    "Capital Expenditures": "Capital Expenditures",
    "Down Payments": "Down Payments",
    "Sale Proceeds": "Sale Proceeds",
    "Transfers Between Accounts": "Inter-Account Transfers",
    "Credit Card Payments": "Credit Card Payments",
    "Owner Contributions/Distributions": "Owner Distributions",
    "Escrow Payments": "Escrow Payments",
    "Security Deposits": "Security Deposits",
    "Mortgage Payments": "Mortgage Payments (P&I)",
    "Other Loan Payments": "Other Loan Payments (P&I)",
    "Mortgage Principal Payments": "Mortgage Principal Payments",
    "Other Loan Principal Payments": "Other Loan Principal Payments",
    "Other Loan Interest Payments": "Other Loan Interest Payments",
    "Mortgage Disbursements Received": "Mortgage Proceeds Received",
    "Other Loan Disbursements Received": "Other Loan Proceeds Received",
}

BASELANE_CATEGORY_ALIASES = {
    "Short Term Rents": "Rents",
    "Long Term Rents": "Rents",
    "Cleaning & Janitorial": "Cleaning & Maintenance",
    "City, State, & Local Taxes": "Taxes",
    "Gas & Electric": "Utilities",
    "Phone, Cable & Internet": "Utilities",
    "Garbage & Recycling": "Utilities",
    "Electric": "Utilities",
    "Appliance Repairs": "Repairs",
    "Electrical Repairs": "Repairs",
    "Legal Fees": "Legal & Other Professional Fees",
    "Plumbing Repairs": "Repairs",
    "Rental Dwelling": "Insurance",
    "Repairs Labor": "Repairs",
    "Gardening & Landscaping": "Cleaning & Maintenance",
}

MERCHANT_OPERATING_INFLOW = {
    "airbnb payments", "airbnb 4977", "evolve vacation", "hospitable, inc",
    "hostshare", "vrbo",
}

MERCHANT_OPERATING_OUTFLOW = {
    "the home depot", "lowe's", "walmart.com", "walmart",
    "hospitable.com", "pricelabsinc*dynaprice",
    "hemlane", "heml", "spectrum", "payless power",
    "epcon lane",
}

DESC_KEYWORD_OPERATING_INFLOW = {"payment", "rent", "payout"}
DESC_KEYWORD_OPERATING_OUTFLOW = {
    "purchase", "subscription", "utility", "electric", "water",
    "gas bill", "insurance", "repair", "maintenance",
}
MERCHANT_FINANCING = {"freedom"}
DESC_KEYWORD_FINANCING = {
    "mortgage pymt", "mtg pymt", "loan payment", "principal",
    "internal_transfer", "ach transfer",
}

ACCRUAL_MARKER_CATEGORIES = {
    "dao": "Legal & Other Professional Fees",
    "insurance": "Insurance",
    "legal": "Legal & Other Professional Fees",
    "mortgage_interest": "Mortgage Interest Payments",
    "pm": "Management Fees",
    "retained_capital": "Retained Earnings",
    "taxes": "Taxes",
}


def parse_amount(val):
    if not val:
        return 0.0
    val = val.strip().replace(",", "")
    if val.startswith("(") and val.endswith(")"):
        return -float(val[1:-1])
    return float(val)


FORMULA_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def parse_cell_number(value, sheet=None):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("=") and sheet is not None:
            return evaluate_formula(stripped, sheet)
        try:
            return parse_amount(stripped.replace("$", ""))
        except Exception:
            return 0.0
    return 0.0


def evaluate_formula_ast(node):
    if isinstance(node, ast.Expression):
        return evaluate_formula_ast(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.BinOp) and type(node.op) in FORMULA_OPERATORS:
        return FORMULA_OPERATORS[type(node.op)](evaluate_formula_ast(node.left), evaluate_formula_ast(node.right))
    if isinstance(node, ast.UnaryOp) and type(node.op) in FORMULA_OPERATORS:
        return FORMULA_OPERATORS[type(node.op)](evaluate_formula_ast(node.operand))
    raise ValueError(f"Unsupported formula expression: {ast.dump(node)}")


def evaluate_formula(formula, sheet):
    """Evaluate the small formula subset used by canonical CF accrual rows."""
    expression = str(formula or "").strip()
    if expression.startswith("="):
        expression = expression[1:]
    expression = expression.upper().replace("$", "")

    def replace_sum(match):
        start, end = match.group(1), match.group(2)
        total = 0.0
        for row in sheet[start:end]:
            for cell in row:
                total += parse_cell_number(cell.value, sheet)
        return str(total)

    expression = re.sub(r"SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", replace_sum, expression)

    def replace_cell(match):
        return str(parse_cell_number(sheet[match.group(1)].value, sheet))

    expression = re.sub(r"(?<![A-Z])([A-Z]+\d+)(?![A-Z])", replace_cell, expression)
    if not re.fullmatch(r"[0-9eE\.\+\-\*/\(\) ]+", expression):
        raise ValueError(f"Unsupported formula after substitution: {expression}")
    return evaluate_formula_ast(ast.parse(expression, mode="eval"))


def parse_date(val):
    val = val.strip().strip('"')
    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d",
                "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"]:
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {val}")


def categorize_transaction(row):
    """Categorize a GL transaction into CF section + category.

    Priority: Sub-category > Category > keyword fallback
    For blank categories, flag as uncategorized.
    """
    category = (row.get("Category") or "").strip()
    subcategory = (row.get("Sub-category") or "").strip()
    merchant = (row.get("Merchant") or "").strip()
    description = (row.get("Description") or "").strip()
    amount = parse_amount(row.get("Amount", "0"))

    # Primary: use Baselane's tagged Category (not Sub-category since it's always blank)
    key = subcategory if subcategory else category
    transaction_type = (row.get("Type") or "").strip()
    notes = (row.get("Notes") or "").strip()

    accrual_marker = re.search(r"AOPS-[A-Z]+-ACCRUAL\|([a-z_]+)\|", notes, flags=re.IGNORECASE)
    if accrual_marker:
        cf_category = ACCRUAL_MARKER_CATEGORIES.get(accrual_marker.group(1).lower())
        if cf_category:
            return "Operating", cf_category, False

    if key == "Landscaping":
        if "capex" in transaction_type.lower() or "loan payments" in transaction_type.lower():
            return "Investing", "Capital Expenditures", False
        return "Operating", "Cleaning & Maintenance", False

    if key in BASELANE_CATEGORY_ALIASES:
        cf_category = BASELANE_CATEGORY_ALIASES[key]
        if cf_category in CF_OPERATING_INFLOW:
            return "Operating", CF_SUBCATEGORY_LABELS.get(cf_category, cf_category), False
        if cf_category in CF_OPERATING_OUTFLOW:
            return "Operating", CF_SUBCATEGORY_LABELS.get(cf_category, cf_category), False
        if cf_category in CF_INVESTING:
            return "Investing", CF_SUBCATEGORY_LABELS.get(cf_category, cf_category), False
        if cf_category in CF_FINANCING:
            return "Financing", CF_SUBCATEGORY_LABELS.get(cf_category, cf_category), False

    if key:
        if key in CF_OPERATING_INFLOW:
            return "Operating", CF_SUBCATEGORY_LABELS.get(key, key), False
        if key in CF_OPERATING_OUTFLOW:
            return "Operating", CF_SUBCATEGORY_LABELS.get(key, key), False
        if key in CF_INVESTING:
            return "Investing", CF_SUBCATEGORY_LABELS.get(key, key), False
        if key in CF_FINANCING:
            return "Financing", CF_SUBCATEGORY_LABELS.get(key, key), False

    # Keyword fallback for untagged transactions
    merchant_lower = merchant.lower()
    desc_lower = description.lower()

    if "internal_transfer" in desc_lower or "internal_transfer" in merchant_lower:
        return "Financing", "Inter-Account Transfers", True
    for m in MERCHANT_FINANCING:
        if m in merchant_lower:
            return "Financing", "Mortgage Payments (P&I)", True
    for kw in DESC_KEYWORD_FINANCING:
        if kw in desc_lower:
            return "Financing", "Mortgage Payments (P&I)", True
    for m in MERCHANT_OPERATING_INFLOW:
        if m in merchant_lower:
            return "Operating", "Rents", True
    for kw in DESC_KEYWORD_OPERATING_INFLOW:
        if kw in desc_lower:
            return "Operating", "Rents", True
    for m in MERCHANT_OPERATING_OUTFLOW:
        if m in merchant_lower:
            return "Operating", "Operating Expenses", True
    for kw in DESC_KEYWORD_OPERATING_OUTFLOW:
        if kw in desc_lower:
            return "Operating", "Operating Expenses", True

    if "owner" in desc_lower and ("distribution" in desc_lower or "contribution" in desc_lower):
        return "Financing", "Owner Distributions", True
    if "baselane" in merchant_lower:
        if amount > 0:
            return "Financing", "Inter-Account Transfers", True
        else:
            return "Operating", "Management Fees", True
    if "osc" in merchant_lower and "risk" in merchant_lower:
        return "Operating", "Insurance", True
    if "arcadia" in merchant_lower:
        return "Operating", "Utilities", True

    if amount > 0:
        return "Operating", "Uncategorized Income", True
    return "Operating", "Uncategorized Expense", True


def load_gl_data(gl_path):
    if not gl_path.exists():
        raise FileNotFoundError(f"GL file not found: {gl_path}")
    transactions = []
    untagged_count = 0
    with open(gl_path, "r", encoding="utf-8-sig", errors="ignore") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                row["_date"] = parse_date(row.get("Date", ""))
            except (ValueError, KeyError):
                continue
            row["_amount"] = parse_amount(row.get("Amount", "0"))
            prop = (row.get("Property") or "").strip()
            if not prop or prop in ("N/A", "None"):
                account = (row.get("Account") or "").strip()
                if "-" in account:
                    m = re.search(r'DAO LLC[\s-]+(.+?)(?:\s+Operations|\s+Reserves|\s+Capital)', account)
                    if m:
                        prop = m.group(1).strip()
                    else:
                        parts = account.split("-")
                        if len(parts) >= 2:
                            prop = parts[1].strip()
            row["_property"] = prop
            section, cf_cat, is_fallback = categorize_transaction(row)
            row["_section"] = section
            row["_cf_category"] = cf_cat
            row["_is_fallback"] = is_fallback
            if is_fallback:
                untagged_count += 1
            transactions.append(row)
    print(f"  {untagged_count}/{len(transactions)} transactions used keyword fallback (unflagged in Baselane)")
    return transactions


def filter_by_month(transactions, year, month):
    return [t for t in transactions if t["_date"].year == year and t["_date"].month == month]


def is_month_end_accounting_row(transaction, year, month):
    row_date = transaction.get("_date")
    notes = str(transaction.get("Notes") or "").strip().upper()
    return (
        row_date is not None
        and row_date.year == year
        and row_date.month == month
        and notes.startswith("AOPS-")
    )


def filter_through_date(transactions, cutoff, accounting_month=None):
    if cutoff is None:
        return transactions
    return [
        transaction
        for transaction in transactions
        if (
            (
                transaction.get("_date") is not None
                and transaction["_date"].date() <= cutoff
            )
            or (
                accounting_month is not None
                and is_month_end_accounting_row(
                    transaction, accounting_month[0], accounting_month[1]
                )
            )
        )
    ]


def filter_by_property(transactions, property_name):
    """Return every transaction for the guarded alias-equivalent property identity.

    Baselane cash transactions often carry a short address while typed accruals
    carry the workbook's full city/state address.  A raw substring comparison
    silently drops one side of that same property ledger.
    """
    return [
        transaction
        for transaction in transactions
        if normalized_property_is_match(transaction.get("_property", ""), property_name)
    ]


def has_hemlane_direct_split_pm_void(transactions, cf_value):
    """Return whether a typed zero-cash Hemlane PM void proves this CF amount.

    Hemlane withholds these fees before depositing net rent, so the ledger's
    zero-cash correction is evidence of the expense rather than a zero expense.
    """
    if cf_value is None or abs(round(float(cf_value), 2)) <= CONFLICT_THRESHOLD:
        return False
    expected_amount = round(abs(float(cf_value)), 2)
    for transaction in transactions:
        if round(float(transaction.get("_amount") or 0), 2) != 0:
            continue
        category = str(transaction.get("_cf_category") or "")
        if category not in {"Management Fees", "Property Management"}:
            continue
        evidence = " ".join(
            str(transaction.get(field) or "")
            for field in ("Merchant", "Description", "Notes")
        )
        if "Hemlane Direct-Split PM Void" not in evidence:
            continue
        legacy_amount = re.search(r"Voided legacy manual PM row of \$(\d+(?:\.\d+)?)", evidence)
        if legacy_amount and round(float(legacy_amount.group(1)), 2) == expected_amount:
            return True
    return False


def is_earldao_interest_transaction(transaction):
    text = " ".join(
        str(transaction.get(field) or "")
        for field in ("Account", "Merchant", "Description", "Type", "Category", "Sub-category", "Notes")
    ).upper()
    return "EARLDAO" in text and "INTEREST" in text


def is_non_cash_accrual_overlay(transaction):
    """Manual P&L overlays affect accrual reporting but never bank cash."""
    return "AOPS-PNL-ACCRUAL" in str(transaction.get("Notes") or "")


def source_cash_balance(transactions, year=None, month=None):
    """Return full ECO Net DAO Funds from every canonical property-GL row.

    This is a general-ledger balance, not physical bank cash. Accrual journals
    remain included because they are DAO liabilities and assets. Their later
    settlement rows must also remain in the GL; balanced accrual/settlement
    accounting prevents double counting without discarding either side.
    """
    included_total = 0.0
    included_count = 0
    included_accrual_count = 0
    included_accrual_total = 0.0
    for transaction in transactions or []:
        amount = float(transaction.get("_amount") or 0.0)
        if is_non_cash_accrual_overlay(transaction):
            included_accrual_count += 1
            included_accrual_total += amount
        included_total += amount
        included_count += 1
    return {
        "expected": round(included_total, 2),
        "included_count": included_count,
        "included_manual_accrual_count": included_accrual_count,
        "included_manual_accrual_total": round(included_accrual_total, 2),
        "excluded_manual_accrual_count": 0,
        "excluded_manual_accrual_total": 0.0,
        "excluded_earldao_interest_count": 0,
        "excluded_earldao_interest_total": 0.0,
        "as_of_date": "full_property_split_gl",
        "scope": "full_property_split_gl_column_e",
        "month_filter_applied": False,
        "balance_semantics": "full_eco_gl_column_e_balance_not_monthly_cashflow",
    }


def source_cash_balance_as_of(transactions, year, month):
    """Return the ECO GL Column E balance through a reporting month-end.

    This is intentionally separate from ``source_cash_balance``. The latter
    remains the authoritative current full-balance calculation used for DAO
    reconciliation; this helper is for closed historical CF columns only.
    """
    cutoff = date(year, month, 1) + timedelta(days=32)
    cutoff = cutoff.replace(day=1) - timedelta(days=1)
    dated_rows = []
    for transaction in transactions or []:
        row_date = transaction.get("_date")
        if isinstance(row_date, datetime):
            row_date = row_date.date()
        if row_date is not None and row_date <= cutoff:
            dated_rows.append(transaction)
    balance = source_cash_balance(dated_rows, year, month)
    return {
        **balance,
        "as_of_date": cutoff.isoformat(),
        "scope": "property_split_gl_column_e_through_month_end",
        "month_filter_applied": True,
        "source_cash_balance_mode": "as_of_month_end",
        "balance_semantics": "historical_eco_gl_column_e_balance_through_month_end",
    }


def source_cash_balance_for_month(transactions, year, month, mode="full_column_e"):
    """Select explicit current or closed-historical ECO cash semantics."""
    if mode == "as_of_month_end":
        return source_cash_balance_as_of(transactions, year, month)
    if mode != "full_column_e":
        raise ValueError(f"unknown source cash balance mode: {mode}")
    balance = source_cash_balance(transactions, year, month)
    return {**balance, "source_cash_balance_mode": mode}


def money_to_float(text):
    text = str(text or "").strip().replace("$", "").replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        return -float(text[1:-1])
    return float(text)


def extract_pdf_text(path):
    try:
        result = subprocess.run(
            ["pdftotext", str(path), "-"],
            capture_output=True,
            text=True,
            timeout=20,
            check=False,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except Exception:
        pass
    try:
        return Path(path).read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""


def public_dir_for_cf_workbook(xlsx_path):
    for parent in Path(xlsx_path).resolve().parents:
        if parent.name == "Public":
            return parent
    return None


def mortgage_statement_candidates(xlsx_path, year, month):
    public_dir = public_dir_for_cf_workbook(xlsx_path)
    if public_dir is None:
        return []
    loan_dir = public_dir / "04 - Loan Documents"
    loan_dirs = [
        candidate
        for candidate in (
            loan_dir / str(year),
            loan_dir / "Mortgage Statements" / str(year),
            loan_dir / "Mortgage Statements",
            loan_dir,
        )
        if candidate.is_dir()
    ]
    candidates = []
    for loan_year_dir in loan_dirs:
        for path in loan_year_dir.rglob("*"):
            if not path.is_file() or "statement" not in path.name.lower():
                continue
            match = re.search(r"(20\d{2})[-_]?(\d{2})(?:[-_]?(\d{2}))?", path.name, flags=re.I)
            if not match:
                continue
            statement_year = int(match.group(1))
            statement_month = int(match.group(2))
            statement_day = int(match.group(3) or "31")
            if statement_year == year and statement_month == month:
                candidates.append((statement_day, path))
    seen = []
    seen_paths = set()
    for _, path in sorted(candidates, reverse=True):
        if path in seen_paths:
            continue
        seen_paths.add(path)
        seen.append(path)
    return seen


def parse_citadel_statement_text(text):
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    money_re = re.compile(r"^\(?\$?-?\d[\d,]*\.\d{2}\)?$")
    money_any_re = re.compile(r"\(?\$-?\d[\d,]*\.\d{2}\)?")
    principal_balance = None
    paid_last_month = []

    for index, line in enumerate(lines):
        lowered = line.lower()
        if principal_balance is None and "outstanding principal balance" in lowered:
            inline_match = money_any_re.search(line)
            if inline_match:
                principal_balance = money_to_float(inline_match.group(0))
                continue
            for candidate in lines[index + 1 : index + 20]:
                if money_re.match(candidate):
                    principal_balance = money_to_float(candidate)
                    break
        if "paid last month" in lowered:
            labeled_values = {}
            for candidate in lines[index + 1 : index + 12]:
                candidate_lower = candidate.lower()
                values = money_any_re.findall(candidate)
                if values:
                    for label in ("principal", "interest", "escrow", "fees", "total"):
                        if candidate_lower.startswith(label):
                            labeled_values[label] = money_to_float(values[0])
                            break
                if money_re.match(candidate):
                    paid_last_month.append(money_to_float(candidate))
                if len(paid_last_month) >= 5:
                    break
            if len(labeled_values) >= 3:
                paid_last_month = [
                    labeled_values.get("principal", 0.0),
                    labeled_values.get("interest", 0.0),
                    labeled_values.get("escrow", 0.0),
                    labeled_values.get("fees", 0.0),
                    labeled_values.get("total", 0.0),
                ]
                break

    if principal_balance is None or len(paid_last_month) < 3:
        return None
    return {
        "principal_balance": round(principal_balance, 2),
        "paid_principal": round(paid_last_month[0], 2),
        "paid_interest": round(paid_last_month[1], 2),
        "paid_escrow": round(paid_last_month[2], 2),
        "paid_fees": round(paid_last_month[3], 2) if len(paid_last_month) > 3 else 0.0,
        "paid_total": round(paid_last_month[4], 2) if len(paid_last_month) > 4 else None,
    }


def parse_loandepot_statement_text(text):
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    money_any_re = re.compile(r"\(?\$-?\d[\d,]*\.\d{2}\)?")
    principal_balance = None
    explanation = {}
    past_payments = {}
    for index, line in enumerate(lines):
        lowered = line.lower()
        if principal_balance is None and "outstanding principal balance" in lowered:
            inline_match = money_any_re.search(line)
            if inline_match:
                principal_balance = money_to_float(inline_match.group(0))
        if lowered.startswith(("principal", "interest", "escrow")):
            values = money_any_re.findall(line)
            if not values:
                continue
            key = "escrow" if lowered.startswith("escrow") else lowered.split()[0]
            if "paid year to date" in " ".join(lines[max(0, index - 4) : index]).lower():
                past_payments[key] = money_to_float(values[0])
            else:
                explanation.setdefault(key, money_to_float(values[0]))
        if lowered.startswith(("regular monthly payment", "current amount due")):
            values = money_any_re.findall(line)
            if values:
                explanation.setdefault("total", money_to_float(values[0]))
    principal = past_payments.get("principal", explanation.get("principal"))
    interest = past_payments.get("interest", explanation.get("interest"))
    escrow = past_payments.get("escrow", explanation.get("escrow"))
    total = explanation.get("total")
    if principal_balance is None or principal is None or interest is None or escrow is None:
        return None
    if total is None:
        total = round(float(principal) + float(interest) + float(escrow), 2)
    return {
        "principal_balance": round(principal_balance, 2),
        "paid_principal": round(principal, 2),
        "paid_interest": round(interest, 2),
        "paid_escrow": round(escrow, 2),
        "paid_fees": 0.0,
        "paid_total": round(total, 2),
    }


def parse_generic_mortgage_statement_text(text):
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    money_any_re = re.compile(r"\(?\$-?\d[\d,]*\.\d{2}\)?")
    principal_balance = None
    interest_rate = None
    due_year = None
    due_month = None
    due_components = {}
    paid_components = {}
    in_past_payments = False

    for line in lines:
        lowered = line.lower()
        if principal_balance is None and (
            "outstanding principal" in lowered
            or "outstanding balance" in lowered
        ):
            values = money_any_re.findall(line)
            if values:
                principal_balance = max(money_to_float(value) for value in values)
        if interest_rate is None and "interest rate" in lowered:
            rate_match = re.search(r"(\d+(?:\.\d+)?)\s*%", line)
            if rate_match:
                interest_rate = float(rate_match.group(1))
        if due_year is None and ("payment due date" in lowered or "next due date" in lowered):
            date_match = re.search(r"\b(\d{1,2})/(\d{1,2})/(20\d{2}|\d{2})\b", line)
            if date_match:
                due_month = int(date_match.group(1))
                year_value = int(date_match.group(3))
                due_year = 2000 + year_value if year_value < 100 else year_value
        if "past payments breakdown" in lowered:
            in_past_payments = True
            continue
        if in_past_payments and lowered.startswith(("transaction activity", "important messages", "amount due")):
            in_past_payments = False

        values = money_any_re.findall(line)
        if not values:
            continue
        key = None
        if lowered.startswith("principal"):
            key = "principal"
        elif "principal" in lowered and not ("balance" in lowered and len(values) == 1):
            key = "principal"
        elif lowered.startswith("interest"):
            key = "interest"
        elif "interest" in lowered and "rate" not in lowered and not ("balance" in lowered and len(values) == 1):
            key = "interest"
        elif lowered.startswith("escrow"):
            key = "escrow"
        elif "escrow" in lowered and "balance" not in lowered:
            key = "escrow"
        elif lowered.startswith("fees") or lowered.startswith("total fees"):
            key = "fees"
        elif lowered.startswith("total"):
            key = "total"
        elif "regular monthly payment" in lowered:
            key = "total"
        if key is None:
            continue
        if key == "interest" and lowered.startswith("interest rate"):
            continue
        if in_past_payments:
            paid_components.setdefault(key, money_to_float(values[0]))
        else:
            due_components.setdefault(key, money_to_float(values[-1]))

    principal = paid_components.get("principal", due_components.get("principal"))
    interest = paid_components.get("interest", due_components.get("interest"))
    escrow = paid_components.get("escrow", due_components.get("escrow"))
    if principal_balance is None or principal is None or interest is None or escrow is None:
        return None
    paid_total = paid_components.get("total")
    if paid_total is None:
        paid_total = round(float(principal) + float(interest) + float(escrow) + float(paid_components.get("fees") or 0.0), 2)
    result = {
        "principal_balance": round(principal_balance, 2),
        "paid_principal": round(principal, 2),
        "paid_interest": round(interest, 2),
        "paid_escrow": round(escrow, 2),
        "paid_fees": round(float(paid_components.get("fees") or 0.0), 2),
        "paid_total": round(float(paid_total), 2),
    }
    if interest_rate is not None:
        result["interest_rate"] = round(interest_rate, 6)
    if due_year and due_month:
        result["due_year"] = due_year
        result["due_month"] = due_month
    if {"principal", "interest", "escrow"} <= due_components.keys():
        due_total = due_components.get("total")
        if due_total is None:
            due_total = round(
                float(due_components["principal"])
                + float(due_components["interest"])
                + float(due_components["escrow"])
                + float(due_components.get("fees") or 0.0),
                2,
            )
        result.update(
            {
                "due_principal": round(float(due_components["principal"]), 2),
                "due_interest": round(float(due_components["interest"]), 2),
                "due_escrow": round(float(due_components["escrow"]), 2),
                "due_fees": round(float(due_components.get("fees") or 0.0), 2),
                "due_total": round(float(due_total), 2),
            }
        )
    return result


def parse_mortgage_statement_text(text):
    generic = parse_generic_mortgage_statement_text(text)
    parsed = parse_citadel_statement_text(text) or parse_loandepot_statement_text(text) or generic
    if not parsed:
        return None
    if generic:
        merged = dict(generic)
        merged.update({key: value for key, value in parsed.items() if value is not None})
        return merged
    return parsed


def select_statement_amount_variant(statement, row_amount=None, prefer_due=False):
    if row_amount is None:
        return statement
    target = abs(float(row_amount))
    variants = (("due", statement.get("due_total")), ("paid", statement.get("paid_total"))) if prefer_due else (
        ("paid", statement.get("paid_total")),
        ("due", statement.get("due_total")),
    )
    for prefix, total in variants:
        if total is None or abs(target - float(total)) > 1.0:
            continue
        if prefix == "paid":
            selected = dict(statement)
        else:
            selected = {
                **statement,
                "paid_principal": statement.get("due_principal"),
                "paid_interest": statement.get("due_interest"),
                "paid_escrow": statement.get("due_escrow"),
                "paid_fees": statement.get("due_fees", 0.0),
                "paid_total": statement.get("due_total"),
            }
        selected["statement_component_source"] = prefix
        return selected
    return statement


def citadel_statement_for_month(xlsx_path, year, month, row_amount=None):
    for path in mortgage_statement_candidates(xlsx_path, year, month):
        parsed = parse_mortgage_statement_text(extract_pdf_text(path))
        if parsed:
            parsed = select_statement_amount_variant(parsed, row_amount, prefer_due=False)
            return {
                **parsed,
                "statement_path": str(path),
                "statement_found": True,
            }
    return {"statement_found": False}


def text_contains_citadel(transaction):
    haystack = " ".join(
        str(transaction.get(field) or "")
        for field in ("Merchant", "Description", "Category", "Sub-category", "Notes")
    )
    return bool(CITADEL_TEXT_RE.search(haystack))


def citadel_mortgage_statement_split(gl_month_data, xlsx_path, year, month):
    raw_citadel_rows = [
        transaction
        for transaction in gl_month_data
        if text_contains_citadel(transaction)
        and transaction.get("_cf_category") in {"Mortgage Payments (P&I)", "Mortgage Payments"}
    ]
    split_principal_total = round(
        sum(float(transaction.get("_amount") or 0.0) for transaction in gl_month_data if transaction.get("_cf_category") == MORTGAGE_PRINCIPAL_PAYMENTS_LABEL),
        2,
    )
    split_interest_total = round(
        sum(float(transaction.get("_amount") or 0.0) for transaction in gl_month_data if transaction.get("_cf_category") == MORTGAGE_INTEREST_PAYMENTS_LABEL),
        2,
    )
    evidence = {
        "checked": bool(raw_citadel_rows),
        "status": "not_applicable",
        "raw_citadel_payment_count": len(raw_citadel_rows),
        "existing_split_principal_total": split_principal_total,
        "existing_split_interest_total": split_interest_total,
        "statement_found": False,
    }
    if not raw_citadel_rows:
        return evidence
    if abs(split_principal_total) > CONFLICT_THRESHOLD or abs(split_interest_total) > CONFLICT_THRESHOLD:
        evidence["status"] = "existing_split_rows_present"
        return evidence

    row_amount = raw_citadel_rows[0].get("_amount") if raw_citadel_rows else None
    statement = citadel_statement_for_month(xlsx_path, year, month, row_amount)
    evidence.update(statement)
    if not statement.get("statement_found"):
        evidence["status"] = "missing_statement"
        return evidence

    evidence.update(
        {
            "status": "applied",
            "principal_payment": round(-float(statement["paid_principal"]), 2),
            "interest_payment": round(-float(statement["paid_interest"]), 2),
            "escrow_payment": round(-float(statement["paid_escrow"]), 2),
        }
    )
    return evidence


def find_row_by_label(sheet, target_label):
    for row_number in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row_number, column=1).value or "").strip() == target_label:
            return row_number
    return None


def citadel_escrow_component_values(sheet, month_col, statement_split):
    if statement_split.get("status") != "applied":
        return None
    paid_escrow = abs(float(statement_split.get("escrow_payment") or 0.0))
    if paid_escrow <= CONFLICT_THRESHOLD:
        return None
    insurance_row = find_row_by_label(sheet, "Insurance")
    taxes_row = find_row_by_label(sheet, "Taxes")
    if insurance_row is None or taxes_row is None:
        return None
    for source_col in range(month_col, 1, -1):
        insurance_value = numeric_cell_value(sheet.cell(row=insurance_row, column=source_col).value)
        taxes_value = numeric_cell_value(sheet.cell(row=taxes_row, column=source_col).value)
        if insurance_value is None or taxes_value is None:
            continue
        if abs(insurance_value) <= CONFLICT_THRESHOLD and abs(taxes_value) <= CONFLICT_THRESHOLD:
            continue
        insurance_component = round(abs(insurance_value), 2)
        taxes_component = round(paid_escrow - insurance_component, 2)
        if taxes_component < 0:
            continue
        if abs(taxes_component - abs(taxes_value)) <= 1.0:
            return {
                "insurance": round(-insurance_component, 2),
                "taxes": round(-taxes_component, 2),
                "source_column": source_col,
            }
    return None


def overlay_citadel_statement_split(gl_by_category, statement_split, sheet=None, month_col=None):
    if statement_split.get("status") != "applied":
        return
    gl_by_category[MORTGAGE_PRINCIPAL_PAYMENTS_LABEL] = statement_split["principal_payment"]
    gl_by_category[MORTGAGE_INTEREST_PAYMENTS_LABEL] = statement_split["interest_payment"]
    if sheet is not None and month_col is not None:
        escrow_components = citadel_escrow_component_values(sheet, month_col, statement_split)
        if escrow_components and abs(gl_by_category.get("Insurance", 0.0)) <= CONFLICT_THRESHOLD:
            gl_by_category["Insurance"] = escrow_components["insurance"]
        if escrow_components and abs(gl_by_category.get("Taxes", 0.0)) <= CONFLICT_THRESHOLD:
            gl_by_category["Taxes"] = escrow_components["taxes"]
        if escrow_components:
            statement_split["escrow_component_insurance"] = escrow_components["insurance"]
            statement_split["escrow_component_taxes"] = escrow_components["taxes"]
            statement_split["escrow_component_source_column"] = escrow_components["source_column"]


def audit_source_cash_balance_row(sheet, source_cash_data, year, month, source_cash_mode="full_column_e"):
    balance = source_cash_balance_for_month(source_cash_data, year, month, source_cash_mode)
    violations = []
    row_number = None
    for candidate_row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=candidate_row, column=1).value or "").strip() in ECO_GL_NET_CASH_BALANCE_LABELS:
            row_number = candidate_row
            break
    if row_number is None:
        violations.append(
            {
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "value_type": "missing_row",
                "expected": balance["expected"],
            }
        )
        return {**balance, "checked": True, "row": None, "violation_count": len(violations), "violations_bounded": violations}

    column_number = get_month_column(sheet, year, month)
    if not column_number:
        return {**balance, "checked": False, "row": row_number, "violation_count": 0, "violations_bounded": []}

    cell = sheet.cell(row=row_number, column=column_number)
    numeric_value = numeric_cell_value(cell.value)
    if isinstance(cell.value, str) and cell.value.strip().startswith("="):
        violations.append(
            {
                "row": row_number,
                "column": column_number,
                "cell": cell.coordinate,
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "expected": balance["expected"],
                "actual": cell.value,
                "value_type": "formula",
                "action": "replace_with_source_gl_reporting_balance",
            }
        )
    elif numeric_value is None:
        violations.append(
            {
                "row": row_number,
                "column": column_number,
                "cell": cell.coordinate,
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "expected": balance["expected"],
                "actual": cell.value,
                "value_type": "non_numeric",
                "action": "replace_with_source_gl_reporting_balance",
            }
        )
    elif abs(round(numeric_value, 2) - balance["expected"]) > CONFLICT_THRESHOLD:
        violations.append(
            {
                "row": row_number,
                "column": column_number,
                "cell": cell.coordinate,
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "expected": balance["expected"],
                "actual": round(numeric_value, 2),
                "diff": round(abs(round(numeric_value, 2) - balance["expected"]), 2),
                "value_type": "number",
                "action": "replace_with_source_gl_reporting_balance",
            }
        )
    return {
        **balance,
        "checked": True,
        "row": row_number,
        "column": column_number,
        "cell": cell.coordinate,
        "actual": cell.value,
        "violation_count": len(violations),
        "violations_bounded": violations[:25],
    }


def find_source_cash_balance_insert_row(sheet):
    for row_number in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row=row_number, column=1).value or "").strip()
        if label.startswith("Total Market Cap"):
            for candidate_row in range(row_number + 1, min(sheet.max_row + 2, row_number + 8)):
                if all(sheet.cell(row=candidate_row, column=column).value in (None, "") for column in range(1, min(sheet.max_column, 14) + 1)):
                    return candidate_row
            return sheet.max_row + 1
    for row_number in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row=row_number, column=1).value or "").strip()
        if label == "Balance Sheet":
            return row_number + 1
    return sheet.max_row + 1


def ensure_source_cash_balance_row(sheet):
    for row_number in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row=row_number, column=1).value or "").strip()
        if label in ECO_GL_NET_CASH_BALANCE_LABELS:
            renamed = label != ECO_GL_NET_CASH_BALANCE_LABEL
            if renamed:
                sheet.cell(row=row_number, column=1, value=ECO_GL_NET_CASH_BALANCE_LABEL)
            return row_number, False, renamed
    row_number = find_source_cash_balance_insert_row(sheet)
    sheet.cell(row=row_number, column=1, value=ECO_GL_NET_CASH_BALANCE_LABEL)
    return row_number, True, False


# ── CF xlsx operations ──

def parse_month_header(value):
    """Parse CF statement month headers such as May-26, May 2026, or 2026-05."""
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%b-%y", "%B-%y", "%b %y", "%B %y", "%b-%Y", "%B-%Y", "%b %Y", "%B %Y", "%Y-%m"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.year, parsed.month
        except ValueError:
            continue
    return None


def get_month_column(sheet, year, month):
    """Find the column for a given month in the CF statement.

    Canonical workbooks use row 1. Legacy Net Cash Flow exports use row 4,
    but can still be audited once the sheet has been copied into a year tab.
    """
    for header_row in (1, 4):
        for col in range(2, 14):
            cell = sheet.cell(row=header_row, column=col)
            if cell.value:
                if isinstance(cell.value, datetime):
                    if cell.value.year == year and cell.value.month == month:
                        return col
                elif isinstance(cell.value, str):
                    if parse_month_header(cell.value) == (year, month):
                        return col
    return None


def get_year_sheet(wb, year):
    """Get the sheet for a given year."""
    sheet_name = str(year)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return None


def numeric_cell_value(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0.0
        if text.startswith("="):
            return None
        try:
            return float(text.replace(",", "").replace("$", "").replace("(", "-").replace(")", ""))
        except ValueError:
            return None
    return None


def formula_row_label(label):
    return (
        label in FORMULA_ROWS
        or label.startswith("Total")
        or label.startswith("Net")
        or label.startswith("Percentage")
    )


def template_formula_row_should_clear(label):
    return str(label or "").strip() == MORTGAGE_PRINCIPAL_BALANCE_LABEL


def restore_blank_formula_from_neighbor(sheet, row_num, col, dry_run=True):
    """Restore a formula cell cleared for a future month when that month is synced."""
    target = sheet.cell(row=row_num, column=col)
    if target.value is not None:
        return None
    max_col = min(sheet.max_column, 14)
    source_cols = list(range(col - 1, 1, -1)) + list(range(col + 1, max_col + 1))
    for source_col in source_cols:
        source = sheet.cell(row=row_num, column=source_col)
        source_value = source.value
        if not (isinstance(source_value, str) and source_value.strip().startswith("=")):
            continue
        try:
            translated = Translator(source_value, origin=source.coordinate).translate_formula(target.coordinate)
        except Exception:
            translated = source_value
        if not dry_run:
            target.value = translated
        return {
            "row": row_num,
            "label": str(sheet.cell(row=row_num, column=1).value or "").strip(),
            "action": "restore_formula",
            "old_value": None,
            "new_value": translated,
            "source_cell": source.coordinate,
        }
    return None


def audit_no_mortgage_debt_rows(wb, xlsx_path):
    state = property_state_from_path(xlsx_path)
    if not no_mortgage_debt_policy_applies(xlsx_path):
        return {
            "state": state,
            "checked": False,
            "checked_cell_count": 0,
            "violation_count": 0,
            "violations_bounded": [],
        }

    violations = []
    checked_cell_count = 0
    for sheet_name in wb.sheetnames:
        if not str(sheet_name).isdigit():
            continue
        sheet = wb[sheet_name]
        for row_number in range(2, sheet.max_row + 1):
            label = str(sheet.cell(row=row_number, column=1).value or "").strip()
            if label not in NO_MORTGAGE_DEBT_ROW_LABELS:
                continue
            if no_mortgage_debt_row_is_exempt(xlsx_path, label):
                continue
            for column_number in range(2, min(sheet.max_column, 15) + 1):
                checked_cell_count += 1
                cell = sheet.cell(row=row_number, column=column_number)
                value = cell.value
                numeric_value = numeric_cell_value(value)
                if isinstance(value, str) and value.strip().startswith("="):
                    violations.append(
                        {
                            "sheet": sheet_name,
                            "row": row_number,
                            "column": column_number,
                            "cell": cell.coordinate,
                            "label": label,
                            "value_type": "formula",
                        }
                    )
                elif numeric_value is not None and abs(numeric_value) > CONFLICT_THRESHOLD:
                    violations.append(
                        {
                            "sheet": sheet_name,
                            "row": row_number,
                            "column": column_number,
                            "cell": cell.coordinate,
                            "label": label,
                            "value": round(numeric_value, 2),
                            "value_type": "number",
                        }
                    )
    return {
        "state": state,
        "checked": True,
        "checked_cell_count": checked_cell_count,
        "violation_count": len(violations),
        "violations_bounded": violations[:25],
    }


def clear_no_mortgage_debt_rows(wb, xlsx_path):
    """Force no-mortgage state debt cells to zero, including copied formulas."""
    state = property_state_from_path(xlsx_path)
    if not no_mortgage_debt_policy_applies(xlsx_path):
        return {"state": state, "checked": False, "cleared_cell_count": 0}

    cleared_cell_count = 0
    for sheet_name in wb.sheetnames:
        if not str(sheet_name).isdigit():
            continue
        sheet = wb[sheet_name]
        for row_number in range(2, sheet.max_row + 1):
            label = str(sheet.cell(row=row_number, column=1).value or "").strip()
            if label not in NO_MORTGAGE_DEBT_ROW_LABELS:
                continue
            if no_mortgage_debt_row_is_exempt(xlsx_path, label):
                continue
            for column_number in range(2, min(sheet.max_column, 15) + 1):
                cell = sheet.cell(row=row_number, column=column_number)
                if cell.value != 0:
                    cell.value = 0
                    cleared_cell_count += 1
    return {"state": state, "checked": True, "cleared_cell_count": cleared_cell_count}


def is_incomplete_reporting_month(year, month, as_of=None):
    """Return true when a requested monthly actual period has not closed yet."""
    current = as_of or date.today()
    return (year, month) >= (current.year, current.month)


def create_year_sheet_from_template(wb, year, template_sheet=None):
    """Create a new year sheet by copying structure from the latest existing sheet.

    Copies row labels and formulas, clears all data cells for the new year.
    """
    if template_sheet is None:
        # Use the latest year sheet as template
        year_sheets = [s for s in wb.sheetnames if s.isdigit()]
        if year_sheets:
            latest_year = max(year_sheets, key=int)
            template_sheet = wb[latest_year]
        else:
            return None

    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = str(year)

    # Update month headers in row 1
    for col in range(2, 14):
        cell = new_sheet.cell(row=1, column=col)
        if cell.value and isinstance(cell.value, datetime):
            # Set to same month but new year
            cell.value = cell.value.replace(year=year)
        elif cell.value and isinstance(cell.value, str):
            parsed = parse_month_header(cell.value)
            if parsed:
                _, parsed_month = parsed
                cell.value = datetime(year, parsed_month, 1).strftime("%b-%y")

    # Clear all data cells (keep formulas)
    for row in range(2, new_sheet.max_row + 1):
        for col in range(2, 15):  # B through N (incl YTD)
            cell = new_sheet.cell(row=row, column=col)
            label = new_sheet.cell(row=row, column=1).value
            if template_formula_row_should_clear(label):
                cell.value = None
                continue
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                # Fix cross-sheet formula references
                cell.value = cell.value.replace(f"'{template_sheet.title}'!", f"'{new_sheet.title}'!")
                cell.value = cell.value.replace(f"'{int(template_sheet.title)-1}'!", f"'{year-1}'!")
                cell.value = cell.value.replace(f"'{int(template_sheet.title)+1}'!", f"'{year+1}'!")
                continue
            if cell.value is not None:
                cell.value = None

    return new_sheet


def clear_template_workbook_data(wb, year):
    """Clear copied template values so new CF workbooks start blank, not with another property's numbers."""
    target_sheet = get_year_sheet(wb, year)
    if target_sheet is None:
        target_sheet = create_year_sheet_from_template(wb, year)
    cleared_cell_count = 0
    for sheet in wb.worksheets:
        if not str(sheet.title).isdigit():
            continue
        for row_number in range(2, sheet.max_row + 1):
            for column_number in range(2, min(sheet.max_column, 14) + 1):
                cell = sheet.cell(row=row_number, column=column_number)
                label = sheet.cell(row=row_number, column=1).value
                if template_formula_row_should_clear(label):
                    if cell.value is not None:
                        cell.value = None
                        cleared_cell_count += 1
                    continue
                if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                    continue
                if cell.value is not None:
                    cell.value = None
                    cleared_cell_count += 1
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    return cleared_cell_count


def display_property_name_from_owner_statement_dir(owner_statement_dir):
    owner_statement_dir = Path(owner_statement_dir)
    if owner_statement_dir.parent.name == LEGACY_PUBLIC_DIR_PART:
        return owner_statement_dir.parent.parent.name
    return owner_statement_dir.parent.name


def safe_cf_filename(display_name):
    display_name = re.sub(r"[\\/:*?\"<>|]+", " ", str(display_name or "")).strip()
    display_name = re.sub(r"\s+", " ", display_name)
    return f"Cash Flow Statement - {display_name}.xlsx"


def select_cf_template(cf_files):
    if TEMPLATE_PATH.exists():
        return TEMPLATE_PATH
    for path in sorted(cf_files.values(), key=lambda item: str(item).lower()):
        if Path(path).is_file():
            return Path(path)
    return None


def create_missing_cf_workbooks(cf_files, discovery_metadata, year, dry_run=True):
    """Create missing canonical CF workbooks from template structure only."""
    owner_statement_dirs = discovery_metadata.get("canonical_owner_statement_dirs") or {}
    skipped = discovery_metadata.get("skipped") or {}
    template = select_cf_template(cf_files)
    results = []
    for key in sorted(owner_statement_dirs):
        if key in cf_files or key in skipped:
            continue
        owner_statement_dir = Path(owner_statement_dirs[key])
        display_name = display_property_name_from_owner_statement_dir(owner_statement_dir)
        target_path = owner_statement_dir / safe_cf_filename(display_name)
        result = {
            "property": key,
            "owner_statement_dir": str(owner_statement_dir),
            "target_path": str(target_path),
            "template_path": str(template) if template else None,
            "dry_run": dry_run,
        }
        if target_path.exists():
            result["status"] = "exists"
            cf_files[key] = target_path
            results.append(result)
            continue
        if template is None:
            result["status"] = "blocked"
            result["reason"] = "template_missing"
            results.append(result)
            continue
        if dry_run:
            result["status"] = "would_create"
            results.append(result)
            continue
        try:
            owner_statement_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(template, target_path)
        except Exception as exc:  # noqa: BLE001
            result["status"] = "failed"
            result["reason"] = "copy_failed"
            result["error"] = f"{type(exc).__name__}: {exc}"
            results.append(result)
            continue
        try:
            wb = openpyxl.load_workbook(target_path)
            result["cleared_template_cell_count"] = clear_template_workbook_data(wb, year)
            no_mortgage_clear = clear_no_mortgage_debt_rows(wb, target_path)
            result["no_mortgage_debt_state"] = no_mortgage_clear["state"]
            result["no_mortgage_debt_checked"] = no_mortgage_clear["checked"]
            result["no_mortgage_debt_cleared_cell_count"] = no_mortgage_clear["cleared_cell_count"]
            wb.save(target_path)
            wb.close()
        except Exception as exc:  # noqa: BLE001
            result["status"] = "failed"
            result["error"] = f"{type(exc).__name__}: {exc}"
            results.append(result)
            continue
        result["status"] = "created"
        cf_files[key] = target_path
        results.append(result)
    return results


def backfill_created_workbooks(created_results, transactions, source_cash_transactions, gl_properties, year, through_month, dry_run=True):
    results = []
    for created in created_results:
        if created.get("status") not in {"created", "exists"}:
            continue
        property_key = created.get("property")
        target_path = Path(created.get("target_path") or "")
        matched_gl = match_gl_property(property_key, gl_properties)
        if not matched_gl:
            matched_gl = match_gl_property(property_name_from_cf_file(target_path), gl_properties)
        result = {
            "property": property_key,
            "target_path": str(target_path),
            "matched_gl_property": matched_gl,
            "month_count": 0,
            "change_count": 0,
            "dry_run": dry_run,
        }
        if not matched_gl:
            result["status"] = "blocked"
            result["reason"] = "no_gl_property_match"
            results.append(result)
            continue
        prop_source_cash_data = (
            filter_by_property(source_cash_transactions, matched_gl)
            if source_cash_transactions is not None
            else None
        )
        for backfill_month in range(1, through_month + 1):
            month_rows = filter_by_property(filter_by_month(transactions, year, backfill_month), matched_gl)
            changes = update_xlsx(
                target_path,
                matched_gl,
                month_rows,
                year,
                backfill_month,
                dry_run=dry_run,
                source_cash_data=prop_source_cash_data,
            )
            result["month_count"] += 1
            result["change_count"] += sum(1 for change in changes if not change.get("error"))
            errors = [change for change in changes if change.get("error")]
            if errors:
                result.setdefault("errors", []).extend(errors[:5])
        result["status"] = "ok" if not result.get("errors") else "review"
        results.append(result)
    return results


def audit_xlsx(
    xlsx_path,
    gl_property,
    gl_month_data,
    year,
    month,
    source_cash_data=None,
    source_cash_mode="full_column_e",
):
    """
    Audit a CF xlsx against GL data.
    Returns a list of conflicts and a summary.
    """
    conflicts = []
    summary = {
        "file": str(xlsx_path), "property": gl_property, "rows": 0,
        "matches": 0, "conflicts": 0, "missing_in_gl": 0, "formulas": 0,
        "formula_matches": 0, "formula_eval_errors": 0,
        "balance_sheet_skipped": 0, "untagged_gl_rows": 0,
        "no_mortgage_debt_state": property_state_from_path(xlsx_path),
        "no_mortgage_debt_checked": False,
        "no_mortgage_debt_checked_cell_count": 0,
        "no_mortgage_debt_violation_count": 0,
        "no_mortgage_debt_violations_bounded": [],
        "source_cash_balance_policy": (
            "Current CF columns use the full property-split ECO GL Column E balance; "
            "closed historical columns use the same balance through month-end."
        ),
        "source_cash_balance_checked": False,
        "source_cash_balance_expected": None,
        "source_cash_balance_actual": None,
        "source_cash_balance_month_filter_applied": source_cash_mode == "as_of_month_end",
        "source_cash_balance_semantics": (
            "historical_eco_gl_column_e_balance_through_month_end"
            if source_cash_mode == "as_of_month_end"
            else "full_eco_gl_column_e_balance_not_monthly_cashflow"
        ),
        "source_cash_balance_mode": source_cash_mode,
        "source_cash_balance_violation_count": 0,
        "source_cash_balance_violations_bounded": [],
        "mortgage_statement_split_checked": False,
        "mortgage_statement_split_status": "not_applicable",
        "mortgage_statement_split_raw_payment_count": 0,
        "mortgage_statement_split_statement_path": None,
    }

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return {"summary": summary, "error": f"Cannot open: {e}", "conflicts": []}

    no_mortgage_debt_audit = audit_no_mortgage_debt_rows(wb, xlsx_path)
    summary.update(
        {
            "no_mortgage_debt_state": no_mortgage_debt_audit["state"],
            "no_mortgage_debt_checked": no_mortgage_debt_audit["checked"],
            "no_mortgage_debt_checked_cell_count": no_mortgage_debt_audit["checked_cell_count"],
            "no_mortgage_debt_violation_count": no_mortgage_debt_audit["violation_count"],
            "no_mortgage_debt_violations_bounded": no_mortgage_debt_audit["violations_bounded"],
        }
    )

    sheet = get_year_sheet(wb, year)
    if not sheet:
        wb.close()
        return {"summary": summary, "error": f"No sheet for year {year}", "conflicts": []}

    col = get_month_column(sheet, year, month)
    if not col:
        wb.close()
        return {"summary": summary, "error": f"No column for {year}-{month:02d}", "conflicts": []}

    if source_cash_data is not None:
        source_cash_audit = audit_source_cash_balance_row(
            sheet, source_cash_data, year, month, source_cash_mode=source_cash_mode
        )
        summary.update(
            {
                "source_cash_balance_checked": source_cash_audit["checked"],
                "source_cash_balance_expected": source_cash_audit["expected"],
                "source_cash_balance_actual": source_cash_audit.get("actual"),
                "source_cash_balance_month_filter_applied": source_cash_audit.get("month_filter_applied", False),
                "source_cash_balance_semantics": source_cash_audit.get(
                    "balance_semantics", "full_eco_gl_column_e_balance_not_monthly_cashflow"
                ),
                "source_cash_balance_included_transaction_count": source_cash_audit["included_count"],
                "source_cash_balance_excluded_earldao_interest_count": source_cash_audit["excluded_earldao_interest_count"],
                "source_cash_balance_excluded_earldao_interest_total": source_cash_audit["excluded_earldao_interest_total"],
                "source_cash_balance_violation_count": source_cash_audit["violation_count"],
                "source_cash_balance_violations_bounded": source_cash_audit["violations_bounded"],
            }
        )
        for violation in source_cash_audit["violations_bounded"]:
            summary["conflicts"] += 1
            conflicts.append(
                {
                    "row": violation.get("row"),
                    "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                    "cf_value": violation.get("actual"),
                    "gl_total": source_cash_audit["expected"],
                    "diff": violation.get("diff"),
                    "type": "source_cash_balance_mismatch",
                    "action": "replace_with_source_gl_cumulative_balance",
                    "value_type": violation.get("value_type"),
                }
            )

    # Sum GL data by category for this property
    gl_by_category = defaultdict(float)
    untagged_count = 0
    for t in gl_month_data:
        cat = t["_cf_category"]
        gl_by_category[cat] += t["_amount"]
        if t.get("_is_fallback"):
            untagged_count += 1
    summary["untagged_gl_rows"] = untagged_count
    mortgage_statement_split = citadel_mortgage_statement_split(gl_month_data, xlsx_path, year, month)
    if no_mortgage_debt_policy_applies(xlsx_path):
        mortgage_statement_split["status"] = "no_dao_mortgage_policy"
    else:
        overlay_citadel_statement_split(gl_by_category, mortgage_statement_split, sheet, col)
    summary.update(
        {
            "mortgage_statement_split_checked": mortgage_statement_split["checked"],
            "mortgage_statement_split_status": mortgage_statement_split["status"],
            "mortgage_statement_split_raw_payment_count": mortgage_statement_split["raw_citadel_payment_count"],
            "mortgage_statement_split_statement_path": mortgage_statement_split.get("statement_path"),
            "mortgage_statement_split_principal_payment": mortgage_statement_split.get("principal_payment"),
            "mortgage_statement_split_interest_payment": mortgage_statement_split.get("interest_payment"),
            "mortgage_statement_split_principal_balance": mortgage_statement_split.get("principal_balance"),
            "mortgage_statement_split_escrow_component_insurance": mortgage_statement_split.get("escrow_component_insurance"),
            "mortgage_statement_split_escrow_component_taxes": mortgage_statement_split.get("escrow_component_taxes"),
            "mortgage_statement_split_escrow_component_source_column": mortgage_statement_split.get("escrow_component_source_column"),
        }
    )

    for row_num in range(2, sheet.max_row + 1):
        label_cell = sheet.cell(row=row_num, column=1)
        label = str(label_cell.value or "").strip()
        if not label:
            continue

        if label == MORTGAGE_PRINCIPAL_BALANCE_LABEL and mortgage_statement_split.get("status") == "applied":
            value_cell = sheet.cell(row=row_num, column=col)
            cf_numeric = numeric_cell_value(value_cell.value)
            expected_balance = round(float(mortgage_statement_split["principal_balance"]), 2)
            if cf_numeric is None or abs(round(cf_numeric, 2) - expected_balance) > CONFLICT_THRESHOLD:
                summary["conflicts"] += 1
                conflicts.append(
                    {
                        "row": row_num,
                        "label": label,
                        "cf_value": value_cell.value,
                        "gl_total": expected_balance,
                        "diff": None if cf_numeric is None else round(abs(round(cf_numeric, 2) - expected_balance), 2),
                        "type": "mortgage_statement_balance_mismatch",
                        "action": "set_mortgage_principal_balance_from_statement",
                        "statement_path": mortgage_statement_split.get("statement_path"),
                    }
                )
            else:
                summary["matches"] += 1
            continue

        # Skip balance sheet rows
        if label in BALANCE_SHEET_ROWS:
            summary["balance_sheet_skipped"] += 1
            continue

        # Skip section headers and totals
        if label in SKIP_ROWS or label in FORMULA_ROWS:
            continue
        if label == "Retained Earnings" and retained_earnings_reconciliation_exempt(gl_property):
            continue
        if label.startswith("Total") or label.startswith("Net") or label.startswith("Percentage"):
            continue

        value_cell = sheet.cell(row=row_num, column=col)
        cf_value = value_cell.value

        is_formula = isinstance(cf_value, str) and cf_value.startswith("=")

        # Find matching GL categories
        gl_categories = ROW_TO_GL_CATEGORIES.get(label, None)
        if gl_categories is None:
            for key in ROW_TO_GL_CATEGORIES:
                norm_key = key.lower().replace("&", "").replace("  ", " ")
                norm_label = label.lower().replace("&", "").replace("  ", " ")
                if norm_key in norm_label or norm_label in norm_key:
                    gl_categories = ROW_TO_GL_CATEGORIES[key]
                    break

        if gl_categories is None:
            continue

        gl_total = round(sum(gl_by_category.get(cat, 0) for cat in gl_categories), 2)
        summary["rows"] += 1

        if is_formula:
            summary["formulas"] += 1
            formula_value = None
            try:
                formula_value = round(evaluate_formula(cf_value, sheet), 2)
            except Exception:
                summary["formula_eval_errors"] += 1
            if formula_value is not None and abs(formula_value - gl_total) <= CONFLICT_THRESHOLD:
                summary["formula_matches"] += 1
                summary["matches"] += 1
                continue
            if formula_value is None and abs(gl_total) <= CONFLICT_THRESHOLD:
                continue
            summary["conflicts"] += 1
            diff = abs((formula_value or 0.0) - gl_total)
            conflicts.append({
                "row": row_num, "label": label,
                "cf_value": cf_value, "gl_total": gl_total,
                "diff": round(diff, 2),
                "type": "formula_vs_gl",
                "action": "overwrite_formula_from_gl",
                "formula_value": formula_value,
            })
            continue

        # Parse numeric value
        if cf_value is None:
            cf_numeric = 0.0
        elif isinstance(cf_value, (int, float)):
            cf_numeric = float(cf_value)
        elif isinstance(cf_value, str):
            try:
                cf_numeric = float(cf_value.replace(",", "").replace("$", "").replace("(", "-").replace(")", ""))
            except ValueError:
                continue
        else:
            continue

        if (
            "Management Fees" in gl_categories
            and abs(gl_total) <= CONFLICT_THRESHOLD
            and has_hemlane_direct_split_pm_void(gl_month_data, cf_numeric)
        ):
            summary["matches"] += 1
            summary["hemlane_direct_split_pm_void_match_count"] = summary.get(
                "hemlane_direct_split_pm_void_match_count", 0
            ) + 1
            continue

        diff = abs(cf_numeric - gl_total)
        if diff > CONFLICT_THRESHOLD:
            summary["conflicts"] += 1
            action = "overwrite"
            if abs(gl_total) < CONFLICT_THRESHOLD and abs(cf_numeric) > CONFLICT_THRESHOLD:
                action = "cf_has_value_gl_empty"  # Accrual/manual entry not in Baselane
            elif abs(cf_numeric) < CONFLICT_THRESHOLD and abs(gl_total) > CONFLICT_THRESHOLD:
                action = "fill_from_gl"
            conflicts.append({
                "row": row_num, "label": label,
                "cf_value": round(cf_numeric, 2), "gl_total": round(gl_total, 2),
                "diff": round(diff, 2), "type": "value_mismatch",
                "action": action
            })
        else:
            summary["matches"] += 1

    wb.close()
    return {"summary": summary, "conflicts": conflicts}


def update_xlsx(
    xlsx_path,
    gl_property,
    gl_month_data,
    year,
    month,
    dry_run=True,
    source_cash_data=None,
    source_cash_only=False,
    only_rows=None,
    source_cash_mode="full_column_e",
):
    """Update a CF xlsx from GL data. Overwrites GL-sourced cells; preserves only formula summary rows."""
    changes = []
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return [{"error": f"Cannot open {xlsx_path}: {e}"}]

    no_mortgage_policy = no_mortgage_debt_policy_applies(xlsx_path)
    selected_rows = {str(label).strip().casefold() for label in (only_rows or []) if str(label).strip()}
    if no_mortgage_policy and not selected_rows:
        no_mortgage_clear = clear_no_mortgage_debt_rows(wb, xlsx_path)
        if no_mortgage_clear.get("cleared_cell_count"):
            changes.append({
                "action": "clear_no_dao_mortgage_debt_rows",
                "cleared_cell_count": no_mortgage_clear["cleared_cell_count"],
                "state": no_mortgage_clear.get("state"),
            })

    sheet = get_year_sheet(wb, year)
    if not sheet:
        # Try to create the year sheet
        template = None
        for sn in wb.sheetnames:
            if sn.isdigit():
                template = wb[sn]
                break
        if template:
            sheet = create_year_sheet_from_template(wb, year, template)
            if sheet:
                changes.append({"action": "created_year_sheet", "year": year})
        if not sheet:
            wb.close()
            return [{"error": f"No sheet for year {year} in {xlsx_path}"}]

    col = get_month_column(sheet, year, month)
    if not col:
        wb.close()
        return [{"error": f"No column for {year}-{month:02d} in {xlsx_path}"}]

    principal_row_change = ensure_madison_90_principal_payments_row(
        sheet, xlsx_path, dry_run=dry_run
    )
    if principal_row_change:
        changes.append(principal_row_change)

    source_cash_update = None
    if source_cash_data is not None and (
        not selected_rows or ECO_GL_NET_CASH_BALANCE_LABEL.casefold() in selected_rows
    ):
        source_cash_update = source_cash_balance_for_month(
            source_cash_data, year, month, source_cash_mode
        )
        source_cash_row, source_cash_row_created, source_cash_row_renamed = ensure_source_cash_balance_row(sheet)
        if source_cash_row_created:
            changes.append({
                "row": source_cash_row,
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "action": "create_source_cash_balance_row",
            })
        elif source_cash_row_renamed:
            changes.append({
                "row": source_cash_row,
                "label": ECO_GL_NET_CASH_BALANCE_LABEL,
                "action": "rename_source_cash_balance_row",
            })

    gl_by_category = defaultdict(float)
    for t in gl_month_data:
        cat = t["_cf_category"]
        gl_by_category[cat] += t["_amount"]
    mortgage_statement_split = citadel_mortgage_statement_split(gl_month_data, xlsx_path, year, month)
    if no_mortgage_policy:
        mortgage_statement_split["status"] = "no_dao_mortgage_policy"
    else:
        overlay_citadel_statement_split(gl_by_category, mortgage_statement_split, sheet, col)

    for row_num in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row=row_num, column=1).value or "").strip()
        if not label:
            continue
        if selected_rows and label.casefold() not in selected_rows:
            continue
        source_cash_labels = {
            ECO_GL_NET_CASH_BALANCE_LABEL,
            *ECO_GL_NET_CASH_BALANCE_LEGACY_LABELS,
        }
        if source_cash_only:
            source_cash_labels.add("ECO General Ledger (ECO GL Column E Total)")
        if label in source_cash_labels and source_cash_update is not None:
            value_cell = sheet.cell(row=row_num, column=col)
            old_value = value_cell.value
            new_value = source_cash_update["expected"]
            old_numeric = numeric_cell_value(old_value)
            old_is_formula = isinstance(old_value, str) and old_value.strip().startswith("=")
            if not old_is_formula and old_value is not None and old_numeric is not None and abs(round(old_numeric, 2) - new_value) <= CONFLICT_THRESHOLD:
                continue
            if not dry_run:
                value_cell.value = new_value
            changes.append({
                "row": row_num,
                "label": label,
                "action": "set_source_cash_balance",
                "old_value": old_value,
                "new_value": new_value,
                "source_included_transaction_count": source_cash_update["included_count"],
                "source_excluded_earldao_interest_count": source_cash_update["excluded_earldao_interest_count"],
                "source_excluded_earldao_interest_total": source_cash_update["excluded_earldao_interest_total"],
                "source_cash_balance_mode": source_cash_update.get(
                    "source_cash_balance_mode", source_cash_mode
                ),
                "source_cash_balance_as_of_date": source_cash_update.get("as_of_date"),
            })
            continue
        if source_cash_only:
            continue
        if (
            no_mortgage_policy
            and label in NO_MORTGAGE_DEBT_ROW_LABELS
            and not no_mortgage_debt_row_is_exempt(xlsx_path, label)
        ):
            value_cell = sheet.cell(row=row_num, column=col)
            old_value = value_cell.value
            old_numeric = numeric_cell_value(old_value)
            old_is_formula = isinstance(old_value, str) and old_value.strip().startswith("=")
            if old_is_formula or old_numeric is None or abs(round(old_numeric, 2)) > CONFLICT_THRESHOLD:
                if not dry_run:
                    value_cell.value = 0
                changes.append({
                    "row": row_num,
                    "label": label,
                    "action": "set_no_dao_mortgage_debt_zero",
                    "old_value": old_value,
                    "new_value": 0,
                })
            continue
        if label == MORTGAGE_PRINCIPAL_BALANCE_LABEL and mortgage_statement_split.get("status") == "applied":
            value_cell = sheet.cell(row=row_num, column=col)
            old_value = value_cell.value
            new_value = round(float(mortgage_statement_split["principal_balance"]), 2)
            old_numeric = numeric_cell_value(old_value)
            if old_numeric is not None and abs(round(old_numeric, 2) - new_value) <= CONFLICT_THRESHOLD:
                continue
            if not dry_run:
                value_cell.value = new_value
            changes.append({
                "row": row_num,
                "label": label,
                "action": "set_mortgage_principal_balance_from_statement",
                "old_value": old_value,
                "new_value": new_value,
                "statement_path": mortgage_statement_split.get("statement_path"),
            })
            continue
        if label in BALANCE_SHEET_ROWS:
            continue
        if label in SKIP_ROWS:
            continue
        if label == "Retained Earnings" and retained_earnings_reconciliation_exempt(gl_property):
            continue
        if formula_row_label(label):
            restored = restore_blank_formula_from_neighbor(sheet, row_num, col, dry_run=dry_run)
            if restored:
                changes.append(restored)
            continue

        value_cell = sheet.cell(row=row_num, column=col)
        cf_value = value_cell.value
        is_formula = isinstance(cf_value, str) and str(cf_value).startswith("=")

        gl_categories = ROW_TO_GL_CATEGORIES.get(label, None)
        if gl_categories is None:
            for key in ROW_TO_GL_CATEGORIES:
                norm_key = key.lower().replace("&", "").replace("  ", " ")
                norm_label = label.lower().replace("&", "").replace("  ", " ")
                if norm_key in norm_label or norm_label in norm_key:
                    gl_categories = ROW_TO_GL_CATEGORIES[key]
                    break
        if gl_categories is None:
            continue

        gl_total = round(sum(gl_by_category.get(cat, 0) for cat in gl_categories), 2)

        if is_formula:
            formula_value = None
            try:
                formula_value = round(evaluate_formula(cf_value, sheet), 2)
            except Exception:
                pass
            if not dry_run:
                value_cell.value = gl_total
            changes.append({
                "row": row_num, "label": label, "action": "overwrite_formula",
                "old_value": cf_value, "new_value": gl_total, "formula_value": formula_value
            })
            continue

        if abs(gl_total) < CONFLICT_THRESHOLD and (cf_value is None or (isinstance(cf_value, (int, float)) and abs(float(cf_value)) < CONFLICT_THRESHOLD)):
            continue

        cf_numeric = numeric_cell_value(cf_value)
        if (
            "Management Fees" in gl_categories
            and cf_numeric is not None
            and abs(gl_total) <= CONFLICT_THRESHOLD
            and has_hemlane_direct_split_pm_void(gl_month_data, cf_numeric)
        ):
            changes.append({
                "row": row_num,
                "label": label,
                "action": "preserve_hemlane_direct_split_pm_void",
                "old_value": cf_value,
                "new_value": cf_value,
            })
            continue
        if cf_numeric is not None and abs(round(cf_numeric, 2) - gl_total) <= CONFLICT_THRESHOLD:
            continue

        old_value = cf_value
        if not dry_run:
            value_cell.value = gl_total
        changes.append({
            "row": row_num, "label": label, "action": "overwrite",
            "old_value": old_value, "new_value": gl_total
        })

    if not dry_run and any(c.get("action") in {"overwrite", "overwrite_formula", "restore_formula", "set_source_cash_balance", "create_source_cash_balance_row", "rename_source_cash_balance_row", "rename_legacy_90_madison_principal_row", "set_mortgage_principal_balance_from_statement", "set_no_dao_mortgage_debt_zero", "clear_no_dao_mortgage_debt_rows"} for c in changes):
        wb.save(xlsx_path)

    wb.close()
    return changes


def format_discord_report(audit_results, year, month, dry_run=True):
    """Format audit results as a Discord-friendly message."""
    lines = [f"📊 **CF Statement Audit: {year}-{month:02d}** {'(DRY RUN)' if dry_run else '(LIVE)'}\n"]

    def money(value):
        if value is None:
            return "n/a"
        try:
            return f"${float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)

    total_conflicts = 0
    total_matches = 0
    total_formula_flags = 0
    total_balance_skipped = 0
    total_untagged = 0
    conflict_properties = []
    accrual_flags = []

    for result in audit_results:
        if "error" in result:
            summary = result.get("summary") or {}
            prop = summary.get("property") or result.get("property") or "Unknown"
            lines.append(f"❌ **{prop}**: {result['error']}")
            continue

        summary = result.get("summary", {})
        conflicts = result.get("conflicts", [])
        prop = summary.get("property", "Unknown")

        total_conflicts += summary.get("conflicts", 0)
        total_matches += summary.get("matches", 0)
        total_formula_flags += summary.get("formulas", 0)
        total_balance_skipped += summary.get("balance_sheet_skipped", 0)
        total_untagged += summary.get("untagged_gl_rows", 0)

        if conflicts:
            conflict_properties.append(prop)
            lines.append(f"\n⚠️ **{prop}** — {len(conflicts)} conflicts:")
            for c in conflicts[:5]:
                action = c.get("action")
                action_emoji = "🔄" if action == "overwrite" else "📝" if action == "review_accrual_in_baselane" else "➕"
                if c.get("type") == "formula_vs_gl":
                    lines.append(f"  • {c.get('label')}: CF formula has value, GL={money(c.get('gl_total'))} → needs accrual in Baselane")
                else:
                    lines.append(
                        f"  • {c.get('label')}: CF={money(c.get('cf_value'))} "
                        f"vs GL={money(c.get('gl_total'))} (Δ={money(c.get('diff'))}) {action_emoji}"
                    )
            if len(conflicts) > 5:
                lines.append(f"  … and {len(conflicts) - 5} more")

            # Collect accrual flags (CF has value, GL empty)
            for c in conflicts:
                if c.get("action") == "cf_has_value_gl_empty":
                    accrual_flags.append((prop, c["label"], c["cf_value"]))

    lines.append(f"\n**Summary**: {len(audit_results)} properties audited")
    lines.append(f"✅ {total_matches} matching rows")
    lines.append(f"⚠️ {total_conflicts} conflicts")
    lines.append(f"📝 {total_formula_flags} formula cells (preserved)")
    lines.append(f"🔢 {total_balance_skipped} balance sheet rows (skipped, not GL-sourced)")
    lines.append(f"❓ {total_untagged} GL transactions used keyword fallback (need Baselane tagging)")

    if accrual_flags:
        lines.append(f"\n🔍 **Accruals missing from Baselane** (CF has values, GL is empty):")
        for prop, label, value in accrual_flags[:20]:
            lines.append(f"  • {prop}: {label} = ${value:.2f}")

    if conflict_properties:
        lines.append(f"\n🔍 Properties with conflicts: {', '.join(conflict_properties)}")

    return "\n".join(lines)


def main():
    global GL_PATH, REAL_ESTATE_BASE, OUTPUT_DIR, CONFLICT_THRESHOLD

    parser = argparse.ArgumentParser(description="Update CF statements from Baselane GL")
    parser.add_argument("--month", help="Month in YYYY-MM format (default: previous month)")
    parser.add_argument(
        "--cutoff-date",
        type=date.fromisoformat,
        default=None,
        help="Include source transactions dated on or before YYYY-MM-DD.",
    )
    parser.add_argument("--property", help="Process specific property (fuzzy match)")
    parser.add_argument(
        "--only-row",
        action="append",
        default=[],
        help="Write only the exact CF row label; repeat for multiple rows",
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    parser.add_argument("--audit", action="store_true", help="Audit only, don't write")
    parser.add_argument("--create-missing", action="store_true", help="Create CF xlsx for properties without one")
    parser.add_argument("--include-excluded-properties", action="store_true", help="Include sold/closed/manual excluded properties in discovery")
    parser.add_argument("--report", action="store_true", help="Post conflict report to Discord")
    parser.add_argument("--gl-csv", type=Path, default=None, help="Override Baselane GL CSV path")
    parser.add_argument("--source-cash-gl-csv", type=Path, default=None, help="Raw Baselane GL CSV path for cumulative ECO cash balance")
    parser.add_argument(
        "--policy-root",
        type=Path,
        default=None,
        help="Workspace root used for exclusion and listing policy lookup",
    )
    parser.add_argument(
        "--report-output-dir",
        type=Path,
        default=None,
        help="Directory for audit/discovery reports; does not change policy lookup scope",
    )
    parser.add_argument(
        "--skip-source-cash-balance-row",
        action="store_true",
        help=(
            "Do not create, update, or audit the ECO GL Net Cash Balance balance-sheet row. "
            "Use when that row is owned by the guarded balance-sheet cash updater."
        ),
    )
    parser.add_argument("--output-dir", type=Path, default=None, help="Override Real Estate root for CF file discovery")
    parser.add_argument("--conflict-threshold", type=float, default=None, help="Override conflict threshold")
    parser.add_argument("--alert-file", type=Path, default=None, help="Append a text conflict summary to this file")
    parser.add_argument(
        "--allow-incomplete-month",
        action="store_true",
        help="Allow writing current/future monthly actual cells. Default blocks these periods to prevent template-looking actuals.",
    )
    args = parser.parse_args()

    if args.gl_csv:
        GL_PATH = args.gl_csv
    if args.output_dir:
        REAL_ESTATE_BASE = args.output_dir
    if args.report_output_dir:
        OUTPUT_DIR = args.report_output_dir
    elif args.output_dir:
        OUTPUT_DIR = Path(os.environ.get("ROOT") or WORKSPACE_ROOT) / "reports" / "cf_statement_sync"
    if args.conflict_threshold is not None:
        CONFLICT_THRESHOLD = args.conflict_threshold

    if args.month:
        year, month = map(int, args.month.split("-"))
    else:
        today = datetime.now()
        first_of_month = today.replace(day=1)
        last_month = first_of_month - timedelta(days=1)
        year, month = last_month.year, last_month.month

    if not (args.audit or args.dry_run or args.allow_incomplete_month) and is_incomplete_reporting_month(year, month):
        print(
            f"ERROR: refusing to write incomplete monthly actuals for {year}-{month:02d}. "
            "Run the prior closed month, use --audit/--dry-run, or pass --allow-incomplete-month intentionally.",
            file=sys.stderr,
        )
        return 2

    print(f"Loading GL data for {year}-{month:02d}...")
    transactions = load_gl_data(GL_PATH)
    transactions = filter_through_date(
        transactions, args.cutoff_date, accounting_month=(year, month)
    )
    source_cash_gl_path = None
    source_cash_transactions = None
    if args.skip_source_cash_balance_row:
        print("Skipping ECO GL Net Cash Balance row updates/audit; owned by guarded balance-sheet cash updater.")
    else:
        source_cash_gl_path = args.source_cash_gl_csv or GL_PATH
        if source_cash_gl_path == GL_PATH:
            source_cash_transactions = transactions
        else:
            print(f"Loading raw source-cash GL data from {source_cash_gl_path}...")
            source_cash_transactions = load_gl_data(source_cash_gl_path)
            source_cash_transactions = filter_through_date(
                source_cash_transactions, args.cutoff_date, accounting_month=(year, month)
            )
    month_data = filter_by_month(transactions, year, month)
    print(f"  {len(month_data)} transactions for {year}-{month:02d}")

    print("Discovering CF xlsx files...")
    excluded_properties = (
        set()
        if args.include_excluded_properties
        else load_excluded_property_names(args.policy_root)
    )
    property_scope_exclusion_overrides = []
    if args.property and excluded_properties:
        retained_exclusions = set()
        for excluded_property in excluded_properties:
            if property_scope_matches_exclusion(args.property, excluded_property):
                property_scope_exclusion_overrides.append(excluded_property)
            else:
                retained_exclusions.add(excluded_property)
        excluded_properties = retained_exclusions
    if excluded_properties:
        print(f"  Loaded {len(excluded_properties)} excluded property names")
    if property_scope_exclusion_overrides:
        print(f"  Included {len(property_scope_exclusion_overrides)} excluded property names matching --property")
    cf_files, discovery_metadata = discover_cf_files(
        include_metadata=True,
        excluded_properties=excluded_properties,
        property_scope=args.property,
    )
    discovery_metadata["property_scope_exclusion_overrides"] = sorted(property_scope_exclusion_overrides)
    print(f"  Found {len(cf_files)} properties with CF files")
    if discovery_metadata.get("excluded_property_count"):
        print(f"  Skipped {discovery_metadata.get('excluded_property_count')} excluded property folders")
    skipped_count = sum(len(paths) for paths in (discovery_metadata.get("skipped") or {}).values())
    if skipped_count:
        print(f"  Ignored {skipped_count} noncanonical CF candidates")

    gl_properties = set()
    for t in transactions:
        if t["_property"]:
            gl_properties.add(t["_property"])
    month_gl_properties = {t["_property"] for t in month_data if t["_property"]}
    print(f"  {len(month_gl_properties)} properties in GL data for {year}-{month:02d}")
    print(f"  {len(gl_properties)} properties in full GL data for matching")

    create_missing_results = []
    backfill_created_results = []
    if args.create_missing:
        create_missing_results = create_missing_cf_workbooks(
            cf_files,
            discovery_metadata,
            year,
            dry_run=(args.audit or args.dry_run),
        )
        created_count = sum(1 for result in create_missing_results if result.get("status") == "created")
        would_create_count = sum(1 for result in create_missing_results if result.get("status") == "would_create")
        if args.audit or args.dry_run:
            print(f"  Would create {would_create_count} missing CF workbooks")
        else:
            print(f"  Created {created_count} missing CF workbooks")
            backfill_created_results = backfill_created_workbooks(
                create_missing_results,
                transactions,
                source_cash_transactions,
                gl_properties,
                year,
                month,
                dry_run=False,
            )
            print(f"  Backfilled {sum(1 for result in backfill_created_results if result.get('status') == 'ok')} created CF workbooks")

    audit_results = build_discovery_review_rows(cf_files, discovery_metadata, gl_properties)

    for cf_name, cf_path in sorted(cf_files.items()):
        # Skip removed properties
        if cf_name in SKIP_PROPERTIES:
            print(f"  Skipping {cf_name} (no longer managed)")
            continue

        # Find matching GL property without substring-only address-number matches.
        matched_gl = match_gl_property(cf_name, gl_properties)
        if not matched_gl:
            file_label = property_name_from_cf_file(cf_path)
            matched_gl = match_gl_property(file_label, gl_properties)

        if args.property and matched_gl:
            if args.property.lower() not in matched_gl.lower():
                continue
        elif args.property and not matched_gl:
            if args.property.lower() not in str(cf_path).lower():
                continue

        prop_data = filter_by_property(month_data, matched_gl) if matched_gl else []
        prop_source_cash_data = (
            filter_by_property(source_cash_transactions, matched_gl)
            if source_cash_transactions is not None and matched_gl
            else None
        )
        if matched_gl:
            update_changes = []
            if not args.audit and not args.dry_run:
                update_changes = update_xlsx(
                    cf_path,
                    matched_gl,
                    prop_data,
                    year,
                    month,
                    dry_run=False,
                    source_cash_data=prop_source_cash_data,
                    only_rows=args.only_row,
                )
            result = audit_xlsx(cf_path, matched_gl, prop_data, year, month, source_cash_data=prop_source_cash_data)
            result["cf_statement_update_changes"] = [
                change for change in update_changes if change.get("action") in {"overwrite", "overwrite_formula"}
            ]
            result["source_cash_balance_update_changes"] = [
                change for change in update_changes if change.get("action") == "set_source_cash_balance"
            ]
            result["summary"]["cf_statement_update_count"] = len(result["cf_statement_update_changes"])
            result["summary"]["cf_statement_zero_fill_count"] = sum(
                1
                for change in result["cf_statement_update_changes"]
                if abs(numeric_cell_value(change.get("new_value")) or 0.0) <= CONFLICT_THRESHOLD
            )
            result["summary"]["source_cash_balance_update_count"] = len(result["source_cash_balance_update_changes"])
            result["summary"]["source_cash_balance_source_gl_csv"] = str(source_cash_gl_path) if source_cash_gl_path else None
        else:
            synthetic_property = property_name_from_cf_file(cf_path) or cf_name
            update_changes = []
            if not args.audit and not args.dry_run:
                update_changes = update_xlsx(
                    cf_path,
                    synthetic_property,
                    [],
                    year,
                    month,
                    dry_run=False,
                    source_cash_data=[],
                    only_rows=args.only_row,
                )
            result = audit_xlsx(cf_path, synthetic_property, [], year, month, source_cash_data=[])
            result["no_gl_property_match"] = True
            result["warning"] = "No matching property found in Baselane GL; workbook reconciled against zero raw GL rows for this month"
            result["cf_statement_update_changes"] = [
                change for change in update_changes if change.get("action") in {"overwrite", "overwrite_formula"}
            ]
            result["source_cash_balance_update_changes"] = [
                change for change in update_changes if change.get("action") == "set_source_cash_balance"
            ]
            result["summary"]["matched_gl_property"] = None
            result["summary"]["no_gl_property_match"] = True
            result["summary"]["cf_statement_update_count"] = len(result["cf_statement_update_changes"])
            result["summary"]["cf_statement_zero_fill_count"] = sum(
                1
                for change in result["cf_statement_update_changes"]
                if abs(numeric_cell_value(change.get("new_value")) or 0.0) <= CONFLICT_THRESHOLD
            )
            result["summary"]["source_cash_balance_update_count"] = len(result["source_cash_balance_update_changes"])
            result["summary"]["source_cash_balance_source_gl_csv"] = str(source_cash_gl_path) if source_cash_gl_path else None
        audit_results.append(result)

        if matched_gl:
            print(f"  {cf_name}: {len(prop_data)} GL transactions, {result.get('summary', {}).get('conflicts', '?')} conflicts")
        else:
            print(f"  {cf_name}: No GL match found")

    report = format_discord_report(audit_results, year, month, dry_run=(args.dry_run or args.audit))
    print(report)
    if args.alert_file:
        args.alert_file.parent.mkdir(parents=True, exist_ok=True)
        with args.alert_file.open("a", encoding="utf-8") as handle:
            handle.write(f"\n[{datetime.now().isoformat(timespec='seconds')}] CF statement sync audit {year}-{month:02d}\n")
            handle.write(report)
            handle.write("\n")

    if not args.audit:
        for result in audit_results:
            if "error" in result:
                continue
            summary = result.get("summary", {})
            if summary.get("conflicts", 0) > 0:
                # TODO: implement live update pass
                pass

    report_dir = OUTPUT_DIR
    report_dir.mkdir(parents=True, exist_ok=True)
    report_file = report_dir / f"audit_{year}-{month:02d}.json"
    with open(report_file, "w") as f:
        json.dump(audit_results, f, indent=2, default=str)
    discovery_file = report_dir / f"discovery_{year}-{month:02d}.json"
    with open(discovery_file, "w") as f:
        json.dump(discovery_report(discovery_metadata, cf_files, year, month, gl_properties), f, indent=2, default=str)
    if args.create_missing:
        create_missing_file = report_dir / f"create_missing_{year}-{month:02d}.json"
        with open(create_missing_file, "w") as f:
            json.dump(
                {
                    "month": f"{year}-{month:02d}",
                    "dry_run": bool(args.audit or args.dry_run),
                    "create_missing_results": create_missing_results,
                    "backfill_created_results": backfill_created_results,
                    "created_count": sum(1 for result in create_missing_results if result.get("status") == "created"),
                    "would_create_count": sum(1 for result in create_missing_results if result.get("status") == "would_create"),
                    "failed_count": sum(1 for result in create_missing_results if result.get("status") in {"failed", "blocked"}),
                },
                f,
                indent=2,
                default=str,
            )
    print(f"\nReport saved to {report_file}")
    print(f"Discovery saved to {discovery_file}")
    if args.create_missing:
        print(f"Create-missing report saved to {create_missing_file}")


if __name__ == "__main__":
    sys.exit(main() or 0)
